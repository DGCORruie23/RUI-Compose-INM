from bokeh.plotting import figure
from bokeh.models import GeoJSONDataSource, HoverTool, TapTool, CustomJS, LinearColorMapper, FactorRange, ColumnDataSource, NumeralTickFormatter, RangeTool, DatetimeTickFormatter, LabelSet
from bokeh.layouts import column
from bokeh.embed import components
from bokeh.palettes import Greens256
from bokeh.transform import cumsum
from math import pi, cos, sin

from .models import (Estado, Nacionalidad, Repatriados, Recibidos,
                    ExtRescatados, Ingresos, Tramites, Retornados, Inadmitidos,
                    PuntosInternacionEstacion, CatalogoOR, Encuentros, TipoPRH,
                    RegistroMexExtPunto, CATALOGO_PUNTOS_MEX_EXT,
                    PRHs, Titular, Estudio, GradoAcademico, TelefonoTitular, CorreoTitular, 
                    TipoNombramiento, TrayectoriaLaboral, ExperienciaProfesional, TipoProcendencia,
                    Comodato, FiguraOcupacion, TipoInmueble, SituacionActual, TipoActividad, Inmueble, HistoricoComentarios, TipoOficina,
                    ProgramaIPC, PersonalINM, OrganigramaF, EstatusPersonal, TipoPlaza, TipoDependencia)
from usuarioL.models import usuarioL
from usuario.models import RescatePunto

from datetime import datetime
import json
import random
import os
import unicodedata
import base64
import openpyxl
import requests
import urllib3
from collections import Counter
from urllib.parse import quote
from concurrent.futures import ThreadPoolExecutor

from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from weasyprint import HTML

from django.contrib import messages
from django.contrib.staticfiles import finders
from django.shortcuts import render, redirect
from django.urls import reverse
from django.apps import apps
from django.core.files.base import ContentFile
from django.conf import settings
from django.core.cache import cache
from django.db import transaction, models, connection
from django.db.models import Sum, Count, Max, Q
from django.db.models.functions import TruncDay, TruncMonth, TruncWeek, Upper
from datetime import date, timedelta
from django.http import Http404, HttpResponse, JsonResponse
from django.template.loader import get_template
from django.views.decorators.http import require_GET

from .models import (
    Capufe,
    CombustibleExt,
    Kilometraje,
    PrestadoDe,
    Siniestros,
    SituacionVeh,
    TipoAsignacionVeh,
    TipoVeh,
    VehiculosOR,
)


def normalizar_nombre(texto):
    if not texto: return ""
    texto = unicodedata.normalize('NFD', texto)
    texto = texto.encode('ascii', 'ignore').decode("utf-8")
    return str(texto).strip().upper()

def parse_curp_details(curp_str):
    if not curp_str or len(curp_str) < 10:
        return None, None
    
    curp_str = curp_str.strip().upper()
    
    birth_date = None
    try:
        yy_str = curp_str[4:6]
        mm_str = curp_str[6:8]
        dd_str = curp_str[8:10]
        
        # Determine century
        year_prefix = 1900
        if len(curp_str) >= 17:
            century_char = curp_str[16]
            if not(century_char.isdigit()):
                year_prefix = 2000
        else:
            yy_int = int(yy_str)
            if yy_int < 30:
                year_prefix = 2000
                
        year = year_prefix + int(yy_str)
        month = int(mm_str)
        day = int(dd_str)
        
        from datetime import date
        birth_date = date(year, month, day)
    except Exception:
        birth_date = None
        
    gender = None
    if len(curp_str) >= 11:
        g_char = curp_str[10]
        if g_char == 'H':
            gender = 'M' # Masculino
        elif g_char == 'M':
            gender = 'F' # Femenino
            
    return birth_date, gender

def get_user_state(request):
    """Retorna el objeto Estado asociado al usuario o None si es superusuario."""
    if request.user.is_superuser:
        return None
    
    profile = getattr(request.user, 'usuarioL', None)
    if not profile:
        return None
        
    # Mapeo especial para nombres comunes en usuarioL vs Estado
    mapping = {
        'CDMX': 'CIUDAD DE MEXICO',
        'EDOMEX': 'ESTADO DE MEXICO',
    }
    
    oficina = profile.oficinaR
    target_name = mapping.get(oficina, oficina)
    target_name = normalizar_nombre(target_name)
    
    try:
        return Estado.objects.get(nombre__iexact=target_name)
    except Estado.DoesNotExist:
        # Intento de búsqueda por normalización si falla el iexact
        for edo in Estado.objects.all():
            if normalizar_nombre(edo.nombre) == target_name:
                return edo
    return None

# --- FUNCIONES DE AGREGACIÓN (EXTRACTADAS PARA REUTILIZACIÓN) ---

def get_totals_by_period(start, end):
    data_by_state = {}
    estados = Estado.objects.all()

    # Función auxiliar para convertir Decimal a int (necesario para JSON)
    def val_int(d, key):
        return int(d.get(key) or 0)

    datos_rep = {}
    datos_rec = {}
    datos_res = {}
    datos_ing = {}
    datos_tra = {}
    datos_ret = {}
    datos_ina = {}
    datos_t = {}

    # Agrupar y resumir datos en una sola consulta por modelo
    rep_raw = Repatriados.objects.filter(fecha__range=[start, end]) \
        .values('estado_id') \
        .annotate(
            total=Sum('mex_rep'),
            adultos=Sum('adultos'),
            menores=Sum('menores'),
            nna_solo=Sum('nna_solo'),
            nna_acom=Sum('nna_acom'),
            terrestres=Sum('terrestres'),
            vuelos=Sum('vuelos')
        )
    rep_map = {item['estado_id']: item for item in rep_raw}

    rec_raw = Recibidos.objects.filter(fecha__range=[start, end]) \
        .values('estado_id') \
        .annotate(
            total=Sum('ext_rec'),
            adultos=Sum('adultos'),
            menores=Sum('menores')
        )
    rec_map = {item['estado_id']: item for item in rec_raw}

    res_raw = ExtRescatados.objects.filter(fecha__range=[start, end]) \
        .values('estado_id') \
        .annotate(
            total=Sum('rescatados'),
            una_vez=Sum('una_vez'),
            reincidente=Sum('reincidente'),
            estacion=Sum('estacion'),
            dif=Sum('dif'),
            conduccion=Sum('conduccion')
        )
    res_map = {item['estado_id']: item for item in res_raw}

    ing_raw = Ingresos.objects.filter(fecha__range=[start, end]) \
        .values('estado_id') \
        .annotate(
            total=Sum('ingresos_total'),
            aereos=Sum('aereos'),
            maritimos=Sum('maritimos'),
            terrestres=Sum('terrestres')
        )
    ing_map = {item['estado_id']: item for item in ing_raw}

    tra_raw = Tramites.objects.filter(fecha__range=[start, end]) \
        .values('estado_id') \
        .annotate(
            total=Sum('total_documentos'),
            res_perm=Sum('residente_permanente'),
            res_temp=Sum('residente_temporal'),
            res_est=Sum('residente_temp_estudio'),
            vis_hum=Sum('visitante_humanitario'),
            vis_adop=Sum('visitante_adopcion'),
            vis_reg=Sum('visitante_regional'),
            vis_trab=Sum('visitante_trabajador')
        )
    tra_map = {item['estado_id']: item for item in tra_raw}

    ret_raw = Retornados.objects.filter(fecha__range=[start, end]) \
        .values('estado_id') \
        .annotate(
            total=Sum('retornados_total'),
            deportado=Sum('deportado'),
            retornado=Sum('retornado')
        )
    ret_map = {item['estado_id']: item for item in ret_raw}

    ina_raw = Inadmitidos.objects.filter(fecha__range=[start, end]) \
        .values('estado_id') \
        .annotate(
            total=Sum('inadmitidos_total')
        )
    ina_map = {item['estado_id']: item for item in ina_raw}

    for edo in estados:
        key = normalizar_nombre(edo.nombre)
        edo_id = edo.id
        
        rep = rep_map.get(edo_id, {})
        rec = rec_map.get(edo_id, {})
        res = res_map.get(edo_id, {})
        ing = ing_map.get(edo_id, {})
        tra = tra_map.get(edo_id, {})
        ret = ret_map.get(edo_id, {})
        ina = ina_map.get(edo_id, {})

        datos_rep[key] = val_int(rep, 'total')
        datos_rec[key] = val_int(rec, 'total')
        datos_res[key] = val_int(res, 'total')
        datos_ing[key] = val_int(ing, 'total')
        datos_tra[key] = val_int(tra, 'total')
        datos_ret[key] = val_int(ret, 'total')
        datos_ina[key] = val_int(ina, 'total')

        data_by_state[key] = {
            'todos': val_int(rep, 'total') + val_int(rec, 'total') + val_int(res, 'total') + val_int(ing, 'total') + val_int(tra, 'total') + val_int(ret, 'total') + val_int(ina, 'total'),
            'color_t': 32, 'color_rep': 32, 'color_rec': 32, 'color_res': 32, 'color_ing': 32, 'color_tra': 32, 'color_ret': 32, 'color_ina': 32,
            'repatriados': val_int(rep, 'total'),
            'rep_adultos': val_int(rep, 'adultos'),
            'rep_menores': val_int(rep, 'menores'),
            'rep_nna_solo': val_int(rep, 'nna_solo'),
            'rep_nna_acom': val_int(rep, 'nna_acom'),
            'rep_terrestres': val_int(rep, 'terrestres'),
            'rep_vuelos': val_int(rep, 'vuelos'),
            'recibidos': val_int(rec, 'total'),
            'rec_adultos': val_int(rec, 'adultos'),
            'rec_menores': val_int(rec, 'menores'),
            'rescatados': val_int(res, 'total'),
            'res_una_vez': val_int(res, 'una_vez'),
            'res_reincidente': val_int(res, 'reincidente'),
            'res_estacion': val_int(res, 'estacion'),
            'res_dif': val_int(res, 'dif'),
            'res_conduccion': val_int(res, 'conduccion'),
            'ingresos': val_int(ing, 'total'),
            'ing_aereos': val_int(ing, 'aereos'),
            'ing_maritimos': val_int(ing, 'maritimos'),
            'ing_terrestres': val_int(ing, 'terrestres'),
            'tramites': val_int(tra, 'total'),
            'tra_res_perm': val_int(tra, 'res_perm'),
            'tra_res_temp': val_int(tra, 'res_temp'),
            'tra_res_est': val_int(tra, 'res_est'),
            'tra_vis_hum': val_int(tra, 'vis_hum'),
            'tra_vis_adop': val_int(tra, 'vis_adop'),
            'tra_vis_reg': val_int(tra, 'vis_reg'),
            'tra_vis_trab': val_int(tra, 'vis_trab'),
            'retornados': val_int(ret, 'total'),
            'ret_deportado': val_int(ret, 'deportado'),
            'ret_retornado': val_int(ret, 'retornado'),
            'inadmitidos': val_int(ina, 'total'),
        }

    # Rankings para Mapa de Calor
    ordenados_rep = sorted(datos_rep.items(), key=lambda x: x[1], reverse=True)
    ordenados_rec = sorted(datos_rec.items(), key=lambda x: x[1], reverse=True)
    ordenados_res = sorted(datos_res.items(), key=lambda x: x[1], reverse=True)
    ordenados_ing = sorted(datos_ing.items(), key=lambda x: x[1], reverse=True)
    ordenados_tra = sorted(datos_tra.items(), key=lambda x: x[1], reverse=True)
    ordenados_ret = sorted(datos_ret.items(), key=lambda x: x[1], reverse=True)
    ordenados_ina = sorted(datos_ina.items(), key=lambda x: x[1], reverse=True)

    for rank, (k, value) in enumerate(ordenados_rep, start=1):
        data_by_state[k]['color_rep'] = 32 if value == 0 else rank
    for rank, (k, value) in enumerate(ordenados_rec, start=1):
        data_by_state[k]['color_rec'] = 32 if value == 0 else rank
    for rank, (k, value) in enumerate(ordenados_res, start=1):
        data_by_state[k]['color_res'] = 32 if value == 0 else rank
    for rank, (k, value) in enumerate(ordenados_ing, start=1):
        data_by_state[k]['color_ing'] = 32 if value == 0 else rank
    for rank, (k, value) in enumerate(ordenados_tra, start=1):
        data_by_state[k]['color_tra'] = 32 if value == 0 else rank
    for rank, (k, value) in enumerate(ordenados_ret, start=1):
        data_by_state[k]['color_ret'] = 32 if value == 0 else rank
    for rank, (k, value) in enumerate(ordenados_ina, start=1):
        data_by_state[k]['color_ina'] = 32 if value == 0 else rank

    for edo in estados:
        k = normalizar_nombre(edo.nombre)
        datos_t[k] = data_by_state[k].get('color_rep', 32) + data_by_state[k].get('color_rec', 32) + \
                     data_by_state[k].get('color_res', 32) + data_by_state[k].get('color_ing', 32) + \
                     data_by_state[k].get('color_tra', 32) + data_by_state[k].get('color_ret', 32) + \
                     data_by_state[k].get('color_ina', 32)
    
    ordenados_t = sorted(datos_t.items(), key=lambda x: x[1])
    for rank, (k, value) in enumerate(ordenados_t, start=1):
        if k in data_by_state: data_by_state[k]['color_t'] = rank

    return data_by_state

def calc_national(totals_dict):
    keys = [
        'todos', 'repatriados', 'rep_adultos', 'rep_menores', 'rep_nna_solo', 'rep_nna_acom', 'rep_terrestres', 'rep_vuelos',
        'recibidos', 'rec_adultos', 'rec_menores', 'rescatados', 'res_una_vez', 'res_reincidente', 'res_estacion', 
        'res_dif', 'res_conduccion', 'ingresos', 'ing_aereos', 'ing_maritimos', 'ing_terrestres',
        'tramites', 'tra_res_perm', 'tra_res_temp', 'tra_res_est', 'tra_vis_hum', 'tra_vis_adop', 'tra_vis_reg', 'tra_vis_trab',
        'retornados', 'ret_deportado', 'ret_retornado', 'inadmitidos'
    ]
    national = {k: 0 for k in keys}
    for state_data in totals_dict.values():
        for k in keys:
            if k in state_data: national[k] += state_data[k]
    return national

#retomar la colocaciones de las fechas de actualizacion en el template

def get_global_update_date():
    """Calcula el mínimo de los máximos para determinar la fecha de integridad total."""
    from .models import Repatriados, Recibidos, ExtRescatados, Ingresos, Tramites, Retornados, Inadmitidos
    
    models_to_check = [
        Repatriados, Recibidos, ExtRescatados, Ingresos, 
        Tramites, Retornados, Inadmitidos
    ]
    
    max_dates = []
    for model in models_to_check:
        res = model.objects.aggregate(max_f=Max('fecha'))['max_f']
        if res:
            max_dates.append(res)
    
    if not max_dates:
        return None
    
    # Retornamos el Mínimo de los Máximos (Integridad Total)
    return min(max_dates)

def get_all_update_dates():
    """Retorna un diccionario con la última fecha de cada modelo."""
    from .models import Repatriados, Recibidos, ExtRescatados, Ingresos, Tramites, Retornados, Inadmitidos, Encuentros
    models_available = {
        'Repatriados': Repatriados,
        'Recibidos': Recibidos,
        'Rescatados': ExtRescatados,
        'Ingresos': Ingresos,
        'Tramites': Tramites,
        'Retornados': Retornados,
        'Inadmitidos': Inadmitidos,
        'Encuentros': Encuentros,
    }
    
    results = []
    for name, model in models_available.items():
        max_f = model.objects.aggregate(max_f=Max('fecha'))['max_f']
        results.append({
            'name': name,
            'date': max_f
        })
    return results

# --- VIEW PRINCIPAL ---

def mapa_informacion(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    
    user_state = get_user_state(request)
    if not request.user.is_superuser and not user_state:
        return render(request, 'base/error404.html')

    fecha_act = get_global_update_date() or date.today()

    CS_START = date(2024, 10, 1)
    DT_START = date(2025, 1, 20)

    totals_cs = get_totals_by_period(CS_START, fecha_act)
    totals_dt = get_totals_by_period(DT_START, fecha_act)
    
    # --- Capa de Instrucciones (Consolidado desde API externa) ---
    import requests
    import urllib3
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    
    api_instrucciones = {}
    try:
        api_res = requests.get('https://172.16.16.167/api/mapa-datos/', verify=False, timeout=2.5)
        if api_res.status_code == 200:
            api_instrucciones = api_res.json()
    except Exception as e:
        print(f"Error al obtener instrucciones de la API: {str(e)}")

    instrucciones_avance = {}
    instrucciones_totales_dict = {}
    
    for edo in Estado.objects.all():
        edo_key = normalizar_nombre(edo.nombre)
        api_key = None
        for key in api_instrucciones.keys():
            if normalizar_nombre(key) == edo_key:
                api_key = key
                break
        
        visits = api_instrucciones.get(api_key, []) if api_key else []
        active_visits = [v for v in visits if v.get('pendiente', 0) > 0]
        
        if active_visits:
            avg_avance = sum(v.get('avance', 0) for v in active_visits) // len(active_visits)
            total_pendientes = sum(v.get('pendiente', 0) for v in visits)
            total_atendidos = sum(v.get('atendido', 0) for v in visits)
            total_acuerdos = sum(v.get('total', 0) for v in visits)
            
            instrucciones_avance[edo_key] = avg_avance
            instrucciones_totales_dict[edo_key] = {
                'avance': avg_avance,
                'pendiente': total_pendientes,
                'atendido': total_atendidos,
                'total': total_acuerdos,
                'has_pending': True
            }
        else:
            if visits:
                total_atendidos = sum(v.get('atendido', 0) for v in visits)
                total_acuerdos = sum(v.get('total', 0) for v in visits)
                instrucciones_avance[edo_key] = 100
                instrucciones_totales_dict[edo_key] = {
                    'avance': 100,
                    'pendiente': 0,
                    'atendido': total_atendidos,
                    'total': total_acuerdos,
                    'has_pending': False
                }
            else:
                instrucciones_avance[edo_key] = 100
                instrucciones_totales_dict[edo_key] = {
                    'avance': 100,
                    'pendiente': 0,
                    'atendido': 0,
                    'total': 0,
                    'has_pending': False
                }

    estados_para_rankear = [item for item in instrucciones_totales_dict.items() if item[1]['has_pending']]
    estados_ordenados = sorted(estados_para_rankear, key=lambda x: x[1]['avance'])
    
    instrucciones_color_rank = {}
    for rank, (edo_key, data_val) in enumerate(estados_ordenados, start=1):
        instrucciones_color_rank[edo_key] = rank
    
    # Etiqueta centralizada para la escala global
    LABEL_NACIONAL = "Total Nacional"

    # Diccionario maestro de etiquetas de métricas
    METRIC_LABELS = {
        'todos': 'Todos',
        'repatriados': 'Mexicanos Recibidos',
        'recibidos': 'Extranjeros Recibidos',
        'rescatados': 'Rescates',
        'ingresos': 'Internaciones',
        'tramites': 'Trámites',
        'retornados': 'Retornados',
        'inadmitidos': 'Inadmitidos',
        'recibidos_total': 'Recibidos',
        'instrucciones': 'Supervision'
    }

    # --- Recopilación de Infraestructura y Titulares ---
    infra_raw = PuntosInternacionEstacion.objects.values('estado__nombre', 'tipo').annotate(total=Count('id'))
    titulares_raw = Titular.objects.all().select_related('estado', 'tipo_nombramiento')
    
    infra_data = {}
    # Estructura base para todos los estados
    for edo in Estado.objects.all():
        infra_data[normalizar_nombre(edo.nombre)] = {
            'estado_id': edo.id,
            'AEREO': 0, 'MARITIMO': 0, 'TERRESTRE': 0, 'ESTACION': 0,
            'PRH': 0,
            'personal_total': 0,
            'personal_activo': 0,
            'personal_vacante': 0,
            'subrep_federal': 0,
            'subrep_local': 0,
            'rep_local': 0,
            'titular': 'Sin titular asignado',
            'titular_id': None,
            'foto': None,
            'tipo_nombramiento': None
        }
    
    # Población con datos reales
    for item in infra_raw:
        edo_name = normalizar_nombre(item['estado__nombre'])
        if edo_name in infra_data:
            infra_data[edo_name][item['tipo']] = item['total']
            
    for t in titulares_raw:
        edo_name = normalizar_nombre(t.estado.nombre)
        if edo_name in infra_data:
            infra_data[edo_name]['titular'] = f"{t.nombre} {t.apellido_paterno} {t.apellido_materno}"
            infra_data[edo_name]['titular_id'] = t.id
            infra_data[edo_name]['tipo_nombramiento'] = t.tipo_nombramiento.nombre if t.tipo_nombramiento else None
            if t.fotografia:
                infra_data[edo_name]['foto'] = t.fotografia.url

    # PRHs por estado
    prh_raw = PRHs.objects.values('estado__nombre').annotate(total=Count('id'))
    for item in prh_raw:
        edo_name = normalizar_nombre(item['estado__nombre'])
        if edo_name in infra_data:
            infra_data[edo_name]['PRH'] = item['total']

    # Personal por estado
    personal_qs = PersonalINM.objects.all().select_related('estado', 'estatus', 'tipo_plaza')
    for p in personal_qs:
        if not p.estado:
            continue
        
        tipo_plaza = (p.tipo_plaza.plazaT if p.tipo_plaza else '').upper()
        if tipo_plaza not in ['BASE', 'CONFIANZA']:
            continue

        edo_name = normalizar_nombre(p.estado.nombre)
        if edo_name in infra_data:
            infra_data[edo_name]['personal_total'] += 1
            estatus_name = (p.estatus.estatus if p.estatus else '').upper()
            if estatus_name == 'ACTIVO':
                infra_data[edo_name]['personal_activo'] += 1
            elif estatus_name == 'VACANTE':
                infra_data[edo_name]['personal_vacante'] += 1
            
            puesto_clean = normalizar_nombre(p.puesto_especifico)
            if 'SUB REPRESENTACION FEDERAL' in puesto_clean:
                infra_data[edo_name]['subrep_federal'] += 1
            elif 'SUB REPRESENTACION LOCAL' in puesto_clean:
                infra_data[edo_name]['subrep_local'] += 1
            elif 'REPRESENTACION LOCAL' in puesto_clean:
                infra_data[edo_name]['rep_local'] += 1

    # Totales Nacionales
    subrep_federal_nat = 0
    subrep_local_nat = 0
    rep_local_nat = 0
    for key, val in infra_data.items():
        subrep_federal_nat += val.get('subrep_federal', 0)
        subrep_local_nat += val.get('subrep_local', 0)
        rep_local_nat += val.get('rep_local', 0)

    infra_data[LABEL_NACIONAL] = {
        'estado_id': None,
        'AEREO': PuntosInternacionEstacion.objects.filter(tipo='AEREO').count(),
        'MARITIMO': PuntosInternacionEstacion.objects.filter(tipo='MARITIMO').count(),
        'TERRESTRE': PuntosInternacionEstacion.objects.filter(tipo='TERRESTRE').count(),
        'ESTACION': PuntosInternacionEstacion.objects.filter(tipo='ESTACION').count(),
        'PRH': PRHs.objects.count(),
        'personal_total': PersonalINM.objects.count(),
        'personal_activo': PersonalINM.objects.filter(estatus__estatus__iexact='ACTIVO').count(),
        'personal_vacante': PersonalINM.objects.filter(estatus__estatus__iexact='VACANTE').count(),
        'subrep_federal': subrep_federal_nat,
        'subrep_local': subrep_local_nat,
        'rep_local': rep_local_nat,
        'titular': 'Datos Nacionales',
        'foto': None,
        'tipo_nombramiento': None
    }


    # Valores por defecto para estados sin datos
    default_vals = {
        'todos': 0, 'color_t': 32, 'color_rep': 32, 'color_rec': 32, 'color_res': 32, 'color_ing': 32, 'color_tra': 32, 'color_ret': 32, 'color_ina': 32,
        'repatriados': 0, 'rep_adultos': 0, 'rep_menores': 0, 'rep_nna_solo': 0, 'rep_nna_acom': 0, 'rep_terrestres': 0, 'rep_vuelos': 0,
        'recibidos': 0, 'rec_adultos': 0, 'rec_menores': 0,
        'rescatados': 0, 'res_una_vez': 0, 'res_reincidente': 0, 'res_estacion': 0, 'res_dif': 0, 'res_conduccion': 0,
        'ingresos': 0, 'ing_aereos': 0, 'ing_maritimos': 0, 'ing_terrestres': 0,
        'tramites': 0, 'tra_res_perm': 0, 'tra_res_temp': 0, 'tra_res_est': 0, 'tra_vis_hum': 0, 'tra_vis_adop': 0, 'tra_vis_reg': 0, 'tra_vis_trab': 0,
        'retornados': 0, 'ret_deportado': 0, 'ret_retornado': 0, 'inadmitidos': 0,
        'instrucciones': 0, 'color_ins': 32
    }

    # Ruta al archivo geojson descargado
    geojson_path = os.path.join(settings.BASE_DIR, 'mapa', 'static', 'mapa', 'data', 'inegi_latlon_mexico.geojson')
    
    with open(geojson_path, 'r', encoding='utf-8') as f:
        geo_data = json.load(f)

    # Filtrar el GeoJSON por estado si el usuario no es superusuario
    if user_state:
        user_state_name_normalized = normalizar_nombre(user_state.nombre)
        geo_data['features'] = [
            f for f in geo_data['features']
            if normalizar_nombre(f['properties']['name']) == user_state_name_normalized
        ]
    # Inyectar datos reales en cada estado
    for feature in geo_data['features']:
        name_normalized = normalizar_nombre(feature['properties']['name'])
        
        cs = totals_cs.get(name_normalized, default_vals).copy()
        dt = totals_dt.get(name_normalized, default_vals).copy()
        
        # Inyectar datos de instrucciones calculados
        ins_val = instrucciones_totales_dict.get(name_normalized, {'avance': 100, 'has_pending': False})
        ins_rank = instrucciones_color_rank.get(name_normalized, 32)
        
        cs['instrucciones'] = ins_val['avance']
        cs['color_ins'] = ins_rank
        dt['instrucciones'] = ins_val['avance']
        dt['color_ins'] = ins_rank
        
        # Inyectar en GeoJSON (usamos prefijos cs_, dt_ y pe_)
        for k in default_vals:
            feature['properties'][f'cs_{k}'] = cs[k]
            feature['properties'][f'dt_{k}'] = dt[k]
            feature['properties'][f'pe_{k}'] = cs[k] # Inicializar PE con valores de CS
        
        # Cadenas formateadas para el Tooltip Dinámico
        for k in ['todos', 'repatriados', 'recibidos', 'rescatados', 'ingresos', 'tramites', 'retornados', 'inadmitidos', 'instrucciones']:
            if k == 'instrucciones':
                feature['properties'][f'cs_str_{k}'] = f"{cs[k]}%"
                feature['properties'][f'dt_str_{k}'] = f"{dt[k]}%"
                feature['properties'][f'pe_str_{k}'] = f"{cs[k]}%"
            else:
                feature['properties'][f'cs_str_{k}'] = f"{cs[k]:,}"
                feature['properties'][f'dt_str_{k}'] = f"{dt[k]:,}"
                feature['properties'][f'pe_str_{k}'] = f"{cs[k]:,}"
        
    # --- Capa de Infraestructura (Iconos SVG) ---
    infra_points_objs = PuntosInternacionEstacion.objects.all()
    if user_state:
        infra_points_objs = infra_points_objs.filter(estado=user_state)
    infra_pts_data = []
    for pt in infra_points_objs:
        icon_file = 'terrestre2.svg' # Default
        if pt.tipo == 'AEREO': icon_file = 'aereo2.svg'
        elif pt.tipo == 'MARITIMO': icon_file = 'maritimo2.svg'
        elif pt.tipo == 'ESTACION': icon_file = 'estacion2.svg'
        
        infra_pts_data.append({
            'x': float(pt.longitud) if pt.longitud else 0,
            'y': float(pt.latitud) if pt.latitud else 0,
            'nombre': pt.nombre,
            'estado': normalizar_nombre(pt.estado.nombre),
            'tipo': pt.tipo,
            'url': f"{settings.STATIC_URL}mapa/icons/{icon_file}"
        })

    # --- Capa de Puntos de Rescate Humano (PRH) ---
    prh_points = PRHs.objects.all().select_related('modalidad')
    if user_state:
        prh_points = prh_points.filter(estado=user_state)
    prh_pts_data = []
    for pt in prh_points:
        icon = 'agente_activo2.svg' if pt.activo else 'agente_inactivo2.svg'
        prh_pts_data.append({
            'x': float(pt.longitud) if pt.longitud else 0,
            'y': float(pt.latitud) if pt.latitud else 0,
            'nombre': pt.nombre,
            'estado': normalizar_nombre(pt.estado.nombre),
            'modalidad': pt.modalidad.nombre,
            'status': 'Activo' if pt.activo else 'Inactivo',
            'url': f"{settings.STATIC_URL}mapa/icons/{icon}"
        })

    # --- Capa de Inmuebles (Icono OR_ACTIVO) ---
    inmuebles_objs = Inmueble.objects.all().prefetch_related('tipo_oficina')
    if user_state:
        inmuebles_objs = inmuebles_objs.filter(estado=user_state)
    inmuebles_pts_data = []
    for pt in inmuebles_objs:
        inmuebles_pts_data.append({
            'id': pt.id,
            'x': float(pt.longitud) if pt.longitud else 0,
            'y': float(pt.latitud) if pt.latitud else 0,
            'nombre': pt.nombre_inmueble,
            'estado': normalizar_nombre(pt.estado.nombre),
            'tipo': 'INMUEBLE',
            'tipo_oficina': [normalizar_nombre(to.nombre) for to in pt.tipo_oficina.all()],
            'url': f"{settings.STATIC_URL}mapa/icons/OR_ACTIVO.svg"
        })

    total_national_acuerdos = sum(s['total'] for s in instrucciones_totales_dict.values())
    total_national_atendidos = sum(s['atendido'] for s in instrucciones_totales_dict.values())
    national_avance = (total_national_atendidos * 100 // total_national_acuerdos) if total_national_acuerdos > 0 else 100

    if user_state:
        # Totales del estado del usuario
        user_state_key = normalizar_nombre(user_state.nombre)
        ins_val = instrucciones_totales_dict.get(user_state_key, {'avance': 100, 'has_pending': False})
        
        cs = totals_cs.get(user_state_key, default_vals).copy()
        dt = totals_dt.get(user_state_key, default_vals).copy()
        cs['instrucciones'] = ins_val['avance']
        dt['instrucciones'] = ins_val['avance']
        
        national_data = {
            'name': user_state.nombre.upper(),
            'cs': cs,
            'dt': dt,
            'pe': cs
        }
    else:
        national_data = {
            'name': LABEL_NACIONAL,
            'cs': calc_national(totals_cs),
            'dt': calc_national(totals_dt),
            'pe': calc_national(totals_cs)
        }
        national_data['cs']['instrucciones'] = national_avance
        national_data['dt']['instrucciones'] = national_avance
        national_data['pe']['instrucciones'] = national_avance
    
    context = {
        'geo_data_json': json.dumps(geo_data),
        'national_data_json': json.dumps(national_data),
        'infra_data_json': json.dumps(infra_data),
        'infra_pts_data_json': json.dumps(infra_pts_data),
        'prh_pts_data_json': json.dumps(prh_pts_data),
        'inmuebles_pts_data_json': json.dumps(inmuebles_pts_data),
        'label_nacional': LABEL_NACIONAL,
        'metric_labels': METRIC_LABELS,
        'metric_labels_json': json.dumps(METRIC_LABELS),
        'fecha_actualizacion': fecha_act,
        'instrucciones_api_json': json.dumps(api_instrucciones),
        'is_superuser': request.user.is_superuser,
        'user_state_name': user_state.nombre if user_state else '',
        'user_state_name_normalized': normalizar_nombre(user_state.nombre) if user_state else '',
    }
    
    return render(request, 'mapa/informacion.html', context)

def carga_datos(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    
    user_state = get_user_state(request)
    if not request.user.is_superuser and not user_state:
        return render(request, 'base/error404.html')

    models_available = {
        'Repatriados': Repatriados,
        'Recibidos': Recibidos,
        'Extranjeros Rescatados': ExtRescatados,
        'Ingresos': Ingresos,
        'Tramites': Tramites,
        'Retornados': Retornados,
        'Inadmitidos': Inadmitidos,
        'Encuentros': Encuentros,
    }

    if request.method == 'POST':
        model_name = request.POST.get('model_name')
        excel_file = request.FILES.get('excel_file')

        if not model_name or not excel_file:
            messages.error(request, "Por favor seleccione un modelo y un archivo.")
            return redirect('carga_datos')

        if model_name not in models_available:
            messages.error(request, "Modelo no válido.")
            return redirect('carga_datos')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            model_class = models_available[model_name]
            
            # Obtener campos del modelo (específicos para ignorar id y estado FK inicialmente)
            # El orden del Excel debe ser: fecha, estado (nombre), y luego los campos específicos
            fields = [field.name for field in model_class._meta.fields if field.name not in ['id']]
            
            # Pre-cargar catálogos en memoria para evitar miles de consultas
            estados_dict = {normalizar_nombre(e.nombre): e for e in Estado.objects.all()}
            nacionalidades_dict = {normalizar_nombre(n.nombre): n for n in Nacionalidad.objects.all()}
            
            rows_created = 0
            rows_updated = 0
            errors = []
            
            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row): continue
                    
                    fecha_val = row[0]
                    if model_name == 'Encuentros':
                        # Formato: Fecha(0), Agencia(1), CiudadEU(2), EstadoEU(3), EstadoMex(4), Nacionalidad(5), Total(6)
                        agencia_raw = row[1]
                        ciudad_eu_raw = row[2]
                        estado_eu_raw = row[3]
                        estado_nombre = row[4]
                        nacionalidad_nombre = row[5]
                        
                        agencia_norm = normalizar_nombre(agencia_raw)
                        ciudad_eu_norm = normalizar_nombre(ciudad_eu_raw)
                        estado_eu_norm = normalizar_nombre(estado_eu_raw)
                        
                        nacionalidad_norm = normalizar_nombre(nacionalidad_nombre)
                        nac_obj = nacionalidades_dict.get(nacionalidad_norm)
                        
                        if not nac_obj:
                            errors.append(f"Fila {row_idx}: Nacionalidad '{nacionalidad_nombre}' no encontrada.")
                            continue
                        
                        estado_norm = normalizar_nombre(estado_nombre)
                        estado_obj = estados_dict.get(estado_norm)
                        if not estado_obj:
                            errors.append(f"Fila {row_idx}: Estado '{estado_nombre}' no encontrado.")
                            continue
                        if user_state and estado_obj != user_state:
                            errors.append(f"Fila {row_idx}: No tiene permisos para modificar datos del estado '{estado_nombre}'.")
                            continue

                        obj, created = Encuentros.objects.update_or_create(
                            fecha=fecha_val,
                            agencia=agencia_norm,
                            ciudadEU=ciudad_eu_norm,
                            estadoEU=estado_eu_norm,
                            estado=estado_obj,
                            nacionalidad=nac_obj,
                            defaults={'encuentros_total': row[6] if len(row) > 6 else 0}
                        )
                    else:
                        estado_nombre = row[1]
                        nacionalidad_nombre = row[2] # Columna C
                        
                        # Normalizar nacionalidad
                        nacionalidad_norm = normalizar_nombre(nacionalidad_nombre)
                        nac_obj = nacionalidades_dict.get(nacionalidad_norm)
                        
                        if not nac_obj:
                            errors.append(f"Fila {row_idx}: La nacionalidad '{nacionalidad_nombre}' no existe en el catálogo.")
                            continue

                        # Convertir fecha si es necesario
                        if isinstance(fecha_val, str):
                            try:
                                fecha_val = datetime.strptime(fecha_val, '%Y-%m-%d').date()
                            except:
                                errors.append(f"Fila {row_idx}: Formato de fecha inválido (esperado YYYY-MM-DD).")
                                continue
                        
                        # Normalizar el nombre del estado
                        estado_norm_busqueda = normalizar_nombre(estado_nombre)
                        estado_obj = estados_dict.get(estado_norm_busqueda)
                        
                        if not estado_obj:
                            errors.append(f"Fila {row_idx}: Estado '{estado_nombre}' no encontrado.")
                            continue
                        if user_state and estado_obj != user_state:
                            errors.append(f"Fila {row_idx}: No tiene permisos para modificar datos del estado '{estado_nombre}'.")
                            continue

                        # Construir diccionario de datos
                        data = {}
                        fields_to_populate = [f.name for f in model_class._meta.fields if f.name not in ['id', 'fecha', 'estado', 'nacionalidad']]
                        
                        for i, field_name in enumerate(fields_to_populate):
                            excel_idx = i + 3
                            if excel_idx < len(row):
                                val = row[excel_idx]
                                data[field_name] = val if val is not None else 0

                        obj, created = model_class.objects.update_or_create(
                            fecha=fecha_val,
                            estado=estado_obj,
                            nacionalidad=nac_obj,
                            defaults=data
                        )
                    
                    if created:
                        rows_created += 1
                    else:
                        rows_updated += 1

            if errors:
                for error in errors:
                    messages.error(request, error)
            
            if rows_created > 0 or rows_updated > 0:
                messages.success(request, f"Carga completada. Creados: {rows_created}, Actualizados: {rows_updated}")
            
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            
        return redirect('carga_datos')

    user_state = get_user_state(request)
    
    if user_state:
        estados_list = [user_state]
        titulares_list = Titular.objects.filter(estado=user_state).order_by('nombre')
    else:
        estados_list = Estado.objects.all().order_by('nombre')
        titulares_list = Titular.objects.all().order_by('nombre')

    update_dates = get_all_update_dates()
    return render(request, 'mapa/carga_datos.html', {
        'models': models_available.keys(),
        'update_dates': update_dates,
        'estados_list': Estado.objects.all().order_by('nombre'),
    })

def titulares_list(request):
    """Vista para la gestión independiente de expedientes de titulares."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    user_state = get_user_state(request)
    
    if user_state:
        estados_list = [user_state]
        titulares_list = Titular.objects.filter(estado=user_state).order_by('estado__nombre', 'nombre')
    else:
        estados_list = Estado.objects.all().order_by('nombre')
        titulares_list = Titular.objects.all().order_by('estado__nombre', 'nombre')

    return render(request, 'mapa/titulares.html', {
        'estados_list': estados_list,
        'nacionalidades_list': Nacionalidad.objects.all().order_by('nombre'),
        'grados_academicos': GradoAcademico.objects.all().order_by('nombre'),
        'tipos_nombramiento': TipoNombramiento.objects.all().order_by('nombre'),
        'procedencias_list': TipoProcendencia.objects.all().order_by('institucion'),
        'titulares_list': titulares_list,
    })

def api_get_titular(request, titular_id):
    """Retorna los datos de un titular en formato JSON para edición."""
    try:
        user_state = get_user_state(request)
        titular = Titular.objects.get(id=titular_id)
        
        # Validación de seguridad
        if user_state and titular.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        data = {
            'id': titular.id,
            'nombre': titular.nombre,
            'apellido_paterno': titular.apellido_paterno,
            'apellido_materno': titular.apellido_materno,
            'curp': titular.curp,
            'fecha_nacimiento': titular.fecha_nacimiento.isoformat() if titular.fecha_nacimiento else None,
            'sexo': titular.sexo,
            'nivel': titular.nivel,
            'codigo_plaza': titular.codigo_plaza,
            'tipo_nombramiento': titular.tipo_nombramiento.nombre if titular.tipo_nombramiento else None,
            'tipo_nombramiento_id': titular.tipo_nombramiento_id,
            'procedencia': titular.procedencia.institucion if titular.procedencia else None,
            'procedencia_id': titular.procedencia_id,
            'estado': titular.estado.nombre,
            'estado_id': titular.estado_id,
            'nacionalidad_id': titular.nacionalidad_id,
            'foto_url': titular.fotografia.url if titular.fotografia else None,
            'telefonos': list(titular.telefonos.values('tipo', 'numero')),
            'correos': list(titular.correos.values('tipo', 'correo')),
            'estudios': [
                {
                    'grado': e.grado.nombre if e.grado else 'Sin grado',
                    'grado_id': e.grado_id,
                    'carrera': e.carrera
                } for e in titular.estudios.all()
            ],
            'trayectoria': list(titular.trayectoria.values('puesto', 'area', 'fecha_inicio', 'fecha_fin', 'actual')),
            'experiencia': list(titular.experiencia_externa.values('institucion', 'cargo', 'fecha_inicio', 'fecha_fin', 'descripcion')),
        }
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

@transaction.atomic
def eliminar_titular(request, titular_id):
    """Elimina un titular y todos sus datos relacionados."""
    try:
        user_state = get_user_state(request)
        titular = Titular.objects.get(id=titular_id)
        
        # Validación de seguridad
        if user_state and titular.estado != user_state:
            messages.error(request, "No tienes permisos para eliminar este registro.")
            return redirect('titulares_list')
            
        nombre = titular.nombre
        titular.delete()
        messages.success(request, f"Expediente de {nombre} eliminado correctamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar titular: {str(e)}")
    
    return redirect('titulares_list')


def personal_list(request):
    """Vista para la gestión independiente del Personal INM con paginación y filtros en el servidor."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    user_state = get_user_state(request)
    
    # 1. Obtener filtros de la solicitud GET
    nombre_query = request.GET.get('nombre', '').strip()
    estado_id_query = request.GET.get('estado_id', '').strip()
    puesto_query = request.GET.get('puesto', '').strip()
    page_number = request.GET.get('page', 1)
    
    # 2. Filtrar queryset de personal
    personal_qs = PersonalINM.objects.all().select_related('estado', 'lugar_asignado')
    
    if user_state:
        personal_qs = personal_qs.filter(estado=user_state)
        estados_list_all = [user_state]
    else:
        estados_list_all = Estado.objects.all().order_by('nombre')
        if estado_id_query:
            personal_qs = personal_qs.filter(estado_id=estado_id_query)
            
    if nombre_query:
        from django.db.models import Q
        personal_qs = personal_qs.filter(
            Q(nombre__icontains=nombre_query) | 
            Q(apellido__icontains=nombre_query) | 
            Q(num_empleado__icontains=nombre_query) |
            Q(codigo_plaza__icontains=nombre_query)
        )
        
    if puesto_query:
        personal_qs = personal_qs.filter(puesto_especifico__icontains=puesto_query)
        
    personal_qs = personal_qs.order_by('estado__nombre', 'nombre')
    total_matched = personal_qs.count()
    
    # 3. Paginación de resultados (50 por página)
    from django.core.paginator import Paginator
    paginator = Paginator(personal_qs, 50)
    page_obj = paginator.get_page(page_number)
    
    # 4. Listados para formularios de creación/edición
    if user_state:
        estados_list = [user_state]
        inmuebles_list = Inmueble.objects.filter(estado=user_state).order_by('nombre_inmueble')
    else:
        estados_list = Estado.objects.all().order_by('nombre')
        inmuebles_list = Inmueble.objects.all().order_by('nombre_inmueble')
        
    return render(request, 'mapa/personal_list.html', {
        'estados_list': estados_list,
        'estados_list_all': estados_list_all,
        'actividades_list': TipoDependencia.objects.all().order_by('nombre'),
        'inmuebles_list': inmuebles_list,
        'personal_list': page_obj,
        'total_matched': total_matched,
        'nombre_query': nombre_query,
        'estado_id_query': estado_id_query,
        'puesto_query': puesto_query,
        'estatus_list': EstatusPersonal.objects.all().order_by('estatus'),
        'tipo_plaza_list': TipoPlaza.objects.all().order_by('plazaT'),
    })


def api_get_personal(request, personal_id):
    """Retorna los datos de un personal en formato JSON para edición."""
    try:
        user_state = get_user_state(request)
        personal = PersonalINM.objects.get(id=personal_id)
        
        # Validación de seguridad
        if user_state and personal.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        data = {
            'id': personal.id,
            'estado_id': personal.estado_id,
            'estatus_id': personal.estatus_id or '',
            'tipo_plaza_id': personal.tipo_plaza_id or '',
            'codigo_plaza': personal.codigo_plaza,
            'nivel': personal.nivel,
            'num_empleado': personal.num_empleado or '',
            'nombre': personal.nombre or '',
            'apellido': personal.apellido or '',
            'curp': personal.curp or '',
            'fecha_nacimiento': personal.fecha_nacimiento.isoformat() if personal.fecha_nacimiento else '',
            'sexo': personal.sexo or '',
            'tipo_movimiento': personal.tipo_movimiento,
            'fecha_ingreso_inm': personal.fecha_ingreso_inm.isoformat() if personal.fecha_ingreso_inm else None,
            'fecha_ingreso_plaza': personal.fecha_ingreso_plaza.isoformat() if personal.fecha_ingreso_plaza else None,
            'vig_inicio_mov': personal.vig_inicio_mov.isoformat() if personal.vig_inicio_mov else None,
            'vig_termino_mov': personal.vig_termino_mov.isoformat() if personal.vig_termino_mov else None,
            'puesto_especifico': personal.puesto_especifico,
            'sueldo_bruto': personal.sueldo_bruto,
            'sueldo_neto': personal.sueldo_neto,
            'actividad_ids': list(personal.actividad.values_list('id', flat=True)),
            'jefe_oficina': personal.jefe_oficina,
            'lugar_asignado_id': personal.lugar_asignado_id or '',
        }
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@transaction.atomic
def eliminar_personal(request, personal_id):
    """Elimina un registro de personal."""
    try:
        user_state = get_user_state(request)
        personal = PersonalINM.objects.get(id=personal_id)
        
        # Validación de seguridad
        if user_state and personal.estado != user_state:
            messages.error(request, "No tienes permisos para eliminar este registro.")
            return redirect('personal_list')
            
        nombre = f"{personal.nombre or ''} {personal.apellido or ''}"
        personal.delete()
        messages.success(request, f"Registro de {nombre} eliminado correctamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar registro: {str(e)}")
    
    return redirect('personal_list')


@transaction.atomic
def guardar_personal(request):
    if request.method != 'POST':
        return redirect('personal_list')
    
    user_state = get_user_state(request)
    
    try:
        personal_id = request.POST.get('personal_id')
        estado_id = request.POST.get('estado_id')
        
        # Validación de seguridad: el estado enviado debe coincidir con el del usuario
        if user_state and str(user_state.id) != str(estado_id):
            raise PermissionError("No tienes permisos para registrar personal en este estado.")

        # Parsear floats y fechas opcionales
        sueldo_bruto = request.POST.get('sueldo_bruto') or None
        if sueldo_bruto:
            sueldo_bruto = float(sueldo_bruto)
            
        sueldo_neto = request.POST.get('sueldo_neto') or None
        if sueldo_neto:
            sueldo_neto = float(sueldo_neto)

        curp = request.POST.get('curp', '').strip().upper() or None
        fecha_nacimiento = request.POST.get('fecha_nacimiento') or None
        sexo = request.POST.get('sexo') or None
        
        # Derivar fecha_nacimiento y sexo si hay CURP pero no se enviaron
        if curp and (not fecha_nacimiento or not sexo):
            derived_dob, derived_sex = parse_curp_details(curp)
            if not fecha_nacimiento:
                fecha_nacimiento = derived_dob
            if not sexo:
                sexo = derived_sex

        defaults = {
            'estado_id': estado_id,
            'estatus_id': request.POST.get('estatus_id') or None,
            'tipo_plaza_id': request.POST.get('tipo_plaza_id') or None,
            'codigo_plaza': request.POST.get('codigo_plaza', '').strip().upper(),
            'nivel': request.POST.get('nivel', '').strip().upper(),
            'num_empleado': request.POST.get('num_empleado', '').strip().upper() or None,
            'nombre': normalizar_nombre(request.POST.get('nombre', '')),
            'apellido': normalizar_nombre(request.POST.get('apellido', '')),
            'curp': curp,
            'fecha_nacimiento': fecha_nacimiento,
            'sexo': sexo,
            'tipo_movimiento': request.POST.get('tipo_movimiento') == 'on',
            'fecha_ingreso_inm': request.POST.get('fecha_ingreso_inm') or None,
            'fecha_ingreso_plaza': request.POST.get('fecha_ingreso_plaza') or None,
            'vig_inicio_mov': request.POST.get('vig_inicio_mov') or None,
            'vig_termino_mov': request.POST.get('vig_termino_mov') or None,
            'puesto_especifico': normalizar_nombre(request.POST.get('puesto_especifico', '')),
            'sueldo_bruto': sueldo_bruto,
            'sueldo_neto': sueldo_neto,
            'jefe_oficina': request.POST.get('jefe_oficina') == 'on',
            'lugar_asignado_id': request.POST.get('lugar_asignado_id') or None,
        }

        actividad_ids = request.POST.getlist('actividad_ids[]')
        # También aceptar actividad_ids como campo único separado por comas (fallback)
        if not actividad_ids:
            raw = request.POST.get('actividad_ids', '')
            actividad_ids = [v.strip() for v in raw.split(',') if v.strip()]

        if personal_id:
            personal = PersonalINM.objects.get(id=personal_id)
            if user_state and personal.estado != user_state:
                raise PermissionError("No tienes permisos para modificar este registro.")
            for key, val in defaults.items():
                setattr(personal, key, val)
            personal.save()
            personal.actividad.set(actividad_ids)
            messages.success(request, "Registro de personal actualizado correctamente.")
        else:
            personal = PersonalINM.objects.create(**defaults)
            personal.actividad.set(actividad_ids)
            messages.success(request, "Nuevo personal registrado correctamente.")
            
    except Exception as e:
        messages.error(request, f"Error al guardar personal: {str(e)}")
        
    return redirect('personal_list')


@transaction.atomic
def carga_rapida_personal(request):
    if request.method != 'POST':
        return redirect('personal_list')
        
    user_state = get_user_state(request)
    
    # Check if request is JSON (AJAX chunks)
    is_json = False
    if request.content_type == 'application/json':
        try:
            import json
            payload = json.loads(request.body)
            is_json = payload.get('is_ajax_chunk', False)
        except Exception:
            is_json = False
            
    try:
        if is_json:
            headers = payload.get('headers', [])
            rows_data = payload.get('rows', [])
        else:
            excel_file = request.FILES.get('excel_file')
            if not excel_file:
                messages.error(request, "No se ha proporcionado ningún archivo.")
                return redirect('personal_list')
                
            if not excel_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, "El archivo debe ser un Excel (.xlsx o .xls).")
                return redirect('personal_list')
                
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            sheet = wb.active
            
            # Encontrar la fila de encabezados. Buscaremos la primera fila que contenga 'STATUS' o 'CODIGO-PLAZA'
            headers = []
            header_row_idx = None
            
            for r in range(1, 21):
                row_vals = [str(cell.value or '').strip().upper() for cell in sheet[r]]
                if 'STATUS' in row_vals or 'CODIGO-PLAZA\nNUEVO' in row_vals or 'CODIGO-PLAZA NUEVO' in row_vals or 'CODIGO-PLAZA' in row_vals:
                    headers = [str(cell.value or '').strip() for cell in sheet[r]]
                    header_row_idx = r
                    break
                    
            if not header_row_idx:
                header_row_idx = 8
                headers = [str(cell.value or '').strip() for cell in sheet[8]]
                
            rows_data = []
            for r in range(header_row_idx + 1, sheet.max_row + 1):
                row_cells = list(sheet[r])
                if not any(cell.value for cell in row_cells):
                    continue
                row_vals = [cell.value for cell in row_cells]
                rows_data.append(row_vals)

        # Normalizar nombres de columnas a claves consistentes
        col_mapping = {}
        for idx, h in enumerate(headers):
            h_norm = normalizar_nombre(h).replace('\n', ' ').strip()
            if 'STATUS' in h_norm:
                col_mapping['status'] = idx
                # print(f"status {idx}")
            elif 'TIPO DE PLAZA' in h_norm:
                col_mapping['tipo_plaza'] = idx
            elif 'ADSCRIPCION' in h_norm or 'ESTADO' in h_norm:
                col_mapping['adscripcion'] = idx
            elif 'CODIGO' in h_norm:
                col_mapping['codigo_plaza'] = idx
                # print(f"codigo_plaza {idx}")
            elif 'NIVEL' in h_norm:
                col_mapping['nivel'] = idx
            elif 'NUM EMP' in h_norm or 'NUM_EMP' in h_norm or 'EMPLEADO' in h_norm:
                col_mapping['num_empleado'] = idx
            elif 'CURP' in h_norm:
                col_mapping['curp'] = idx
            elif 'NOMBRE' in h_norm:
                col_mapping['nombre'] = idx
            elif 'MOVIMIENTO' in h_norm:
                col_mapping['tipo_movimiento'] = idx
            elif 'FECHA DE ING. INM' in h_norm or 'FECHA ING INM' in h_norm or 'ING. INM' in h_norm:
                col_mapping['fecha_ingreso_inm'] = idx
                # print(f"fecha_ingreso_inm {idx}")
            elif 'FECHA DE ING A LA PLAZA' in h_norm or 'FECHA ING PLAZA' in h_norm or 'ING A LA PLAZA' in h_norm or 'FECHA_INGRESO_PLAZA' in h_norm or 'FECHA INGRESO PLAZA' in h_norm:
                col_mapping['fecha_ingreso_plaza'] = idx
                # print(f"fecha_ingreso_plaza {idx}")
            elif 'VIG. DE INICIO' in h_norm or 'INICIO MOV' in h_norm:
                col_mapping['vig_inicio_mov'] = idx
                # print(f"vig_inicio_mov {idx}")
            elif 'VIG. DE TERMINO' in h_norm or 'TERMINO MOV' in h_norm:
                col_mapping['vig_termino_mov'] = idx
                # print(f"vig_termino_mov {idx}")
            elif 'PUESTO' in h_norm:
                col_mapping['puesto_especifico'] = idx
            elif 'SUELDO BRUTO' in h_norm:
                col_mapping['sueldo_bruto'] = idx
            elif 'SUELDO NETO' in h_norm:
                col_mapping['sueldo_neto'] = idx
            elif 'JEFE' in h_norm:
                col_mapping['jefe_oficina'] = idx
            elif 'LUGAR' in h_norm or 'INMUEBLE' in h_norm or 'UBICACION' in h_norm or 'ASIGNADO' in h_norm:
                col_mapping['lugar_asignado'] = idx
                
        if 'codigo_plaza' not in col_mapping:
            if is_json:
                from django.http import JsonResponse
                return JsonResponse({'status': 'error', 'message': 'No se encontró la columna de Código de Plaza en el Excel.'}, status=400)
            messages.error(request, "No se encontró la columna de Código de Plaza en el Excel.")
            return redirect('personal_list')
            
        # Helpers para limpiar y parsear datos de celda
        def parse_str(val):
            if val is None: return None
            val_s = str(val).strip()
            if val_s in ('...', '..', '-', '', 'None'): return None
            return val_s
            
        def parse_date(val):
            if not val: return None
            if isinstance(val, (datetime, date)):
                return val if isinstance(val, date) else val.date()
            val_s = str(val).strip()
            if val_s in ('...', '..', '-', '', 'None'): return None
            if 'T' in val_s:
                val_s = val_s.split('T')[0]
            for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    return datetime.strptime(val_s, fmt).date()
                except ValueError:
                    continue
            return None
            
        def parse_float(val):
            if val is None: return None
            val_s = str(val).strip()
            if val_s in ('...', '..', '-', '', 'None'): return None
            try:
                return float(val_s)
            except ValueError:
                return None

        # Pre-cargar Estados para optimizar
        estados_dict = {normalizar_nombre(e.nombre): e for e in Estado.objects.all()}
        
        def find_estado(adscripcion_val):
            if not adscripcion_val: return None
            clean = adscripcion_val.strip().upper()
            for prefix in ["O.R. ", "O.R.", "OR.", "OR "]:
                if clean.startswith(prefix):
                    clean = clean[len(prefix):].strip()
            if clean in ("DISTRITO FEDERAL", "DF"):
                clean = "CIUDAD DE MEXICO"
            clean_norm = normalizar_nombre(clean)
            if clean_norm in estados_dict:
                return estados_dict[clean_norm]
            for est_norm, est in estados_dict.items():
                if est_norm in clean_norm or clean_norm in est_norm:
                    return est
            return None

        # Pre-cargar Estatus y Plazas para optimizar y validar
        estatus_dict = {normalizar_nombre(ep.estatus): ep for ep in EstatusPersonal.objects.all()}
        plazas_dict = {normalizar_nombre(tp.plazaT): tp for tp in TipoPlaza.objects.all()}
        
        lineas_estatus_incorrectas = []
        lineas_plaza_incorrectas = []

        creados = 0
        actualizados = 0
        omitidos = 0
        
        for idx_row, row_vals in enumerate(rows_data):
            
            def get_cell_val(key):
                if key in col_mapping:
                    idx = col_mapping[key]
                    if idx < len(row_vals):
                        return row_vals[idx]
                return None
                
            codigo_plaza = parse_str(get_cell_val('codigo_plaza'))
            if not codigo_plaza:
                continue
                
            codigo_plaza = codigo_plaza.upper()
            
            adscripcion = parse_str(get_cell_val('adscripcion'))
            row_estado = find_estado(adscripcion)
            
            if user_state:
                if row_estado and row_estado != user_state:
                    omitidos += 1
                    continue
                row_estado = user_state
            
            if not row_estado:
                omitidos += 1
                continue
                
            # Estatus matching database EstatusPersonal
            estatus_val = None
            estatus_raw = parse_str(get_cell_val('status'))
            if estatus_raw:
                estatus_norm = normalizar_nombre(estatus_raw)
                if estatus_norm in estatus_dict:
                    estatus_val = estatus_dict[estatus_norm]
                else:
                    sheet_row_num = (header_row_idx + 1 + idx_row) if not is_json else (idx_row + 2)
                    lineas_estatus_incorrectas.append(sheet_row_num)
                    
            # Tipo Plaza matching database TipoPlaza
            tipo_plaza_val = None
            tipo_plaza_raw = parse_str(get_cell_val('tipo_plaza'))
            if tipo_plaza_raw:
                tipo_plaza_norm = normalizar_nombre(tipo_plaza_raw)
                if tipo_plaza_norm in plazas_dict:
                    tipo_plaza_val = plazas_dict[tipo_plaza_norm]
                else:
                    sheet_row_num = (header_row_idx + 1 + idx_row) if not is_json else (idx_row + 2)
                    lineas_plaza_incorrectas.append(sheet_row_num)
                
            nivel = parse_str(get_cell_val('nivel')) or 'N/A'
            nivel = nivel.upper()
            
            puesto_especifico = parse_str(get_cell_val('puesto_especifico')) or 'SIN PUESTO'
            puesto_especifico = normalizar_nombre(puesto_especifico)
            
            sueldo_bruto = parse_float(get_cell_val('sueldo_bruto'))
            sueldo_neto = parse_float(get_cell_val('sueldo_neto'))
            
            num_empleado = None
            nombre = None
            apellido = None
            curp = None
            fecha_nacimiento = None
            sexo = None
            tipo_movimiento = None
            fecha_ingreso_inm = None
            fecha_ingreso_plaza = None
            vig_inicio_mov = None
            vig_termino_mov = None
            
            is_vacant = (estatus_val and estatus_val.estatus.strip().upper() == 'VACANTE')
            if not is_vacant:
                num_empleado = parse_str(get_cell_val('num_empleado'))
                
                full_name = parse_str(get_cell_val('nombre'))
                if full_name:
                    parts = full_name.split()
                    if len(parts) >= 3:
                        apellido = " ".join(parts[:2])
                        nombre = " ".join(parts[2:])
                    elif len(parts) == 2:
                        apellido = parts[0]
                        nombre = parts[1]
                    else:
                        nombre = full_name
                        apellido = ""
                    nombre = normalizar_nombre(nombre)
                    apellido = normalizar_nombre(apellido)
                
                curp = parse_str(get_cell_val('curp'))
                if curp:
                    curp = curp.strip().upper()
                    fecha_nacimiento, sexo = parse_curp_details(curp)

                mov_raw = parse_str(get_cell_val('tipo_movimiento'))
                if mov_raw:
                    mov_norm = mov_raw.strip().upper()
                    if mov_norm == 'DEFINITIVO':
                        tipo_movimiento = True
                    elif mov_norm == 'INTERINO':
                        tipo_movimiento = False
                    else:
                        tipo_movimiento = None
                else:
                    tipo_movimiento = None
                    
                fecha_ingreso_inm = parse_date(get_cell_val('fecha_ingreso_inm'))
                fecha_ingreso_plaza = parse_date(get_cell_val('fecha_ingreso_plaza'))
                vig_inicio_mov = parse_date(get_cell_val('vig_inicio_mov'))
                vig_termino_mov = parse_date(get_cell_val('vig_termino_mov'))
                
            # jefe_oficina se mantiene siempre en False por solicitud del usuario
            jefe_oficina = False
            
            # lugar_asignado (Inmueble) se mantiene siempre en None (blank) por solicitud del usuario
            lugar_asignado = None

            defaults = {
                'estado': row_estado,
                'estatus': estatus_val,
                'tipo_plaza': tipo_plaza_val,
                'nivel': nivel,
                'num_empleado': num_empleado,
                'nombre': nombre,
                'apellido': apellido,
                'curp': curp,
                'fecha_nacimiento': fecha_nacimiento,
                'sexo': sexo,
                'tipo_movimiento': tipo_movimiento,
                'fecha_ingreso_inm': fecha_ingreso_inm,
                'fecha_ingreso_plaza': fecha_ingreso_plaza,
                'vig_inicio_mov': vig_inicio_mov,
                'vig_termino_mov': vig_termino_mov,
                'puesto_especifico': puesto_especifico,
                'sueldo_bruto': sueldo_bruto,
                'sueldo_neto': sueldo_neto,
                'jefe_oficina': jefe_oficina,
                'lugar_asignado': lugar_asignado,
            }
            
            personal_obj, created = PersonalINM.objects.update_or_create(
                codigo_plaza=codigo_plaza,
                defaults=defaults
            )
            
            if created:
                creados += 1
            else:
                actualizados += 1
                
        if is_json:
            from django.http import JsonResponse
            return JsonResponse({
                'status': 'success',
                'creados': creados,
                'actualizados': actualizados,
                'omitidos': omitidos,
                'lineas_estatus_incorrectas': lineas_estatus_incorrectas,
                'lineas_plaza_incorrectas': lineas_plaza_incorrectas,
            })
            
        msg = f"Carga rápida completada: {creados} creados, {actualizados} actualizados, {omitidos} omitidos."
        messages.success(request, msg)
        if lineas_estatus_incorrectas:
            messages.warning(request, f"Líneas con Estatus incorrecto (se guardaron como NULL): {', '.join(map(str, lineas_estatus_incorrectas))}")
        if lineas_plaza_incorrectas:
            messages.warning(request, f"Líneas con Tipo de Plaza incorrecto (se guardaron como NULL): {', '.join(map(str, lineas_plaza_incorrectas))}")
            
    except Exception as e:
        if is_json:
            from django.http import JsonResponse
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        messages.error(request, f"Error al procesar el archivo Excel: {str(e)}")
        
    return redirect('personal_list')


@transaction.atomic
def guardar_titular(request):
    if request.method != 'POST':
        return redirect('carga_datos')
    
    user_state = get_user_state(request)
    
    try:
        curp = request.POST.get('curp', '').strip().upper()
        if not curp:
            raise ValueError("La CURP es obligatoria.")
        
        # Validación de seguridad: el estado enviado debe coincidir con el del usuario
        estado_id = request.POST.get('estado_id')
        if user_state and str(user_state.id) != str(estado_id):
            raise PermissionError("No tienes permisos para registrar titulares en este estado.")

        titular, created = Titular.objects.update_or_create(
            curp=curp,
            defaults={
                'nombre': normalizar_nombre(request.POST.get('nombre', '')),
                'apellido_paterno': normalizar_nombre(request.POST.get('apellido_paterno', '')),
                'apellido_materno': normalizar_nombre(request.POST.get('apellido_materno', '')),
                'fecha_nacimiento': request.POST.get('fecha_nacimiento'),
                'sexo': request.POST.get('sexo'),
                'nivel': request.POST.get('nivel', '').strip().upper(),
                'codigo_plaza': request.POST.get('codigo_plaza', '').strip().upper() or None,
                'tipo_nombramiento_id': request.POST.get('tipo_nombramiento_id') or None,
                'procedencia_id': request.POST.get('procedencia_id') or None,
                'estado_id': request.POST.get('estado_id'),
                'nacionalidad_id': request.POST.get('nacionalidad_id') or None,
            }
        )

        if 'fotografia' in request.FILES:
            titular.fotografia = request.FILES['fotografia']
            titular.save()
        
        # Procesar imagen recortada (Base64)
        cropped_data = request.POST.get('cropped_image')
        if cropped_data and ';base64,' in cropped_data:
            format, imgstr = cropped_data.split(';base64,')
            ext = format.split('/')[-1]
            data = ContentFile(base64.b64decode(imgstr), name=f"titular_{titular.curp}.{ext}")
            titular.fotografia = data
            titular.save()

        # 2. Teléfonos
        tels_tipo = request.POST.getlist('tel_tipo[]')
        tels_num = request.POST.getlist('tel_numero[]')
        TelefonoTitular.objects.filter(titular=titular).delete()
        for t, n in zip(tels_tipo, tels_num):
            if n.strip():
                TelefonoTitular.objects.create(titular=titular, tipo=t, numero=n.strip())

        # 3. Correos
        emails_tipo = request.POST.getlist('email_tipo[]')
        emails_val = request.POST.getlist('email_valor[]')
        CorreoTitular.objects.filter(titular=titular).delete()
        for t, v in zip(emails_tipo, emails_val):
            if v.strip():
                CorreoTitular.objects.create(titular=titular, tipo=t, correo=v.strip())

        # 4. Estudios
        edu_grados = request.POST.getlist('edu_grado[]')
        edu_carreras = request.POST.getlist('edu_carrera[]')
        Estudio.objects.filter(titular=titular).delete()
        for g, c in zip(edu_grados, edu_carreras):
            if c.strip():
                Estudio.objects.create(titular=titular, grado_id=g, carrera=normalizar_nombre(c))

        # 5. Trayectoria Institucional (INM)
        tray_puestos = request.POST.getlist('tray_puesto[]')
        tray_areas = request.POST.getlist('tray_area[]')
        tray_inicios = request.POST.getlist('tray_inicio[]')
        tray_fins = request.POST.getlist('tray_fin[]')
        tray_actual_idx = request.POST.get('tray_actual_index')
        
        TrayectoriaLaboral.objects.filter(titular=titular).delete()
        for idx, (p, a, i, f) in enumerate(zip(tray_puestos, tray_areas, tray_inicios, tray_fins)):
            if p.strip() and a.strip() and i:
                is_actual = str(idx) == tray_actual_idx
                TrayectoriaLaboral.objects.create(
                    titular=titular, 
                    puesto=normalizar_nombre(p), 
                    area=normalizar_nombre(a), 
                    fecha_inicio=i,
                    fecha_fin=f if f else None,
                    actual=is_actual
                )

        # 6. Experiencia Profesional Previa
        exp_insts = request.POST.getlist('exp_inst[]')
        exp_cargos = request.POST.getlist('exp_cargo[]')
        exp_inicios = request.POST.getlist('exp_inicio[]')
        exp_fins = request.POST.getlist('exp_fin[]')
        exp_descs = request.POST.getlist('exp_desc[]')
        ExperienciaProfesional.objects.filter(titular=titular).delete()
        for inst, cargo, ini, fin, desc in zip(exp_insts, exp_cargos, exp_inicios, exp_fins, exp_descs):
            if inst.strip() and cargo.strip() and ini:
                ExperienciaProfesional.objects.create(
                    titular=titular,
                    institucion=normalizar_nombre(inst),
                    cargo=normalizar_nombre(cargo),
                    fecha_inicio=ini,
                    fecha_fin=fin if fin else None,
                    descripcion=desc.strip()
                )

        messages.success(request, f"Expediente de {titular.nombre} {'creado' if created else 'actualizado'} correctamente.")
    except Exception as e:
        messages.error(request, f"Error al guardar expediente: {str(e)}")
    
    return redirect('titulares_list')

def api_periodo_custom(request):
    """API para obtener datos en un rango de fechas personalizado."""
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Acceso denegado'}, status=403)

    start_str = request.GET.get('start')
    end_str = request.GET.get('end')

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()

        # Restricciones de Rango
        MIN_DATE = date(2024, 10, 1)
        MAX_DATE = get_global_update_date() or date.today()

        if start_date < MIN_DATE:
            start_date = MIN_DATE
        if end_date > MAX_DATE:
            end_date = MAX_DATE
        if start_date > end_date:
            start_date = end_date

        # Obtener datos
        totals_custom = get_totals_by_period(start_date, end_date)
        national_custom = calc_national(totals_custom)

        return JsonResponse({
            'status': 'success',
            'data': totals_custom,
            'national': national_custom
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def api_nacionalidad_ranking(request):
    """API para obtener el ranking de nacionalidades por estado y métrica."""
    if not request.user.is_superuser:
        return JsonResponse({'status': 'error', 'message': 'Acceso denegado'}, status=403)

    estado_norm = request.GET.get('estado')
    metric = request.GET.get('metric')
    start_str = request.GET.get('start')
    end_str = request.GET.get('end')

    if not all([estado_norm, metric, start_str, end_str]):
        return JsonResponse({'status': 'error', 'message': 'Faltan parámetros'}, status=400)

    try:
        start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_str, '%Y-%m-%d').date()

        # Mapeo de métricas a modelos y campos de total
        metric_map = {
            'repatriados': (Repatriados, 'mex_rep'),
            'recibidos': (Recibidos, 'ext_rec'),
            'rescatados': (ExtRescatados, 'rescatados'),
            'ingresos': (Ingresos, 'ingresos_total'),
            'tramites': (Tramites, 'total_documentos'),
            'retornados': (Retornados, 'retornados_total'),
            'inadmitidos': (Inadmitidos, 'inadmitidos_total'),
        }

        if metric not in metric_map:
            return JsonResponse({'status': 'error', 'message': 'Métrica no válida'}, status=400)

        model_class, total_field = metric_map[metric]
        
        # Buscar el estado por nombre normalizado
        estados = Estado.objects.all()
        target_estado = None
        for edo in estados:
            if normalizar_nombre(edo.nombre) == estado_norm:
                target_estado = edo
                break
        
        if not target_estado:
            return JsonResponse({'status': 'error', 'message': 'Estado no encontrado'}, status=404)

        # Agregación por Nacionalidad
        ranking = model_class.objects.filter(
            estado=target_estado,
            fecha__range=[start_date, end_date]
        ).values('nacionalidad__nombre').annotate(
            total=Sum(total_field)
        ).order_by('-total')[:12] # Top 12 nacionalidades

        data = []
        for item in ranking:
            nombre_nac = item['nacionalidad__nombre']
            if not nombre_nac: continue # Saltar si no hay nacionalidad

            data.append({
                'name': nombre_nac,
                'value': int(item['total'] or 0)
            })

        return JsonResponse({
            'status': 'success',
            'data': data,
            'state_name': target_estado.nombre
        })

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def api_reporte_nacionalidades(request):
    """API para obtener el ranking de nacionalidades por rubro en un periodo."""
    rubro = request.GET.get('rubro', 'Encuentros')
    start_str = request.GET.get('start')
    end_str = request.GET.get('end')

    if not all([start_str, end_str]):
        return JsonResponse({'status': 'error', 'message': 'Faltan fechas'}, status=400)

    try:
        # Convertir timestamps de JS (ms) o strings ISO a date
        if start_str.isdigit():
            start_date = datetime.fromtimestamp(int(start_str)/1000.0).date()
            end_date = datetime.fromtimestamp(int(end_str)/1000.0).date()
        else:
            start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_str, '%Y-%m-%d').date()

        data = []
        if rubro == 'Encuentros':
            ranking = Encuentros.objects.filter(
                fecha__range=[start_date, end_date]
            ).values('nacionalidad__nombre').annotate(
                total=Sum('encuentros_total')
            ).order_by('-total')[:10]
        elif rubro == 'Rescatados':
            ranking = ExtRescatados.objects.filter(
                fecha__range=[start_date, end_date]
            ).values('nacionalidad__nombre').annotate(
                total=Sum('rescatados')
            ).order_by('-total')[:10]
        else: # Recibidos
            # Combinar total de Repatriados (México) con el ranking de Recibidos
            res_rec = Recibidos.objects.filter(
                fecha__range=[start_date, end_date]
            ).values('nacionalidad__nombre').annotate(
                total=Sum('ext_rec')
            ).order_by('-total')[:10]
            
            # Obtener total de mexicanos
            mex_total = Repatriados.objects.filter(
                fecha__range=[start_date, end_date]
            ).aggregate(total=Sum('mex_rep'))['total'] or 0
            
            data.append({'name': 'MÉXICO', 'value': int(mex_total)})
            for item in res_rec:
                data.append({'name': item['nacionalidad__nombre'], 'value': int(item['total'] or 0)})
            
            # Ordenar de nuevo por si algún país superó a México (poco probable pero posible)
            data = sorted(data, key=lambda x: x['value'], reverse=True)[:10]
            return JsonResponse({'status': 'success', 'data': data})

        for item in ranking:
            data.append({
                'name': item['nacionalidad__nombre'] or 'DESCONOCIDA',
                'value': int(item['total'] or 0)
            })

        return JsonResponse({'status': 'success', 'data': data})

    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)

def carga_datos_batch(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'No autenticado'}, status=401)
        
    user_state = get_user_state(request)
    if not request.user.is_superuser and not user_state:
        return JsonResponse({'status': 'error', 'message': 'Acceso denegado'}, status=403)

    if request.method == 'POST':
        try:
            import json
            payload = json.loads(request.body)
            model_name = payload.get('model_name')
            rows = payload.get('data', [])

            models_available = {
                'Repatriados': Repatriados,
                'Recibidos': Recibidos,
                'Extranjeros Rescatados': ExtRescatados,
                'Ingresos': Ingresos,
                'Tramites': Tramites,
                'Retornados': Retornados,
                'Inadmitidos': Inadmitidos,
                'Encuentros': Encuentros,
            }

            if model_name not in models_available:
                return JsonResponse({'status': 'error', 'message': 'Modelo no válido'}, status=400)

            model_class = models_available[model_name]
            
            # Pre-cargar catálogos en memoria
            estados_dict = {normalizar_nombre(e.nombre): e for e in Estado.objects.all()}
            nacionalidades_dict = {normalizar_nombre(n.nombre): n for n in Nacionalidad.objects.all()}
            
            rows_created = 0
            rows_updated = 0
            errors = []
            
            # Campos del modelo a poblar dinámicamente
            fields_to_populate = [f.name for f in model_class._meta.fields if f.name not in ['id', 'fecha', 'estado', 'nacionalidad']]

            with transaction.atomic():
                for row_idx, row in enumerate(rows):
                    if not any(row): continue
                    
                    fecha_val = row[0]
                    if model_name == 'Encuentros':
                        # Formato: Fecha(0), Agencia(1), CiudadEU(2), EstadoEU(3), EstadoMex(4), Nacionalidad(5), Total(6)
                        agencia_raw = row[1]
                        ciudad_eu_raw = row[2]
                        estado_eu_raw = row[3]
                        estado_nombre = row[4]
                        nacionalidad_nombre = row[5]
                        
                        agencia_norm = normalizar_nombre(agencia_raw)
                        ciudad_eu_norm = normalizar_nombre(ciudad_eu_raw)
                        estado_eu_norm = normalizar_nombre(estado_eu_raw)
                        
                        nacionalidad_norm = normalizar_nombre(nacionalidad_nombre)
                        nac_obj = nacionalidades_dict.get(nacionalidad_norm)
                        
                        if not nac_obj:
                            errors.append(f"Reg {row_idx}: Nacionalidad '{nacionalidad_nombre}' no existe.")
                            continue
                        
                        estado_norm = normalizar_nombre(estado_nombre)
                        estado_obj = estados_dict.get(estado_norm)
                        if not estado_obj:
                            errors.append(f"Reg {row_idx}: Estado '{estado_nombre}' no encontrado.")
                            continue
                        if user_state and estado_obj != user_state:
                            errors.append(f"Reg {row_idx}: No tiene permisos para modificar datos del estado '{estado_nombre}'.")
                            continue

                        data_dict = {'encuentros_total': row[6] if len(row) > 6 else 0}

                        obj, created = Encuentros.objects.update_or_create(
                            fecha=fecha_val,
                            agencia=agencia_norm,
                            ciudadEU=ciudad_eu_norm,
                            estadoEU=estado_eu_norm,
                            estado=estado_obj,
                            nacionalidad=nac_obj,
                            defaults=data_dict
                        )
                    else:
                        estado_nombre = row[1]
                        nacionalidad_nombre = row[2]
                        
                        # Normalizar nacionalidad
                        nacionalidad_norm = normalizar_nombre(nacionalidad_nombre)
                        nac_obj = nacionalidades_dict.get(nacionalidad_norm)
                        
                        if not nac_obj:
                            errors.append(f"Reg {row_idx}: Nacionalidad '{nacionalidad_nombre}' no existe.")
                            continue

                        # Normalizar estado
                        estado_norm_busqueda = normalizar_nombre(estado_nombre)
                        estado_obj = estados_dict.get(estado_norm_busqueda)
                        
                        if not estado_obj:
                            errors.append(f"Reg {row_idx}: Estado '{estado_nombre}' no encontrado.")
                            continue
                        if user_state and estado_obj != user_state:
                            errors.append(f"Reg {row_idx}: No tiene permisos para modificar datos del estado '{estado_nombre}'.")
                            continue

                        # Preparar datos
                        data_dict = {}
                        for i, field_name in enumerate(fields_to_populate):
                            excel_idx = i + 3
                            val = row[excel_idx] if excel_idx < len(row) else 0
                            data_dict[field_name] = val if val is not None else 0

                        obj, created = model_class.objects.update_or_create(
                            fecha=fecha_val,
                            estado=estado_obj,
                            nacionalidad=nac_obj,
                            defaults=data_dict
                        )
                    
                    if created:
                        rows_created += 1
                    else:
                        rows_updated += 1

            return JsonResponse({
                'status': 'success',
                'created': rows_created,
                'updated': rows_updated,
                'errors': errors
            })

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)

def carga_nacionalidades(request):
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, "Por favor seleccione un archivo Excel.")
            return redirect('carga_datos')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            created_count = 0
            for row in sheet.iter_rows(min_row=1, values_only=True):
                nombre_raw = row[0]
                if not nombre_raw: continue
                
                nombre_norm = normalizar_nombre(str(nombre_raw))
                if nombre_norm:
                    obj, created = Nacionalidad.objects.get_or_create(nombre=nombre_norm)
                    if created:
                        created_count += 1
            
            messages.success(request, f"Catálogo actualizado. Se agregaron {created_count} nuevas nacionalidades.")
        except Exception as e:
            messages.error(request, f"Error al procesar el catálogo: {str(e)}")
            
    return redirect('carga_datos')

def carga_oficinas(request):
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    if request.method == 'POST':
        model_name = request.POST.get('model_name')
        excel_file = request.FILES.get('excel_file')
        tipo_punto = request.POST.get('tipo', '').upper()

        if not model_name or not excel_file:
            messages.error(request, "Por favor seleccione un modelo y un archivo.")
            return redirect('carga_datos')

        try:
            wb = openpyxl.load_workbook(excel_file)
            sheet = wb.active
            
            # Pre-cargar catálogos para eficiencia
            estados_dict = {normalizar_nombre(e.nombre): e for e in Estado.objects.all()}
            tipos_prh_dict = {normalizar_nombre(t.nombre): t for t in TipoPRH.objects.all()}
            
            created_count = 0
            updated_count = 0
            updated_rows = []
            errors = []

            with transaction.atomic():
                for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row): continue
                    
                    estado_raw = row[0]
                    estado_norm = normalizar_nombre(estado_raw)
                    estado_obj = estados_dict.get(estado_norm)

                    if not estado_obj:
                        errors.append(f"Fila {row_idx}: Estado '{estado_raw}' no encontrado.")
                        continue

                    if model_name == 'PuntosInternacionEstacion':
                        if not tipo_punto:
                            messages.error(request, "Debe seleccionar un tipo para Puntos de Internación.")
                            return redirect('carga_datos')
                        
                        nombre_raw = row[1]
                        lat = row[2]
                        lon = row[3]
                        
                        # Normalizar Nombre
                        nombre_norm = normalizar_nombre(nombre_raw)

                        obj, created = PuntosInternacionEstacion.objects.update_or_create(
                            nombre=nombre_norm,
                            tipo=tipo_punto,
                            defaults={
                                'estado': estado_obj,
                                'latitud': float(lat) if lat is not None else 0.0,
                                'longitud': float(lon) if lon is not None else 0.0
                            }
                        )
                        if created: 
                            created_count += 1
                        else: 
                            updated_count += 1
                            updated_rows.append(str(row_idx))

                    elif model_name == 'CatalogoOR':
                        titular_raw = row[1]
                        domicilio_raw = row[2]
                        correo_raw = row[3]

                        # Normalizar campos de texto
                        titular_norm = normalizar_nombre(titular_raw)
                        domicilio_norm = normalizar_nombre(domicilio_raw)
                        correo_norm = str(correo_raw).strip().lower() # Correo se mantiene con formato, pero sin espacios

                        obj, created = CatalogoOR.objects.update_or_create(
                            titular=titular_norm,
                            defaults={
                                'estado': estado_obj,
                                'domicilio': domicilio_norm,
                                'correo': correo_norm
                            }
                        )
                        if created: 
                            created_count += 1
                        else: 
                            updated_count += 1
                            updated_rows.append(str(row_idx))

                    elif model_name == 'PRHs':
                        # Formato: Estado(0), Nombre(1), Modalidad(2), Activo(3), Coordenadas(4), Lat(5), Lon(6)
                        if len(row) < 7:
                            errors.append(f"Fila {row_idx}: Faltan columnas (se requieren 7).")
                            continue

                        nombre_raw = row[1]
                        modalidad_raw = row[2]
                        activo_raw = str(row[3]).strip().upper()
                        coordenadas_raw = row[4]
                        lat = row[5]
                        lon = row[6]

                        # Normalizar campos
                        nombre_norm = normalizar_nombre(nombre_raw)
                        modalidad_norm = normalizar_nombre(modalidad_raw)
                        coordenadas_norm = normalizar_nombre(coordenadas_raw)

                        # Validar Modalidad
                        tipo_obj = tipos_prh_dict.get(modalidad_norm)
                        if not tipo_obj:
                            errors.append(f"Fila {row_idx}: Modalidad '{modalidad_raw}' no existe en el catálogo.")
                            continue

                        # Convertir Activo a Boolean
                        activo_bool = True if activo_raw == "ACTIVO" else False

                        obj, created = PRHs.objects.update_or_create(
                            nombre=nombre_norm,
                            estado=estado_obj,
                            modalidad=tipo_obj,
                            defaults={
                                'activo': activo_bool,
                                'coordenadasTexto': coordenadas_norm,
                                'latitud': float(lat) if lat is not None else 0.0,
                                'longitud': float(lon) if lon is not None else 0.0
                            }
                        )
                        if created: 
                            created_count += 1
                        else: 
                            updated_count += 1
                            updated_rows.append(str(row_idx))

            if errors:
                for err in errors: messages.warning(request, err)
            
            res_msg = f"Carga de oficinas completada. Creados: {created_count}, Actualizados: {updated_count}"
            if updated_rows:
                res_msg += f" (Filas: {', '.join(updated_rows)})"
            
            messages.success(request, res_msg)
        except Exception as e:
            messages.error(request, f"Error al procesar el archivo: {str(e)}")
            
    return redirect('carga_datos')

def reportes(request):
    """Vista para el tablero de reportes con datos reales y gráficas Bokeh."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    rubro = request.GET.get('rubro', 'Encuentros')
    fecha_max = get_global_update_date() or date.today()
    
    # Periodos Definidos
    CSP_START = date(2024, 10, 1)
    TRUMP_START = date(2025, 1, 20)
    SEMANA_START = fecha_max - timedelta(days=6)
    INICIO_2026 = date(2026, 1, 1)

    # Función auxiliar para obtener totales
    def get_data(start, end):
        if rubro == 'Encuentros':
            val = Encuentros.objects.filter(fecha__range=[start, end]).aggregate(t=Sum('encuentros_total'))['t'] or 0
            return val, None, None
        elif rubro == 'Rescatados':
            val = ExtRescatados.objects.filter(fecha__range=[start, end]).aggregate(t=Sum('rescatados'))['t'] or 0
            return val, None, None
        elif rubro == 'Recibidos':
            mex = Repatriados.objects.filter(fecha__range=[start, end]).aggregate(t=Sum('mex_rep'))['t'] or 0
            ext = Recibidos.objects.filter(fecha__range=[start, end]).aggregate(t=Sum('ext_rec'))['t'] or 0
            return mex + ext, mex, ext
        return 0, 0, 0

    # Texto del rubro
    rubro_text = "Encuentros" if rubro == 'Encuentros' else ("Recibidos" if rubro == 'Recibidos' else "Rescatados")

    # Cálculos para tarjetas
    def build_card(start, end, label):
        if rubro == 'Encuentros':
            qs_period = Encuentros.objects.filter(fecha__range=[start, end])
            total = qs_period.aggregate(total=Sum('encuentros_total'))['total'] or 0
            
            # Intentar identificar mexicanos por nombre (MEXICO, MEXICA, etc.)
            mex = qs_period.filter(
                Q(nacionalidad__nombre__icontains='MEXICO') | 
                Q(nacionalidad__nombre__icontains='MEXICA')
            ).aggregate(total=Sum('encuentros_total'))['total'] or 0
            ext = total - mex
            
            p_mex = round((mex / total * 100)) if total > 0 else 0
            p_ext = 100 - p_mex if total > 0 else 0
        elif rubro == 'Rescatados':
            total = ExtRescatados.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('rescatados'))['total'] or 0
            mex, ext, p_mex, p_ext = 0, total, 0, 100 # Rescatados suele ser solo para extranjeros en este modelo
        else: # Recibidos
            mex = Repatriados.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('mex_rep'))['total'] or 0
            ext = Recibidos.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('ext_rec'))['total'] or 0
            total = mex + ext
            p_mex = round((mex / total * 100)) if total > 0 else 0
            p_ext = 100 - p_mex if total > 0 else 0

        days = (end - start).days + 1
        avg = round(total / days) if days > 0 else 0
        
        # Formatear periodo
        if label == "Semana":
            fmt_periodo = f"del {start.strftime('%d')} al {end.strftime('%d de %b.')} de {end.year}"
        else:
            fmt_periodo = f"{start.strftime('%d/%b/%y')} a {end.strftime('%d/%b/%y')}"

        return {
            'total': f"{total:,}",
            'promedio': f"{avg:,}",
            'periodo': fmt_periodo,
            'subtitulo': f"{rubro}",
            'mex': f"{mex:,}",
            'ext': f"{ext:,}",
            'p_mex': p_mex,
            'p_ext': p_ext
        }

    card_semana = build_card(SEMANA_START, fecha_max, "Semana")
    card_csp = build_card(CSP_START, fecha_max, "CSP")
    card_trump = build_card(TRUMP_START, fecha_max, "Trump")

    # --- GRÁFICA DE BARRAS (EVOLUCIÓN DIARIA) ---
    # Obtenemos datos agrupados por día
    if rubro == 'Encuentros':
        qs = Encuentros.objects.filter(fecha__range=[CSP_START, fecha_max]) \
            .annotate(day=TruncDay('fecha')) \
            .values('day') \
            .annotate(total=Sum('encuentros_total')) \
            .order_by('day')
    elif rubro == 'Rescatados':
        qs = ExtRescatados.objects.filter(fecha__range=[CSP_START, fecha_max]) \
            .annotate(day=TruncDay('fecha')) \
            .values('day') \
            .annotate(total=Sum('rescatados')) \
            .order_by('day')
    else: # Recibidos
        rep_qs = Repatriados.objects.filter(fecha__range=[CSP_START, fecha_max]) \
            .annotate(day=TruncDay('fecha')) \
            .values('day') \
            .annotate(total=Sum('mex_rep'))
        rec_qs = Recibidos.objects.filter(fecha__range=[CSP_START, fecha_max]) \
            .annotate(day=TruncDay('fecha')) \
            .values('day') \
            .annotate(total=Sum('ext_rec'))
        
        combined = {}
        for item in rep_qs:
            combined[item['day']] = item['total']
        for item in rec_qs:
            combined[item['day']] = combined.get(item['day'], 0) + item['total']
        
        qs = [{'day': d, 'total': combined[d]} for d in sorted(combined.keys())]

    x_data = []
    y_data = []
    for d in qs:
        day_val = d['day'] if isinstance(d, dict) else d.day
        if isinstance(day_val, date) and not isinstance(day_val, datetime):
            day_val = datetime.combine(day_val, datetime.min.time())
        x_data.append(day_val)
        y_data.append(d['total'] if isinstance(d, dict) else d.total)

    source_bar = ColumnDataSource(data=dict(x=x_data, y=y_data))

    # Normalizar fechas a datetime para comparaciones seguras
    def to_datetime(d):
        if isinstance(d, date) and not isinstance(d, datetime):
            return datetime.combine(d, datetime.min.time())
        return d

    dt_fecha_max = to_datetime(fecha_max)
    dt_csp_start = to_datetime(CSP_START)
    
    # Determinar rango inicial para p1 (últimos 180 días)
    initial_range_start = dt_fecha_max - timedelta(days=180)
    if x_data and x_data[0] > initial_range_start:
        initial_range_start = x_data[0]
    elif not x_data:
        initial_range_start = dt_csp_start

    p1_opts = {
        'height': 300, 
        'sizing_mode': "stretch_width",
        'x_axis_type': "datetime",
        'x_axis_location': "above",
        'x_range': (initial_range_start, dt_fecha_max),
        'toolbar_location': "right", 
        'tools': "pan,box_zoom,xwheel_zoom,reset,tap",
        'background_fill_color': "#efefef",
        'border_fill_color': None,
        'outline_line_color': "#666666"
    }
    
    p1 = figure(**p1_opts)
    p1.y_range.start = 0
    p1.line(x='x', y='y', line_width=2, color="#285C4D", source=source_bar)
    
    p1.xgrid.grid_line_color = "#ffffff"
    p1.ygrid.grid_line_color = "#ffffff"
    p1.yaxis.visible = True
    p1.yaxis.major_label_text_font_size = "9pt"
    p1.yaxis.formatter = NumeralTickFormatter(format="0a")
    p1.xaxis.major_label_text_font_size = "9pt"
    p1.xaxis.formatter = DatetimeTickFormatter(
        days="%d %b",
        months="%b %Y",
        years="%Y"
    )
    
    hover_bar = HoverTool(tooltips=[("Fecha", "@x{%d/%b/%y}"), ("Valor", "@y{0,0}")], formatters={'@x': 'datetime'})
    p1.add_tools(hover_bar)
    
    # --- GRÁFICA DE SELECCIÓN (NAVIGATOR) ---
    select = figure(
        title="Arrastra el recuadro para navegar por el tiempo",
        height=100, sizing_mode="stretch_width",
        x_axis_type="datetime", y_axis_type=None,
        tools="", toolbar_location=None, 
        background_fill_color="#f9f9f9",
        outline_line_color="#e5e7eb"
    )
    
    select.line(x='x', y='y', color="#285C4D", alpha=0.5, source=source_bar)
    select.ygrid.grid_line_color = None
    select.xgrid.grid_line_color = None
    select.xaxis.major_label_text_font_size = "7pt"
    
    range_tool = RangeTool(x_range=p1.x_range)
    range_tool.overlay.fill_color = "#285C4D"
    range_tool.overlay.fill_alpha = 0.2
    select.add_tools(range_tool)
    
    # Empaquetamos ambas en una columna
    layout_p1 = column(p1, select, sizing_mode="stretch_width")

    tap_bar_js = CustomJS(args=dict(source=source_bar), code="""
        const indices = source.selected.indices;
        if (indices.length > 0) {
            const idx = indices[0];
            const date = new Date(source.data['x'][idx]);
            const period = date.toLocaleDateString('es-MX', {day: 'numeric', month: 'short', year: 'numeric'});
            const val = source.data['y'][idx];
            const valFmt = new Intl.NumberFormat('en-US').format(val);
            if (window.showTouchToast) window.showTouchToast(period, valFmt);
        }
    """)
    source_bar.selected.js_on_change('indices', tap_bar_js)

    # --- GRÁFICA DE LÍNEAS (SEMANAL 2026) ---
    if rubro == 'Encuentros':
        qs_w = Encuentros.objects.filter(fecha__range=[INICIO_2026, fecha_max]) \
            .annotate(week=TruncWeek('fecha')) \
            .values('week') \
            .annotate(total=Sum('encuentros_total')) \
            .order_by('week')
    elif rubro == 'Rescatados':
        qs_w = ExtRescatados.objects.filter(fecha__range=[INICIO_2026, fecha_max]) \
            .annotate(week=TruncWeek('fecha')) \
            .values('week') \
            .annotate(total=Sum('rescatados')) \
            .order_by('week')
    else: # Recibidos
        rep_w = Repatriados.objects.filter(fecha__range=[INICIO_2026, fecha_max]) \
            .annotate(week=TruncWeek('fecha')) \
            .values('week') \
            .annotate(total=Sum('mex_rep'))
        rec_w = Recibidos.objects.filter(fecha__range=[INICIO_2026, fecha_max]) \
            .annotate(week=TruncWeek('fecha')) \
            .values('week') \
            .annotate(total=Sum('ext_rec'))
        
        combined_w = {}
        for item in rep_w: combined_w[item['week']] = item['total']
        for item in rec_w: combined_w[item['week']] = combined_w.get(item['week'], 0) + item['total']
        
        qs_w = [{'week': w, 'total': combined_w[w]} for w in sorted(combined_w.keys())]

    x_line = []
    y_line = []
    for d in qs_w:
        w_date = d['week'] if isinstance(d, dict) else d['week']
        x_line.append(w_date.strftime('%d-%m'))
        y_line.append(d['total'])

    source_line = ColumnDataSource(data=dict(x=x_line, y=y_line))

    # P2 usa FactorRange, si x_line está vacío, Bokeh puede fallar.
    p2_args = {
        'height': 300, 
        'sizing_mode': "stretch_width",
        'toolbar_location': None,
        'tools': "tap",
        'background_fill_color': None,
        'border_fill_color': None,
        'outline_line_color': None
    }
    if x_line:
        p2_args['x_range'] = x_line
    
    p2 = figure(**p2_args)
    p2.line(x='x', y='y', line_width=2, color="#86895D", source=source_line)
    p2.scatter(x='x', y='y', size=8, color="#86895D", fill_color="white", source=source_line)
    
    p2.xgrid.grid_line_color = None
    p2.ygrid.grid_line_color = None
    p2.yaxis.visible = False
    p2.xaxis.major_label_orientation = 1.5708
    p2.xaxis.major_label_text_font_size = "7pt"

    hover_line = HoverTool(tooltips=[("Semana", "@x"), ("Total", "@y{0,0}")])
    p2.add_tools(hover_line)

    # Total 2026 para el footer
    total_2026 = sum(y_line)

    # === GRÁFICA TOP 10 NACIONALIDADES DINÁMICA ===
    start_init = initial_range_start.date() if isinstance(initial_range_start, datetime) else initial_range_start
    end_init = fecha_max
    
    if rubro == 'Encuentros':
        ranking = Encuentros.objects.filter(fecha__range=[start_init, end_init]).values('nacionalidad__nombre').annotate(total=Sum('encuentros_total')).order_by('-total')[:10]
        top_names = [item['nacionalidad__nombre'] or 'OTRO' for item in ranking][::-1]
        top_values = [int(item['total'] or 0) for item in ranking][::-1]
    elif rubro == 'Rescatados':
        ranking = ExtRescatados.objects.filter(fecha__range=[start_init, end_init]).values('nacionalidad__nombre').annotate(total=Sum('rescatados')).order_by('-total')[:10]
        top_names = [item['nacionalidad__nombre'] or 'OTRO' for item in ranking][::-1]
        top_values = [int(item['total'] or 0) for item in ranking][::-1]
    else: # Recibidos
        res_rec = Recibidos.objects.filter(fecha__range=[start_init, end_init]).values('nacionalidad__nombre').annotate(total=Sum('ext_rec')).order_by('-total')[:10]
        mex_total = Repatriados.objects.filter(fecha__range=[start_init, end_init]).aggregate(total=Sum('mex_rep'))['total'] or 0
        r_list = [{'n': 'MÉXICO', 'v': int(mex_total)}]
        for item in res_rec:
            r_list.append({'n': item['nacionalidad__nombre'], 'v': int(item['total'] or 0)})
        r_sorted = sorted(r_list, key=lambda x: x['v'], reverse=True)[:10]
        top_names = [item['n'] for item in r_sorted][::-1]
        top_values = [item['v'] for item in r_sorted][::-1]

    source_top = ColumnDataSource(data=dict(names=top_names, values=top_values))
    p_top = figure(y_range=top_names, height=450, title=None,
                   toolbar_location=None, tools="", sizing_mode="stretch_width")
    p_top.hbar(y='names', right='values', height=0.7, color="#285C4D", source=source_top)
    p_top.x_range.start = 0
    p_top.xaxis.formatter = NumeralTickFormatter(format="0a")
    p_top.outline_line_color = None
    p_top.grid.grid_line_color = None
    p_top.yaxis.major_label_text_font_size = "9pt"
    p_top.yaxis.major_label_text_font_style = "bold"
    
    h_top = HoverTool(tooltips=[("País", "@names"), ("Total", "@values{0,0}")])
    p_top.add_tools(h_top)

    # Callback JS para actualizar Top 10 cuando cambia el rango de p1
    update_top10_js = CustomJS(args=dict(source=source_top, y_range=p_top.y_range, rubro=rubro), code="""
        const start = cb_obj.start;
        const end = cb_obj.end;
        if (window.top10Timeout) clearTimeout(window.top10Timeout);
        window.top10Timeout = setTimeout(() => {
            fetch(`/mapa/api/reporte-nacionalidades?rubro=${rubro}&start=${Math.round(start)}&end=${Math.round(end)}`)
                .then(response => response.json())
                .then(res => {
                    if (res.status === 'success') {
                        const new_names = res.data.map(d => d.name).reverse();
                        const new_values = res.data.map(d => d.value).reverse();
                        source.data = { names: new_names, values: new_values };
                        y_range.factors = new_names;
                        source.change.emit();
                    }
                });
        }, 400);
    """)
    p1.x_range.js_on_change('start', update_top10_js)

    plot_script, plot_divs = components((layout_p1, p2, p_top))
    plot_bar_div, plot_line_div, plot_top_div = plot_divs

    context = {
        'rubro': rubro,
        'card_semana': card_semana,
        'card_csp': card_csp,
        'card_trump': card_trump,
        'plot_script': plot_script,
        'plot_bar_div': plot_bar_div,
        'plot_line_div': plot_line_div,
        'plot_top_div': plot_top_div,
        'total_2026': f"{total_2026:,}",
    }

    return render(request, 'mapa/reportes.html', context)


# @FADAR
# =============================================================================
# Reporte descargable "Mexicanos y Extranjeros" -- mismas reglas de negocio
# EXACTAS que la rama rubro=='Recibidos' de reportes() de arriba (la unica
# parte de ese tablero que es, literalmente, mexicanos vs extranjeros):
#  - Mexicanos = Repatriados.mex_rep (100% nacionalidad MEXICO, verificado).
#  - Extranjeros = Recibidos.ext_rec (80 nacionalidades distintas, ninguna
#    mexicana, verificado).
#  - 3 tarjetas de comparacion con periodos FIJOS (no un rango libre -- son
#    parte de la regla de negocio tal cual esta en el tablero):
#      Semana = ultimos 7 dias con datos, CSP = desde 2024-10-01,
#      Trump = desde 2025-01-20.
#  - Evolucion diaria (desde CSP_START) y semanal (desde 2026-01-01),
#    sumando mex_rep + ext_rec por dia/semana.
#  - Top 10 nacionalidades: se toma el top 10 de Recibidos y se inserta
#    "MEXICO" con el total de Repatriados como una entrada mas, reordenando
#    -- exactamente como en reportes()/api_reporte_nacionalidades.
# El tablero original (reportes()) no tiene boton de descarga -- esa es la
# unica pieza nueva aqui, reutilizando el calculo tal cual.
# =============================================================================

def _build_card_mex_ext(start, end, label):
    mex = Repatriados.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('mex_rep'))['total'] or 0
    ext = Recibidos.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('ext_rec'))['total'] or 0
    total = mex + ext
    p_mex = round((mex / total * 100)) if total > 0 else 0
    p_ext = 100 - p_mex if total > 0 else 0
    days = (end - start).days + 1
    avg = round(total / days) if days > 0 else 0
    if label == "Semana":
        fmt_periodo = f"del {start.strftime('%d')} al {end.strftime('%d de %b.')} de {end.year}"
    else:
        fmt_periodo = f"{start.strftime('%d/%b/%y')} a {end.strftime('%d/%b/%y')}"
    return {
        'total': total, 'total_fmt': f"{total:,}",
        'promedio': f"{avg:,}", 'periodo': fmt_periodo, 'label': label,
        'mex': mex, 'ext': ext, 'mex_fmt': f"{mex:,}", 'ext_fmt': f"{ext:,}",
        'p_mex': p_mex, 'p_ext': p_ext,
    }


def _top10_mex_ext(start, end):
    res_rec = Recibidos.objects.filter(fecha__range=[start, end]) \
        .values('nacionalidad__nombre').annotate(total=Sum('ext_rec')).order_by('-total')[:10]
    mex_total = Repatriados.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('mex_rep'))['total'] or 0
    lista = [{'n': 'MÉXICO', 'v': int(mex_total)}]
    for item in res_rec:
        lista.append({'n': item['nacionalidad__nombre'] or 'DESCONOCIDA', 'v': int(item['total'] or 0)})
    return sorted(lista, key=lambda x: x['v'], reverse=True)[:10]


# @FADAR -- entidades tal como aparecen en FORMATO_MEX-.xlsm ("REPORTE
# GENERAL"), no las 32. Los puntos (12 ciudades fijas) ya no se duplican
# aqui -- vienen de CATALOGO_PUNTOS_MEX_EXT (mapa/models.py), unica fuente
# compartida con RegistroMexExtPunto.
RESCATES_ESTADOS_MEX_EXT = list(CATALOGO_PUNTOS_MEX_EXT.keys())


def _puntos_por_estado():
    return CATALOGO_PUNTOS_MEX_EXT


# @FADAR -- captura real por punto (RegistroMexExtPunto). Hombres+Mujeres
# se suman como "Adultos" y Ninos+Ninas como "Menores" para que la fila del
# punto use las mismas columnas que la fila del estado (Adultos/Menores/
# Total) -- el detalle fino (H/M/Ninos/Ninas) sigue intacto en la tabla,
# solo se resume asi para esta tabulacion.
def _capturas_por_punto(start, end):
    filas = RegistroMexExtPunto.objects.filter(fecha__range=[start, end], estado__in=RESCATES_ESTADOS_MEX_EXT) \
        .values('estado', 'punto', 'categoria') \
        .annotate(hombres=Sum('hombres'), mujeres=Sum('mujeres'), ninos=Sum('ninos'), ninas=Sum('ninas'))
    resultado = {}
    for f in filas:
        clave = (f['estado'], f['punto'])
        adultos = (f['hombres'] or 0) + (f['mujeres'] or 0)
        menores = (f['ninos'] or 0) + (f['ninas'] or 0)
        resultado.setdefault(clave, {})[f['categoria']] = {'ad': adultos, 'me': menores, 'total': adultos + menores}
    return resultado


def _tabla_mex_ext_por_estado(start, end):
    rep = Repatriados.objects.filter(fecha__range=[start, end], estado__nombre__in=RESCATES_ESTADOS_MEX_EXT) \
        .values('estado__nombre').annotate(mex=Sum('mex_rep'), mex_ad=Sum('adultos'), mex_me=Sum('menores'))
    rec = Recibidos.objects.filter(fecha__range=[start, end], estado__nombre__in=RESCATES_ESTADOS_MEX_EXT) \
        .values('estado__nombre').annotate(ext=Sum('ext_rec'), ext_ad=Sum('adultos'), ext_me=Sum('menores'))
    filas = {e: {'estado': e, 'mex': 0, 'mex_ad': 0, 'mex_me': 0, 'ext': 0, 'ext_ad': 0, 'ext_me': 0} for e in RESCATES_ESTADOS_MEX_EXT}
    for r in rep:
        e = r['estado__nombre']
        filas[e]['mex'] = r['mex'] or 0
        filas[e]['mex_ad'] = r['mex_ad'] or 0
        filas[e]['mex_me'] = r['mex_me'] or 0
    for r in rec:
        e = r['estado__nombre']
        filas[e]['ext'] = r['ext'] or 0
        filas[e]['ext_ad'] = r['ext_ad'] or 0
        filas[e]['ext_me'] = r['ext_me'] or 0

    catalogo = _puntos_por_estado()
    capturas = _capturas_por_punto(start, end)
    for f in filas.values():
        f['total'] = f['mex'] + f['ext']
        puntos_estado = []
        for nombre in catalogo.get(f['estado'], []):
            dato = capturas.get((f['estado'], nombre))
            mex_p = dato.get('MEX') if dato else None
            ext_p = dato.get('EXT') if dato else None
            puntos_estado.append({
                'nombre': nombre,
                'con_datos': dato is not None,
                'mex_ad': mex_p['ad'] if mex_p else None, 'mex_me': mex_p['me'] if mex_p else None, 'mex': mex_p['total'] if mex_p else None,
                'ext_ad': ext_p['ad'] if ext_p else None, 'ext_me': ext_p['me'] if ext_p else None, 'ext': ext_p['total'] if ext_p else None,
                'total': (mex_p['total'] if mex_p else 0) + (ext_p['total'] if ext_p else 0) if dato else None,
            })
        f['puntos'] = puntos_estado
    return sorted(filas.values(), key=lambda f: f['total'], reverse=True)


def _tabla_mex_ext_nacionalidades(start, end, limite=15):
    rec = Recibidos.objects.filter(fecha__range=[start, end], estado__nombre__in=RESCATES_ESTADOS_MEX_EXT) \
        .values('nacionalidad__nombre').annotate(ext=Sum('ext_rec'), ext_ad=Sum('adultos'), ext_me=Sum('menores')).order_by('-ext')[:limite]
    return [
        {'nacionalidad': r['nacionalidad__nombre'] or 'DESCONOCIDA', 'ext': r['ext'] or 0,
         'ext_ad': r['ext_ad'] or 0, 'ext_me': r['ext_me'] or 0}
        for r in rec
    ]


# @FADAR
def _sumar_filas_mex_ext(filas, campos):
    total = {c: 0 for c in campos}
    for f in filas:
        for c in campos:
            total[c] += f[c]
    return total


def _leer_rango_mex_ext(request, fecha_max):
    fecha_inicio = request.GET.get('fecha_inicio', (fecha_max - timedelta(days=6)).isoformat())
    fecha_fin = request.GET.get('fecha_fin', fecha_max.isoformat())
    fi = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    ff = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    return fecha_inicio, fecha_fin, fi, ff


# @FADAR
def _datos_mex_extranjeros():
    fecha_max = get_global_update_date() or date.today()
    CSP_START = date(2024, 10, 1)
    TRUMP_START = date(2025, 1, 20)
    SEMANA_START = fecha_max - timedelta(days=6)
    INICIO_2026 = date(2026, 1, 1)

    card_semana = _build_card_mex_ext(SEMANA_START, fecha_max, "Semana")
    card_csp = _build_card_mex_ext(CSP_START, fecha_max, "CSP")
    card_trump = _build_card_mex_ext(TRUMP_START, fecha_max, "Trump")

    # Evolucion diaria (desde CSP_START) -- mex_rep + ext_rec combinados.
    rep_qs = Repatriados.objects.filter(fecha__range=[CSP_START, fecha_max]) \
        .annotate(day=TruncDay('fecha')).values('day').annotate(total=Sum('mex_rep'))
    rec_qs = Recibidos.objects.filter(fecha__range=[CSP_START, fecha_max]) \
        .annotate(day=TruncDay('fecha')).values('day').annotate(total=Sum('ext_rec'))
    combined = {}
    for item in rep_qs:
        combined[item['day']] = item['total']
    for item in rec_qs:
        combined[item['day']] = combined.get(item['day'], 0) + item['total']
    evolucion_diaria = [(d, combined[d]) for d in sorted(combined.keys())]

    # Evolucion semanal 2026.
    rep_w = Repatriados.objects.filter(fecha__range=[INICIO_2026, fecha_max]) \
        .annotate(week=TruncWeek('fecha')).values('week').annotate(total=Sum('mex_rep'))
    rec_w = Recibidos.objects.filter(fecha__range=[INICIO_2026, fecha_max]) \
        .annotate(week=TruncWeek('fecha')).values('week').annotate(total=Sum('ext_rec'))
    combined_w = {}
    for item in rep_w:
        combined_w[item['week']] = item['total']
    for item in rec_w:
        combined_w[item['week']] = combined_w.get(item['week'], 0) + item['total']
    evolucion_semanal = [(w, combined_w[w]) for w in sorted(combined_w.keys())]
    total_2026 = sum(t for _, t in evolucion_semanal)

    start_top = max(CSP_START, fecha_max - timedelta(days=180))
    top_nacionalidades = _top10_mex_ext(start_top, fecha_max)

    return {
        'fecha_max': fecha_max,
        'card_semana': card_semana, 'card_csp': card_csp, 'card_trump': card_trump,
        'evolucion_diaria': evolucion_diaria,
        'evolucion_semanal': evolucion_semanal,
        'total_2026': total_2026,
        'top_nacionalidades': top_nacionalidades,
        'start_top': start_top,
    }


def reporte_mex_extranjeros(request):
    """Vista previa en pantalla -- version descargable del rubro 'Recibidos'
    del tablero /mapa/reportes (que no tiene boton de descarga)."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    datos = _datos_mex_extranjeros()
    fecha_inicio, fecha_fin, fi, ff = _leer_rango_mex_ext(request, datos['fecha_max'])
    card_personalizado = _build_card_mex_ext(fi, ff, "Personalizado")
    top_personalizado = _top10_mex_ext(fi, ff)
    tabla_estados = _tabla_mex_ext_por_estado(fi, ff)
    tabla_nacionalidades = _tabla_mex_ext_nacionalidades(fi, ff)
    total_estados = _sumar_filas_mex_ext(tabla_estados, ['mex_ad', 'mex_me', 'mex', 'ext_ad', 'ext_me', 'ext', 'total'])
    total_nacionalidades = _sumar_filas_mex_ext(tabla_nacionalidades, ['ext_ad', 'ext_me', 'ext'])

    # --- Grafica de evolucion diaria (linea + navegador) ---
    x_data = [datetime.combine(d, datetime.min.time()) for d, _ in datos['evolucion_diaria']]
    y_data = [t for _, t in datos['evolucion_diaria']]
    source_bar = ColumnDataSource(data=dict(x=x_data, y=y_data))
    fecha_max_dt = datetime.combine(datos['fecha_max'], datetime.min.time())
    rango_inicial = datetime.combine(datos['start_top'], datetime.min.time())
    p1 = figure(
        height=300, sizing_mode="stretch_width", x_axis_type="datetime", x_axis_location="above",
        x_range=(rango_inicial, fecha_max_dt), toolbar_location="right", tools="pan,box_zoom,xwheel_zoom,reset",
        background_fill_color="#efefef", border_fill_color=None, outline_line_color="#666666",
        title="Evolución diaria — Mexicanos repatriados + Extranjeros recibidos",
    )
    p1.y_range.start = 0
    p1.line(x='x', y='y', line_width=2, color="#285C4D", source=source_bar)
    p1.xgrid.grid_line_color = "#ffffff"
    p1.ygrid.grid_line_color = "#ffffff"
    p1.yaxis.formatter = NumeralTickFormatter(format="0a")
    p1.xaxis.formatter = DatetimeTickFormatter(days="%d %b", months="%b %Y", years="%Y")
    p1.add_tools(HoverTool(tooltips=[("Fecha", "@x{%d/%b/%y}"), ("Total", "@y{0,0}")], formatters={'@x': 'datetime'}))

    select = figure(
        title="Arrastra el recuadro para navegar por el tiempo", height=100, sizing_mode="stretch_width",
        x_axis_type="datetime", y_axis_type=None, tools="", toolbar_location=None,
        background_fill_color="#f9f9f9", outline_line_color="#e5e7eb",
    )
    select.line(x='x', y='y', color="#285C4D", alpha=0.5, source=source_bar)
    select.ygrid.grid_line_color = None
    select.xgrid.grid_line_color = None
    range_tool = RangeTool(x_range=p1.x_range)
    range_tool.overlay.fill_color = "#285C4D"
    range_tool.overlay.fill_alpha = 0.2
    select.add_tools(range_tool)
    layout_p1 = column(p1, select, sizing_mode="stretch_width")

    # --- Grafica semanal 2026 ---
    x_line = [w.strftime('%d-%m') for w, _ in datos['evolucion_semanal']]
    y_line = [t for _, t in datos['evolucion_semanal']]
    source_line = ColumnDataSource(data=dict(x=x_line, y=y_line))
    p2_args = dict(height=300, sizing_mode="stretch_width", toolbar_location=None, tools="",
                    background_fill_color=None, border_fill_color=None, outline_line_color=None,
                    title="Evolución semanal 2026")
    if x_line:
        p2_args['x_range'] = x_line
    p2 = figure(**p2_args)
    p2.line(x='x', y='y', line_width=2, color="#86895D", source=source_line)
    p2.scatter(x='x', y='y', size=8, color="#86895D", fill_color="white", source=source_line)
    p2.xgrid.grid_line_color = None
    p2.ygrid.grid_line_color = None
    p2.yaxis.visible = False
    p2.xaxis.major_label_orientation = 1.5708
    p2.xaxis.major_label_text_font_size = "7pt"
    p2.add_tools(HoverTool(tooltips=[("Semana", "@x"), ("Total", "@y{0,0}")]))

    # --- Grafica Top 10 nacionalidades (Mexico incluido, rango filtrado) ---
    top_names = [d['n'] for d in top_personalizado][::-1]
    top_values = [d['v'] for d in top_personalizado][::-1]
    source_top = ColumnDataSource(data=dict(names=top_names, values=top_values))
    p_top = figure(y_range=top_names, height=380, toolbar_location=None, tools="", sizing_mode="stretch_width",
                   title=f"Top 10 nacionalidades (México incluido) — {fecha_inicio} a {fecha_fin}")
    p_top.hbar(y='names', right='values', height=0.7, color="#285C4D", source=source_top)
    p_top.x_range.start = 0
    p_top.xaxis.formatter = NumeralTickFormatter(format="0a")
    p_top.grid.grid_line_color = None
    p_top.add_tools(HoverTool(tooltips=[("País", "@names"), ("Total", "@values{0,0}")]))

    plot_script, plot_divs = components((layout_p1, p2, p_top))
    plot_bar_div, plot_line_div, plot_top_div = plot_divs

    context = {
        'card_semana': datos['card_semana'], 'card_csp': datos['card_csp'], 'card_trump': datos['card_trump'],
        'card_personalizado': card_personalizado,
        'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin,
        'total_2026': f"{datos['total_2026']:,}",
        'plot_script': plot_script, 'plot_bar_div': plot_bar_div,
        'plot_line_div': plot_line_div, 'plot_top_div': plot_top_div,
        'tabla_estados': tabla_estados, 'tabla_nacionalidades': tabla_nacionalidades,
        'total_estados': total_estados, 'total_nacionalidades': total_nacionalidades,
    }
    return render(request, 'mapa/reporte_mex_extranjeros.html', context)


def reporte_mex_extranjeros_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    fecha_max = get_global_update_date() or date.today()
    fecha_inicio, fecha_fin, fi, ff = _leer_rango_mex_ext(request, fecha_max)
    tabla_estados = _tabla_mex_ext_por_estado(fi, ff)
    tabla_nacionalidades = _tabla_mex_ext_nacionalidades(fi, ff)
    context = {
        'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin,
        'tabla_estados': tabla_estados, 'tabla_nacionalidades': tabla_nacionalidades,
        'total_estados': _sumar_filas_mex_ext(tabla_estados, ['mex_ad', 'mex_me', 'mex', 'ext_ad', 'ext_me', 'ext', 'total']),
        'total_nacionalidades': _sumar_filas_mex_ext(tabla_nacionalidades, ['ext_ad', 'ext_me', 'ext']),
    }
    template = get_template("mapa/_reporte_mex_extranjeros_pdf.html")
    html_string = template.render(context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="mexicanos_extranjeros_{fecha_max.isoformat()}.pdf"'
    return response


def reporte_mex_extranjeros_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    datos = _datos_mex_extranjeros()
    fecha_inicio, fecha_fin, fi, ff = _leer_rango_mex_ext(request, datos['fecha_max'])
    card_personalizado = _build_card_mex_ext(fi, ff, "Personalizado")
    top_personalizado = _top10_mex_ext(fi, ff)
    tabla_estados = _tabla_mex_ext_por_estado(fi, ff)
    tabla_nacionalidades = _tabla_mex_ext_nacionalidades(fi, ff)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.merge_cells("A1:E1")
    _rescates_excel_celda(ws1, 1, 1, "Mexicanos Repatriados vs Extranjeros Recibidos", negrita=True, tam=13)
    fila = 3
    _rescates_excel_fila(ws1, fila, ["Periodo", "Rango", "Mexicanos", "Extranjeros", "Total", "% Mex", "% Ext"][:5], RESCATES_LETRA_T1, negrita=True)
    _rescates_excel_celda(ws1, fila, 6, "% Mex", bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], negrita=True)
    _rescates_excel_celda(ws1, fila, 7, "% Ext", bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], negrita=True)
    fila += 1
    for card in (datos['card_semana'], datos['card_csp'], datos['card_trump'], card_personalizado):
        _rescates_excel_fila(ws1, fila, [card['label'], card['periodo'], card['mex'], card['ext'], card['total'], f"{card['p_mex']}%", f"{card['p_ext']}%"], RESCATES_LETRA_T2)
        fila += 1
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 30
    for col in ("C", "D", "E", "F", "G"):
        ws1.column_dimensions[col].width = 14

    ws3 = wb.create_sheet("Evolución diaria")
    _rescates_excel_fila(ws3, 1, ["Fecha", "Total (Mex + Ext)"], RESCATES_LETRA_T1, negrita=True)
    f = 2
    for d, t in datos['evolucion_diaria']:
        _rescates_excel_fila(ws3, f, [d.strftime('%d/%m/%Y'), t], RESCATES_LETRA_T2)
        f += 1
    ws3.column_dimensions["A"].width = 14

    ws4 = wb.create_sheet(f"Top 10 ({fecha_inicio} a {fecha_fin})"[:31])
    _rescates_excel_fila(ws4, 1, ["Nacionalidad", "Total"], RESCATES_LETRA_T1, negrita=True)
    f = 2
    for item in top_personalizado:
        _rescates_excel_celda(ws4, f, 1, item['n'], bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], centrado=False)
        _rescates_excel_celda(ws4, f, 2, item['v'], bg=RESCATES_LETRA_T0[0], color_texto=RESCATES_LETRA_T0[1], negrita=True)
        f += 1
    ws4.column_dimensions["A"].width = 28

    # Hojas "Reporte General"/"Nacionalidades" -- mismo formato visual exacto
    # que FORMATO_MEX-.xlsm (negro/blanco en titulos y encabezados, gris en
    # filas de dato, verde #235B4E en el total), tabulado al nivel real
    # disponible (estado, no punto; adultos/menores, no H/M/ninos/ninas).
    NEGRO, BLANCO, VERDE, GRIS = "000000", "FFFFFF", "235B4E", "C9B182"

    ws5 = wb.create_sheet("Reporte General")
    ws5.merge_cells("A1:H1")
    _rescates_excel_celda(ws5, 1, 1, "MEXICANOS - EXTRANJEROS RECIBIDOS", bg=NEGRO, color_texto=BLANCO, tam=20, centrado=True)
    ws5.merge_cells("A2:H2")
    _rescates_excel_celda(ws5, 2, 1, f"{fecha_inicio} a {fecha_fin}", tam=12, centrado=True)

    ws5.merge_cells("A4:A5")
    _rescates_excel_celda(ws5, 4, 1, "ESTADO", bg=NEGRO, color_texto=BLANCO, tam=14, negrita=True)
    ws5.merge_cells("B4:D4")
    _rescates_excel_celda(ws5, 4, 2, "MEXICANOS", bg=NEGRO, color_texto=BLANCO)
    ws5.merge_cells("E4:G4")
    _rescates_excel_celda(ws5, 4, 5, "EXTRANJEROS", bg=NEGRO, color_texto=BLANCO)
    ws5.merge_cells("H4:H5")
    _rescates_excel_celda(ws5, 4, 8, "TOTAL GENERAL", bg=NEGRO, color_texto=BLANCO, negrita=True)
    _rescates_excel_fila(ws5, 5, ["ADULTOS", "MENORES", "TOTAL", "ADULTOS", "MENORES", "TOTAL"], (None, "000000"), col_inicial=2)

    f = 6
    total_gral = {'mex_ad': 0, 'mex_me': 0, 'mex': 0, 'ext_ad': 0, 'ext_me': 0, 'ext': 0, 'total': 0}
    for e in tabla_estados:
        _rescates_excel_celda(ws5, f, 1, e['estado'], bg=GRIS)
        for i, campo in enumerate(('mex_ad', 'mex_me', 'mex', 'ext_ad', 'ext_me', 'ext', 'total')):
            _rescates_excel_celda(ws5, f, 2 + i, e[campo], bg=GRIS)
            total_gral[campo] = total_gral.get(campo, 0) + e[campo]
        f += 1
        for punto in e.get('puntos', []):
            _rescates_excel_celda(ws5, f, 1, punto['nombre'], color_texto="6B7280", tam=9, centrado=False)
            if punto['con_datos']:
                valores_punto = (punto['mex_ad'], punto['mex_me'], punto['mex'], punto['ext_ad'], punto['ext_me'], punto['ext'], punto['total'])
                for i, v in enumerate(valores_punto):
                    _rescates_excel_celda(ws5, f, 2 + i, v, color_texto="6B7280", tam=9)
            else:
                for col in range(2, 9):
                    _rescates_excel_celda(ws5, f, col, "—", color_texto="6B7280", tam=9)
            f += 1
    _rescates_excel_celda(ws5, f, 1, "TOTAL", bg=VERDE, color_texto=BLANCO, negrita=True)
    for i, campo in enumerate(('mex_ad', 'mex_me', 'mex', 'ext_ad', 'ext_me', 'ext', 'total')):
        _rescates_excel_celda(ws5, f, 2 + i, total_gral[campo], bg=VERDE, color_texto=BLANCO, negrita=True)
    ws5.column_dimensions["A"].width = 22

    ws6 = wb.create_sheet("Nacionalidades")
    ws6.merge_cells("A1:D1")
    _rescates_excel_celda(ws6, 1, 1, "NACIONALIDADES", bg=NEGRO, color_texto=BLANCO, tam=20, centrado=True)
    ws6.merge_cells("A2:D2")
    _rescates_excel_celda(ws6, 2, 1, f"{fecha_inicio} a {fecha_fin}", tam=12, centrado=True)
    ws6.merge_cells("A4:A5")
    _rescates_excel_celda(ws6, 4, 1, "PAÍS", bg=NEGRO, color_texto=BLANCO, tam=14, negrita=True)
    ws6.merge_cells("B4:D4")
    _rescates_excel_celda(ws6, 4, 2, "EXTRANJEROS", bg=NEGRO, color_texto=BLANCO)
    _rescates_excel_fila(ws6, 5, ["ADULTOS", "MENORES", "TOTAL"], (None, "000000"), col_inicial=2)
    f = 6
    total_nac = {'ext_ad': 0, 'ext_me': 0, 'ext': 0}
    for n in tabla_nacionalidades:
        _rescates_excel_celda(ws6, f, 1, n['nacionalidad'], bg=GRIS)
        _rescates_excel_celda(ws6, f, 2, n['ext_ad'], bg=GRIS)
        _rescates_excel_celda(ws6, f, 3, n['ext_me'], bg=GRIS)
        _rescates_excel_celda(ws6, f, 4, n['ext'], bg=GRIS)
        total_nac['ext_ad'] += n['ext_ad']
        total_nac['ext_me'] += n['ext_me']
        total_nac['ext'] += n['ext']
        f += 1
    _rescates_excel_celda(ws6, f, 1, "TOTAL", bg=VERDE, color_texto=BLANCO, negrita=True)
    _rescates_excel_celda(ws6, f, 2, total_nac['ext_ad'], bg=VERDE, color_texto=BLANCO, negrita=True)
    _rescates_excel_celda(ws6, f, 3, total_nac['ext_me'], bg=VERDE, color_texto=BLANCO, negrita=True)
    _rescates_excel_celda(ws6, f, 4, total_nac['ext'], bg=VERDE, color_texto=BLANCO, negrita=True)
    ws6.column_dimensions["A"].width = 22

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="mexicanos_extranjeros_{datos["fecha_max"].isoformat()}.xlsx"'
    wb.save(response)
    return response


def mapa_ejemplo(request):
    """Vista de ejemplo para MapLibre GL JS."""
    return render(request, 'mapa/mapa_ejemplo.html')


# -------------------------------------------------------
# --- -------  --- VIEW MAPA DR --- ------ ------- ------
# -------------------------------------------------------

def mapa_interactivo(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    
    user_state = get_user_state(request)
    if not request.user.is_superuser and not user_state:
        return render(request, 'base/error404.html')

    fecha_act = get_global_update_date() or date.today()

    # No database queries for periods to avoid delays
    totals_cs = {}
    totals_dt = {}
    
    urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
    api_instrucciones = {}
    try:
        api_res = requests.get('https://172.16.16.167/api/mapa-datos/', verify=False, timeout=2.5)
        if api_res.status_code == 200:
            api_instrucciones = api_res.json()
    except Exception as e:
        print(f"Error al obtener instrucciones de la API: {str(e)}")
    
    instrucciones_avance = {}
    instrucciones_totales_dict = {}
    instrucciones_color_rank = {}

    LABEL_NACIONAL = "Total Nacional"

    METRIC_LABELS = {
        'todos': 'Todos'
    }

    # --- Recopilación de Infraestructura y Titulares ---
    infra_raw = PuntosInternacionEstacion.objects.values('estado__nombre', 'tipo').annotate(total=Count('id'))
    titulares_raw = Titular.objects.all().select_related('estado', 'tipo_nombramiento')
    
    infra_data = {}
    # Estructura base para todos los estados
    for edo in Estado.objects.all():
        infra_data[normalizar_nombre(edo.nombre)] = {
            'estado_id': edo.id,
            'AEREO': 0, 'MARITIMO': 0, 'TERRESTRE': 0, 'ESTACION': 0,
            'PRH': 0,
            'personal_total': 0,
            'personal_activo': 0,
            'personal_vacante': 0,
            'subrep_federal': 0,
            'subrep_local': 0,
            'rep_local': 0,
            'titular': 'Sin titular asignado',
            'titular_id': None,
            'foto': None,
            'tipo_nombramiento': None
        }
    
    for item in infra_raw:
        edo_name = normalizar_nombre(item['estado__nombre'])
        if edo_name in infra_data:
            infra_data[edo_name][item['tipo']] = item['total']
            
    for t in titulares_raw:
        edo_name = normalizar_nombre(t.estado.nombre)
        if edo_name in infra_data:
            infra_data[edo_name]['titular'] = f"{t.nombre} {t.apellido_paterno} {t.apellido_materno}"
            infra_data[edo_name]['titular_id'] = t.id
            infra_data[edo_name]['tipo_nombramiento'] = t.tipo_nombramiento.nombre if t.tipo_nombramiento else None
            if t.fotografia:
                infra_data[edo_name]['foto'] = t.fotografia.url

    # PRHs por estado
    prh_raw = PRHs.objects.values('estado__nombre').annotate(total=Count('id'))
    for item in prh_raw:
        edo_name = normalizar_nombre(item['estado__nombre'])
        if edo_name in infra_data:
            infra_data[edo_name]['PRH'] = item['total']

    # Personal por estado
    personal_qs = PersonalINM.objects.all().select_related('estado', 'estatus', 'tipo_plaza')
    for p in personal_qs:
        if not p.estado:
            continue
        
        tipo_plaza = (p.tipo_plaza.plazaT if p.tipo_plaza else '').upper()
        if tipo_plaza not in ['BASE', 'CONFIANZA']:
            continue

        edo_name = normalizar_nombre(p.estado.nombre)
        if edo_name in infra_data:
            infra_data[edo_name]['personal_total'] += 1
            estatus_name = (p.estatus.estatus if p.estatus else '').upper()
            if estatus_name == 'ACTIVO':
                infra_data[edo_name]['personal_activo'] += 1
            elif estatus_name == 'VACANTE':
                infra_data[edo_name]['personal_vacante'] += 1
            
            puesto_clean = normalizar_nombre(p.puesto_especifico)
            if 'SUB REPRESENTACION FEDERAL' in puesto_clean:
                infra_data[edo_name]['subrep_federal'] += 1
            elif 'SUB REPRESENTACION LOCAL' in puesto_clean:
                infra_data[edo_name]['subrep_local'] += 1
            elif 'REPRESENTACION LOCAL' in puesto_clean:
                infra_data[edo_name]['rep_local'] += 1

    # Totales Nacionales
    subrep_federal_nat = 0
    subrep_local_nat = 0
    rep_local_nat = 0
    for key, val in infra_data.items():
        subrep_federal_nat += val.get('subrep_federal', 0)
        subrep_local_nat += val.get('subrep_local', 0)
        rep_local_nat += val.get('rep_local', 0)

    infra_data[LABEL_NACIONAL] = {
        'estado_id': None,
        'AEREO': PuntosInternacionEstacion.objects.filter(tipo='AEREO').count(),
        'MARITIMO': PuntosInternacionEstacion.objects.filter(tipo='MARITIMO').count(),
        'TERRESTRE': PuntosInternacionEstacion.objects.filter(tipo='TERRESTRE').count(),
        'ESTACION': PuntosInternacionEstacion.objects.filter(tipo='ESTACION').count(),
        'PRH': PRHs.objects.count(),
        'personal_total': PersonalINM.objects.count(),
        'personal_activo': PersonalINM.objects.filter(estatus__estatus__iexact='ACTIVO').count(),
        'personal_vacante': PersonalINM.objects.filter(estatus__estatus__iexact='VACANTE').count(),
        'subrep_federal': subrep_federal_nat,
        'subrep_local': subrep_local_nat,
        'rep_local': rep_local_nat,
        'titular': 'Datos Nacionales',
        'foto': None,
        'tipo_nombramiento': None
    }

    # Valores por defecto para estados sin datos
    default_vals = {
        'todos': 0, 'color_t': 32
    }

    # Ruta al archivo geojson descargado
    geojson_path = os.path.join(settings.BASE_DIR, 'mapa', 'static', 'mapa', 'data', 'inegi_latlon_mexico.geojson')
    
    with open(geojson_path, 'r', encoding='utf-8') as f:
        geo_data = json.load(f)

    # Filtrar el GeoJSON por estado si el usuario no es superusuario
    if user_state:
        user_state_name_normalized = normalizar_nombre(user_state.nombre)
        geo_data['features'] = [
            f for f in geo_data['features']
            if normalizar_nombre(f['properties']['name']) == user_state_name_normalized
        ]
        
    for feature in geo_data['features']:
        name_normalized = normalizar_nombre(feature['properties']['name'])
        cs = default_vals.copy()
        dt = default_vals.copy()
        
        for k in default_vals:
            feature['properties'][f'cs_{k}'] = cs[k]
            feature['properties'][f'dt_{k}'] = dt[k]
            feature['properties'][f'pe_{k}'] = cs[k]
        
        feature['properties']['cs_str_todos'] = '0'
        feature['properties']['dt_str_todos'] = '0'
        feature['properties']['pe_str_todos'] = '0'
        
    # --- Capa de Infraestructura (Iconos SVG) ---
    infra_points_objs = PuntosInternacionEstacion.objects.all()
    if user_state:
        infra_points_objs = infra_points_objs.filter(estado=user_state)
    infra_pts_data = []
    for pt in infra_points_objs:
        icon_file = 'terrestre2.svg' # Default
        if pt.tipo == 'AEREO': icon_file = 'aereo2.svg'
        elif pt.tipo == 'MARITIMO': icon_file = 'maritimo2.svg'
        elif pt.tipo == 'ESTACION': icon_file = 'estacion2.svg'
        
        infra_pts_data.append({
            'x': float(pt.longitud) if pt.longitud else 0,
            'y': float(pt.latitud) if pt.latitud else 0,
            'nombre': pt.nombre,
            'estado': normalizar_nombre(pt.estado.nombre),
            'tipo': pt.tipo,
            'url': f"{settings.STATIC_URL}mapa/icons/{icon_file}"
        })

    # --- Capa de Puntos de Rescate Humano (PRH) ---
    prh_points = PRHs.objects.all().select_related('modalidad')
    if user_state:
        prh_points = prh_points.filter(estado=user_state)
    prh_pts_data = []
    for pt in prh_points:
        icon = 'agente_activo2.svg' if pt.activo else 'agente_inactivo2.svg'
        prh_pts_data.append({
            'x': float(pt.longitud) if pt.longitud else 0,
            'y': float(pt.latitud) if pt.latitud else 0,
            'nombre': pt.nombre,
            'estado': normalizar_nombre(pt.estado.nombre),
            'modalidad': pt.modalidad.nombre,
            'status': 'Activo' if pt.activo else 'Inactivo',
            'url': f"{settings.STATIC_URL}mapa/icons/{icon}"
        })

    # --- Capa de Inmuebles (Icono OR_ACTIVO) ---
    inmuebles_objs = Inmueble.objects.all().prefetch_related('tipo_oficina')
    if user_state:
        inmuebles_objs = inmuebles_objs.filter(estado=user_state)
    inmuebles_pts_data = []
    for pt in inmuebles_objs:
        inmuebles_pts_data.append({
            'id': pt.id,
            'x': float(pt.longitud) if pt.longitud else 0,
            'y': float(pt.latitud) if pt.latitud else 0,
            'nombre': pt.nombre_inmueble,
            'estado': normalizar_nombre(pt.estado.nombre),
            'tipo': 'INMUEBLE',
            'tipo_oficina': [normalizar_nombre(to.nombre) for to in pt.tipo_oficina.all()],
            'url': f"{settings.STATIC_URL}mapa/icons/OR_ACTIVO.svg"
        })

    if user_state:
        national_data = {
            'name': user_state.nombre.upper(),
            'cs': default_vals.copy(),
            'dt': default_vals.copy(),
            'pe': default_vals.copy()
        }
    else:
        national_data = {
            'name': LABEL_NACIONAL,
            'cs': default_vals.copy(),
            'dt': default_vals.copy(),
            'pe': default_vals.copy()
        }
    
    context = {
        'geo_data_json': json.dumps(geo_data),
        'national_data_json': json.dumps(national_data),
        'infra_data_json': json.dumps(infra_data),
        'infra_pts_data_json': json.dumps(infra_pts_data),
        'prh_pts_data_json': json.dumps(prh_pts_data),
        'inmuebles_pts_data_json': json.dumps(inmuebles_pts_data),
        'label_nacional': LABEL_NACIONAL,
        'metric_labels': METRIC_LABELS,
        'metric_labels_json': json.dumps(METRIC_LABELS),
        'fecha_actualizacion': fecha_act,
        'instrucciones_api_json': json.dumps(api_instrucciones),
        'is_superuser': request.user.is_superuser,
        'user_state_name': user_state.nombre if user_state else '',
        'user_state_name_normalized': normalizar_nombre(user_state.nombre) if user_state else '',
    }
    
    return render(request, 'mapa/mapa_activo.html', context)


# =====================================================================
# GESTIÓN DE INMUEBLES
# =====================================================================

def inmuebles_list(request):
    """Muestra la lista de inmuebles y gestiona sus catálogos."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    user_state = get_user_state(request)
    
    if request.user.is_superuser:
        inmuebles_list = Inmueble.objects.all().order_by('estado__nombre', 'nombre_inmueble')
        estados_list = Estado.objects.all().order_by('nombre')
    elif user_state:
        inmuebles_list = Inmueble.objects.filter(estado=user_state).order_by('nombre_inmueble')
        estados_list = [user_state]
    else:
        inmuebles_list = Inmueble.objects.none()
        estados_list = []
        
    return render(request, 'mapa/inmuebles.html', {
        'inmuebles_list': inmuebles_list,
        'estados_list': estados_list,
        'tipos_inmueble': TipoInmueble.objects.all().order_by('nombre'),
        'situaciones_actuales': SituacionActual.objects.all().order_by('nombre'),
        'tipos_actividad': TipoActividad.objects.all().order_by('nombre'),
        'tipos_oficina': TipoOficina.objects.all().order_by('nombre'),
        'figuras_ocupacion': FiguraOcupacion.objects.all().order_by('tipo'),
        'comodatos_list': Comodato.objects.all().order_by('nombre'),
    })


@transaction.atomic
def guardar_inmueble(request):
    if request.method != 'POST':
        return redirect('inmuebles_list')
        
    user_state = get_user_state(request)
    
    try:
        inmueble_id = request.POST.get('inmueble_id')
        estado_id = request.POST.get('estado_id')
        
        # Validación de seguridad: el estado enviado debe coincidir con el del usuario
        if not request.user.is_superuser:
            if not user_state or str(user_state.id) != str(estado_id):
                raise PermissionError("No tienes permisos para registrar inmuebles en este estado.")
        
        # Obtener o crear objeto Comodato si viene
        comodato_id = request.POST.get('comodato_id') or None
        
        # Campos de superficie
        try:
            sup_total = float(request.POST.get('superficie_total') or 0)
            sup_const = float(request.POST.get('superficie_construida') or 0)
            sup_util = float(request.POST.get('superficie_utilizada') or 0)
        except ValueError:
            raise ValueError("Las superficies deben ser valores numéricos válidos.")

        # Coordenadas
        try:
            lat = float(request.POST.get('latitud') or 0)
            lng = float(request.POST.get('longitud') or 0)
        except ValueError:
            raise ValueError("La latitud y longitud deben ser coordenadas numéricas válidas.")

        # Fechas
        fecha_ocup = request.POST.get('fecha_ocupacion') or None
        anio_const = request.POST.get('anio_construccion') or None

        # Monto Renta
        monto_renta = request.POST.get('monto_renta') or None
        if monto_renta:
            monto_renta = monto_renta.replace('$', '').replace(',', '').strip()
            if not monto_renta:
                monto_renta = None

        defaults = {
            'estado_id': estado_id,
            'nombre_inmueble': normalizar_nombre(request.POST.get('nombre_inmueble', '')),
            'calle': normalizar_nombre(request.POST.get('calle', '')),
            'numero_exterior': request.POST.get('numero_exterior', '').strip(),
            'numero_interior': request.POST.get('numero_interior', '').strip(),
            'colonia': normalizar_nombre(request.POST.get('colonia', '')),
            'municipio': normalizar_nombre(request.POST.get('municipio', '')),
            'codigo_postal': request.POST.get('codigo_postal', '').strip(),
            'latitud': lat,
            'longitud': lng,
            'situacion_actual_id': request.POST.get('situacion_actual_id') or None,
            'tipo_inmueble_id': request.POST.get('tipo_inmueble_id') or None,
            'superficie_total': sup_total,
            'superficie_construida': sup_const,
            'superficie_utilizada': sup_util,
            'numero_de_niveles': int(request.POST.get('numero_de_niveles') or 0),
            'anio_construccion': anio_const,
            'fecha_ocupacion': fecha_ocup,
            'figura_ocupacion_id': request.POST.get('figura_ocupacion_id') or None,
            'monto_renta': monto_renta,
            'comodato_id': comodato_id,
        }

        if inmueble_id:
            inmueble = Inmueble.objects.get(id=inmueble_id)
            # Validación adicional de seguridad para edición
            if not request.user.is_superuser and inmueble.estado != user_state:
                raise PermissionError("No tienes permisos para modificar este inmueble.")
            for key, val in defaults.items():
                setattr(inmueble, key, val)
            inmueble.save()
            created = False
        else:
            inmueble = Inmueble.objects.create(**defaults)
            created = True

        # Guardar/Actualizar ProgramaIPC
        programa_ipc, _ = ProgramaIPC.objects.get_or_create(inmueble=inmueble)
        programa_ipc.inm_pipc = request.POST.get('inm_pipc') == 'on'
        programa_ipc.fecha_inm = request.POST.get('fecha_inm') or None
        programa_ipc.comodante_pipc = request.POST.get('comodante_pipc') == 'on'
        programa_ipc.fecha_comodante = request.POST.get('fecha_comodante') or None
        programa_ipc.plan_emergencia = request.POST.get('plan_emergencia') == 'on'
        programa_ipc.fecha_inicio_plan = request.POST.get('fecha_inicio_plan') or None
        programa_ipc.save()

        # Guardar ManyToMany tipo_actividad
        tipo_actividades_ids = request.POST.getlist('tipo_actividad[]')
        inmueble.tipo_actividad.set(tipo_actividades_ids)

        # Guardar ManyToMany tipo_oficina
        tipo_oficina_ids = request.POST.getlist('tipo_oficina[]')
        inmueble.tipo_oficina.set(tipo_oficina_ids)

        # Guardar nuevo comentario en el histórico si existe
        comentario_texto = request.POST.get('observaciones', '').strip()
        if comentario_texto:
            HistoricoComentarios.objects.create(
                inmueble=inmueble,
                comentario=comentario_texto
            )

        messages.success(request, f"Inmueble '{inmueble.nombre_inmueble}' {'creado' if created else 'actualizado'} correctamente.")
    except Exception as e:
        messages.error(request, f"Error al guardar inmueble: {str(e)}")
        
    return redirect('inmuebles_list')


@transaction.atomic
def eliminar_inmueble(request, inmueble_id):
    """Elimina un inmueble si tiene permisos por estado."""
    try:
        user_state = get_user_state(request)
        inmueble = Inmueble.objects.get(id=inmueble_id)
        
        # Validación de seguridad
        if not request.user.is_superuser:
            if not user_state or inmueble.estado != user_state:
                messages.error(request, "No tienes permisos para eliminar este registro.")
                return redirect('inmuebles_list')
                
        nombre = inmueble.nombre_inmueble
        inmueble.delete()
        messages.success(request, f"Inmueble '{nombre}' eliminado correctamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar inmueble: {str(e)}")
        
    return redirect('inmuebles_list')


def api_get_inmueble(request, inmueble_id):
    """Retorna los datos de un inmueble en formato JSON."""
    try:
        user_state = get_user_state(request)
        inmueble = Inmueble.objects.get(id=inmueble_id)
        
        # Validación de seguridad
        if not request.user.is_superuser:
            if not user_state or inmueble.estado != user_state:
                return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
                
        data = {
            'id': inmueble.id,
            'nombre_inmueble': inmueble.nombre_inmueble,
            'estado_id': inmueble.estado_id,
            'calle': inmueble.calle,
            'numero_exterior': inmueble.numero_exterior,
            'numero_interior': inmueble.numero_interior,
            'colonia': inmueble.colonia,
            'municipio': inmueble.municipio,
            'codigo_postal': inmueble.codigo_postal,
            'latitud': inmueble.latitud,
            'longitud': inmueble.longitud,
            'situacion_actual_id': inmueble.situacion_actual_id,
            'tipo_inmueble_id': inmueble.tipo_inmueble_id,
            'superficie_total': inmueble.superficie_total,
            'superficie_construida': inmueble.superficie_construida,
            'superficie_utilizada': inmueble.superficie_utilizada,
            'numero_de_niveles': inmueble.numero_de_niveles,
            'anio_construccion': inmueble.anio_construccion.isoformat() if inmueble.anio_construccion else None,
            'fecha_ocupacion': inmueble.fecha_ocupacion.isoformat() if inmueble.fecha_ocupacion else None,
            'figura_ocupacion_id': inmueble.figura_ocupacion_id,
            'monto_renta': float(inmueble.monto_renta) if inmueble.monto_renta else None,
            'comodato_id': inmueble.comodato_id,
            'inm_pipc': inmueble.pipc.first().inm_pipc if inmueble.pipc.exists() else False,
            'fecha_inm': inmueble.pipc.first().fecha_inm.isoformat() if inmueble.pipc.exists() and inmueble.pipc.first().fecha_inm else None,
            'comodante_pipc': inmueble.pipc.first().comodante_pipc if inmueble.pipc.exists() else False,
            'fecha_comodante': inmueble.pipc.first().fecha_comodante.isoformat() if inmueble.pipc.exists() and inmueble.pipc.first().fecha_comodante else None,
            'plan_emergencia': inmueble.pipc.first().plan_emergencia if inmueble.pipc.exists() else False,
            'fecha_inicio_plan': inmueble.pipc.first().fecha_inicio_plan.isoformat() if inmueble.pipc.exists() and inmueble.pipc.first().fecha_inicio_plan else None,
            'comentarios': [
                {
                    'id': c.id,
                    'comentario': c.comentario,
                    'fecha_creacion': c.fecha_creacion.strftime('%d/%m/%Y %H:%M')
                }
                for c in inmueble.comentarios.all().order_by('-fecha_creacion')
            ],
            'tipo_actividad_ids': list(inmueble.tipo_actividad.values_list('id', flat=True)),
            'tipo_oficina_ids': list(inmueble.tipo_oficina.values_list('id', flat=True)),
        }
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


@transaction.atomic
def api_guardar_comodato(request):
    """Crea un nuevo comodato mediante petición AJAX y lo retorna."""
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
    try:
        nombre = request.POST.get('nombre', '').strip().upper()
        if not nombre:
            return JsonResponse({'status': 'error', 'message': 'El nombre es obligatorio.'}, status=400)
            
        comodato, created = Comodato.objects.get_or_create(nombre=nombre)
        return JsonResponse({'status': 'success', 'data': {'id': comodato.id, 'nombre': comodato.nombre}})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def organigramas_list(request):
    """Vista para la gestión de Organigramas por Estado."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    user_state = get_user_state(request)
    
    if user_state:
        estados_list = [user_state]
        organigramas = OrganigramaF.objects.filter(estado=user_state).order_by('estado__nombre')
    else:
        estados_list = Estado.objects.all().order_by('nombre')
        organigramas = OrganigramaF.objects.all().order_by('estado__nombre')

    return render(request, 'mapa/organigramas.html', {
        'estados_list': estados_list,
        'organigramas_list': organigramas,
    })


@transaction.atomic
def guardar_organigrama(request):
    """Crea o actualiza la Estructura Orgánica para un Estado."""
    if request.method != 'POST':
        return redirect('organigramas_list')
        
    user_state = get_user_state(request)
    
    try:
        estado_id = request.POST.get('estado_id')
        vigencia = request.POST.get('vigencia') or None
        archivo = request.FILES.get('archivo')
        
        # Seguridad: verificar que coincida con el estado del usuario si no es superusuario
        if user_state and str(user_state.id) != str(estado_id):
            raise PermissionError("No tienes permisos para registrar la Estructura Orgánica en este estado.")
            
        defaults = {
            'vigencia': vigencia,
        }
        if archivo:
            defaults['archivo'] = archivo
            
        OrganigramaF.objects.update_or_create(
            estado_id=estado_id,
            defaults=defaults
        )
        
        messages.success(request, "Estructura Orgánica guardada correctamente.")
    except Exception as e:
        messages.error(request, f"Error al guardar la Estructura Orgánica: {str(e)}")
        
    return redirect('organigramas_list')


@transaction.atomic
def eliminar_organigrama(request, org_id):
    """Elimina la Estructura Orgánica de un Estado."""
    try:
        user_state = get_user_state(request)
        org = OrganigramaF.objects.get(id=org_id)
        
        if user_state and org.estado != user_state:
            messages.error(request, "No tienes permisos para eliminar este registro.")
            return redirect('organigramas_list')
            
        estado_nombre = org.estado.nombre
        org.delete()
        messages.success(request, f"Estructura Orgánica de {estado_nombre} eliminada correctamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar registro: {str(e)}")
        
    return redirect('organigramas_list')


def api_get_organigrama(request, estado_id):
    """Retorna la información del organigrama de un estado en formato JSON."""
    try:
        org = OrganigramaF.objects.get(estado_id=estado_id)
        return JsonResponse({
            'status': 'success',
            'data': {
                'pdf_url': org.archivo.url if org.archivo else '',
                'vigencia': org.vigencia.strftime('%d/%m/%Y') if org.vigencia else 'Sin vigencia'
            }
        })
    except OrganigramaF.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'No hay organigrama cargado para este estado.'
        })


def api_get_inmueble_detalle(request, inmueble_id):
    """Retorna toda la información detallada de un inmueble en formato JSON para el modal interactivo."""
    try:
        inmueble = Inmueble.objects.select_related('estado', 'figura_ocupacion', 'comodato').get(id=inmueble_id)
        
        # 1. Jefe de Oficina
        jefe = PersonalINM.objects.filter(lugar_asignado=inmueble, jefe_oficina=True).first()
        if jefe:
            nombre_jefe = f"{jefe.nombre or ''} {jefe.apellido or ''}".strip()
            if not nombre_jefe:
                nombre_jefe = "S/A"
        else:
            nombre_jefe = "S/A"
            
        # 2. Personal conteo
        personal_qs = PersonalINM.objects.filter(lugar_asignado=inmueble)
        total_personal = personal_qs.count()
        activos = personal_qs.filter(estatus__estatus__iexact='ACTIVO').count()
        inactivos = personal_qs.filter(estatus__estatus__iexact='VACANTE').count()
        
        # 2.5. Vehículos conteo
        from .models import VehiculosOR
        veh_qs = VehiculosOR.objects.filter(inmueble=inmueble)
        total_veh = veh_qs.count()
        activos_veh = veh_qs.filter(situacion__situacion__iexact='ACTIVO').count()
        inactivos_veh = total_veh - activos_veh
        pipc = inmueble.pipc.first()
        fecha_inm_str = "sin programa"
        fecha_comodante_str = "sin programa"
        fecha_inicio_plan_str = "sin programa"
        
        if pipc:
            if pipc.inm_pipc and pipc.fecha_inm:
                fecha_inm_str = pipc.fecha_inm.strftime('%d/%m/%Y')
            if pipc.comodante_pipc and pipc.fecha_comodante:
                fecha_comodante_str = pipc.fecha_comodante.strftime('%d/%m/%Y')
            if pipc.plan_emergencia and pipc.fecha_inicio_plan:
                fecha_inicio_plan_str = pipc.fecha_inicio_plan.strftime('%d/%m/%Y')

        # 4. Dirección completa
        num_ext = f"No. {inmueble.numero_exterior}" if inmueble.numero_exterior else "S/N"
        num_int = f" Int. {inmueble.numero_interior}" if inmueble.numero_interior and inmueble.numero_interior.strip() != "" else ""
        colonia = f", Col. {inmueble.colonia}" if inmueble.colonia else ""
        municipio = f", {inmueble.municipio}" if inmueble.municipio else ""
        cp = f", C.P. {inmueble.codigo_postal}" if inmueble.codigo_postal else ""
        estado_nombre = inmueble.estado.nombre
        
        direccion_completa = f"{inmueble.calle or ''} {num_ext}{num_int}{colonia}{municipio}{cp}, {estado_nombre}".strip()
        
        # 5. General info
        superficie_util = inmueble.superficie_utilizada if inmueble.superficie_utilizada else 0.0
        niveles = inmueble.numero_de_niveles if inmueble.numero_de_niveles is not None else 0
        anio_const = inmueble.anio_construccion.strftime('%Y') if inmueble.anio_construccion else "S/D"
        fecha_ocu = inmueble.fecha_ocupacion.strftime('%d/%m/%Y') if inmueble.fecha_ocupacion else "S/D"
        
        # 6. Figura de Ocupación
        figura_tipo = inmueble.figura_ocupacion.tipo.upper() if inmueble.figura_ocupacion else ""
        
        arrendado_activo = "ARRENDADO" in figura_tipo
        propio_activo = "PROPIO" in figura_tipo
        terreno_activo = "TERRENO" in figura_tipo
        comodato_activo = "COMODATO" in figura_tipo
        
        # 7. Renta
        renta_str = "S/D"
        if arrendado_activo and inmueble.monto_renta is not None:
            renta_str = f"${inmueble.monto_renta:,.2f}"
            
        # 8. Tipos de Oficina (las que tiene asignadas el inmueble)
        oficinas_asignadas = list(inmueble.tipo_oficina.values_list('nombre', flat=True))
        
        # 9. Actividades (todas y marcar activas)
        actividades_all = TipoActividad.objects.all().order_by('nombre')
        actividades_data = []
        for act in actividades_all:
            is_active = inmueble.tipo_actividad.filter(id=act.id).exists()
            actividades_data.append({
                'id': act.id,
                'nombre': act.nombre,
                'activo': is_active
            })
            
        data = {
            'id': inmueble.id,
            'nombre_inmueble': inmueble.nombre_inmueble,
            'estado_nombre': estado_nombre,
            'estado_id': inmueble.estado.id,
            'nombre_jefe': nombre_jefe,
            'personal': {
                'total': total_personal,
                'activos': activos,
                'inactivos': inactivos
            },
            'vehiculos': {
                'total': total_veh,
                'activos': activos_veh,
                'inactivos': inactivos_veh
            },
            'pipc': {
                'fecha_inm': fecha_inm_str,
                'fecha_comodante': fecha_comodante_str,
                'fecha_inicio_plan': fecha_inicio_plan_str
            },
            'direccion_completa': direccion_completa,
            'superficie_utilizada': f"{superficie_util:,.2f}" if superficie_util else "0.00",
            'numero_de_niveles': niveles,
            'anio_construccion': anio_const,
            'fecha_ocupacion': fecha_ocu,
            'figura_ocupacion': {
                'tipo': figura_tipo,
                'arrendado_activo': arrendado_activo,
                'propio_activo': propio_activo,
                'terreno_activo': terreno_activo,
                'comodato_activo': comodato_activo
            },
            'renta': renta_str,
            'oficinas_asignadas': oficinas_asignadas,
            'actividades': actividades_data
        }
        
        return JsonResponse({
            'status': 'success',
            'data': data
        })
    except Inmueble.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'El inmueble solicitado no existe.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


def api_get_estado_detalle(request, estado_id):
    """Retorna la información agregada de todos los inmuebles de un estado en formato JSON para el modal interactivo de estado."""
    try:
        if estado_id == 0:
            inmuebles = Inmueble.objects.all()
            inmuebles_ids = list(inmuebles.values_list('id', flat=True))
            inmueble_count = inmuebles.count()
            
            # 1. Personal
            personal_qs = PersonalINM.objects.all()
            total_personal = personal_qs.count()
            activos = personal_qs.filter(estatus__estatus__iexact='ACTIVO').count()
            inactivos = personal_qs.filter(estatus__estatus__iexact='VACANTE').count()
            
            # 1.5. Vehículos
            from .models import VehiculosOR
            veh_qs = VehiculosOR.objects.all()
            total_veh = veh_qs.count()
            activos_veh = veh_qs.filter(situacion__situacion__iexact='ACTIVO').count()
            inactivos_veh = total_veh - activos_veh
            
            # 2. PIPC (Programa IPC)
            from .models import ProgramaIPC
            pipc_qs = ProgramaIPC.objects.filter(inmueble_id__in=inmuebles_ids)
            count_inm_pipc = pipc_qs.filter(inm_pipc=True).count()
            count_comodante_pipc = pipc_qs.filter(comodante_pipc=True).count()
            count_plan_emergencia = pipc_qs.filter(plan_emergencia=True).count()
            
            inm_pipc_list = list(Inmueble.objects.filter(id__in=list(pipc_qs.filter(inm_pipc=True).values_list('inmueble_id', flat=True))).values('id', 'nombre_inmueble'))
            comodante_pipc_list = list(Inmueble.objects.filter(id__in=list(pipc_qs.filter(comodante_pipc=True).values_list('inmueble_id', flat=True))).values('id', 'nombre_inmueble'))
            plan_emergencia_list = list(Inmueble.objects.filter(id__in=list(pipc_qs.filter(plan_emergencia=True).values_list('inmueble_id', flat=True))).values('id', 'nombre_inmueble'))
            
            # 3. Tipos de Oficina
            from django.db.models import Count
            from .models import TipoOficina
            oficinas_conteo = TipoOficina.objects.filter(inmueble__isnull=False).annotate(total=Count('inmueble')).filter(total__gt=0).order_by('-total')
            oficinas_data = []
            for of in oficinas_conteo:
                inms_list = list(Inmueble.objects.filter(tipo_oficina=of).values('id', 'nombre_inmueble'))
                oficinas_data.append({
                    'nombre': of.nombre,
                    'total': of.total,
                    'inmuebles': inms_list
                })
                
            # 4. Actividades
            from .models import TipoActividad
            actividades_conteo = TipoActividad.objects.filter(inmueble__isnull=False).annotate(total=Count('inmueble')).filter(total__gt=0).order_by('-total')
            actividades_data = []
            for act in actividades_conteo:
                inms_list = list(Inmueble.objects.filter(tipo_actividad=act).values('id', 'nombre_inmueble'))
                actividades_data.append({
                    'nombre': act.nombre,
                    'total': act.total,
                    'inmuebles': inms_list
                })
                
            # 5. Superficies y Renta
            from django.db.models import Sum
            sums = inmuebles.aggregate(
                total_construida=Sum('superficie_construida'),
                total_utilizada=Sum('superficie_utilizada'),
                total_renta=Sum('monto_renta')
            )
            
            superficie_const = sums['total_construida'] or 0.0
            superficie_util = sums['total_utilizada'] or 0.0
            total_renta = sums['total_renta'] or 0.0
            
            # 6. Figura de Ocupación
            from .models import FiguraOcupacion
            figuras_conteo = FiguraOcupacion.objects.filter(inmueble__isnull=False).annotate(total=Count('inmueble')).filter(total__gt=0)
            figuras_dict = {
                'ARRENDADO': 0,
                'PROPIO': 0,
                'TERRENO': 0,
                'COMODATO': 0
            }
            for fig in figuras_conteo:
                fig_tipo = fig.tipo.upper()
                if 'ARRENDADO' in fig_tipo:
                    figuras_dict['ARRENDADO'] += fig.total
                elif 'PROPIO' in fig_tipo:
                    figuras_dict['PROPIO'] += fig.total
                elif 'TERRENO' in fig_tipo:
                    figuras_dict['TERRENO'] += fig.total
                elif 'COMODATO' in fig_tipo:
                    figuras_dict['COMODATO'] += fig.total
                    
            data = {
                'estado_nombre': 'TOTAL NACIONAL',
                'inmueble_count': inmueble_count,
                'personal': {
                    'total': total_personal,
                    'activos': activos,
                    'inactivos': inactivos
                },
                'vehiculos': {
                    'total': total_veh,
                    'activos': activos_veh,
                    'inactivos': inactivos_veh
                },
                'pipc': {
                    'inm_pipc_count': count_inm_pipc,
                    'inm_pipc_inmuebles': inm_pipc_list,
                    'comodante_pipc_count': count_comodante_pipc,
                    'comodante_pipc_inmuebles': comodante_pipc_list,
                    'plan_emergencia_count': count_plan_emergencia,
                    'plan_emergencia_inmuebles': plan_emergencia_list
                },
                'superficie_construida': f"{superficie_const:,.2f}" if superficie_const else "0.00",
                'superficie_utilizada': f"{superficie_util:,.2f}" if superficie_util else "0.00",
                'figura_ocupacion': figuras_dict,
                'figura_ocupacion_inmuebles': {
                    'ARRENDADO': list(Inmueble.objects.filter(figura_ocupacion__tipo__icontains='ARRENDADO').values('id', 'nombre_inmueble')),
                    'PROPIO': list(Inmueble.objects.filter(figura_ocupacion__tipo__icontains='PROPIO').values('id', 'nombre_inmueble')),
                    'TERRENO': list(Inmueble.objects.filter(figura_ocupacion__tipo__icontains='TERRENO').values('id', 'nombre_inmueble')),
                    'COMODATO': list(Inmueble.objects.filter(figura_ocupacion__tipo__icontains='COMODATO').values('id', 'nombre_inmueble')),
                },
                'renta': f"${total_renta:,.2f}" if total_renta else "S/D",
                'oficinas': oficinas_data,
                'actividades': actividades_data
            }
            
            return JsonResponse({
                'status': 'success',
                'data': data
            })

        estado = Estado.objects.get(id=estado_id)
        
        # Inmuebles en este estado
        inmuebles = Inmueble.objects.filter(estado=estado)
        inmuebles_ids = list(inmuebles.values_list('id', flat=True))
        inmueble_count = inmuebles.count()
        
        # 1. Personal
        personal_qs = PersonalINM.objects.filter(estado=estado)
        total_personal = personal_qs.count()
        activos = personal_qs.filter(estatus__estatus__iexact='ACTIVO').count()
        inactivos = personal_qs.filter(estatus__estatus__iexact='VACANTE').count()
        
        # 1.5. Vehículos
        from .models import VehiculosOR
        veh_qs = VehiculosOR.objects.filter(estado=estado)
        total_veh = veh_qs.count()
        activos_veh = veh_qs.filter(situacion__situacion__iexact='ACTIVO').count()
        inactivos_veh = total_veh - activos_veh
        
        # 2. PIPC (Programa IPC) - Cuenta de inmuebles con programas activos y sus listados
        from .models import ProgramaIPC
        pipc_qs = ProgramaIPC.objects.filter(inmueble_id__in=inmuebles_ids)
        count_inm_pipc = pipc_qs.filter(inm_pipc=True).count()
        count_comodante_pipc = pipc_qs.filter(comodante_pipc=True).count()
        count_plan_emergencia = pipc_qs.filter(plan_emergencia=True).count()
        
        inm_pipc_list = list(Inmueble.objects.filter(estado=estado, id__in=list(pipc_qs.filter(inm_pipc=True).values_list('inmueble_id', flat=True))).values('id', 'nombre_inmueble'))
        comodante_pipc_list = list(Inmueble.objects.filter(estado=estado, id__in=list(pipc_qs.filter(comodante_pipc=True).values_list('inmueble_id', flat=True))).values('id', 'nombre_inmueble'))
        plan_emergencia_list = list(Inmueble.objects.filter(estado=estado, id__in=list(pipc_qs.filter(plan_emergencia=True).values_list('inmueble_id', flat=True))).values('id', 'nombre_inmueble'))
        
        # 3. Tipos de Oficina (Suma de inmuebles por tipo de oficina)
        from django.db.models import Count
        from .models import TipoOficina
        oficinas_conteo = TipoOficina.objects.filter(inmueble__estado=estado).annotate(total=Count('inmueble')).filter(total__gt=0).order_by('-total')
        oficinas_data = []
        for of in oficinas_conteo:
            inms_list = list(Inmueble.objects.filter(estado=estado, tipo_oficina=of).values('id', 'nombre_inmueble'))
            oficinas_data.append({
                'nombre': of.nombre,
                'total': of.total,
                'inmuebles': inms_list
            })
            
        # 4. Actividades (Suma de inmuebles por tipo de actividad)
        from .models import TipoActividad
        actividades_conteo = TipoActividad.objects.filter(inmueble__estado=estado).annotate(total=Count('inmueble')).filter(total__gt=0).order_by('-total')
        actividades_data = []
        for act in actividades_conteo:
            inms_list = list(Inmueble.objects.filter(estado=estado, tipo_actividad=act).values('id', 'nombre_inmueble'))
            actividades_data.append({
                'nombre': act.nombre,
                'total': act.total,
                'inmuebles': inms_list
            })
            
        # 5. Superficies y Renta
        from django.db.models import Sum
        sums = inmuebles.aggregate(
            total_construida=Sum('superficie_construida'),
            total_utilizada=Sum('superficie_utilizada'),
            total_renta=Sum('monto_renta')
        )
        
        superficie_const = sums['total_construida'] or 0.0
        superficie_util = sums['total_utilizada'] or 0.0
        total_renta = sums['total_renta'] or 0.0
        
        # 6. Figura de Ocupación (Conteo por figura de ocupación)
        from .models import FiguraOcupacion
        figuras_conteo = FiguraOcupacion.objects.filter(inmueble__estado=estado).annotate(total=Count('inmueble')).filter(total__gt=0)
        figuras_dict = {
            'ARRENDADO': 0,
            'PROPIO': 0,
            'TERRENO': 0,
            'COMODATO': 0
        }
        for fig in figuras_conteo:
            fig_tipo = fig.tipo.upper()
            if 'ARRENDADO' in fig_tipo:
                figuras_dict['ARRENDADO'] += fig.total
            elif 'PROPIO' in fig_tipo:
                figuras_dict['PROPIO'] += fig.total
            elif 'TERRENO' in fig_tipo:
                figuras_dict['TERRENO'] += fig.total
            elif 'COMODATO' in fig_tipo:
                figuras_dict['COMODATO'] += fig.total
                
        data = {
            'estado_nombre': estado.nombre,
            'inmueble_count': inmueble_count,
            'personal': {
                'total': total_personal,
                'activos': activos,
                'inactivos': inactivos
            },
            'vehiculos': {
                'total': total_veh,
                'activos': activos_veh,
                'inactivos': inactivos_veh
            },
            'pipc': {
                'inm_pipc_count': count_inm_pipc,
                'inm_pipc_inmuebles': inm_pipc_list,
                'comodante_pipc_count': count_comodante_pipc,
                'comodante_pipc_inmuebles': comodante_pipc_list,
                'plan_emergencia_count': count_plan_emergencia,
                'plan_emergencia_inmuebles': plan_emergencia_list
            },
            'superficie_construida': f"{superficie_const:,.2f}" if superficie_const else "0.00",
            'superficie_utilizada': f"{superficie_util:,.2f}" if superficie_util else "0.00",
            'figura_ocupacion': figuras_dict,
            'figura_ocupacion_inmuebles': {
                'ARRENDADO': list(Inmueble.objects.filter(estado=estado, figura_ocupacion__tipo__icontains='ARRENDADO').values('id', 'nombre_inmueble')),
                'PROPIO': list(Inmueble.objects.filter(estado=estado, figura_ocupacion__tipo__icontains='PROPIO').values('id', 'nombre_inmueble')),
                'TERRENO': list(Inmueble.objects.filter(estado=estado, figura_ocupacion__tipo__icontains='TERRENO').values('id', 'nombre_inmueble')),
                'COMODATO': list(Inmueble.objects.filter(estado=estado, figura_ocupacion__tipo__icontains='COMODATO').values('id', 'nombre_inmueble')),
            },
            'renta': f"${total_renta:,.2f}" if total_renta else "S/D",
            'oficinas': oficinas_data,
            'actividades': actividades_data
        }
        
        return JsonResponse({
            'status': 'success',
            'data': data
        })
    except Estado.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': 'El estado solicitado no existe.'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        }, status=500)


def api_get_personal_stats(request, estado_id):
    try:
        from datetime import date
        import unicodedata
        today = date.today()
        
        inmueble_id = request.GET.get('inmueble_id')
        inmueble = None
        if inmueble_id:
            try:
                inmueble = Inmueble.objects.get(id=inmueble_id)
                qs_all = PersonalINM.objects.filter(lugar_asignado=inmueble).select_related('estatus', 'tipo_plaza')
                if estado_id == 0:
                    estado_nombre_display = f"TOTAL NACIONAL - {inmueble.nombre_inmueble}"
                else:
                    estado = Estado.objects.get(id=estado_id)
                    estado_nombre_display = f"{estado.nombre} - {inmueble.nombre_inmueble}"
            except Inmueble.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': 'El inmueble solicitado no existe.'}, status=404)
        else:
            if estado_id == 0:
                qs_all = PersonalINM.objects.all().select_related('estatus', 'tipo_plaza')
                estado_nombre_display = "TOTAL NACIONAL"
            else:
                estado = Estado.objects.get(id=estado_id)
                qs_all = PersonalINM.objects.filter(estado=estado).select_related('estatus', 'tipo_plaza')
                estado_nombre_display = estado.nombre
            
        def clean_text(text):
            if not text:
                return ""
            normalized = unicodedata.normalize('NFD', text)
            cleaned = "".join(c for c in normalized if not unicodedata.combining(c))
            return cleaned.upper().strip()

        MANDOS_MEDIOS_KEYWORDS = [
            ('OFICINA DE REPRESENTACION', 'OFICINA DE REPRESENTACION'),
            ('SUB REPRESENTACION FEDERAL', 'SUB REPRESENTACIÓN FEDERAL'),
            ('SUB REPRESENTACION LOCAL', 'SUB REPRESENTACIÓN LOCAL'),
            ('REPRESENTACION LOCAL', 'REPRESENTACIÓN LOCAL'),
            ('SUBDIRECCION', 'SUBDIRECCIÓN'),
            ('COORDINACION', 'COORDINACIÓN'),
            ('DEPARTAMENTO', 'DEPARTAMENTO'),
            ('DIRECCION', 'DIRECCIÓN'),
        ]

        def get_stats_for_qs(qs):
            total = qs.count()
            
            # Plaza type totals
            base = qs.filter(tipo_plaza__plazaT__iexact='BASE').count()
            confianza = qs.filter(tipo_plaza__plazaT__iexact='CONFIANZA').count()
            eventual = qs.filter(tipo_plaza__plazaT__iexact='EVENTUAL').count()
            
            # Categories lists
            enlace_operativo_groups = {}
            mandos_medios_groups = {}
            
            enlace_levels = {'2', '3', '5', '6', '7', '11', 'P11', 'P12', 'P13'}
            mandos_levels = {'O11', 'O21', 'O23', 'M11', 'M23', 'N11', 'N22', 'M41', 'M43'}
            
            for p in qs:
                lvl = (p.nivel or '').strip().upper()
                puesto_raw = p.puesto_especifico or ''
                puesto_upper = puesto_raw.upper().strip()
                
                # Check if it belongs to Enlace y Operativo
                if lvl in enlace_levels:
                    # Group by puesto_especifico (exact upper text)
                    if puesto_upper not in enlace_operativo_groups:
                        enlace_operativo_groups[puesto_upper] = []
                    enlace_operativo_groups[puesto_upper].append(p)
                    
                # Check if it belongs to Mandos Medios
                elif lvl in mandos_levels:
                    # Find keyword group
                    puesto_clean = clean_text(puesto_raw)
                    group_name = None
                    for kw_clean, kw_display in MANDOS_MEDIOS_KEYWORDS:
                        if kw_clean in puesto_clean:
                            group_name = kw_display
                            break
                    if not group_name:
                        group_name = puesto_upper if puesto_upper else "S/D"
                    
                    if group_name not in mandos_medios_groups:
                        mandos_medios_groups[group_name] = []
                    mandos_medios_groups[group_name].append(p)

                else:
                    # Group by puesto_especifico (exact upper text)
                    if puesto_upper not in enlace_operativo_groups:
                        enlace_operativo_groups[puesto_upper] = []
                    enlace_operativo_groups[puesto_upper].append(p)

            def get_row_stats(group_name, plist):
                t = len(plist)
                b = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'BASE')
                c = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'CONFIANZA')
                ev = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'EVENTUAL')
                
                # Gender breakdown per plaza type
                b_m = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'BASE' and p.sexo == 'F')
                b_h = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'BASE' and p.sexo == 'M')
                c_m = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'CONFIANZA' and p.sexo == 'F')
                c_h = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'CONFIANZA' and p.sexo == 'M')
                ev_m = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'EVENTUAL' and p.sexo == 'F')
                ev_h = sum(1 for p in plist if p.tipo_plaza and p.tipo_plaza.plazaT.upper() == 'EVENTUAL' and p.sexo == 'M')
                
                # Puestos especificos breakdown per plaza type
                base_puestos = {}
                confianza_puestos = {}
                eventual_puestos = {}
                for p in plist:
                    puesto = p.puesto_especifico or 'SIN ESPECIFICAR'
                    plaza = p.tipo_plaza.plazaT.upper() if (p.tipo_plaza and p.tipo_plaza.plazaT) else ''
                    if plaza == 'BASE':
                        base_puestos[puesto] = base_puestos.get(puesto, 0) + 1
                    elif plaza == 'CONFIANZA':
                        confianza_puestos[puesto] = confianza_puestos.get(puesto, 0) + 1
                    elif plaza == 'EVENTUAL':
                        eventual_puestos[puesto] = eventual_puestos.get(puesto, 0) + 1
                
                return {
                    'group_name': group_name,
                    'total': t,
                    'base': b,
                    'confianza': c,
                    'eventual': ev,
                    'base_mujeres': b_m,
                    'base_hombres': b_h,
                    'confianza_mujeres': c_m,
                    'confianza_hombres': c_h,
                    'eventual_mujeres': ev_m,
                    'eventual_hombres': ev_h,
                    'base_puestos': base_puestos,
                    'confianza_puestos': confianza_puestos,
                    'eventual_puestos': eventual_puestos,
                }

            enlace_list = [get_row_stats(name, plist) for name, plist in sorted(enlace_operativo_groups.items())]
            mandos_list = [get_row_stats(name, plist) for name, plist in sorted(mandos_medios_groups.items())]
            
            # Age ranges for active staff
            age_ranges = {'18-26': 0, '26-34': 0, '34-42': 0, '42-50': 0, '50-58': 0}
            women_count = sum(1 for p in qs if p.sexo == 'F')
            men_count = sum(1 for p in qs if p.sexo == 'M')
            
            for p in qs:
                if p.fecha_nacimiento:
                    born = p.fecha_nacimiento
                    age = today.year - born.year - ((today.month, today.day) < (born.month, born.day))
                    if 18 <= age <= 26: age_ranges['18-26'] += 1
                    elif 27 <= age <= 34: age_ranges['26-34'] += 1
                    elif 35 <= age <= 42: age_ranges['34-42'] += 1
                    elif 43 <= age <= 50: age_ranges['42-50'] += 1
                    elif 51 <= age <= 58: age_ranges['50-58'] += 1
                    
            return {
                'total': total,
                'base': base,
                'confianza': confianza,
                'eventual': eventual,
                'enlace_operativo': enlace_list,
                'mandos_medios': mandos_list,
                'mujeres': women_count,
                'hombres': men_count,
                'edades': age_ranges
            }
            
        # Segment querysets
        qs_total = qs_all
        qs_vacantes = qs_all.filter(estatus__estatus__iexact='VACANTE')
        qs_activos = qs_all.filter(estatus__estatus__iexact='ACTIVO')
        
        data = {
            'estado_nombre': estado_nombre_display,
            'total': get_stats_for_qs(qs_total),
            'vacantes': get_stats_for_qs(qs_vacantes),
            'activos': get_stats_for_qs(qs_activos),
        }
        
        return JsonResponse({'status': 'success', 'data': data})
    except Estado.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'El estado solicitado no existe.'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)


def vehiculos_list(request):
    """Vista para la gestión de Vehículos con paginación y filtros en el servidor."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    
    from .models import VehiculosOR, TipoVeh, TipoAsignacionVeh, Estado, Inmueble, SituacionVeh
    
    user_state = get_user_state(request)
    
    # 1. Obtener filtros de la solicitud GET
    placa_query = request.GET.get('placa', '').strip()
    marca_query = request.GET.get('marca', '').strip()
    estado_id_query = request.GET.get('estado_id', '').strip()
    page_number = request.GET.get('page', 1)
    
    # 2. Filtrar queryset de vehículos
    vehiculos_qs = VehiculosOR.objects.all().select_related('tipoVeh', 'asignacion', 'estado', 'inmueble', 'fotografias', 'situacion')
    
    if user_state:
        vehiculos_qs = vehiculos_qs.filter(estado=user_state)
        estados_list_all = [user_state]
    else:
        estados_list_all = Estado.objects.all().order_by('nombre')
        if estado_id_query:
            vehiculos_qs = vehiculos_qs.filter(estado_id=estado_id_query)
            
    if placa_query:
        vehiculos_qs = vehiculos_qs.filter(placa__icontains=placa_query)
        
    if marca_query:
        vehiculos_qs = vehiculos_qs.filter(marca__icontains=marca_query)
        
    vehiculos_qs = vehiculos_qs.order_by('estado__nombre', 'marca', 'modelo')
    total_matched = vehiculos_qs.count()
    
    # 3. Paginación de resultados (50 por página)
    from django.core.paginator import Paginator
    paginator = Paginator(vehiculos_qs, 50)
    page_obj = paginator.get_page(page_number)
    
    # 4. Listados para formularios de creación/edición
    tipo_veh_list = TipoVeh.objects.all().order_by('tipo_veh')
    asignacion_list = TipoAsignacionVeh.objects.all().order_by('tipo')
    situacion_list = SituacionVeh.objects.all().order_by('situacion')
    
    if user_state:
        estados_list = [user_state]
        inmuebles_list = Inmueble.objects.filter(estado=user_state).order_by('nombre_inmueble')
    else:
        estados_list = Estado.objects.all().order_by('nombre')
        inmuebles_list = Inmueble.objects.all().order_by('nombre_inmueble')
        
    return render(request, 'mapa/vehiculos_list.html', {
        'estados_list': estados_list,
        'estados_list_all': estados_list_all,
        'inmuebles_list': inmuebles_list,
        'tipo_veh_list': tipo_veh_list,
        'asignacion_list': asignacion_list,
        'situacion_list': situacion_list,
        'vehiculos_list': page_obj,
        'total_matched': total_matched,
        'placa_query': placa_query,
        'marca_query': marca_query,
        'estado_id_query': estado_id_query,
    })


def guardar_vehiculo(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/')
        
    if request.method != 'POST':
        return redirect('vehiculos_list')
    
    user_state = get_user_state(request)
    from .models import VehiculosOR, FotosVeh
    from django.contrib import messages
    
    try:
        vehiculo_id = request.POST.get('vehiculo_id')
        estado_id = request.POST.get('estado_id')
        
        # Validación de seguridad: el estado enviado debe coincidir con el del usuario
        if user_state and str(user_state.id) != str(estado_id):
            raise PermissionError("No tienes permisos para registrar vehículos en este estado.")
            
        anio_year = request.POST.get('anio') or None
        anio = None
        if anio_year and anio_year.strip().isdigit():
            from datetime import date
            anio = date(int(anio_year), 1, 1)
            
        fecha_disp_comb = request.POST.get('fecha_disp_comb') or None
        
        # Procesar fotos de vehículos
        fotos_id = request.POST.get('fotografias_id') or None
        fotos_obj = None
        
        frente_file = request.FILES.get('foto_frente')
        lateral_file = request.FILES.get('foto_lateral')
        trasera_file = request.FILES.get('foto_trasera')
        
        if frente_file or lateral_file or trasera_file or fotos_id:
            if fotos_id:
                fotos_obj = FotosVeh.objects.get(id=fotos_id)
            else:
                fotos_obj = FotosVeh()
                
            if frente_file:
                fotos_obj.frente = frente_file
            if lateral_file:
                fotos_obj.lateral = lateral_file
            if trasera_file:
                fotos_obj.trasera = trasera_file
                
            fotos_obj.save()
            
        monto_raw = request.POST.get('monto') or '0'
        # Limpiar caracteres como signo de pesos o comas
        monto_raw = monto_raw.replace('$', '').replace(',', '').strip()
        monto = float(monto_raw) if monto_raw else 0.0
            
        fecha_asignacion = request.POST.get('fecha_asignacion') or None
        balizado = request.POST.get('balizado') == 'on'

        defaults = {
            'marca': request.POST.get('marca', '').strip().upper(),
            'modelo': request.POST.get('modelo', '').strip().upper(),
            'anio': anio,
            'placa': request.POST.get('placa', '').strip().upper(),
            'no_motor': request.POST.get('no_motor', '').strip().upper(),
            'tarjeta_asig': request.POST.get('tarjeta_asig', '').strip().upper() or None,
            'fecha_disp_comb': fecha_disp_comb,
            'monto': monto,
            'tipoVeh_id': request.POST.get('tipoVeh_id') or None,
            'asignacion_id': request.POST.get('asignacion_id'),
            'estado_id': estado_id,
            'inmueble_id': request.POST.get('inmueble_id') or None,
            'fotografias': fotos_obj,
            'situacion_id': request.POST.get('situacion_id') or None,
            'fecha_asignacion': fecha_asignacion,
            'balizado': balizado,
        }
        
        if vehiculo_id:
            vehiculo = VehiculosOR.objects.get(id=vehiculo_id)
            if user_state and vehiculo.estado != user_state:
                raise PermissionError("No tienes permisos para modificar este registro.")
            for key, value in defaults.items():
                setattr(vehiculo, key, value)
            vehiculo.save()
            messages.success(request, "Vehículo actualizado exitosamente.")
        else:
            vehiculo = VehiculosOR.objects.create(**defaults)
            
            # --- REGISTROS INICIALES OPCIONALES ---
            from .models import Kilometraje, PrestadoDe, Siniestros, Capufe, CombustibleExt
            
            # 1. Kilometraje
            init_odometro = request.POST.get('init_odometro')
            if init_odometro:
                tipo_unidad = request.POST.get('init_tipo_unidad', 'KM')
                evidencia = request.FILES.get('init_evidencia')
                Kilometraje.objects.create(
                    vehiculo=vehiculo,
                    fecha=fecha_disp_comb or anio,
                    tipo=tipo_unidad,
                    odometro=float(init_odometro),
                    evidencia=evidencia
                )
                
            # 2. Préstamo
            init_prestado_estado_id = request.POST.get('init_prestado_estado_id')
            if init_prestado_estado_id:
                PrestadoDe.objects.create(
                    vehiculo=vehiculo,
                    estado_id=init_prestado_estado_id,
                    inmueble_id=request.POST.get('init_prestado_inmueble_id') or None,
                    fecha_prestamo=request.POST.get('init_prestado_fecha') or None
                )
                
            # 3. Siniestro
            init_siniestro_fecha = request.POST.get('init_siniestro_fecha')
            if init_siniestro_fecha:
                Siniestros.objects.create(
                    vehiculo=vehiculo,
                    fecha=init_siniestro_fecha,
                    folio=request.POST.get('init_siniestro_folio', '').strip().upper() or None
                )
                
            # 4. Capufe
            init_capufe_fecha_inicio = request.POST.get('init_capufe_fecha_inicio')
            if init_capufe_fecha_inicio:
                Capufe.objects.create(
                    vehiculo=vehiculo,
                    fecha_inicio=init_capufe_fecha_inicio,
                    fecha_termino=request.POST.get('init_capufe_fecha_termino') or None
                )
                
            # 5. Combustible Extra
            init_combustible_monto = request.POST.get('init_combustible_monto')
            if init_combustible_monto:
                monto_comb = init_combustible_monto.replace('$', '').replace(',', '').strip()
                if monto_comb:
                    CombustibleExt.objects.create(
                        vehiculo=vehiculo,
                        fecha=request.POST.get('init_combustible_fecha') or None,
                        monto=float(monto_comb)
                    )
            
            messages.success(request, "Vehículo registrado exitosamente con sus datos iniciales.")
            
    except Exception as e:
        messages.error(request, f"Error al guardar el vehículo: {str(e)}")
        
    return redirect('vehiculos_list')


def eliminar_vehiculo(request, vehiculo_id):
    if not request.user.is_authenticated:
        return redirect('/log-in/')
        
    from .models import VehiculosOR
    from django.contrib import messages
    
    user_state = get_user_state(request)
    
    try:
        vehiculo = VehiculosOR.objects.get(id=vehiculo_id)
        if user_state and vehiculo.estado != user_state:
            raise PermissionError("No tienes permisos para eliminar este registro.")
            
        fotos_obj = vehiculo.fotografias
        vehiculo.delete()
        
        # Eliminar registro de fotos física y de BD si existe
        if fotos_obj:
            fotos_obj.delete()
            
        messages.success(request, "Vehículo eliminado exitosamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar el vehículo: {str(e)}")
        
    return redirect('vehiculos_list')


def api_get_vehiculo(request, vehiculo_id):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'No autenticado'}, status=401)
        
    from .models import VehiculosOR
    user_state = get_user_state(request)
    
    try:
        vehiculo = VehiculosOR.objects.select_related('fotografias').get(id=vehiculo_id)
        if user_state and vehiculo.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        data = {
            'id': vehiculo.id,
            'marca': vehiculo.marca,
            'modelo': vehiculo.modelo,
            'anio': str(vehiculo.anio.year) if vehiculo.anio else '',
            'placa': vehiculo.placa,
            'no_motor': vehiculo.no_motor,
            'tarjeta_asig': vehiculo.tarjeta_asig or '',
            'fecha_disp_comb': vehiculo.fecha_disp_comb.strftime('%Y-%m-%d') if vehiculo.fecha_disp_comb else '',
            'monto': str(vehiculo.monto),
            'tipoVeh_id': vehiculo.tipoVeh_id or '',
            'asignacion_id': vehiculo.asignacion_id,
            'estado_id': vehiculo.estado_id,
            'inmueble_id': vehiculo.inmueble_id or '',
            'fotografias_id': vehiculo.fotografias_id or '',
            'foto_frente_url': vehiculo.fotografias.frente.url if (vehiculo.fotografias and vehiculo.fotografias.frente) else '',
            'foto_lateral_url': vehiculo.fotografias.lateral.url if (vehiculo.fotografias and vehiculo.fotografias.lateral) else '',
            'foto_trasera_url': vehiculo.fotografias.trasera.url if (vehiculo.fotografias and vehiculo.fotografias.trasera) else '',
            'situacion_id': vehiculo.situacion_id or '',
            'fecha_asignacion': vehiculo.fecha_asignacion.strftime('%Y-%m-%d') if vehiculo.fecha_asignacion else '',
            'balizado': vehiculo.balizado,
        }
        return JsonResponse({'status': 'success', 'data': data})
    except VehiculosOR.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Vehículo no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def api_get_vehiculo_historial(request, vehiculo_id):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'No autenticado'}, status=401)
        
    from .models import VehiculosOR, Kilometraje, PrestadoDe, Siniestros, Capufe, CombustibleExt
    user_state = get_user_state(request)
    
    try:
        vehiculo = VehiculosOR.objects.get(id=vehiculo_id)
        if user_state and vehiculo.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        kms = Kilometraje.objects.filter(vehiculo_id=vehiculo_id).order_by('-fecha', '-id')
        prestados = PrestadoDe.objects.filter(vehiculo_id=vehiculo_id).select_related('estado', 'inmueble').order_by('-fecha_prestamo')
        siniestros = Siniestros.objects.filter(vehiculo_id=vehiculo_id).order_by('-fecha')
        capufes = Capufe.objects.filter(vehiculo_id=vehiculo_id).order_by('-fecha_inicio')
        combustibles = CombustibleExt.objects.filter(vehiculo_id=vehiculo_id).order_by('-fecha')
        
        data = {
            'vehiculo': {
                'id': vehiculo.id,
                'placa': vehiculo.placa,
                'marca': vehiculo.marca,
                'modelo': vehiculo.modelo,
            },
            'kilometraje': [{
                'id': k.id,
                'fecha': k.fecha.strftime('%Y-%m-%d') if k.fecha else 'S/F',
                'tipo': k.tipo,
                'odometro': str(k.odometro),
                'evidencia_url': k.evidencia.url if k.evidencia else ''
            } for k in kms],
            'prestados': [{
                'id': p.id,
                'fecha_prestamo': p.fecha_prestamo.strftime('%Y-%m-%d') if p.fecha_prestamo else 'S/F',
                'estado': p.estado.nombre,
                'inmueble': p.inmueble.nombre_inmueble if p.inmueble else 'Sin Asignar'
            } for p in prestados],
            'siniestros': [{
                'id': s.id,
                'fecha': s.fecha.strftime('%Y-%m-%d') if s.fecha else 'S/F',
                'folio': s.folio or 'S/F'
            } for s in siniestros],
            'capufes': [{
                'id': c.id,
                'fecha_inicio': c.fecha_inicio.strftime('%Y-%m-%d') if c.fecha_inicio else 'S/F',
                'fecha_termino': c.fecha_termino.strftime('%Y-%m-%d') if c.fecha_termino else 'S/F'
            } for c in capufes],
            'combustibles': [{
                'id': co.id,
                'fecha': co.fecha.strftime('%Y-%m-%d') if co.fecha else 'S/F',
                'monto': str(co.monto)
            } for co in combustibles],
        }
        return JsonResponse({'status': 'success', 'data': data})
    except VehiculosOR.DoesNotExist:
        return JsonResponse({'status': 'error', 'message': 'Vehículo no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def guardar_kilometraje(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
        
    from .models import VehiculosOR, Kilometraje
    user_state = get_user_state(request)
    
    try:
        vehiculo_id = request.POST.get('vehiculo_id')
        vehiculo = VehiculosOR.objects.get(id=vehiculo_id)
        if user_state and vehiculo.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        fecha = request.POST.get('fecha') or None
        tipo = request.POST.get('tipo', 'KM')
        odometro = float(request.POST.get('odometro') or 0.0)
        evidencia = request.FILES.get('evidencia')
        
        Kilometraje.objects.create(
            vehiculo=vehiculo,
            fecha=fecha,
            tipo=tipo,
            odometro=odometro,
            evidencia=evidencia
        )
        return JsonResponse({'status': 'success', 'message': 'Kilometraje registrado exitosamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def guardar_prestado(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
        
    from .models import VehiculosOR, PrestadoDe
    user_state = get_user_state(request)
    
    try:
        vehiculo_id = request.POST.get('vehiculo_id')
        vehiculo = VehiculosOR.objects.get(id=vehiculo_id)
        if user_state and vehiculo.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        estado_id = request.POST.get('estado_id')
        inmueble_id = request.POST.get('inmueble_id') or None
        fecha_prestamo = request.POST.get('fecha_prestamo') or None
        
        PrestadoDe.objects.create(
            vehiculo=vehiculo,
            estado_id=estado_id,
            inmueble_id=inmueble_id,
            fecha_prestamo=fecha_prestamo
        )
        return JsonResponse({'status': 'success', 'message': 'Préstamo registrado exitosamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def guardar_siniestro(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
        
    from .models import VehiculosOR, Siniestros
    user_state = get_user_state(request)
    
    try:
        vehiculo_id = request.POST.get('vehiculo_id')
        vehiculo = VehiculosOR.objects.get(id=vehiculo_id)
        if user_state and vehiculo.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        fecha = request.POST.get('fecha') or None
        folio = request.POST.get('folio', '').strip().upper() or None
        
        Siniestros.objects.create(
            vehiculo=vehiculo,
            fecha=fecha,
            folio=folio
        )
        return JsonResponse({'status': 'success', 'message': 'Siniestro registrado exitosamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def guardar_capufe(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
        
    from .models import VehiculosOR, Capufe
    user_state = get_user_state(request)
    
    try:
        vehiculo_id = request.POST.get('vehiculo_id')
        vehiculo = VehiculosOR.objects.get(id=vehiculo_id)
        if user_state and vehiculo.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        fecha_inicio = request.POST.get('fecha_inicio') or None
        fecha_termino = request.POST.get('fecha_termino') or None
        
        Capufe.objects.create(
            vehiculo=vehiculo,
            fecha_inicio=fecha_inicio,
            fecha_termino=fecha_termino
        )
        return JsonResponse({'status': 'success', 'message': 'Registro de Capufe guardado exitosamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


def guardar_combustible(request):
    if not request.user.is_authenticated:
        return JsonResponse({'status': 'error', 'message': 'No autenticado'}, status=401)
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Método no permitido'}, status=405)
        
    from .models import VehiculosOR, CombustibleExt
    user_state = get_user_state(request)
    
    try:
        vehiculo_id = request.POST.get('vehiculo_id')
        vehiculo = VehiculosOR.objects.get(id=vehiculo_id)
        if user_state and vehiculo.estado != user_state:
            return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
            
        fecha = request.POST.get('fecha') or None
        monto_raw = request.POST.get('monto') or '0'
        monto_raw = monto_raw.replace('$', '').replace(',', '').strip()
        monto = float(monto_raw) if monto_raw else 0.0
        
        CombustibleExt.objects.create(
            vehiculo=vehiculo,
            fecha=fecha,
            monto=monto
        )
        return JsonResponse({'status': 'success', 'message': 'Registro de Combustible guardado exitosamente'})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# =====================================================================
# GESTIÓN DE PRHS (PUNTOS DE RESCATE HUMANO)
# =====================================================================

def prhs_list(request):
    """Muestra la lista de PRHs y gestiona sus catálogos."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    user_state = get_user_state(request)
    
    from .models import PRHs, TipoPRH, Estado
    if request.user.is_superuser:
        prhs_list_qs = PRHs.objects.all().select_related('estado', 'modalidad').order_by('estado__nombre', 'nombre')
        estados_list = Estado.objects.all().order_by('nombre')
    elif user_state:
        prhs_list_qs = PRHs.objects.filter(estado=user_state).select_related('estado', 'modalidad').order_by('nombre')
        estados_list = [user_state]
    else:
        prhs_list_qs = PRHs.objects.none()
        estados_list = []
        
    return render(request, 'mapa/prhs.html', {
        'prhs_list': prhs_list_qs,
        'estados_list': estados_list,
        'modalidades_list': TipoPRH.objects.all().order_by('nombre'),
    })


@transaction.atomic
def guardar_prh(request):
    if request.method != 'POST':
        return redirect('prhs_list')
        
    user_state = get_user_state(request)
    from .models import PRHs
    
    try:
        prh_id = request.POST.get('prh_id')
        estado_id = request.POST.get('estado_id')
        
        # Validación de seguridad: el estado enviado debe coincidir con el del usuario
        if not request.user.is_superuser:
            if not user_state or str(user_state.id) != str(estado_id):
                raise PermissionError("No tienes permisos para registrar PRH en este estado.")
        
        # Coordenadas
        try:
            lat = float(request.POST.get('latitud') or 0)
            lng = float(request.POST.get('longitud') or 0)
        except ValueError:
            raise ValueError("La latitud y longitud deben ser coordenadas numéricas válidas.")

        activo = request.POST.get('activo') == 'on' or request.POST.get('activo') == 'true'

        defaults = {
            'estado_id': estado_id,
            'nombre': normalizar_nombre(request.POST.get('nombre', '')),
            'modalidad_id': request.POST.get('modalidad_id'),
            'activo': activo,
            'coordenadasTexto': f"{lat}, {lng}",
            'latitud': lat,
            'longitud': lng,
        }

        if prh_id:
            prh = PRHs.objects.get(id=prh_id)
            if not request.user.is_superuser and prh.estado != user_state:
                raise PermissionError("No tienes permisos para modificar este PRH.")
            for key, val in defaults.items():
                setattr(prh, key, val)
            prh.save()
            created = False
        else:
            prh = PRHs.objects.create(**defaults)
            created = True

        messages.success(request, f"Punto de Rescate Humano '{prh.nombre}' {'creado' if created else 'actualizado'} correctamente.")
    except Exception as e:
        messages.error(request, f"Error al guardar PRH: {str(e)}")
        
    return redirect('prhs_list')


@transaction.atomic
def eliminar_prh(request, prh_id):
    """Elimina un PRH si tiene permisos por estado."""
    from .models import PRHs
    try:
        user_state = get_user_state(request)
        prh = PRHs.objects.get(id=prh_id)
        
        # Validación de seguridad
        if not request.user.is_superuser:
            if not user_state or prh.estado != user_state:
                messages.error(request, "No tienes permisos para eliminar este registro.")
                return redirect('prhs_list')
                
        nombre = prh.nombre
        prh.delete()
        messages.success(request, f"Punto de Rescate Humano '{nombre}' eliminado correctamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar PRH: {str(e)}")
        
    return redirect('prhs_list')


def api_get_prh(request, prh_id):
    """Retorna los datos de un PRH en formato JSON."""
    from .models import PRHs
    try:
        user_state = get_user_state(request)
        prh = PRHs.objects.get(id=prh_id)
        
        # Validación de seguridad
        if not request.user.is_superuser:
            if not user_state or prh.estado != user_state:
                return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
                
        data = {
            'id': prh.id,
            'nombre': prh.nombre,
            'estado_id': prh.estado_id,
            'modalidad_id': prh.modalidad_id,
            'activo': prh.activo,
            'latitud': prh.latitud,
            'longitud': prh.longitud,
        }
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# =====================================================================
# GESTIÓN DE PUNTOS DE INTERNACIÓN Y ESTACIONES
# =====================================================================

def puntos_internacion_list(request):
    """Muestra la lista de puntos de internación y estaciones."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    user_state = get_user_state(request)
    
    from .models import PuntosInternacionEstacion, Estado
    if request.user.is_superuser:
        puntos_list = PuntosInternacionEstacion.objects.all().select_related('estado').order_by('estado__nombre', 'nombre')
        estados_list = Estado.objects.all().order_by('nombre')
    elif user_state:
        puntos_list = PuntosInternacionEstacion.objects.filter(estado=user_state).select_related('estado').order_by('nombre')
        estados_list = [user_state]
    else:
        puntos_list = PuntosInternacionEstacion.objects.none()
        estados_list = []
        
    return render(request, 'mapa/puntos_internacion.html', {
        'puntos_list': puntos_list,
        'estados_list': estados_list,
        'tipos_list': [('AEREO', 'AEREO'), ('MARITIMO', 'MARITIMO'), ('TERRESTRE', 'TERRESTRE'), ('ESTACION', 'ESTACION')],
    })


@transaction.atomic
def guardar_punto_internacion(request):
    if request.method != 'POST':
        return redirect('puntos_internacion_list')
        
    user_state = get_user_state(request)
    from .models import PuntosInternacionEstacion
    
    try:
        punto_id = request.POST.get('punto_id')
        estado_id = request.POST.get('estado_id')
        
        # Validación de seguridad: el estado enviado debe coincidir con el del usuario
        if not request.user.is_superuser:
            if not user_state or str(user_state.id) != str(estado_id):
                raise PermissionError("No tienes permisos para registrar puntos de internación en este estado.")
        
        # Coordenadas
        try:
            lat = float(request.POST.get('latitud') or 0)
            lng = float(request.POST.get('longitud') or 0)
        except ValueError:
            raise ValueError("La latitud y longitud deben ser coordenadas numéricas válidas.")

        defaults = {
            'estado_id': estado_id,
            'nombre': normalizar_nombre(request.POST.get('nombre', '')),
            'tipo': request.POST.get('tipo'),
            'latitud': lat,
            'longitud': lng,
        }

        if punto_id:
            punto = PuntosInternacionEstacion.objects.get(id=punto_id)
            if not request.user.is_superuser and punto.estado != user_state:
                raise PermissionError("No tienes permisos para modificar este punto de internación.")
            for key, val in defaults.items():
                setattr(punto, key, val)
            punto.save()
            created = False
        else:
            punto = PuntosInternacionEstacion.objects.create(**defaults)
            created = True

        messages.success(request, f"Punto de Internación/Estación '{punto.nombre}' {'creado' if created else 'actualizado'} correctamente.")
    except Exception as e:
        messages.error(request, f"Error al guardar punto de internación: {str(e)}")
        
    return redirect('puntos_internacion_list')


@transaction.atomic
def eliminar_punto_internacion(request, punto_id):
    """Elimina un punto de internación/estación si tiene permisos por estado."""
    from .models import PuntosInternacionEstacion
    try:
        user_state = get_user_state(request)
        punto = PuntosInternacionEstacion.objects.get(id=punto_id)
        
        # Validación de seguridad
        if not request.user.is_superuser:
            if not user_state or punto.estado != user_state:
                messages.error(request, "No tienes permisos para eliminar este registro.")
                return redirect('puntos_internacion_list')
                
        nombre = punto.nombre
        punto.delete()
        messages.success(request, f"Punto de Internación/Estación '{nombre}' eliminado correctamente.")
    except Exception as e:
        messages.error(request, f"Error al eliminar punto de internación: {str(e)}")
        
    return redirect('puntos_internacion_list')


def api_get_punto_internacion(request, punto_id):
    """Retorna los datos de un punto de internación/estación en formato JSON."""
    from .models import PuntosInternacionEstacion
    try:
        user_state = get_user_state(request)
        punto = PuntosInternacionEstacion.objects.get(id=punto_id)
        
        # Validación de seguridad
        if not request.user.is_superuser:
            if not user_state or punto.estado != user_state:
                return JsonResponse({'status': 'error', 'message': 'Sin permisos'}, status=403)
                
        data = {
            'id': punto.id,
            'nombre': punto.nombre,
            'estado_id': punto.estado_id,
            'tipo': punto.tipo,
            'latitud': punto.latitud,
            'longitud': punto.longitud,
        }
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=400)


# =============================================================================
# Parque Vehicular -- capa de consulta/visualizacion de solo lectura sobre
# VehiculosOR (integrada aqui desde la antigua app 'vehiculos'; el alta/edicion
# real sigue siendo guardar_vehiculo, guardar_kilometraje, etc. mas arriba).
# =============================================================================

# --- Helpers de conversión (VehiculosOR -> dict amigable para templates) --

def _color_situacion(situacion_nombre):
    """Color consistente en todo el módulo: verde para Activo, amarillo
    para Mantenimiento, rojo para cualquier otra cosa (ej. Posible baja,
    o valores no reconocidos). Centralizado aquí para no repetir la
    lógica de comparación de texto en cada template por separado.

    La situación es texto libre en la base real (ej. "ACTIVO",
    "SERVICIO/MANTENIMIENTO", "PARA BAJA" -- no un catálogo cerrado de 3
    valores), por eso "mantenimiento" se busca como substring y no con
    igualdad exacta; cualquier otra cosa cae en "baja"."""
    nombre = (situacion_nombre or "").strip().upper()
    if nombre == "ACTIVO":
        return "activo"
    if "MANTENIMIENTO" in nombre:
        return "mantenimiento"
    return "baja"


# Valores de relleno que captura usa cuando el vehículo no tiene placa real
# (texto libre, sin catálogo cerrado -- de ahí la variedad). Ninguno sirve
# como identificador: varios vehículos distintos comparten el mismo texto
# (ej. 15 con "SIN NUMERO DE PLACA"), así que buscar por placa devolvería
# siempre el mismo vehículo del grupo sin importar cuál se haya clickeado.
# Estos casos deben enlazarse por id, no por placa. ("S/P" y "S/N" no están
# aquí porque ya los captura el chequeo de "/" en _placa_ambigua().)
_PLACAS_SIN_ASIGNAR = {"SIN PLACAS", "SIN NUMERO DE PLACA", "NO APLICA"}


def _placas_duplicadas():
    """Placas (normalizadas) que le pertenecen a mas de un vehiculo en toda
    la tabla -- no solo relleno compartido a propósito (_PLACAS_SIN_ASIGNAR),
    también coincidencias reales de captura (dos vehículos distintos con la
    misma placa, ej. mismo folio de flotilla reusado). _obtener_vehiculo_o_404
    busca sobre TODA la tabla sin filtro, así que la ambigüedad es global:
    no basta con revisar duplicados dentro de un listado ya filtrado."""
    conteo = (
        VehiculosOR.objects.exclude(placa="")
        .annotate(placa_norm=Upper("placa"))
        .values("placa_norm")
        .annotate(n=Count("id"))
        .filter(n__gt=1)
        .values_list("placa_norm", flat=True)
    )
    return set(conteo)


def _placa_ambigua(placa, duplicadas=None):
    if not placa or "/" in placa:
        return True
    normalizada = placa.strip().upper()
    if normalizada in _PLACAS_SIN_ASIGNAR:
        return True
    return bool(duplicadas) and normalizada in duplicadas


def _urls_detalle_vehiculo(vehiculo_id, placa, ambigua):
    """URLs de 'Ver detalle'/'Ver ficha': por id si la placa es ambigua, por
    placa si no. Centralizado aquí para no repetir el mismo {% if
    placa_ambigua %}...{% else %}...{% endif %} con el mismo <a> en cada
    template que lista vehículos (listado, resumen por estado, popovers)."""
    if ambigua:
        return (
            reverse("vehiculos:detalle_id", args=[vehiculo_id]),
            reverse("vehiculos:detalle_fragmento_id", args=[vehiculo_id]),
        )
    return (
        reverse("vehiculos:detalle", args=[placa]),
        reverse("vehiculos:detalle_fragmento", args=[placa]),
    )


def _vehiculo_a_dict(v, duplicadas=None):
    """Convierte una instancia de VehiculosOR a un dict con nombres
    estables, para no acoplar los templates a los nombres de campo reales
    (que pueden tener sus propias particularidades, ej. 'tipoVeh').
    'duplicadas' (ver _placas_duplicadas) se calcula una sola vez por vista
    y se pasa aquí -- no tiene caso reconsultarlo por cada fila."""
    situacion_nombre = v.situacion.situacion if v.situacion_id else "Sin especificar"
    ambigua = _placa_ambigua(v.placa, duplicadas)
    detalle_url, detalle_fragmento_url = _urls_detalle_vehiculo(v.id, v.placa, ambigua)
    return {
        "id": v.id,
        "placa": v.placa,
        "placa_ambigua": ambigua,
        "detalle_url": detalle_url,
        "detalle_fragmento_url": detalle_fragmento_url,
        "marca": v.marca,
        "modelo": v.modelo,
        "anio": v.anio.year if v.anio else "",
        "no_motor": v.no_motor,
        "tipo_vehiculo": v.tipoVeh.tipo_veh if v.tipoVeh else "Sin especificar",
        "tipo_asignacion": v.asignacion.tipo if v.asignacion_id else "Sin especificar",
        "situacion": situacion_nombre,
        "situacion_color": _color_situacion(situacion_nombre),
        "estado": v.estado.nombre if v.estado_id else "",
        "inmueble_destino": v.inmueble.nombre_inmueble if v.inmueble_id else "Sin asignar",
        "tarjeta": v.tarjeta_asig or "",
        "balizado": v.balizado,
        "fecha_asignacion": v.fecha_asignacion,
    }


def _queryset_base():
    return VehiculosOR.objects.select_related("tipoVeh", "asignacion", "situacion", "estado", "inmueble", "fotografias")


def _es_activo(v):
    return bool(v.situacion_id) and v.situacion.situacion.strip().upper() == "ACTIVO"


# --- Filtros compartidos (listado, popover, excel) ----------------------

def _filtrar_por_situacion(qs, situacion):
    """Filtra por el mismo "bucket" que ya usa _color_situacion() (Activo /
    Mantenimiento / lo demas), en vez de un iexact contra el texto libre
    real -- si no, las tarjetas de resumen ("Activos", "Mantenimiento",
    "Posible baja") llevarian a un listado vacio o equivocado en cuanto
    la situacion real no coincidiera con el texto exacto del boton
    (ej. "SERVICIO/MANTENIMIENTO" o "PARA BAJA" en la base real)."""
    nombre = situacion.strip().upper()
    if nombre == "ACTIVO":
        return qs.filter(situacion__situacion__iexact="Activo")
    if nombre == "MANTENIMIENTO":
        return qs.filter(situacion__situacion__icontains="Mantenimiento")
    # "Posible baja" (o cualquier otro valor): todo lo que no cae en los
    # dos buckets de arriba, igual que el "return baja" de _color_situacion.
    return qs.exclude(situacion__situacion__iexact="Activo").exclude(
        situacion__situacion__icontains="Mantenimiento"
    )


def _aplicar_filtros(qs, request):
    tipo = request.GET.get("tipo", "").strip()
    situacion = request.GET.get("situacion", "").strip()
    asignacion = request.GET.get("asignacion", "").strip()
    estado = request.GET.get("estado", "").strip()
    placa = request.GET.get("placa", "").strip()

    if tipo:
        qs = qs.filter(tipoVeh__tipo_veh__iexact=tipo)
    if situacion:
        qs = _filtrar_por_situacion(qs, situacion)
    if asignacion:
        qs = qs.filter(asignacion__tipo__iexact=asignacion)
    if estado:
        qs = qs.filter(estado__nombre__iexact=normalizar_nombre(estado))
    if placa:
        qs = qs.filter(placa__icontains=placa)
    return qs


# --- Conteo para la tarjeta "Parque Vehicular" del menu del mapa ---------

@require_GET
def conteo_total(request):
    """JSON minimo (solo el total) para la tarjeta del menu del mapa, que
    antes mostraba un "32" fijo sin actualizar. Acepta '?estado=' opcional
    por si algun dia se quiere el conteo de un estado en vez del nacional."""
    estado_nombre = request.GET.get("estado", "").strip()
    qs = _queryset_base()
    if estado_nombre:
        qs = qs.filter(estado__nombre__iexact=normalizar_nombre(estado_nombre))
    return JsonResponse({"total": qs.count()})


# --- Vista general (dashboard) -------------------------------------------

def vehiculos_dashboard(request):
    filas = list(_queryset_base())
    total = len(filas)
    activos = sum(1 for v in filas if _es_activo(v))
    mantenimiento = sum(1 for v in filas if _color_situacion(v.situacion.situacion if v.situacion_id else "") == "mantenimiento")
    posible_baja = total - activos - mantenimiento

    # SituacionVeh es texto libre en la base real (sin catálogo fijo de 3
    # valores) — se cuenta todo lo que NO sea "Activo" como un solo grupo,
    # y se muestra el desglose real de esas otras situaciones aparte.
    otras_situaciones = {}
    for v in filas:
        if not _es_activo(v):
            nombre = v.situacion.situacion if v.situacion_id else "Sin especificar"
            otras_situaciones[nombre] = otras_situaciones.get(nombre, 0) + 1

    tarjetas_asignadas = sum(1 for v in filas if v.tarjeta_asig)

    monto_combustible = sum((c.monto for c in CombustibleExt.objects.only("monto")), start=0)

    # "Total kilómetros" = suma de la lectura MÁS RECIENTE de cada
    # vehículo (no la suma de todo el historial, que sobrecontaría).
    ultima_lectura_por_vehiculo = {}
    for k in Kilometraje.objects.order_by("vehiculo_id", "-fecha").only("vehiculo_id", "odometro"):
        if k.vehiculo_id not in ultima_lectura_por_vehiculo:
            ultima_lectura_por_vehiculo[k.vehiculo_id] = k.odometro
    total_km = sum(ultima_lectura_por_vehiculo.values(), start=0)

    conteo_tipos = Counter(v.tipoVeh.tipo_veh for v in filas if v.tipoVeh_id)
    tipos_catalogo = TipoVeh.objects.all().order_by("tipo_veh")
    tipos_vehiculo = [
        {"nombre": t.tipo_veh, "cantidad": conteo_tipos.get(t.tipo_veh, 0)} for t in tipos_catalogo
    ]

    # Igual que con situación: TipoAsignacionVeh es texto libre en la base
    # real, así que no asumimos que solo existen "Propio"/"Arrendado".
    asignacion_conteo = {}
    for v in filas:
        if v.asignacion_id:
            nombre = v.asignacion.tipo
            asignacion_conteo[nombre] = asignacion_conteo.get(nombre, 0) + 1

    context = {
        "resumen_estado": {
            "activos": activos,
            "mantenimiento": mantenimiento,
            "posible_baja": posible_baja,
            "total": total,
            "otras_situaciones": otras_situaciones,
        },
        # Categorias (tipo_vehiculo) en vez de "modelos distintos": el
        # campo "modelo" es texto libre y se infla por errores de captura
        # (typos, mayusculas, marca repetida con otro nombre); tipo_vehiculo
        # es un catalogo cerrado, asi que el conteo es confiable.
        "categorias_distintas": len(conteo_tipos),
        "asignacion_conteo": asignacion_conteo,
        "tarjetas_asignadas": tarjetas_asignadas,
        "monto_combustible": monto_combustible,
        "total_kilometros": total_km,
        "tipos_vehiculo": tipos_vehiculo,
        "situacion_opciones": sorted({v.situacion.situacion for v in filas if v.situacion_id}),
        "asignacion_opciones": sorted(asignacion_conteo.keys()),
        "estado_opciones": sorted({v.estado.nombre for v in filas if v.estado_id}),
    }
    return render(request, "vehiculos/dashboard.html", context)


@require_GET
def filtrar_dashboard(request):
    qs = _aplicar_filtros(_queryset_base(), request)
    conteo_tipos = Counter(v.tipoVeh.tipo_veh for v in qs if v.tipoVeh_id)
    tipos_catalogo = TipoVeh.objects.all().order_by("tipo_veh")
    tipos_vehiculo = [
        {"nombre": t.tipo_veh, "cantidad": conteo_tipos.get(t.tipo_veh, 0)} for t in tipos_catalogo
    ]
    if request.GET.get("solo_con_unidades") == "1":
        tipos_vehiculo = [t for t in tipos_vehiculo if t["cantidad"] > 0]
    return render(request, "vehiculos/_tipos_grid.html", {"tipos_vehiculo": tipos_vehiculo})


# --- Listado filtrable -----------------------------------------------------

def vehiculos_listado(request):
    context = {
        "vehiculos": [],  # se llena vía AJAX al cargar, igual que el resto
        "tipos_opciones": TipoVeh.objects.all().order_by("tipo_veh").values_list("tipo_veh", flat=True),
        "situacion_opciones": SituacionVeh.objects.all().order_by("situacion").values_list("situacion", flat=True),
        "asignacion_opciones": TipoAsignacionVeh.objects.all().order_by("tipo").values_list("tipo", flat=True),
        "estado_opciones": Estado.objects.all().order_by("nombre").values_list("nombre", flat=True),
    }
    return render(request, "vehiculos/listado.html", context)


@require_GET
def filtrar_listado(request):
    qs = _aplicar_filtros(_queryset_base(), request).order_by("estado__nombre", "marca", "modelo")
    duplicadas = _placas_duplicadas()
    vehiculos = [_vehiculo_a_dict(v, duplicadas) for v in qs]
    return render(request, "vehiculos/_listado_filas.html", {"vehiculos": vehiculos})


@require_GET
def listado_fragmento(request):
    """Listado completo, pero como fragmento para abrir DENTRO del mismo
    popup del mapa (VehiculosModal), no como página aparte. Se puede
    llegar con '?estado=<nombre>' para llegar ya filtrado a ese estado
    (por ejemplo, viniendo del botón 'Ver listado completo' del resumen)."""
    estado_inicial = request.GET.get("estado", "").strip()
    qs = _aplicar_filtros(_queryset_base(), request).order_by("estado__nombre", "marca", "modelo")
    duplicadas = _placas_duplicadas()
    vehiculos = [_vehiculo_a_dict(v, duplicadas) for v in qs]

    context = {
        "vehiculos": vehiculos,
        "estado_inicial": estado_inicial,
        "tipos_opciones": TipoVeh.objects.all().order_by("tipo_veh").values_list("tipo_veh", flat=True),
        "situacion_opciones": SituacionVeh.objects.all().order_by("situacion").values_list("situacion", flat=True),
        "asignacion_opciones": TipoAsignacionVeh.objects.all().order_by("tipo").values_list("tipo", flat=True),
        "estado_opciones": Estado.objects.all().order_by("nombre").values_list("nombre", flat=True),
    }
    return render(request, "vehiculos/_listado_modal.html", context)


# --- Ficha de detalle (solo lectura) -------------------------------------

def _historial_vehiculo(vehiculo_obj, fecha_inicio=None, fecha_fin=None):
    # kilometraje/combustible se materializan en listas (no queryset lazy)
    # porque _resumen_km_monto() indexa [0]/[-1] sobre ellas; así se
    # reutiliza la misma consulta en vez de disparar una segunda.
    # El filtro de fecha solo aplica a estas dos -- es lo único que pidió
    # el calendario de la ficha (siniestros/capufe/préstamos quedan igual).
    km_qs = Kilometraje.objects.filter(vehiculo=vehiculo_obj)
    combustible_qs = CombustibleExt.objects.filter(vehiculo=vehiculo_obj)
    if fecha_inicio:
        km_qs = km_qs.filter(fecha__gte=fecha_inicio)
        combustible_qs = combustible_qs.filter(fecha__gte=fecha_inicio)
    if fecha_fin:
        km_qs = km_qs.filter(fecha__lte=fecha_fin)
        combustible_qs = combustible_qs.filter(fecha__lte=fecha_fin)

    return {
        "kilometraje": list(km_qs.order_by("-fecha")),
        "combustible": list(combustible_qs.order_by("-fecha")),
        "siniestros": Siniestros.objects.filter(vehiculo=vehiculo_obj).order_by("-fecha"),
        "capufe": Capufe.objects.filter(vehiculo=vehiculo_obj).order_by("-fecha_inicio"),
        "prestamos": PrestadoDe.objects.filter(vehiculo=vehiculo_obj).select_related("estado", "inmueble").order_by("-fecha_prestamo"),
    }


def _resumen_km_monto(historial):
    """Resumen para la ficha: el kilometraje son lecturas de odómetro
    (no tiene sentido sumarlas), así que se muestra el actual contra el
    inicial. El combustible externo son montos independientes por
    dispersión, así que ahí sí aplica la sumatoria total."""
    km_lecturas = historial["kilometraje"]  # ya viene ordenado -fecha
    km_inicial = km_lecturas[-1].odometro if km_lecturas else None
    km_actual = km_lecturas[0].odometro if km_lecturas else None
    km_recorridos = (
        km_actual - km_inicial if km_inicial is not None and km_actual is not None else None
    )

    combustible_registros = historial["combustible"]
    monto_total = sum((c.monto for c in combustible_registros), start=0) if combustible_registros else None

    return {
        "km_inicial": km_inicial,
        "km_actual": km_actual,
        "km_recorridos": km_recorridos,
        "monto_total": monto_total,
    }


def _fotos_vehiculo(vehiculo_obj):
    """Solo se usa en la ficha de detalle -- por eso no se agrega a
    _vehiculo_a_dict() (que tambien alimenta listado/popover/excel, donde
    no hace falta cargar fotos por cada fila)."""
    fotos = vehiculo_obj.fotografias
    if not fotos:
        return {"foto_frente": None, "foto_lateral": None, "foto_trasera": None}
    return {
        "foto_frente": fotos.frente.url if fotos.frente else None,
        "foto_lateral": fotos.lateral.url if fotos.lateral else None,
        "foto_trasera": fotos.trasera.url if fotos.trasera else None,
    }


def _evidencia_km_actual(historial):
    """Foto de evidencia (si la tiene) de la lectura de kilometraje mas
    reciente -- historial["kilometraje"] ya viene ordenado -fecha."""
    lecturas = historial["kilometraje"]
    if lecturas and lecturas[0].evidencia:
        return lecturas[0].evidencia.url
    return None


def _obtener_vehiculo_o_404(placa):
    vehiculo_obj = _queryset_base().filter(placa__iexact=placa).first()
    if vehiculo_obj is None:
        raise Http404("No se encontró un vehículo con esa placa.")
    return vehiculo_obj


def _obtener_vehiculo_por_id_o_404(vehiculo_id):
    vehiculo_obj = _queryset_base().filter(pk=vehiculo_id).first()
    if vehiculo_obj is None:
        raise Http404("No se encontró un vehículo con ese id.")
    return vehiculo_obj


def _contexto_ficha(request, vehiculo_obj):
    fecha_inicio = request.GET.get("fecha_inicio", "").strip()
    fecha_fin = request.GET.get("fecha_fin", "").strip()
    historial = _historial_vehiculo(vehiculo_obj, fecha_inicio or None, fecha_fin or None)
    return {
        "vehiculo": _vehiculo_a_dict(vehiculo_obj),
        **historial,
        **_resumen_km_monto(historial),
        **_fotos_vehiculo(vehiculo_obj),
        "km_evidencia": _evidencia_km_actual(historial),
        "fecha_inicio_filtro": fecha_inicio,
        "fecha_fin_filtro": fecha_fin,
    }


def detalle_vehiculo(request, placa):
    vehiculo_obj = _obtener_vehiculo_o_404(placa)
    return render(request, "vehiculos/detalle.html", _contexto_ficha(request, vehiculo_obj))


def detalle_fragmento(request, placa):
    vehiculo_obj = _obtener_vehiculo_o_404(placa)
    return render(request, "vehiculos/_detalle_contenido.html", _contexto_ficha(request, vehiculo_obj))


def detalle_vehiculo_por_id(request, vehiculo_id):
    """Misma ficha que detalle_vehiculo, pero por id en vez de placa -- para
    vehículos con placa de relleno (ver _placa_ambigua) donde buscar por
    placa sería ambiguo o rompería la URL."""
    vehiculo_obj = _obtener_vehiculo_por_id_o_404(vehiculo_id)
    return render(request, "vehiculos/detalle.html", _contexto_ficha(request, vehiculo_obj))


def detalle_fragmento_por_id(request, vehiculo_id):
    vehiculo_obj = _obtener_vehiculo_por_id_o_404(vehiculo_id)
    return render(request, "vehiculos/_detalle_contenido.html", _contexto_ficha(request, vehiculo_obj))


# --- Popovers (hold-menu / mapa) -----------------------------------------

@require_GET
def popover_vehiculos(request):
    qs = _aplicar_filtros(_queryset_base(), request)
    total = qs.count()
    duplicadas = _placas_duplicadas()
    vehiculos = [_vehiculo_a_dict(v, duplicadas) for v in qs[:8]]
    return render(request, "vehiculos/_popover_lista.html", {
        "vehiculos": vehiculos,
        "total": total,
        "query": request.META.get("QUERY_STRING", ""),
    })


@require_GET
def popover_kilometraje(request):
    lecturas_qs = Kilometraje.objects.select_related("vehiculo").order_by("-fecha")[:8]
    total = Kilometraje.objects.count()
    duplicadas = _placas_duplicadas()
    lecturas = []
    for k in lecturas_qs:
        ambigua = _placa_ambigua(k.vehiculo.placa, duplicadas)
        detalle_url, detalle_fragmento_url = _urls_detalle_vehiculo(k.vehiculo_id, k.vehiculo.placa, ambigua)
        lecturas.append({
            "placa": k.vehiculo.placa,
            "placa_ambigua": ambigua,
            "detalle_url": detalle_url,
            "detalle_fragmento_url": detalle_fragmento_url,
            "fecha": k.fecha,
            "km": k.odometro,
        })
    return render(request, "vehiculos/_popover_kilometraje.html", {"lecturas": lecturas, "total": total})


@require_GET
def resumen_estado_fragmento(request):
    """Fragmento para el botón 'Parque Vehicular' del mapa orgánico.
    Sin parámetros -> resumen nacional, todos los vehículos.
    Con 'estado=<nombre>' -> solo los de ese estado.
    Con 'inmueble=<nombre>' -> solo los de ese inmueble específico
    (tiene prioridad sobre 'estado' si ambos llegaran a mandarse).
    Se abre dentro del mismo popup (#infoModal / VehiculosModal) que ya
    usa el resto del mapa, así que no hace falta un modal aparte."""
    estado_nombre = request.GET.get("estado", "").strip()
    inmueble_nombre = request.GET.get("inmueble", "").strip()

    # "Total Nacional" (o variantes) no es el nombre real de ningún estado
    # en la base — es la etiqueta que usa el mapa para "sin filtro". Si
    # llega tal cual, se trata como si no hubiera parámetro de estado.
    if estado_nombre.upper() in ("TOTAL NACIONAL", "NACIONAL", "TOTAL_NACIONAL"):
        estado_nombre = ""

    qs = _queryset_base()
    if inmueble_nombre:
        qs = qs.filter(inmueble__nombre_inmueble__iexact=inmueble_nombre)
        titulo = inmueble_nombre
        filtro_extra = "inmueble=" + quote(inmueble_nombre)
    elif estado_nombre:
        qs = qs.filter(estado__nombre__iexact=normalizar_nombre(estado_nombre))
        titulo = estado_nombre
        filtro_extra = "estado=" + quote(estado_nombre)
    else:
        titulo = "Total Nacional"
        filtro_extra = "estado="

    filas = list(qs)
    total = len(filas)
    activos = sum(1 for v in filas if _es_activo(v))
    mantenimiento = sum(1 for v in filas if _color_situacion(v.situacion.situacion if v.situacion_id else "") == "mantenimiento")
    posible_baja = total - activos - mantenimiento

    # Desglose por situación individual (igual que el dashboard general),
    # no solo un bulto de "otras situaciones" — así cada categoría real
    # (Mantenimiento, Posible baja, etc.) tiene su propio botón filtrable.
    otras_situaciones = {}
    for v in filas:
        if not _es_activo(v):
            nombre = v.situacion.situacion if v.situacion_id else "Sin especificar"
            otras_situaciones[nombre] = otras_situaciones.get(nombre, 0) + 1

    # --- Mismas tarjetas que el dashboard general, pero filtradas ---
    conteo_tipos = Counter(v.tipoVeh.tipo_veh for v in filas if v.tipoVeh_id)
    tipos_catalogo = TipoVeh.objects.all().order_by("tipo_veh")
    tipos_vehiculo = [
        {"nombre": t.tipo_veh, "cantidad": conteo_tipos.get(t.tipo_veh, 0)} for t in tipos_catalogo
    ]
    tipos_vehiculo = [t for t in tipos_vehiculo if t["cantidad"] > 0]

    # Categorias (tipo_vehiculo) en vez de "modelos distintos" -- mismo
    # criterio que el dashboard: "modelo" es texto libre e infla el conteo
    # por errores de captura, tipo_vehiculo es un catalogo cerrado.
    categorias_distintas = len(conteo_tipos)
    tarjetas_asignadas = sum(1 for v in filas if v.tarjeta_asig)

    combustible_qs = CombustibleExt.objects.filter(vehiculo__in=filas)
    monto_combustible = sum((c.monto for c in combustible_qs.only("monto")), start=0)

    ultima_lectura_por_vehiculo = {}
    km_qs = Kilometraje.objects.filter(vehiculo__in=filas).order_by("vehiculo_id", "-fecha")
    for k in km_qs.only("vehiculo_id", "odometro"):
        if k.vehiculo_id not in ultima_lectura_por_vehiculo:
            ultima_lectura_por_vehiculo[k.vehiculo_id] = k.odometro
    total_km = sum(ultima_lectura_por_vehiculo.values(), start=0)

    duplicadas = _placas_duplicadas()
    vehiculos = [_vehiculo_a_dict(v, duplicadas) for v in qs.order_by("marca", "modelo")[:15]]

    return render(request, "vehiculos/_resumen_estado.html", {
        "estado_nombre": titulo,
        "filtro_extra": filtro_extra,  # ej. "estado=Jalisco" o "inmueble=Oficina%20X", ya codificado
        "total": total,
        "activos": activos,
        "mantenimiento": mantenimiento,
        "posible_baja": posible_baja,
        "inactivos": total - activos,
        "otras_situaciones": otras_situaciones,
        "categorias_distintas": categorias_distintas,
        "tipos_vehiculo": tipos_vehiculo,
        "tarjetas_asignadas": tarjetas_asignadas,
        "monto_combustible": monto_combustible,
        "total_kilometros": total_km,
        "vehiculos": vehiculos,
        "hay_mas": total > 15,
    })


# --- Exportación a Excel ---------------------------------------------------

ENCABEZADOS_EXCEL_VEHICULOS = [
    "Placa", "Marca", "Modelo", "Año", "No. Motor", "Tipo de vehículo",
    "Tipo de asignación", "Situación", "Estado", "Inmueble", "Tarjeta",
]
CAMPOS_EXCEL_VEHICULOS = [
    "placa", "marca", "modelo", "anio", "no_motor", "tipo_vehiculo",
    "tipo_asignacion", "situacion", "estado", "inmueble_destino", "tarjeta",
]


@require_GET
def exportar_excel_vehiculos(request):
    qs = _aplicar_filtros(_queryset_base(), request).order_by("estado__nombre", "marca")
    filas = [_vehiculo_a_dict(v) for v in qs]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehículos"

    # Encabezado institucional (logo + título), mismo criterio que el PDF
    # de la ficha, para que el reporte se identifique como oficial del
    # INM/Gobernación. Todo vía append() (no escritura directa de celdas)
    # para que el contador interno de filas de openpyxl no se desincronice
    # con las filas de datos que se agregan más abajo.
    ws.append([""])
    ws.append(["Secretaría de Gobernación · Instituto Nacional de Migración · Parque Vehicular"])
    ws.append([])
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(ENCABEZADOS_EXCEL_VEHICULOS))
    ws["A2"].font = Font(bold=True, color="9A0A38", size=11)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ruta_logo = finders.find("mapa/fondos/logo_gob.png")
    if ruta_logo:
        logo = ExcelImage(ruta_logo)
        logo.height = 40
        logo.width = 160
        ws.add_image(logo, "A1")

    fila_encabezados = ws.max_row + 1
    ws.append(ENCABEZADOS_EXCEL_VEHICULOS)
    relleno_encabezado = PatternFill("solid", fgColor="9A0A38")
    for celda in ws[fila_encabezados]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[fila_encabezados].height = 20
    ws.freeze_panes = f"A{fila_encabezados + 1}"

    for fila in filas:
        ws.append([fila.get(campo, "") for campo in CAMPOS_EXCEL_VEHICULOS])

    for indice, encabezado in enumerate(ENCABEZADOS_EXCEL_VEHICULOS, start=1):
        letra = get_column_letter(indice)
        ws.column_dimensions[letra].width = max(len(encabezado), 12) + 2

    partes_nombre = ["vehiculos"]
    for etiqueta in ("tipo", "situacion", "asignacion", "estado"):
        valor = request.GET.get(etiqueta, "")
        if valor:
            partes_nombre.append(valor.replace(" ", "_"))
    nombre_archivo = "_".join(partes_nombre) + ".xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


# --- Descarga de la ficha en PDF -------------------------------------------

@require_GET
def descargar_pdf_ficha(request, placa):
    """Mismo contexto que la ficha en pantalla (_contexto_ficha), pero
    renderizado a un template plano (sin Tailwind/JS, WeasyPrint no los
    procesa) y convertido a PDF -- mismo patron que ya usa 'estadistica'
    para sus reportes."""
    vehiculo_obj = _obtener_vehiculo_o_404(placa)
    context = _contexto_ficha(request, vehiculo_obj)

    template = get_template("vehiculos/_pdf_ficha.html")
    html_string = template.render(context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ficha_{vehiculo_obj.placa}.pdf"'
    return response


# --- Ubicaciones para el mapa (íconos por inmueble) -----------------------

@require_GET
def mapa_ubicaciones_geojson(request):
    """GeoJSON con un punto por cada inmueble que tenga vehículos asignados
    y coordenadas capturadas, con el conteo de vehículos en cada uno — para
    pintar íconos en el mapa (MapLibre). Los vehículos sin inmueble
    asignado, o cuyo inmueble no tenga latitud/longitud, no aparecen aquí
    (a propósito: no hay dónde ubicarlos con precisión).
    Con 'estado=<nombre>' -> solo los de ese estado (igual que
    resumen_estado_fragmento); sin parámetro -> todo el país."""
    qs = _queryset_base().filter(
        inmueble__isnull=False,
        inmueble__latitud__isnull=False,
        inmueble__longitud__isnull=False,
    )

    estado_nombre = request.GET.get("estado", "").strip()
    if estado_nombre.upper() not in ("", "TOTAL NACIONAL", "NACIONAL", "TOTAL_NACIONAL"):
        qs = qs.filter(estado__nombre__iexact=normalizar_nombre(estado_nombre))

    conteo_por_inmueble = {}
    for v in qs:
        inmueble = v.inmueble
        clave = inmueble.id
        if clave not in conteo_por_inmueble:
            conteo_por_inmueble[clave] = {
                "nombre": inmueble.nombre_inmueble,
                "lat": float(inmueble.latitud),
                "lng": float(inmueble.longitud),
                "total": 0,
                "placa_unica": v.placa,  # se usa solo si total termina en 1
            }
        else:
            conteo_por_inmueble[clave]["placa_unica"] = None  # ya hay más de uno
        conteo_por_inmueble[clave]["total"] += 1

    features = [
        {
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [datos["lng"], datos["lat"]]},
            "properties": {
                "nombre": datos["nombre"],
                "total": datos["total"],
                # Si el inmueble tiene un único vehículo, se manda su placa
                # para poder ir directo a la ficha sin pasar por la lista.
                "placa_unica": datos["placa_unica"] if datos["total"] == 1 else None,
            },
        }
        for datos in conteo_por_inmueble.values()
    ]

    return JsonResponse({"type": "FeatureCollection", "features": features})





# Clasificacion de nacionalidades "atipicas" (Medio Oriente / Europa / otras
# poco conocidas), armada a partir de los codigos ISO3 que realmente
# aparecen en la tabla -- no una lista generica adivinada. Para volver a
# verificar contra los datos reales:
#   SELECT iso3, MAX(nacionalidad), COUNT(*) FROM usuario_rescatepunto
#   GROUP BY iso3 ORDER BY 3 DESC;
RESCATES_ISO3_MEDIO_ORIENTE = {
    'TUR', 'AFG', 'EGY', 'JOR', 'IRN', 'IRQ', 'SYR', 'YEM', 'ISR', 'LBN',
    'SAU', 'ARE', 'PSE', 'KWT',
}
RESCATES_ISO3_EUROPA = {
    'POL', 'DEU', 'FRA', 'NLD', 'ITA', 'ESP', 'ROU', 'GBR', 'CHE', 'ALB',
    'BEL', 'SVK', 'CZE', 'HUN', 'SVN', 'DNK', 'IRL', 'AUT', 'MKD', 'PRT',
    'LTU', 'BGR', 'FIN', 'BLR', 'SWE', 'GRC', 'NOR', 'EST', 'HRV', 'LVA',
    'KOS', 'SRB', 'SMR', 'LUX', 'CYP', 'AND', 'MCO', 'RUS', 'UKR', 'MDA',
    'ARM', 'GEO', 'AZE',
}
# El grueso "normal" de la operacion (America Latina/Caribe) -- nunca cuenta
# como atipica, sin importar el volumen.
RESCATES_ISO3_LATAM_CARIBE = {
    'VEN', 'ECU', 'HND', 'GTM', 'COL', 'SLV', 'CUB', 'NIC', 'HTI', 'PER',
    'DOM', 'BOL', 'BRA', 'CHL', 'PAN', 'BLZ', 'CRI', 'URY', 'JAM', 'ARG',
    'DMA', 'GUY', 'TTO', 'MEX', 'ANT', 'PRI', 'GLP', 'SUR', 'VGB', 'VCT',
    'TCA', 'BHS', 'LCA', 'GRD', 'GUF', 'PRY',
}
# Paises grandes/muy conocidos fuera de Medio Oriente, Europa o LATAM (ej.
# India, EUA) -- se excluyen a proposito de "poco conocidas" para no
# etiquetarlos como obscuros solo por tener pocos registros en esta muestra.
# China (CHN) se pidio explicitamente que SI cuente como atipica, por eso
# no esta en este set -- cae en "Otras / poco conocidas" por defecto.
RESCATES_ISO3_CONOCIDOS_NO_ATIPICOS = {
    'IND', 'USA', 'CAN', 'JPN', 'KOR', 'VNM', 'PAK', 'PHL', 'IDN',
    'BGD', 'AUS', 'NZL', 'THA', 'MYS', 'SGP',
}

RESCATES_OFICINAS = [
    "AGUASCALIENTES", "BAJA CALIFORNIA", "BAJA CALIFORNIA SUR", "CAMPECHE",
    "CDMX", "CHIAPAS", "CHIHUAHUA", "COAHUILA", "COLIMA", "DURANGO",
    "EDOMEX", "GUANAJUATO", "GUERRERO", "HIDALGO", "JALISCO", "MICHOACÁN",
    "MORELOS", "NAYARIT", "NUEVO LEÓN", "OAXACA", "PUEBLA", "QUERÉTARO",
    "QUINTANA ROO", "SAN LUIS POTOSÍ", "SINALOA", "SONORA", "TABASCO",
    "TAMAULIPAS", "TLAXCALA", "VERACRUZ", "YUCATÁN", "ZACATECAS",
]

# Las 3 "zonas migratorias" (Rio Bravo/Centro/Suchiate) tal como estan
# definidas en estadistica.views.generar_pdfT (ORs_CECO_N/C/S) -- se copian
# aqui identicas, sin reordenar ni completar/corregir nada, porque la
# seccion "Regiones" del dashboard reutiliza esa clasificacion tal cual.
RESCATES_ZONA_RIO_BRAVO = [
    "BAJA CALIFORNIA", "CHIHUAHUA", "COAHUILA", "DURANGO", "NUEVO LEÓN",
    "SAN LUIS POTOSÍ", "SINALOA", "SONORA", "TAMAULIPAS",
]
RESCATES_ZONA_CENTRO = [
    "AGUASCALIENTES", "BAJA CALIFORNIA SUR", "COLIMA", "CDMX", "GUANAJUATO",
    "GUERRERO", "JALISCO", "MICHOACÁN", "MORELOS", "NAYARIT", "QUERÉTARO",
    "ZACATECAS",
]
RESCATES_ZONA_SUCHIATE = [
    "CAMPECHE", "CHIAPAS", "HIDALGO", "EDOMEX", "OAXACA", "PUEBLA",
    "QUINTANA ROO", "TABASCO", "TLAXCALA", "VERACRUZ", "YUCATÁN",
]


def _rescates_region_atipica(iso3):
    """None si la nacionalidad NO cuenta como atipica (LATAM/Caribe o un
    pais grande/conocido); si no, el nombre de la region atipica."""
    if iso3 in RESCATES_ISO3_MEDIO_ORIENTE:
        return "Medio Oriente"
    if iso3 in RESCATES_ISO3_EUROPA:
        return "Europa"
    if iso3 in RESCATES_ISO3_LATAM_CARIBE or iso3 in RESCATES_ISO3_CONOCIDOS_NO_ATIPICOS:
        return None
    return "Otras / poco conocidas"


def rescates_dashboard(request):
    """Tablero de Rescates/Operación: diagrama de dispersión de rescates por
    día, barras por hora del día, desglose hombres/mujeres/niños/niñas
    (mayoría de edad = 18 años), y pastel de nacionalidades atípicas.
    Filtrable por rango de fechas (default: día actual) y por entidad
    federativa (oficina de representación)."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    hoy = date.today().isoformat()
    fecha_inicio = request.GET.get('fecha_inicio', hoy)
    fecha_fin = request.GET.get('fecha_fin', hoy)
    oficina = request.GET.get('oficina', '').strip()

    filtro_oficina_sql = ''
    params_base = [fecha_inicio, fecha_fin]
    if oficina:
        filtro_oficina_sql = ' AND "oficinaRepre" = %s'
        params_base = params_base + [oficina]

    # Optimización (sin tocar el esquema/índices, a petición explícita):
    # de ~7 consultas secuenciales, cada una escaneando toda la tabla por
    # separado, se pasa a 3 -- usando GROUPING SETS para calcular varias
    # agregaciones en una sola pasada -- corriendo en paralelo (cada una en
    # su propia conexión), en vez de una tras otra. El tiempo total pasa de
    # ser la suma de las 7 a ser el de la más lenta de las 3.

    def _consulta_sexo_edad():
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT "
                f"  COUNT(*) FILTER (WHERE sexo=true  AND edad>=18) AS hombres, "
                f"  COUNT(*) FILTER (WHERE sexo=false AND edad>=18) AS mujeres, "
                f"  COUNT(*) FILTER (WHERE sexo=true  AND edad<18)  AS ninos, "
                f"  COUNT(*) FILTER (WHERE sexo=false AND edad<18)  AS ninas "
                f"FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql}",
                params_base,
            )
            resultado = cur.fetchone()
        connection.close()
        return resultado

    def _consulta_dia_hora():
        # "hora" es texto libre y no siempre viene como HH:MM (hay valores
        # sueltos como "06") -- se descarta al formato inválido dentro del
        # CASE, para no tronar el cast a time por un registro mal capturado.
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT GROUPING(dia) AS g_dia, dia, hr, COUNT(*) FROM ( "
                f"  SELECT TO_DATE(fecha,'DD-MM-YY') AS dia, "
                f"         CASE WHEN hora ~ '^[0-2][0-9]:[0-5][0-9]$' THEN EXTRACT(HOUR FROM hora::time)::int END AS hr "
                f"  FROM usuario_rescatepunto "
                f"  WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f") sub "
                f"GROUP BY GROUPING SETS ((dia), (hr))",
                params_base,
            )
            filas = cur.fetchall()
        connection.close()
        filas_dia_local = sorted(
            [(dia, total) for g_dia, dia, hr, total in filas if g_dia == 0],
            key=lambda x: x[0],
        )
        conteo_por_hora_local = {
            int(hr): total for g_dia, dia, hr, total in filas if g_dia == 1 and hr is not None
        }
        return filas_dia_local, conteo_por_hora_local

    def _consulta_nacionalidad():
        # Nacionalidad por iso3 (nombre en mayúsculas, reduce duplicados de
        # capitalización). (Parentesco y Punto Estratégico "crudos" se
        # quitaron del dashboard a petición explícita -- era vista de
        # análisis interno, no algo para usuario final. numFamilia ahora
        # se resume como conteo de familias, no como distribución cruda.)
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT iso3, UPPER(MAX(nacionalidad)), COUNT(*) FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f"GROUP BY iso3",
                params_base,
            )
            iso_rows_local = cur.fetchall()
        connection.close()
        return iso_rows_local

    def _consulta_reincidentes():
        # Reincidente = coincide (nombre, apellidos, nacionalidad) con otro
        # registro en TODO el historico de usuario_rescatepunto (>=2
        # apariciones), vista mapa_mv_reincidencia_rescates -- reemplaza el
        # conteo que antes venia de mapa_extrescatados.reincidente.
        #
        # Dos apartados pedidos explícitamente:
        #  1. Por día, respetando el filtro de fecha/entidad actual.
        #  2. General histórico: TODOS los registros, sin importar el
        #     filtro de fecha/entidad seleccionado.
        filtro_estado_sql = ""
        params_reinc = [fecha_inicio, fecha_fin]
        if oficina:
            filtro_estado_sql = ' AND r."oficinaRepre" = %s'
            params_reinc = params_reinc + [oficina]
        with connection.cursor() as cur:
            # Reincidentes y primera vez del rango filtrado en UNA sola
            # pasada sobre el JOIN (antes eran 2 consultas identicas salvo
            # por el filtro de clasificacion -- redundante).
            cur.execute(
                f"SELECT "
                f"  COUNT(*) FILTER (WHERE v.clasificacion = 'Reincidente'), "
                f"  COUNT(*) FILTER (WHERE v.clasificacion = 'Rescate primera vez') "
                f"FROM usuario_rescatepunto r "
                f"JOIN {RESCATES_MV_REINCIDENCIA} v "
                f"  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad "
                f"WHERE TO_DATE(r.fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_estado_sql}",
                params_reinc,
            )
            total_reincidentes_local, total_primera_vez_local = cur.fetchone()

            cur.execute(
                f"SELECT TO_DATE(r.fecha,'DD-MM-YY') AS dia, COUNT(*) FROM usuario_rescatepunto r "
                f"JOIN {RESCATES_MV_REINCIDENCIA} v "
                f"  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad "
                f"WHERE v.clasificacion = 'Reincidente' "
                f"AND TO_DATE(r.fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_estado_sql} "
                f"GROUP BY dia ORDER BY dia",
                params_reinc,
            )
            reincidentes_por_dia_local = cur.fetchall()

            # Misma logica (>=2 vs =1) para el historico, tambien en una
            # sola pasada -- la vista ya trae "veces" pre-agregado por
            # persona, asi que ni siquiera toca usuario_rescatepunto.
            cur.execute(
                f"SELECT "
                f"  COALESCE(SUM(veces) FILTER (WHERE clasificacion = 'Reincidente'), 0), "
                f"  COALESCE(SUM(veces) FILTER (WHERE clasificacion = 'Rescate primera vez'), 0) "
                f"FROM {RESCATES_MV_REINCIDENCIA}"
            )
            total_reincidentes_historico_local, total_primera_vez_historico_local = cur.fetchone()
        connection.close()
        return (
            total_reincidentes_local, reincidentes_por_dia_local,
            total_reincidentes_historico_local, total_primera_vez_local,
            total_primera_vez_historico_local,
        )

    def _consulta_por_entidad():
        # Solo tiene sentido comparar entidades cuando NO hay una entidad ya
        # seleccionada -- si se filtró por "TABASCO", no hay nada que
        # comparar. Se calcula aparte (no usa filtro_oficina_sql/params_base
        # porque esos ya vienen con el filtro de oficina fijo).
        with connection.cursor() as cur:
            cur.execute(
                'SELECT "oficinaRepre", COUNT(*) FROM usuario_rescatepunto '
                "WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s "
                'GROUP BY "oficinaRepre" ORDER BY 2 DESC',
                [fecha_inicio, fecha_fin],
            )
            filas = cur.fetchall()
        connection.close()
        return filas

    def _consulta_familias():
        # Agrupa por "numFamilia" (el contador real de familia que ya
        # verificamos) + hora + punto estratégico + agente -- numFamilia

        with connection.cursor() as cur:
            cur.execute(
                f'SELECT * FROM ( '
                f'  SELECT "oficinaRepre" AS oficina, TO_DATE(fecha,\'DD-MM-YY\') AS fecha, hora, '
                f'         "puntoEstra" AS punto, "numFamilia" AS nf, '
                f'         nombre, apellidos, parentesco, "fechaNacimiento" AS fecha_nac, edad, sexo, '
                f"         UPPER(nacionalidad) AS nacionalidad, "
                f'         COUNT(*) OVER (PARTITION BY "oficinaRepre", fecha, hora, "puntoEstra", "numFamilia") AS integrantes '
                f"  FROM usuario_rescatepunto "
                f"  WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f'    AND "numFamilia" > 0 '
                f") sub "
                f"WHERE integrantes >= 2 "
                f"ORDER BY integrantes DESC, oficina, fecha, hora, punto, nf, edad DESC",
                params_base,
            )
            filas = cur.fetchall()
        connection.close()

        # Agrupa las filas planas en núcleos familiares (mismo
        # oficina+fecha+hora+punto+numFamilia). El conteo total de familias
        # se calcula sobre TODAS (no solo las que se muestran); la lista
        # que se despliega se limita a los primeros 15 núcleos -- no a un
        # número fijo de filas, para no cortar una familia a la mitad.
        familias = []
        familia_actual = None
        clave_actual = None
        total_familias = 0
        for oficina, fecha, hora, punto, nf, nombre, apellidos, parentesco, fecha_nac, edad, sexo, nacionalidad, integrantes in filas:
            clave = (oficina, fecha, hora, punto, nf)
            if clave != clave_actual:
                clave_actual = clave
                total_familias += 1
                familia_actual = {
                    "oficina": oficina, "fecha": fecha, "hora": hora, "punto": punto,
                    "num_familia": nf, "integrantes": integrantes, "miembros": [],
                }
                if total_familias <= 15:
                    familias.append(familia_actual)
            if total_familias <= 15:
                familia_actual["miembros"].append({
                    "nombre": nombre, "apellidos": apellidos, "parentesco": parentesco,
                    "fecha_nac": fecha_nac, "edad": edad, "sexo": sexo, "nacionalidad": nacionalidad,
                })
        return familias, total_familias

    def _consulta_tipo_rescate():
        # Medio de rescate (carretero/aéreo/ferroviario/etc.) -- campos
        # booleanos ya existentes en la tabla, en una sola pasada.
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT "
                f"  COUNT(*) FILTER (WHERE carretero) AS carretero, "
                f"  COUNT(*) FILTER (WHERE aeropuerto) AS aereo, "
                f"  COUNT(*) FILTER (WHERE ferrocarril) AS ferroviario, "
                f'  COUNT(*) FILTER (WHERE "centralAutobus") AS central_autobus, '
                f'  COUNT(*) FILTER (WHERE "casaSeguridad") AS casa_seguridad, '
                f"  COUNT(*) FILTER (WHERE hotel) AS hotel, "
                f"  COUNT(*) FILTER (WHERE reclusorio) AS reclusorio "
                f"FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql}",
                params_base,
            )
            resultado = cur.fetchone()
        connection.close()
        return resultado

    def _consulta_retornados():
        # mapa_retornados: tabla distinta a RescatePunto, con fecha real
        # (no texto). Mismo criterio que reincidentes: rango filtrado +
        # general histórico (toda la tabla, sin importar el filtro).
        filtro_estado_sql = ""
        params_ret = [fecha_inicio, fecha_fin]
        if oficina:
            filtro_estado_sql = " AND estado_id = (SELECT id FROM mapa_estado WHERE nombre = %s)"
            params_ret = params_ret + [oficina]
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT COALESCE(SUM(retornados_total),0), COALESCE(SUM(deportado),0), COALESCE(SUM(retornado),0) "
                f"FROM mapa_retornados WHERE fecha BETWEEN %s AND %s{filtro_estado_sql}",
                params_ret,
            )
            total_local, deportado_local, retornado_local = cur.fetchone()

            cur.execute("SELECT COALESCE(SUM(retornados_total),0) FROM mapa_retornados")
            total_historico_local = cur.fetchone()[0]
        connection.close()
        return total_local, deportado_local, retornado_local, total_historico_local

    with ThreadPoolExecutor(max_workers=8) as executor:
        futuro_sexo = executor.submit(_consulta_sexo_edad)
        futuro_dia_hora = executor.submit(_consulta_dia_hora)
        futuro_nac = executor.submit(_consulta_nacionalidad)
        futuro_reincidentes = executor.submit(_consulta_reincidentes)
        futuro_por_entidad = executor.submit(_consulta_por_entidad) if not oficina else None
        futuro_familias = executor.submit(_consulta_familias)
        futuro_tipo_rescate = executor.submit(_consulta_tipo_rescate)
        futuro_retornados = executor.submit(_consulta_retornados)
        # Regiones (Rio Bravo/Centro/Suchiate) -- se calcula tambien aqui
        # para el mini-mapa embebido en el dashboard (la parte mas cara,
        # la deteccion de reincidentes historica, ya vive en la vista
        # materializada mapa_mv_reincidencia_rescates, asi que el costo
        # extra de traerla tambien en el dashboard es bajo). La pagina dedicada
        # /rescates/regiones sigue existiendo aparte para el detalle
        # completo (tablas).
        futuro_regiones = executor.submit(_rescates_regiones, fecha_inicio, fecha_fin, oficina)

        hombres, mujeres, ninos, ninas = futuro_sexo.result()
        filas_dia, conteo_por_hora = futuro_dia_hora.result()
        iso_rows = futuro_nac.result()
        filas_por_entidad = futuro_por_entidad.result() if futuro_por_entidad else []
        (
            total_reincidentes, reincidentes_por_dia,
            total_reincidentes_historico, total_primera_vez,
            total_primera_vez_historico,
        ) = futuro_reincidentes.result()
        familias_detectadas, total_familias = futuro_familias.result()
        carretero, aereo, ferroviario, central_autobus, casa_seguridad, hotel, reclusorio = futuro_tipo_rescate.result()
        total_retornados, retornados_deportado, retornados_retornado, total_retornados_historico = futuro_retornados.result()
        (
            zona_rio_bravo, zona_centro, zona_suchiate,
            subtotal_rio_bravo, subtotal_centro, subtotal_suchiate,
            total_regiones, _nac_1_reinc_sin_usar, _total_nac_1_reinc_sin_usar,
        ) = futuro_regiones.result()

    # El total ya no necesita su propia consulta -- se deriva de la suma de
    # las 4 categorías de sexo/edad (sexo y edad son campos obligatorios en
    # el modelo, así que la suma siempre coincide con el COUNT(*) real).
    total_rango = hombres + mujeres + ninos + ninas

    # --- Clasificación de atípicas (regla geográfica, no de frecuencia) ---
    atipicas_por_region = {"Medio Oriente": 0, "Europa": 0, "Otras / poco conocidas": 0}
    detalle_atipicas = []
    for iso3, nombre, total in iso_rows:
        region = _rescates_region_atipica(iso3)
        if region:
            atipicas_por_region[region] += total
            detalle_atipicas.append({"iso3": iso3, "nombre": nombre, "total": total, "region": region})
    detalle_atipicas.sort(key=lambda d: d["total"], reverse=True)
    total_atipicas = sum(atipicas_por_region.values())

    # --- Desglose (solo números) de hombres/mujeres/niños/niñas y núcleos
    # familiares, pero limitado a las nacionalidades atípicas -- para la
    # sección de detalle, no para el dashboard general. ---
    iso3_atipicas = [d["iso3"] for d in detalle_atipicas]
    if iso3_atipicas:
        placeholders_iso = ",".join(["%s"] * len(iso3_atipicas))
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT "
                f"  COUNT(*) FILTER (WHERE sexo=true  AND edad>=18) AS hombres, "
                f"  COUNT(*) FILTER (WHERE sexo=false AND edad>=18) AS mujeres, "
                f"  COUNT(*) FILTER (WHERE sexo=true  AND edad<18)  AS ninos, "
                f"  COUNT(*) FILTER (WHERE sexo=false AND edad<18)  AS ninas "
                f"FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f"  AND iso3 IN ({placeholders_iso})",
                params_base + iso3_atipicas,
            )
            atip_hombres, atip_mujeres, atip_ninos, atip_ninas = cur.fetchone()

            cur.execute(
                f"SELECT COUNT(DISTINCT (\"oficinaRepre\", fecha, hora, \"puntoEstra\", \"numFamilia\")) "
                f"FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f"  AND iso3 IN ({placeholders_iso}) AND \"numFamilia\" > 0",
                params_base + iso3_atipicas,
            )
            atip_nucleos = cur.fetchone()[0]
    else:
        atip_hombres = atip_mujeres = atip_ninos = atip_ninas = atip_nucleos = 0

    # --- Gráfica 1: diagrama de dispersión (rescates por día) ---
    x_disp = [datetime.combine(d, datetime.min.time()) for d, _ in filas_dia]
    y_disp = [c for _, c in filas_dia]
    etiquetas_disp = [f"{c:,}" for _, c in filas_dia]
    source_disp = ColumnDataSource(data=dict(x=x_disp, y=y_disp, label=etiquetas_disp))
    p_disp = figure(
        height=320, sizing_mode="stretch_width", x_axis_type="datetime",
        toolbar_location="right", tools="pan,box_zoom,reset",
        background_fill_color="#f7f7f7", border_fill_color=None,
        outline_line_color="#666666",
        title="Rescates por día (cada punto = un día)",
    )
    p_disp.scatter(x='x', y='y', size=9, color="#285C4D", alpha=0.75, source=source_disp)
    p_disp.y_range.start = 0
    p_disp.add_layout(LabelSet(
        x='x', y='y', text='label', source=source_disp,
        x_offset=0, y_offset=8, text_font_size="9px", text_color="#285C4D",
        text_align="center",
    ))
    p_disp.xaxis.formatter = DatetimeTickFormatter(days="%d %b", months="%b %Y", years="%Y")
    p_disp.xgrid.grid_line_color = "#ffffff"
    p_disp.ygrid.grid_line_color = "#ffffff"
    p_disp.add_tools(HoverTool(tooltips=[("Día", "@x{%d/%b/%y}"), ("Rescates", "@y{0,0}")], formatters={'@x': 'datetime'}))

    # --- Gráfica 2: barras por hora del día ---
    horas_lbl = [f"{h:02d}:00" for h in range(24)]
    valores_hora = [conteo_por_hora.get(h, 0) for h in range(24)]
    etiquetas_hora = [f"{v:,}" for v in valores_hora]
    source_horas = ColumnDataSource(data=dict(hora=horas_lbl, total=valores_hora, label=etiquetas_hora))
    p_horas = figure(
        x_range=horas_lbl, height=300, sizing_mode="stretch_width",
        toolbar_location=None, tools="",
        background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
        title="Rescates por hora del día (00:00 a 23:59:59)",
    )
    p_horas.vbar(x='hora', top='total', width=0.8, color="#9A0A38", source=source_horas)
    p_horas.y_range.start = 0
    p_horas.add_layout(LabelSet(
        x='hora', y='total', text='label', source=source_horas,
        x_offset=0, y_offset=4, text_font_size="8px", text_color="#9A0A38",
        text_align="center",
    ))
    p_horas.xaxis.major_label_orientation = 0.9
    p_horas.xgrid.grid_line_color = None
    p_horas.ygrid.grid_line_color = "#ffffff"
    p_horas.add_tools(HoverTool(tooltips=[("Hora", "@hora"), ("Rescates", "@total{0,0}")]))

    # --- Gráfica 3: barras hombres / mujeres / niños / niñas ---
    categorias_sexo = ["Hombres", "Mujeres", "Niños", "Niñas"]
    valores_sexo = [hombres, mujeres, ninos, ninas]
    etiquetas_sexo = [f"{v:,}" for v in valores_sexo]
    colores_sexo = ["#0EA5E9", "#EC4899", "#38BDF8", "#F9A8D4"]
    source_sexo = ColumnDataSource(data=dict(
        categoria=categorias_sexo, total=valores_sexo, color=colores_sexo, label=etiquetas_sexo,
    ))
    p_sexo = figure(
        x_range=categorias_sexo, height=300, sizing_mode="stretch_width",
        toolbar_location=None, tools="",
        background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
        title="Hombres, mujeres, niños y niñas (mayoría de edad = 18 años)",
    )
    p_sexo.vbar(x='categoria', top='total', width=0.6, color='color', source=source_sexo)
    p_sexo.y_range.start = 0
    p_sexo.add_layout(LabelSet(
        x='categoria', y='total', text='label', source=source_sexo,
        x_offset=0, y_offset=4, text_font_size="11px", text_color="#374151",
        text_align="center",
    ))
    p_sexo.xgrid.grid_line_color = None
    p_sexo.ygrid.grid_line_color = "#ffffff"
    p_sexo.add_tools(HoverTool(tooltips=[("Categoría", "@categoria"), ("Total", "@total{0,0}")]))

    # --- Gráfica 4: pastel de nacionalidades atípicas ---
    regiones_pie = ["Medio Oriente", "Europa", "Otras / poco conocidas"]
    valores_pie = [atipicas_por_region[r] for r in regiones_pie]
    colores_pie = ["#F59E0B", "#6366F1", "#10B981"]
    angulos = [(v / total_atipicas * 2 * pi) if total_atipicas else 0 for v in valores_pie]
    # Posición de cada etiqueta: a la mitad del ángulo de su rebanada, un
    # poco afuera del radio del pastel, para verse sin pasar el cursor.
    acumulado = 0
    x_etq_pie, y_etq_pie, etiquetas_pie = [], [], []
    for v, ang in zip(valores_pie, angulos):
        medio = acumulado + ang / 2
        x_etq_pie.append(0 + 0.55 * cos(medio))
        y_etq_pie.append(1 + 0.55 * sin(medio))
        etiquetas_pie.append(f"{v:,}")
        acumulado += ang
    source_pie = ColumnDataSource(data=dict(
        region=regiones_pie, total=valores_pie, color=colores_pie, angle=angulos,
    ))
    source_etq_pie = ColumnDataSource(data=dict(x=x_etq_pie, y=y_etq_pie, label=etiquetas_pie))
    p_pie = figure(
        height=320, sizing_mode="stretch_width", toolbar_location=None, tools="",
        background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
        title="Nacionalidades atípicas (Medio Oriente, Europa, otras poco conocidas)",
        x_range=(-0.9, 1.3),
    )
    p_pie.wedge(
        x=0, y=1, radius=0.4,
        start_angle=cumsum('angle', include_zero=True), end_angle=cumsum('angle'),
        line_color="#ffffff", fill_color='color', legend_field='region', source=source_pie,
    )
    p_pie.add_layout(LabelSet(
        x='x', y='y', text='label', source=source_etq_pie,
        text_font_size="11px", text_font_style="bold", text_color="#374151",
        text_align="center", text_baseline="middle",
    ))
    p_pie.axis.visible = False
    p_pie.grid.visible = False
    p_pie.add_tools(HoverTool(tooltips=[("Región", "@region"), ("Total", "@total{0,0}")]))

    # --- Gráfica 5: barras horizontales por entidad, solo cuando "Todas
    # las entidades" está seleccionado (si ya se filtró una, no hay nada
    # que comparar) -- ordenada de mayor a menor. ---
    figuras = {"dispersion": p_disp, "horas": p_horas, "sexo": p_sexo, "pie": p_pie}

    if not oficina and filas_por_entidad:
        entidades_ordenadas = [f[0] for f in filas_por_entidad]  # ya viene ORDER BY 2 DESC
        totales_entidad = [f[1] for f in filas_por_entidad]
        etiquetas_entidad = [f"{v:,}" for v in totales_entidad]
        source_entidad = ColumnDataSource(data=dict(
            entidad=entidades_ordenadas, total=totales_entidad, label=etiquetas_entidad,
        ))
        p_entidad = figure(
            y_range=list(reversed(entidades_ordenadas)),  # mayor arriba
            height=max(320, 24 * len(entidades_ordenadas)), sizing_mode="stretch_width",
            toolbar_location=None, tools="",
            background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
            title="Rescates por entidad federativa (mayor a menor)",
        )
        p_entidad.hbar(y='entidad', right='total', height=0.7, color="#285C4D", source=source_entidad)
        p_entidad.x_range.start = 0
        p_entidad.add_layout(LabelSet(
            x='total', y='entidad', text='label', source=source_entidad,
            x_offset=4, y_offset=-6, text_font_size="9px", text_color="#285C4D",
        ))
        p_entidad.ygrid.grid_line_color = None
        p_entidad.xgrid.grid_line_color = "#ffffff"
        p_entidad.add_tools(HoverTool(tooltips=[("Entidad", "@entidad"), ("Rescates", "@total{0,0}")]))
        figuras["entidad"] = p_entidad

    # --- Gráfica 6: reincidentes por día (dentro del rango filtrado) ---
    x_reinc = [datetime.combine(d, datetime.min.time()) for d, _ in reincidentes_por_dia]
    y_reinc = [int(v or 0) for _, v in reincidentes_por_dia]
    etiquetas_reinc = [f"{v:,}" for v in y_reinc]
    source_reinc = ColumnDataSource(data=dict(x=x_reinc, y=y_reinc, label=etiquetas_reinc))
    p_reinc = figure(
        height=280, sizing_mode="stretch_width", x_axis_type="datetime",
        toolbar_location="right", tools="pan,box_zoom,reset",
        background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
        title="Reincidentes por día (dentro del rango filtrado)",
    )
    p_reinc.vbar(x='x', top='y', width=1000 * 60 * 60 * 20, color="#DC2626", source=source_reinc)
    p_reinc.y_range.start = 0
    p_reinc.add_layout(LabelSet(
        x='x', y='y', text='label', source=source_reinc,
        x_offset=0, y_offset=4, text_font_size="9px", text_color="#DC2626", text_align="center",
    ))
    p_reinc.xaxis.formatter = DatetimeTickFormatter(days="%d %b", months="%b %Y", years="%Y")
    p_reinc.xgrid.grid_line_color = "#ffffff"
    p_reinc.ygrid.grid_line_color = "#ffffff"
    p_reinc.add_tools(HoverTool(tooltips=[("Día", "@x{%d/%b/%y}"), ("Reincidentes", "@y{0,0}")], formatters={'@x': 'datetime'}))
    figuras["reincidentes_dia"] = p_reinc

    nombres = list(figuras.keys())
    plot_script, divs_lista = components(tuple(figuras[n] for n in nombres))
    divs = dict(zip(nombres, divs_lista))
    div_disp = divs["dispersion"]
    div_horas = divs["horas"]
    div_sexo = divs["sexo"]
    div_pie = divs["pie"]
    div_entidad = divs.get("entidad", "")
    div_reincidentes_dia = divs["reincidentes_dia"]

    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "oficina_seleccionada": oficina,
        "oficinas": RESCATES_OFICINAS,
        "total_rango": total_rango,
        "hombres": hombres,
        "mujeres": mujeres,
        "ninos": ninos,
        "ninas": ninas,
        "total_atipicas": total_atipicas,
        "atipicas_por_region": atipicas_por_region,
        "detalle_atipicas": detalle_atipicas[:20],
        "atip_hombres": atip_hombres,
        "atip_mujeres": atip_mujeres,
        "atip_ninos": atip_ninos,
        "atip_ninas": atip_ninas,
        "atip_nucleos": atip_nucleos,
        # Reincidentes (mapa_mv_reincidencia_rescates, >=2 apariciones de
        # nombre+apellidos+nacionalidad en usuario_rescatepunto). Dos
        # apartados: por día (respeta el filtro actual) y general histórico
        # (TODA la tabla, sin importar el filtro de fecha/entidad seleccionado).
        "total_reincidentes": total_reincidentes,
        "total_reincidentes_historico": total_reincidentes_historico,
        "total_primera_vez": total_primera_vez,
        "total_primera_vez_historico": total_primera_vez_historico,
        "familias_detectadas": familias_detectadas,
        "total_familias": total_familias,
        # Medio de rescate (campos booleanos de la tabla).
        "carretero": carretero,
        "aereo": aereo,
        "ferroviario": ferroviario,
        "central_autobus": central_autobus,
        "casa_seguridad": casa_seguridad,
        "hotel": hotel,
        "reclusorio": reclusorio,
        # Retornados (mapa_retornados) -- tabla distinta a RescatePunto.
        "total_retornados": total_retornados,
        "retornados_deportado": retornados_deportado,
        "retornados_retornado": retornados_retornado,
        "total_retornados_historico": total_retornados_historico,
        # Regiones -- mini-mapa embebido; el detalle completo (tablas) vive
        # en su propia pagina (rescates_regiones).
        "subtotal_rio_bravo": subtotal_rio_bravo,
        "subtotal_centro": subtotal_centro,
        "subtotal_suchiate": subtotal_suchiate,
        "total_regiones": total_regiones,
        "geo_data_regiones_json": json.dumps(_rescates_geojson_regiones(zona_rio_bravo, zona_centro, zona_suchiate)),
        "plot_script": plot_script,
        "div_dispersion": div_disp,
        "div_horas": div_horas,
        "div_sexo": div_sexo,
        "div_pie": div_pie,
        "div_entidad": div_entidad,
        "mostrar_seccion_entidad": not oficina,
        "div_reincidentes_dia": div_reincidentes_dia,
    }
    return render(request, "mapa/rescates.html", context)


def rescates_regiones(request):
    """Pagina dedicada a "Regiones" (Rio Bravo/Centro/Suchiate), separada
    del dashboard principal -- mismo filtro de fecha/entidad, con mapa
    interactivo (reutiliza el geojson/MapLibre de mapa_interactivo) que
    pinta cada estado segun su zona y le muestra el total encima."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    hoy = date.today().isoformat()
    fecha_inicio = request.GET.get('fecha_inicio', hoy)
    fecha_fin = request.GET.get('fecha_fin', hoy)
    oficina = request.GET.get('oficina', '').strip()

    (
        zona_rio_bravo, zona_centro, zona_suchiate,
        subtotal_rio_bravo, subtotal_centro, subtotal_suchiate,
        total_regiones, nac_1_reinc, total_nac_1_reinc,
    ) = _rescates_regiones(fecha_inicio, fecha_fin, oficina or None)

    geo_data = _rescates_geojson_regiones(zona_rio_bravo, zona_centro, zona_suchiate)

    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "oficina_seleccionada": oficina,
        "oficinas": RESCATES_OFICINAS,
        "zona_rio_bravo": zona_rio_bravo,
        "zona_centro": zona_centro,
        "zona_suchiate": zona_suchiate,
        "subtotal_rio_bravo": subtotal_rio_bravo,
        "subtotal_centro": subtotal_centro,
        "subtotal_suchiate": subtotal_suchiate,
        "total_regiones": total_regiones,
        "nac_1_reinc": nac_1_reinc,
        "total_nac_1_reinc": total_nac_1_reinc,
        "geo_data_json": json.dumps(geo_data),
    }
    return render(request, "mapa/rescates_regiones.html", context)


# =============================================================================
# Reporte Personalizado -- filtros libres (calendario + menus desplegables)
# para que el usuario arme su propio reporte, en vez de los formatos fijos
# oficiales. Los resultados se agrupan por oficina + nacionalidad (nunca se
# muestran registros individuales/datos personales -- mismo criterio que el
# resto del dashboard).
# =============================================================================

RESCATES_TIPO_RESCATE_CAMPOS = [
    ("carretero", "Carretero"),
    ("aeropuerto", "Aéreo"),
    ("ferrocarril", "Ferroviario"),
    ("centralAutobus", "Central de autobuses"),
    ("casaSeguridad", "Casa de seguridad"),
    ("hotel", "Hotel"),
    ("reclusorio", "Reclusorio"),
]

RESCATES_CATEGORIA_SQL = {
    "hombres": "sexo=true AND edad>=18",
    "mujeres": "sexo=false AND edad>=18",
    "ninos": "sexo=true AND edad<18",
    "ninas": "sexo=false AND edad<18",
}

RESCATES_ZONAS_POR_NOMBRE = {
    "Río Bravo": RESCATES_ZONA_RIO_BRAVO,
    "Centro": RESCATES_ZONA_CENTRO,
    "Suchiate": RESCATES_ZONA_SUCHIATE,
}

RESCATES_CACHE_KEY_NACIONALIDADES = "rescates_nacionalidades_disponibles"
RESCATES_CACHE_TTL_NACIONALIDADES = 1800  # 30 min


def _rescates_nacionalidades_disponibles():
    """Lista de nacionalidades distintas para el menu desplegable -- se
    cachea porque es un DISTINCT sobre 1.5M filas sin indice."""
    resultado = cache.get(RESCATES_CACHE_KEY_NACIONALIDADES)
    if resultado is None:
        with connection.cursor() as cur:
            cur.execute("SELECT DISTINCT UPPER(nacionalidad) FROM usuario_rescatepunto WHERE nacionalidad != '' ORDER BY 1")
            resultado = [fila[0] for fila in cur.fetchall()]
        cache.set(RESCATES_CACHE_KEY_NACIONALIDADES, resultado, RESCATES_CACHE_TTL_NACIONALIDADES)
    return resultado


def _rescates_personalizado(fecha_inicio, fecha_fin, oficina=None, zona=None, tipo_rescate=None, categoria=None, nacionalidad=None):
    """Reporte con filtros libres: rango de fechas (obligatorio) + entidad,
    zona, tipo de rescate, categoria (sexo/edad) y nacionalidad (todos
    opcionales). Agrupa por oficina + nacionalidad con el mismo desglose
    hombres/mujeres/ninos/ninas que ya se usa en el resto del dashboard."""
    filtros_sql = []
    params = [fecha_inicio, fecha_fin]

    if oficina:
        filtros_sql.append('"oficinaRepre" = %s')
        params.append(oficina)
    elif zona and zona in RESCATES_ZONAS_POR_NOMBRE:
        oficinas_zona = RESCATES_ZONAS_POR_NOMBRE[zona]
        placeholders = ",".join(["%s"] * len(oficinas_zona))
        filtros_sql.append(f'"oficinaRepre" IN ({placeholders})')
        params.extend(oficinas_zona)

    if tipo_rescate and tipo_rescate in dict(RESCATES_TIPO_RESCATE_CAMPOS):
        filtros_sql.append(f'"{tipo_rescate}" = true')

    if categoria and categoria in RESCATES_CATEGORIA_SQL:
        filtros_sql.append(RESCATES_CATEGORIA_SQL[categoria])

    if nacionalidad:
        filtros_sql.append("UPPER(nacionalidad) = %s")
        params.append(nacionalidad)

    filtro_extra = "".join(f" AND {f}" for f in filtros_sql)

    with connection.cursor() as cur:
        cur.execute(
            f'SELECT "oficinaRepre", UPPER(nacionalidad), '
            f"  COUNT(*) FILTER (WHERE sexo=true  AND edad>=18) AS hombres, "
            f"  COUNT(*) FILTER (WHERE sexo=false AND edad>=18) AS mujeres, "
            f"  COUNT(*) FILTER (WHERE sexo=true  AND edad<18)  AS ninos, "
            f"  COUNT(*) FILTER (WHERE sexo=false AND edad<18)  AS ninas, "
            f"  COUNT(*) AS total "
            f"FROM usuario_rescatepunto "
            f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_extra} "
            f'GROUP BY "oficinaRepre", UPPER(nacionalidad) '
            f"ORDER BY total DESC",
            params,
        )
        filas = cur.fetchall()

    filas_tabla = [
        {"oficina": of, "nacionalidad": nac, "hombres": h, "mujeres": m, "ninos": n, "ninas": ni, "total": t}
        for of, nac, h, m, n, ni, t in filas
    ]
    total_general = sum(f["total"] for f in filas_tabla)
    return filas_tabla, total_general


def rescates_reporte_personalizado(request):
    """Vista previa en pantalla del reporte personalizado (filtros libres)."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    hoy = date.today().isoformat()
    fecha_inicio = request.GET.get('fecha_inicio', hoy)
    fecha_fin = request.GET.get('fecha_fin', hoy)
    oficina = request.GET.get('oficina', '').strip()
    zona = request.GET.get('zona', '').strip()
    tipo_rescate = request.GET.get('tipo_rescate', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    nacionalidad = request.GET.get('nacionalidad', '').strip()

    filas_tabla, total_general = _rescates_personalizado(
        fecha_inicio, fecha_fin, oficina or None, zona or None,
        tipo_rescate or None, categoria or None, nacionalidad or None,
    )

    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "oficina_seleccionada": oficina,
        "zona_seleccionada": zona,
        "tipo_rescate_seleccionado": tipo_rescate,
        "categoria_seleccionada": categoria,
        "nacionalidad_seleccionada": nacionalidad,
        "oficinas": RESCATES_OFICINAS,
        "zonas": list(RESCATES_ZONAS_POR_NOMBRE.keys()),
        "tipos_rescate": RESCATES_TIPO_RESCATE_CAMPOS,
        "nacionalidades": _rescates_nacionalidades_disponibles(),
        "filas_tabla": filas_tabla[:200],
        "total_filas": len(filas_tabla),
        "total_general": total_general,
    }
    return render(request, "mapa/rescates_reporte_personalizado.html", context)


def rescates_reporte_personalizado_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_inicio = request.GET.get('fecha_inicio', date.today().isoformat())
    fecha_fin = request.GET.get('fecha_fin', date.today().isoformat())
    oficina = request.GET.get('oficina', '').strip()
    zona = request.GET.get('zona', '').strip()
    tipo_rescate = request.GET.get('tipo_rescate', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    nacionalidad = request.GET.get('nacionalidad', '').strip()

    filas_tabla, total_general = _rescates_personalizado(
        fecha_inicio, fecha_fin, oficina or None, zona or None,
        tipo_rescate or None, categoria or None, nacionalidad or None,
    )
    etiquetas_tipo = dict(RESCATES_TIPO_RESCATE_CAMPOS)
    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "oficina": oficina or "Todas",
        "zona": zona or "Todas",
        "tipo_rescate": etiquetas_tipo.get(tipo_rescate, "Todos"),
        "categoria": categoria or "Todas",
        "nacionalidad": nacionalidad or "Todas",
        "filas_tabla": filas_tabla,
        "total_general": total_general,
    }
    template = get_template("mapa/_rescates_personalizado_pdf.html")
    html_string = template.render(context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="reporte_personalizado_{fecha_inicio}_a_{fecha_fin}.pdf"'
    return response


def rescates_reporte_personalizado_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_inicio = request.GET.get('fecha_inicio', date.today().isoformat())
    fecha_fin = request.GET.get('fecha_fin', date.today().isoformat())
    oficina = request.GET.get('oficina', '').strip()
    zona = request.GET.get('zona', '').strip()
    tipo_rescate = request.GET.get('tipo_rescate', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    nacionalidad = request.GET.get('nacionalidad', '').strip()

    filas_tabla, total_general = _rescates_personalizado(
        fecha_inicio, fecha_fin, oficina or None, zona or None,
        tipo_rescate or None, categoria or None, nacionalidad or None,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Personalizado"
    etiquetas_tipo = dict(RESCATES_TIPO_RESCATE_CAMPOS)

    ws.merge_cells("A1:G1")
    _rescates_excel_celda(
        ws, 1, 1,
        f"Reporte Personalizado — {fecha_inicio} a {fecha_fin} — "
        f"Entidad: {oficina or 'Todas'} · Zona: {zona or 'Todas'} · "
        f"Tipo: {etiquetas_tipo.get(tipo_rescate, 'Todos')} · "
        f"Categoría: {categoria or 'Todas'} · Nacionalidad: {nacionalidad or 'Todas'}",
        negrita=True, tam=11, centrado=False,
    )

    fila = 3
    _rescates_excel_fila(ws, fila, ["Oficina", "Nacionalidad", "Hombres", "Mujeres", "Niños", "Niñas", "Total"], RESCATES_LETRA_T1, negrita=True)
    fila += 1
    for f in filas_tabla:
        _rescates_excel_fila(ws, fila, [f["oficina"], f["nacionalidad"], f["hombres"], f["mujeres"], f["ninos"], f["ninas"], f["total"]], RESCATES_LETRA_T2)
        fila += 1
    _rescates_excel_fila(ws, fila, ["", "TOTAL", "", "", "", "", total_general], RESCATES_LETRA_T0, negrita=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 24
    for col in ("C", "D", "E", "F", "G"):
        ws.column_dimensions[col].width = 12

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="reporte_personalizado_{fecha_inicio}_a_{fecha_fin}.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# Reportes de Rescates -- "Cuadro de Datos" (mismo formato/reglas de negocio
# que estadistica.generar_cuadro_diario, pero 100% automatico: nada se
# captura a mano. La unica diferencia real es "Retornados", que en el
# reporte original es un campo de texto libre que alguien llena -- aqui se
# calcula de mapa_retornados, igual que en el resto del dashboard.
#
# Reglas de negocio reutilizadas tal cual del reporte oficial:
#  - "Inadmitido" = un registro de RescatePunto donde NINGUNA bandera de
#    medio de rescate (aereo/carretero/casa de seguridad/central de
#    autobuses/ferroviario/hotel/puestos a disposicion/voluntarios/otro)
#    esta en True.
#  - "Reincidente" = coincide (nombre, apellidos, nacionalidad) con otro
#    registro en TODO el historico de RescatePunto (no solo el dia
#    seleccionado) -- vive en la vista materializada
#    mapa_mv_reincidencia_rescates (ver _rescates_set_duplicados_historicos),
#    refrescada cada 20 min, asi que el costo pesado del GROUP BY ya no se
#    paga por request.
#  - "Alojados en EM" (numFamilia=0/NULL, va solo) vs "Alojados en
#    DIF/Albergue" (numFamilia>0, viene en familia), solo sobre los
#    rescates NUEVOS (no reincidentes) del dia.
#
# Nota: el reporte oficial (estadistica.generar_cuadro_diario) trata a
# CHIAPAS aparte y ahi TODOS sus registros del dia caen en "reincidentes"
# sin comparar contra el historico -- no se replico esa parte porque no
# hay ninguna nota que explique si es una regla real o un descuido del
# reporte original; aqui CHIAPAS se evalua igual que cualquier otra
# entidad. Si el criterio oficial es a proposito, avisar para ajustarlo.
# =============================================================================

RESCATES_BANDERAS_MEDIO = dict(
    aeropuerto=False, carretero=False, casaSeguridad=False, centralAutobus=False,
    ferrocarril=False, hotel=False, puestosADispo=False, voluntarios=False, otro=False,
)

# Nombres de mes en espanol para los reportes -- no se usa strftime("%b")
# porque depende del locale del sistema (no garantizado dentro del
# contenedor) y podria mostrar los meses en ingles.
RESCATES_MESES_ES = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
    7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
}


# @FADAR -- la deteccion de reincidentes (>=2 apariciones de nombre+
# apellidos+nacionalidad en TODO el historico de usuario_rescatepunto) vive
# en la vista materializada mapa_mv_reincidencia_rescates (ver migracion
# 0025). El refresh (CONCURRENTLY, ~168s medido) NO se dispara desde el
# request -- corre aparte cada 20 min en el servicio cron-reincidencia de
# docker-compose.yaml, para que ningun usuario del dashboard tenga que
# esperarlo. Mismo tradeoff que antes: datos nuevos no se ven hasta el
# siguiente refresh del cron.
RESCATES_MV_REINCIDENCIA = "mapa_mv_reincidencia_rescates"


def _rescates_set_duplicados_historicos():
    """Set de (nombre, apellidos, nacionalidad) clasificados 'Reincidente'
    en mapa_mv_reincidencia_rescates -- usada en los 3 lugares que la
    necesitan (dashboard -> seccion "Regiones", reporte Cuadro de Datos,
    reporte Informe Diario). Misma interfaz que antes, ahora respaldada por
    la vista materializada en vez de recalcularse en Python."""
    with connection.cursor() as cur:
        cur.execute(
            f"SELECT nombre, apellidos, nacionalidad FROM {RESCATES_MV_REINCIDENCIA} "
            f"WHERE clasificacion = 'Reincidente'"
        )
        return set(cur.fetchall())


def _rescates_regiones(fecha_inicio, fecha_fin, oficina=None):
    """Reutiliza TAL CUAL la logica de generar_pdfT (estadistica/views.py,
    seccion "Cuadro de Registros" / CECO 2) -- 3 zonas migratorias (Rio
    Bravo/Centro/Suchiate) que cubren las 32 oficinas. Ninguna regla de
    negocio se altera respecto al original:
     - Mismo filtro .exclude(...) del reporte oficial (NO es la regla de
       "inadmitido" que usa el resto de este modulo -- se deja identico a
       proposito, tal como esta en el codigo fuente).
     - CHIAPAS aparte: el 100% de sus registros del rango cuenta como
       reincidente, SIN comparar contra el historico (asi esta en el
       reporte oficial -- comportamiento real y aceptado, no un descuido).
     - "Reincidente" para el resto = coincide (nombre, apellidos,
       nacionalidad) con otro registro en TODO el historico de
       RescatePunto.
    Unico cambio respecto al original: acepta un rango de fechas (el
    reporte original solo trabajaba un dia a la vez) y respeta el filtro
    de entidad si ya hay una oficina seleccionada -- son parte del
    "wrapper" de filtros compartido por el dashboard y los reportes, no de
    la regla de negocio en si. Usada tanto por el dashboard (rango) como
    por el reporte "Regiones" (fecha_inicio == fecha_fin)."""
    fecha_ini_obj = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    fecha_fin_obj = datetime.strptime(fecha_fin, "%Y-%m-%d")
    dias_rango = (fecha_fin_obj - fecha_ini_obj).days
    array_fechas = [
        (fecha_ini_obj + timedelta(days=d)).strftime("%d-%m-%y")
        for d in range(max(dias_rango, 0) + 1)
    ]

    campos_valores = (
        'nombre', 'apellidos', 'nacionalidad', 'oficinaRepre',
        'puntoEstra', 'fecha', 'sexo', 'edad', 'numFamilia',
    )
    exclude_original = dict(
        aeropuerto=False, carretero=True, casaSeguridad=False, centralAutobus=False,
        ferrocarril=False, hotel=False, puestosADispo=False, voluntarios=True, otro=True,
    )

    oficinas_sin_chiapas = [o for o in RESCATES_OFICINAS if o != "CHIAPAS"]
    if oficina:
        oficinas_sin_chiapas = [oficina] if oficina in oficinas_sin_chiapas else []
    datos_no_chiapas = list(
        RescatePunto.objects.filter(fecha__in=array_fechas, oficinaRepre__in=oficinas_sin_chiapas)
        .exclude(**exclude_original)
        .values(*campos_valores)
        .annotate(total=Count('idRescate'))
    ) if oficinas_sin_chiapas else []

    incluir_chiapas = (not oficina) or (oficina == "CHIAPAS")
    datos_chiapas = list(
        RescatePunto.objects.filter(fecha__in=array_fechas, oficinaRepre="CHIAPAS")
        .exclude(**exclude_original)
        .values(*campos_valores)
        .annotate(total=Count('idRescate'))
    ) if incluir_chiapas else []

    set_duplicados_local = _rescates_set_duplicados_historicos()

    nuevos_local, reincidentes_local = [], []
    for dato in datos_no_chiapas:
        clave = (dato['nombre'], dato['apellidos'], dato['nacionalidad'])
        if clave in set_duplicados_local:
            reincidentes_local.append(dato)
        else:
            nuevos_local.append(dato)
    # Chiapas: 100% reincidente, sin excepcion (regla oficial).
    reincidentes_local.extend(datos_chiapas)

    zona_por_oficina_local = {}
    zona_rio_bravo_local = {o: {"nuevos": 0, "reincidentes": 0, "total": 0} for o in RESCATES_ZONA_RIO_BRAVO}
    zona_centro_local = {o: {"nuevos": 0, "reincidentes": 0, "total": 0} for o in RESCATES_ZONA_CENTRO}
    zona_suchiate_local = {o: {"nuevos": 0, "reincidentes": 0, "total": 0} for o in RESCATES_ZONA_SUCHIATE}
    for o in RESCATES_ZONA_RIO_BRAVO:
        zona_por_oficina_local[o] = zona_rio_bravo_local
    for o in RESCATES_ZONA_CENTRO:
        zona_por_oficina_local[o] = zona_centro_local
    for o in RESCATES_ZONA_SUCHIATE:
        zona_por_oficina_local[o] = zona_suchiate_local

    subtotal_rio_bravo_local = {"nuevos": 0, "reincidentes": 0, "total": 0}
    subtotal_centro_local = {"nuevos": 0, "reincidentes": 0, "total": 0}
    subtotal_suchiate_local = {"nuevos": 0, "reincidentes": 0, "total": 0}
    subtotal_por_oficina_local = {}
    for o in RESCATES_ZONA_RIO_BRAVO:
        subtotal_por_oficina_local[o] = subtotal_rio_bravo_local
    for o in RESCATES_ZONA_CENTRO:
        subtotal_por_oficina_local[o] = subtotal_centro_local
    for o in RESCATES_ZONA_SUCHIATE:
        subtotal_por_oficina_local[o] = subtotal_suchiate_local

    nac_1_reinc_local = {}
    total_nac_1_reinc_local = {"nuevos": 0, "reincidentes": 0, "total": 0}

    def _clasificar(lista, campo):
        # @FADAR -- "dato['total']" (no +=1): los datos vienen de un
        # .values(...).annotate(total=Count(...)), asi que una fila
        # agrupada puede representar >1 registro real (ej. un duplicado
        # tecnico con los mismos 9 campos) -- sumar solo +=1 los contaba
        # como uno solo. Por ahora se cuentan como registros reales
        # independientes (sin deduplicar), a peticion explicita.
        for dato in lista:
            cantidad = dato['total']
            of = dato['oficinaRepre']
            tabla_zona = zona_por_oficina_local.get(of)
            if tabla_zona is not None:
                tabla_zona[of][campo] += cantidad
                tabla_zona[of]['total'] += cantidad
                subtotal_por_oficina_local[of][campo] += cantidad
                subtotal_por_oficina_local[of]['total'] += cantidad
            nac = str(dato['nacionalidad']).upper()
            if nac not in nac_1_reinc_local:
                nac_1_reinc_local[nac] = {"nuevos": 0, "reincidentes": 0, "total": 0}
            nac_1_reinc_local[nac][campo] += cantidad
            nac_1_reinc_local[nac]['total'] += cantidad
            total_nac_1_reinc_local[campo] += cantidad
            total_nac_1_reinc_local['total'] += cantidad

    _clasificar(nuevos_local, 'nuevos')
    _clasificar(reincidentes_local, 'reincidentes')

    nac_1_reinc_ordenado_local = dict(
        sorted(nac_1_reinc_local.items(), key=lambda x: x[1]['total'], reverse=True)
    )
    total_regiones_local = {
        "nuevos": subtotal_rio_bravo_local['nuevos'] + subtotal_centro_local['nuevos'] + subtotal_suchiate_local['nuevos'],
        "reincidentes": subtotal_rio_bravo_local['reincidentes'] + subtotal_centro_local['reincidentes'] + subtotal_suchiate_local['reincidentes'],
        "total": subtotal_rio_bravo_local['total'] + subtotal_centro_local['total'] + subtotal_suchiate_local['total'],
    }
    return (
        zona_rio_bravo_local, zona_centro_local, zona_suchiate_local,
        subtotal_rio_bravo_local, subtotal_centro_local, subtotal_suchiate_local,
        total_regiones_local, nac_1_reinc_ordenado_local, total_nac_1_reinc_local,
    )


def _rescates_regiones_reporte(fecha_str, oficina=None):
    """Version 'reporte' (una sola fecha) de _rescates_regiones -- mismo
    patron que _rescates_cuadro_datos / _rescates_informe_diario: regresa
    un dict ya listo para el template/PDF/Excel."""
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    (
        zona_rio_bravo, zona_centro, zona_suchiate,
        subtotal_rio_bravo, subtotal_centro, subtotal_suchiate,
        total_regiones, nac_1_reinc, total_nac_1_reinc,
    ) = _rescates_regiones(fecha_str, fecha_str, oficina)
    return {
        "fecha_actual": f"{fecha_obj.day:02d} {RESCATES_MESES_ES[fecha_obj.month]} {fecha_obj.year}",
        "fecha_iso": fecha_str,
        "oficina": oficina or "Nacional",
        "zona_rio_bravo": zona_rio_bravo,
        "zona_centro": zona_centro,
        "zona_suchiate": zona_suchiate,
        "subtotal_rio_bravo": subtotal_rio_bravo,
        "subtotal_centro": subtotal_centro,
        "subtotal_suchiate": subtotal_suchiate,
        "total_regiones": total_regiones,
        "nac_1_reinc": nac_1_reinc,
        "total_nac_1_reinc": total_nac_1_reinc,
    }


# Los 2 unicos casos donde el nombre de RESCATES_OFICINAS no coincide (una
# vez sin acentos/mayusculas) con la propiedad "name" del geojson de
# estados -- el resto empata directo via normalizar_nombre().
RESCATES_ALIAS_GEOJSON = {
    "CDMX": "Ciudad de México",
    "EDOMEX": "Estado de México",
}

RESCATES_ZONA_COLOR = {
    "Río Bravo": "#1D4ED8",
    "Centro": "#7C3AED",
    "Suchiate": "#B45309",
}


def _rescates_geojson_regiones(zona_rio_bravo, zona_centro, zona_suchiate):
    """Reutiliza el mismo geojson de estados que ya usa mapa_interactivo/
    mapa_informacion (mapa/static/mapa/data/inegi_latlon_mexico.geojson) y
    le anota, por estado, la zona (Rio Bravo/Centro/Suchiate) y sus
    numeros -- para que el mapa se pinte y etiquete directo desde las
    propiedades del geojson, sin logica aparte del lado del cliente."""
    geojson_path = os.path.join(settings.BASE_DIR, 'mapa', 'static', 'mapa', 'data', 'inegi_latlon_mexico.geojson')
    with open(geojson_path, 'r', encoding='utf-8') as f:
        geo_data = json.load(f)

    datos_por_oficina = {}
    for of, d in zona_rio_bravo.items():
        datos_por_oficina[of] = {"zona": "Río Bravo", **d}
    for of, d in zona_centro.items():
        datos_por_oficina[of] = {"zona": "Centro", **d}
    for of, d in zona_suchiate.items():
        datos_por_oficina[of] = {"zona": "Suchiate", **d}

    datos_por_clave = {}
    for of, d in datos_por_oficina.items():
        alias = RESCATES_ALIAS_GEOJSON.get(of, of)
        datos_por_clave[normalizar_nombre(alias)] = {"oficina": of, **d}

    for feature in geo_data.get("features", []):
        props = feature.setdefault("properties", {})
        clave = normalizar_nombre(props.get("name", ""))
        info = datos_por_clave.get(clave)
        if info:
            props["oficina_rescates"] = info["oficina"]
            props["zona_rescates"] = info["zona"]
            props["color_zona_rescates"] = RESCATES_ZONA_COLOR[info["zona"]]
            props["nuevos_rescates"] = info["nuevos"]
            props["reincidentes_rescates"] = info["reincidentes"]
            props["total_rescates"] = info["total"]
        else:
            props["oficina_rescates"] = props.get("name", "")
            props["zona_rescates"] = None
            props["color_zona_rescates"] = "#D1D5DB"
            props["nuevos_rescates"] = 0
            props["reincidentes_rescates"] = 0
            props["total_rescates"] = 0

    return geo_data


# @FADAR -- detalle de Retornados (mapa_retornados), independiente del
# numero que ya resta dentro de la tabla principal del Cuadro de Datos /
# Informe Diario. Por estado si no hay entidad filtrada, por nacionalidad
# si ya se filtro una entidad especifica (misma entidad ya fija el estado).
def _rescates_retornados_detalle(fecha_str, oficina=None):
    with connection.cursor() as cur:
        if oficina:
            cur.execute(
                "SELECT n.nombre, SUM(r.deportado), SUM(r.retornado), SUM(r.retornados_total) "
                "FROM mapa_retornados r "
                "JOIN mapa_estado e ON e.id = r.estado_id "
                "JOIN mapa_nacionalidad n ON n.id = r.nacionalidad_id "
                "WHERE r.fecha = %s AND e.nombre = %s "
                "GROUP BY n.nombre ORDER BY SUM(r.retornados_total) DESC",
                [fecha_str, oficina],
            )
            columna = "Nacionalidad"
        else:
            cur.execute(
                "SELECT e.nombre, SUM(r.deportado), SUM(r.retornado), SUM(r.retornados_total) "
                "FROM mapa_retornados r "
                "JOIN mapa_estado e ON e.id = r.estado_id "
                "WHERE r.fecha = %s "
                "GROUP BY e.nombre ORDER BY SUM(r.retornados_total) DESC",
                [fecha_str],
            )
            columna = "Entidad"
        filas = cur.fetchall()

    filas_tabla = [{"nombre": n, "deportado": d, "retornado": r, "total": t} for n, d, r, t in filas]
    return {
        "columna": columna,
        "filas": filas_tabla,
        "total_deportado": sum(f["deportado"] for f in filas_tabla),
        "total_retornado": sum(f["retornado"] for f in filas_tabla),
        "total_general": sum(f["total"] for f in filas_tabla),
    }


def _rescates_cuadro_datos(fecha_str, oficina=None):
    """fecha_str: 'YYYY-MM-DD'. Devuelve el dict con todos los datos del
    Cuadro de Datos para esa fecha (y opcionalmente una sola entidad)."""
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    fecha_rescate_fmt = fecha_obj.strftime("%d-%m-%y")

    qs_dia = RescatePunto.objects.filter(fecha=fecha_rescate_fmt)
    if oficina:
        qs_dia = qs_dia.filter(oficinaRepre=oficina)

    total_inadmitidos = qs_dia.filter(**RESCATES_BANDERAS_MEDIO).count()

    datos_dia = list(
        qs_dia.exclude(**RESCATES_BANDERAS_MEDIO)
        .values("nombre", "apellidos", "nacionalidad", "sexo", "edad", "numFamilia")
    )
    total_rescatados = len(datos_dia) + total_inadmitidos

    # Reincidencia: coincidencia de (nombre, apellidos, nacionalidad) en
    # TODO el historico de RescatePunto (consulta pesada, cacheada -- ver
    # _rescates_set_duplicados_historicos).
    set_duplicados = _rescates_set_duplicados_historicos()

    reincidentes = []
    nuevos = []
    for d in datos_dia:
        clave = (d["nombre"], d["apellidos"], d["nacionalidad"])
        if clave in set_duplicados:
            reincidentes.append(d)
        else:
            nuevos.append(d)

    conteo_reincidentes = len(reincidentes)
    conteo_nuevos = len(nuevos)
    subtotal1 = total_rescatados - conteo_reincidentes
    subtotal2 = subtotal1 - total_inadmitidos

    # Retornados: automatico desde mapa_retornados (no manual).
    filtro_estado_sql = ""
    params_ret = [fecha_str, fecha_str]
    if oficina:
        filtro_estado_sql = " AND estado_id = (SELECT id FROM mapa_estado WHERE nombre = %s)"
        params_ret = params_ret + [oficina]
    with connection.cursor() as cur:
        cur.execute(
            f"SELECT COALESCE(SUM(retornados_total),0) FROM mapa_retornados "
            f"WHERE fecha BETWEEN %s AND %s{filtro_estado_sql}",
            params_ret,
        )
        total_retornados = cur.fetchone()[0]

    rescates_nuevos_neto = conteo_nuevos - total_retornados

    # EM vs DIF/Albergue, por nacionalidad, solo sobre los "nuevos".
    nacionalidades = {d["nacionalidad"] for d in nuevos}
    datos_em = {n: 0 for n in nacionalidades}
    datos_dif = {n: 0 for n in nacionalidades}
    for d in nuevos:
        va_solo = d["numFamilia"] is None or d["numFamilia"] == 0
        if va_solo:
            datos_em[d["nacionalidad"]] += 1
        else:
            datos_dif[d["nacionalidad"]] += 1

    def _top_y_otros(datos, corte=5):
        ordenado = dict(sorted(datos.items(), key=lambda x: x[1], reverse=True))
        if len(ordenado) <= 7:
            corte = 7
        top = dict(list(ordenado.items())[:corte])
        resto = sum(list(ordenado.values())[corte:])
        if len(ordenado) > corte:
            top["Otras nacionalidades"] = resto
        total = sum(top.values())
        return top, total

    top_em, total_em = _top_y_otros(datos_em)
    top_dif, total_dif = _top_y_otros(datos_dif)

    return {
        "fecha": f"{fecha_obj.day:02d} {RESCATES_MESES_ES[fecha_obj.month]} {fecha_obj.year}",
        "fecha_iso": fecha_str,
        "oficina": oficina or "Nacional",
        "rescatados": total_rescatados,
        "reincidentes": conteo_reincidentes,
        "subtotal1": subtotal1,
        "inadmitidos": total_inadmitidos,
        "subtotal2": subtotal2,
        "retornados": total_retornados,
        "rescates_nuevos": rescates_nuevos_neto,
        "rescates_EM": top_em,
        "rescates_DIF": top_dif,
        "EM_total": total_em,
        "DIF_total": total_dif,
        "retornados_detalle": _rescates_retornados_detalle(fecha_str, oficina),
    }


def rescates_reporte_cuadro(request):
    """Vista previa en pantalla del Cuadro de Datos, con selector de fecha
    y entidad (menus desplegables) -- sin nada manual."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()

    datos = _rescates_cuadro_datos(fecha_str, oficina or None)
    datos["oficinas"] = RESCATES_OFICINAS
    datos["oficina_seleccionada"] = oficina
    return render(request, "mapa/rescates_reporte_cuadro.html", datos)


def rescates_reporte_cuadro_pdf(request):
    """Mismo cálculo que la vista previa, pero devuelto como PDF (mismo
    patrón WeasyPrint que ya usa el resto del proyecto)."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_cuadro_datos(fecha_str, oficina or None)

    template = get_template("mapa/_rescates_cuadro_pdf.html")
    html_string = template.render(datos)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="cuadro_datos_{fecha_str}.pdf"'
    return response


# Mismos colores exactos del PDF/pantalla oficiales, para los 3 formatos
# (pantalla, PDF, Excel) se vean consistentes.
RESCATES_COLOR_GUINDA = "621132"      # rgb(98,17,50) -- resumen del Cuadro de Datos
RESCATES_COLOR_NEGATIVO = "9D2449"    # rgb(157,36,73) -- reincidentes/inadmitidos/retornados
RESCATES_COLOR_GRIS = "D9D9D9"        # rgb(217,217,217) -- filas de subtotal
RESCATES_COLOR_EM = "285C4D"          # rgb(40,92,77) -- Alojados en EM
RESCATES_COLOR_EM_DATO = "8B6B41"     # rgb(139,107,65) -- texto/fondo de datos EM
RESCATES_COLOR_DIF = "7030A0"         # rgb(112,48,160) -- Alojados en DIF/Albergue
RESCATES_COLOR_CELESTE = "0EA5E9"     # rgb(14,165,233) -- Retornados (apartado independiente)

RESCATES_BORDE_DELGADO = Border(*(Side(style="thin", color="000000"),) * 4)


def _rescates_excel_celda(ws, fila, col, valor, bg=None, color_texto="000000", negrita=False, tam=11, centrado=True, fuente=None):
    celda = ws.cell(row=fila, column=col, value=valor)
    kwargs_fuente = {"name": fuente} if fuente else {}
    celda.font = Font(bold=negrita, size=tam, color=color_texto, **kwargs_fuente)
    if bg:
        celda.fill = PatternFill("solid", fgColor=bg)
    celda.border = RESCATES_BORDE_DELGADO
    if centrado:
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Formato de excel real (no texto) para cualquier valor numerico -- asi
    # la columna se puede sumar/importar en Excel en vez de quedar como
    # texto (p.ej. una celda con el valor -5 en vez del texto "-5").
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        celda.number_format = "#,##0"
    return celda


def rescates_reporte_cuadro_excel(request):
    """Mismo cálculo y mismos colores que la vista previa/PDF (guinda,
    verde EM, morado DIF, gris de subtotales), para que los tres formatos
    se vean consistentes."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_cuadro_datos(fecha_str, oficina or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuadro de Datos"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    ws.merge_cells("A1:B1")
    _rescates_excel_celda(ws, 1, 1, "INSTITUTO NACIONAL DE MIGRACIÓN", negrita=True, tam=13)
    ws.merge_cells("A2:B2")
    _rescates_excel_celda(ws, 2, 1, "DIRECCIÓN GENERAL DE COORDINACIÓN DE OFICINAS DE REPRESENTACION", negrita=True, tam=10)
    ws.merge_cells("A3:B3")
    _rescates_excel_celda(ws, 3, 1, f"Cuadro de Datos — {datos['fecha']} — {datos['oficina']}", negrita=True, tam=12)

    # Los valores numericos van como numero real de Excel (incluido el
    # signo negativo), no como texto formateado ("-5") -- para que la
    # columna se pueda sumar/importar directamente en Excel. "FECHA" y "De
    # estos" son las unicas dos filas inherentemente textuales (encabezado
    # y descripcion), el resto son datos.
    filas_resumen = [
        ("FECHA", datos["fecha"], None, "FFFFFF", True),
        ("Rescatados:", datos["rescatados"], None, "FFFFFF", False),
        ("Reincidentes:", -datos["reincidentes"], RESCATES_COLOR_NEGATIVO, "FFFFFF", False),
        ("Subtotal:", datos["subtotal1"], RESCATES_COLOR_GRIS, "000000", False),
        ("Inadmitidos:", -datos["inadmitidos"], RESCATES_COLOR_NEGATIVO, "FFFFFF", False),
        ("Subtotal:", datos["subtotal2"], RESCATES_COLOR_GRIS, "000000", False),
        ("Retornados:", -datos["retornados"], RESCATES_COLOR_NEGATIVO, "FFFFFF", False),
        ("Subtotal:", datos["rescates_nuevos"], RESCATES_COLOR_GRIS, "000000", True),
    ]
    fila = 5
    for etiqueta, valor, bg, color_txt, negrita in filas_resumen:
        _rescates_excel_celda(ws, fila, 1, etiqueta, bg=bg or None, color_texto=color_txt if bg else "000000", negrita=negrita, centrado=False)
        _rescates_excel_celda(ws, fila, 2, valor, bg=bg or None, color_texto=color_txt if bg else "000000", negrita=negrita)
        fila += 1

    # "De estos": en el PDF es una sola celda de texto descriptivo, pero
    # aqui se separa en 2 filas con valor numerico real (EM_total/DIF_total
    # como numero, no incrustados en una oracion) -- mismo dato, formato
    # importable.
    _rescates_excel_celda(ws, fila, 1, "De estos, alojados en EM:", bg=RESCATES_COLOR_GUINDA, color_texto="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws, fila, 2, datos["EM_total"], bg=RESCATES_COLOR_GUINDA, color_texto="FFFFFF", negrita=True)
    fila += 1
    _rescates_excel_celda(ws, fila, 1, "De estos, canalizados al DIF:", bg=RESCATES_COLOR_GUINDA, color_texto="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws, fila, 2, datos["DIF_total"], bg=RESCATES_COLOR_GUINDA, color_texto="FFFFFF", negrita=True)
    fila += 1

    fila += 1
    fila_em_inicio = fila
    _rescates_excel_celda(ws, fila, 1, "Alojados en EM", bg=RESCATES_COLOR_EM, color_texto="FFFFFF", negrita=True)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    fila += 1
    _rescates_excel_celda(ws, fila, 1, "NACIONALIDAD", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 2, "TOTAL", bg="FFFFFF", negrita=True)
    fila += 1
    for nombre, total in datos["rescates_EM"].items():
        _rescates_excel_celda(ws, fila, 1, nombre, color_texto=RESCATES_COLOR_EM_DATO, centrado=False)
        _rescates_excel_celda(ws, fila, 2, total, color_texto=RESCATES_COLOR_EM_DATO)
        fila += 1

    fila += 2
    _rescates_excel_celda(ws, fila, 1, "Alojados en DIF/Albergue", bg=RESCATES_COLOR_DIF, color_texto="FFFFFF", negrita=True)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    fila += 1
    _rescates_excel_celda(ws, fila, 1, "NACIONALIDAD", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 2, "TOTAL", bg="FFFFFF", negrita=True)
    fila += 1
    for nombre, total in datos["rescates_DIF"].items():
        _rescates_excel_celda(ws, fila, 1, nombre, color_texto=RESCATES_COLOR_DIF, centrado=False)
        _rescates_excel_celda(ws, fila, 2, total, color_texto=RESCATES_COLOR_DIF)
        fila += 1

    # @FADAR -- Retornados: hoja aparte, independiente del cuadro principal.
    ws2 = wb.create_sheet("Retornados")
    ws2.column_dimensions["A"].width = 28
    for col in "BCD":
        ws2.column_dimensions[col].width = 14
    _rescates_excel_celda(ws2, 1, 1, f"Retornados — detalle por {datos['retornados_detalle']['columna']}", bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True, centrado=False)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    _rescates_excel_celda(ws2, 2, 1, "Retornados:", bg="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws2, 2, 2, datos["retornados_detalle"]["total_retornado"], bg="FFFFFF", color_texto=RESCATES_COLOR_CELESTE, negrita=True)
    fila_hdr = 4
    _rescates_excel_celda(ws2, fila_hdr, 1, datos["retornados_detalle"]["columna"].upper(), bg="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws2, fila_hdr, 2, "DEPORTADOS", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws2, fila_hdr, 3, "RETORNADOS", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws2, fila_hdr, 4, "TOTAL", bg="FFFFFF", negrita=True)
    fila2 = fila_hdr + 1
    for f in datos["retornados_detalle"]["filas"]:
        _rescates_excel_celda(ws2, fila2, 1, f["nombre"], centrado=False)
        _rescates_excel_celda(ws2, fila2, 2, f["deportado"])
        _rescates_excel_celda(ws2, fila2, 3, f["retornado"])
        _rescates_excel_celda(ws2, fila2, 4, f["total"])
        fila2 += 1
    _rescates_excel_celda(ws2, fila2, 1, "TOTAL", bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws2, fila2, 2, datos["retornados_detalle"]["total_deportado"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws2, fila2, 3, datos["retornados_detalle"]["total_retornado"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws2, fila2, 4, datos["retornados_detalle"]["total_general"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="cuadro_datos_{fecha_str}.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# Reportes de Rescates -- "Informe Diario de Rescatados" (mismo formato y
# reglas de negocio que estadistica.generar_pdf/reporteA.html, tambien
# automatico: Retornados sale de mapa_retornados, no se captura a mano).
#
# 8 categorias (igual criterio que el reporte oficial, definidas solo con
# sexo + edad + numFamilia):
#   H_AS/M_AS = Hombre/Mujer Adulto(a) Solo(a)        (numFamilia 0/NULL)
#   H_AA/M_AA = Hombre/Mujer Adulto(a) Acompaña NNA    (numFamilia > 0)
#   H_mA/M_mA = NNA Acompañado(a), Hombre/Mujer        (numFamilia > 0)
#   H_mS/M_mS = NNA No Acompañado(a), Hombre/Mujer     (numFamilia 0/NULL)
# =============================================================================

RESCATES_ETIQUETAS_CATEGORIA = [
    ("H_AS", "Adultos Hombres Solos"),
    ("M_AS", "Adultas Mujeres Solas"),
    ("H_AA", "Adultos Hombres Acompañan NNA"),
    ("M_AA", "Adultas Mujeres Acompañan NNA"),
    ("H_mA", "NNA Acompañados Hombres"),
    ("M_mA", "NNA Acompañados Mujeres"),
    ("H_mS", "NNA No Acompañados Hombres"),
    ("M_mS", "NNA No Acompañados Mujeres"),
]


def _rescates_clasificar_categoria(sexo, edad, num_familia):
    solo = num_familia is None or num_familia == 0
    if sexo and edad >= 18:
        return "H_AS" if solo else "H_AA"
    if not sexo and edad >= 18:
        return "M_AS" if solo else "M_AA"
    if sexo and edad < 18:
        return "H_mS" if solo else "H_mA"
    return "M_mS" if solo else "M_mA"


def _rescates_informe_diario(fecha_str):
    """fecha_str: 'YYYY-MM-DD'. Reporte nacional (las 32 entidades, sin
    tratamiento especial para ninguna -- ver nota de Chiapas en el Cuadro
    de Datos)."""
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    fecha_rescate_fmt = fecha_obj.strftime("%d-%m-%y")

    qs_dia = RescatePunto.objects.filter(fecha=fecha_rescate_fmt)

    # --- Tabla 1: rescates validos por oficina y medio de rescate ---
    por_oficina_qs = (
        qs_dia.exclude(**RESCATES_BANDERAS_MEDIO)
        .values("oficinaRepre")
        .annotate(
            total=Count("idRescate"),
            total_aereos=Count("idRescate", filter=Q(aeropuerto=True)),
            total_carreteros=Count("idRescate", filter=Q(carretero=True)),
            total_central=Count("idRescate", filter=Q(centralAutobus=True)),
            total_ferro=Count("idRescate", filter=Q(ferrocarril=True)),
            total_puestos=Count("idRescate", filter=Q(puestosADispo=True)),
            total_otros=(
                Count("idRescate", filter=Q(voluntarios=True))
                + Count("idRescate", filter=Q(otro=True))
                + Count("idRescate", filter=Q(casaSeguridad=True))
                + Count("idRescate", filter=Q(hotel=True))
            ),
        )
    )
    rescates_por_oficina = {o: {"total": 0, "total_aereos": 0, "total_carreteros": 0, "total_central": 0,
                                 "total_ferro": 0, "total_puestos": 0, "total_otros": 0} for o in RESCATES_OFICINAS}
    rescates_por_oficina["Total"] = dict(rescates_por_oficina[RESCATES_OFICINAS[0]])
    for fila in por_oficina_qs:
        of = fila["oficinaRepre"]
        if of not in rescates_por_oficina:
            continue
        for campo in ("total", "total_aereos", "total_carreteros", "total_central", "total_ferro", "total_puestos", "total_otros"):
            rescates_por_oficina[of][campo] = fila[campo]
            rescates_por_oficina["Total"][campo] += fila[campo]

    # --- Datos persona por persona (para reincidencia y las 8 categorias) ---
    datos_dia = list(
        qs_dia.exclude(**RESCATES_BANDERAS_MEDIO)
        .values("nombre", "apellidos", "nacionalidad", "oficinaRepre", "sexo", "edad", "numFamilia")
    )

    set_duplicados = _rescates_set_duplicados_historicos()

    reincidentes, nuevos = [], []
    for d in datos_dia:
        clave = (d["nombre"], d["apellidos"], d["nacionalidad"])
        (reincidentes if clave in set_duplicados else nuevos).append(d)

    # --- Tabla 2: reincidentes / nuevos por oficina ---
    reincidentes_por_oficina_ct = Counter(d["oficinaRepre"] for d in reincidentes)
    nuevos_por_oficina_ct = Counter(d["oficinaRepre"] for d in nuevos)
    reincidentes_por_oficina = {o: {"total_reincidentes": 0, "total_nuevos": 0, "total": 0} for o in RESCATES_OFICINAS}
    reincidentes_por_oficina["Total"] = {"total_reincidentes": 0, "total_nuevos": 0, "total": 0}
    for of in RESCATES_OFICINAS:
        r = reincidentes_por_oficina_ct.get(of, 0)
        n = nuevos_por_oficina_ct.get(of, 0)
        reincidentes_por_oficina[of] = {"total_reincidentes": r, "total_nuevos": n, "total": r + n}
        reincidentes_por_oficina["Total"]["total_reincidentes"] += r
        reincidentes_por_oficina["Total"]["total_nuevos"] += n
        reincidentes_por_oficina["Total"]["total"] += r + n

    # --- Tablas 3 y 4: nacionalidad x 8 categorias, nuevos y reincidentes ---
    def _tabla_nacionalidad_categorias(lista):
        nacionalidades = {d["nacionalidad"] for d in lista}
        tabla = {n: {clave: 0 for clave, _ in RESCATES_ETIQUETAS_CATEGORIA} for n in nacionalidades}
        for n in tabla:
            tabla[n]["total"] = 0
        for d in lista:
            cat = _rescates_clasificar_categoria(d["sexo"], d["edad"], d["numFamilia"])
            tabla[d["nacionalidad"]][cat] += 1
            tabla[d["nacionalidad"]]["total"] += 1
        return dict(sorted(tabla.items(), key=lambda x: x[1]["total"], reverse=True))

    nacionalidades_nuevos = _tabla_nacionalidad_categorias(nuevos)
    nacionalidades_reincidentes = _tabla_nacionalidad_categorias(reincidentes)

    # --- Inadmitidos: nacionalidad x 4 categorias (sexo x adulto/menor) ---
    datos_inadmitidos = list(
        qs_dia.filter(**RESCATES_BANDERAS_MEDIO).values("nacionalidad", "sexo", "edad")
    )
    nacionalidades_inadm = {d["nacionalidad"] for d in datos_inadmitidos}
    tabla_inadm = {n: {"H_A": 0, "M_A": 0, "H_m": 0, "M_m": 0, "total": 0} for n in nacionalidades_inadm}
    for d in datos_inadmitidos:
        tabla_inadm[d["nacionalidad"]]["total"] += 1
        if d["sexo"] and d["edad"] >= 18:
            tabla_inadm[d["nacionalidad"]]["H_A"] += 1
        elif not d["sexo"] and d["edad"] >= 18:
            tabla_inadm[d["nacionalidad"]]["M_A"] += 1
        elif d["sexo"] and d["edad"] < 18:
            tabla_inadm[d["nacionalidad"]]["H_m"] += 1
        else:
            tabla_inadm[d["nacionalidad"]]["M_m"] += 1
    nacionalidades_inadmitidos = dict(sorted(tabla_inadm.items(), key=lambda x: x[1]["total"], reverse=True))

    # --- Retornados: automatico desde mapa_retornados (no manual) ---
    with connection.cursor() as cur:
        cur.execute(
            "SELECT COALESCE(SUM(retornados_total),0) FROM mapa_retornados WHERE fecha = %s",
            [fecha_str],
        )
        total_retornados = cur.fetchone()[0]

    return {
        "fecha_actual": fecha_obj.strftime("%d/%m/%Y"),
        "fecha_iso": fecha_str,
        "rescates": rescates_por_oficina,
        "reincidentes": reincidentes_por_oficina,
        "nacionalidades": nacionalidades_nuevos,
        "nacionalidades_re": nacionalidades_reincidentes,
        "nacionalidades_inadm": nacionalidades_inadmitidos,
        "dato": total_retornados,
        "categorias": RESCATES_ETIQUETAS_CATEGORIA,
        "retornados_detalle": _rescates_retornados_detalle(fecha_str),
    }


def rescates_reporte_informe(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_informe_diario(fecha_str)
    return render(request, "mapa/rescates_reporte_informe.html", datos)


def rescates_reporte_informe_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_informe_diario(fecha_str)

    template = get_template("mapa/_rescates_informe_pdf.html")
    html_string = template.render(datos)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="informe_diario_{fecha_str}.pdf"'
    return response


# Mismos colores exactos del PDF ("letraT0".."letraT6") para el Informe
# Diario, igual criterio que el Cuadro de Datos: pantalla/PDF/Excel
# consistentes entre si.
RESCATES_LETRA_T0 = ("B38E5D", "E9E8E8")  # fondo, texto
RESCATES_LETRA_T1 = ("761B36", "FFFFFF")
RESCATES_LETRA_T2 = ("D9D9D9", "000000")
RESCATES_LETRA_T3 = ("4BACC6", "000000")
RESCATES_LETRA_T4 = ("13322B", "EDEDED")
RESCATES_LETRA_T6 = ("7030A0", "EAEAEA")
RESCATES_COLOR_FONDO = ("4E1224", "F2F2F2")


def _rescates_excel_fila(ws, fila, valores, estilo, negrita=False, col_inicial=1):
    bg, color_txt = estilo
    for i, valor in enumerate(valores):
        _rescates_excel_celda(ws, fila, col_inicial + i, valor, bg=bg, color_texto=color_txt, negrita=negrita)


def rescates_reporte_informe_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_informe_diario(fecha_str)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Por oficina"
    oficinas_cols = list(datos["rescates"].keys())
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(oficinas_cols) + 1)
    _rescates_excel_celda(ws1, 1, 1, "INSTITUTO NACIONAL DE MIGRACIÓN — DIRECCIÓN GENERAL DE COORDINACIÓN DE OFICINAS DE REPRESENTACION", negrita=True, tam=10)
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(oficinas_cols) + 1)
    _rescates_excel_celda(ws1, 2, 1, f"Informe Diario de Rescatados — {datos['fecha_actual']}", bg=RESCATES_COLOR_FONDO[0], color_texto=RESCATES_COLOR_FONDO[1], negrita=True, tam=12)

    fila = 4
    _rescates_excel_fila(ws1, fila, ["Rubro"] + oficinas_cols, RESCATES_LETRA_T1, negrita=True)
    fila += 1
    for campo, etiqueta in [("total_aereos", "Aeropuertos"), ("total_carreteros", "Carreteros"),
                             ("total_central", "Central de Autobús"), ("total_ferro", "Ferroviarios"),
                             ("total_puestos", "Puestos a Disposición"), ("total_otros", "Otros")]:
        _rescates_excel_fila(ws1, fila, [etiqueta] + [datos["rescates"][o][campo] for o in oficinas_cols], RESCATES_LETRA_T2)
        fila += 1
    _rescates_excel_fila(ws1, fila, ["TOTAL MIGRANTES RESCATADOS"] + [datos["rescates"][o]["total"] for o in oficinas_cols], RESCATES_LETRA_T1, negrita=True)
    fila += 2

    _rescates_excel_fila(ws1, fila, ["Rubro"] + oficinas_cols, RESCATES_LETRA_T4, negrita=True)
    fila += 1
    for campo, etiqueta in [("total_reincidentes", "Reincidentes"), ("total_nuevos", "Registros Nuevos")]:
        _rescates_excel_fila(ws1, fila, [etiqueta] + [datos["reincidentes"][o][campo] for o in oficinas_cols], RESCATES_LETRA_T2)
        fila += 1
    _rescates_excel_fila(ws1, fila, ["TOTAL MIGRANTES RESCATADOS"] + [datos["reincidentes"][o]["total"] for o in oficinas_cols], RESCATES_LETRA_T0, negrita=True)

    ws1.column_dimensions["A"].width = 26
    for i in range(len(oficinas_cols)):
        ws1.column_dimensions[get_column_letter(2 + i)].width = 12

    def _hoja_nacionalidad(nombre, tabla):
        ws = wb.create_sheet(nombre)
        encabezados = ["Nacionalidad"] + [e for _, e in RESCATES_ETIQUETAS_CATEGORIA] + ["Total"]
        _rescates_excel_fila(ws, 1, encabezados, RESCATES_LETRA_T1, negrita=True)
        f = 2
        for nac, datos_nac in tabla.items():
            _rescates_excel_celda(ws, f, 1, nac, bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], centrado=False)
            for i, (c, _) in enumerate(RESCATES_ETIQUETAS_CATEGORIA):
                _rescates_excel_celda(ws, f, 2 + i, datos_nac[c], bg=RESCATES_LETRA_T2[0], color_texto=RESCATES_LETRA_T2[1])
            _rescates_excel_celda(ws, f, 2 + len(RESCATES_ETIQUETAS_CATEGORIA), datos_nac["total"], bg=RESCATES_LETRA_T0[0], color_texto=RESCATES_LETRA_T0[1], negrita=True)
            f += 1
        ws.column_dimensions["A"].width = 28
        return ws

    _hoja_nacionalidad("Nacionalidades Nuevos", datos["nacionalidades"])
    _hoja_nacionalidad("Nacionalidades Reincidentes", datos["nacionalidades_re"])

    ws4 = wb.create_sheet("Inadmitidos y Retornados")
    _rescates_excel_fila(ws4, 1, ["Nacionalidad", "Adultos Hombres", "Adultas Mujeres", "Menores Hombres", "Menores Mujeres", "Total"], RESCATES_LETRA_T1, negrita=True)
    f = 2
    for nac, d in datos["nacionalidades_inadm"].items():
        _rescates_excel_celda(ws4, f, 1, nac, bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], centrado=False)
        for i, campo in enumerate(("H_A", "M_A", "H_m", "M_m")):
            _rescates_excel_celda(ws4, f, 2 + i, d[campo], bg=RESCATES_LETRA_T3[0], color_texto=RESCATES_LETRA_T3[1])
        _rescates_excel_celda(ws4, f, 6, d["total"], bg=RESCATES_LETRA_T0[0], color_texto=RESCATES_LETRA_T0[1], negrita=True)
        f += 1
    f += 1
    _rescates_excel_fila(ws4, f, ["TOTAL DE RETORNADOS A SU PAÍS DE ORIGEN", datos["dato"]], RESCATES_LETRA_T6, negrita=True)
    ws4.column_dimensions["A"].width = 32

    # @FADAR -- Retornados: hoja aparte, independiente de "Inadmitidos y Retornados".
    ws5 = wb.create_sheet("Retornados (detalle)")
    ws5.column_dimensions["A"].width = 28
    for col in "BCD":
        ws5.column_dimensions[col].width = 14
    rd = datos["retornados_detalle"]
    _rescates_excel_celda(ws5, 1, 1, f"Retornados — detalle por {rd['columna']}", bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True, centrado=False)
    ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    _rescates_excel_celda(ws5, 2, 1, "Retornados:", bg="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws5, 2, 2, rd["total_retornado"], bg="FFFFFF", color_texto=RESCATES_COLOR_CELESTE, negrita=True)
    fila_hdr5 = 4
    _rescates_excel_celda(ws5, fila_hdr5, 1, rd["columna"].upper(), bg="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws5, fila_hdr5, 2, "DEPORTADOS", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws5, fila_hdr5, 3, "RETORNADOS", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws5, fila_hdr5, 4, "TOTAL", bg="FFFFFF", negrita=True)
    f5 = fila_hdr5 + 1
    for fila_r in rd["filas"]:
        _rescates_excel_celda(ws5, f5, 1, fila_r["nombre"], centrado=False)
        _rescates_excel_celda(ws5, f5, 2, fila_r["deportado"])
        _rescates_excel_celda(ws5, f5, 3, fila_r["retornado"])
        _rescates_excel_celda(ws5, f5, 4, fila_r["total"])
        f5 += 1
    _rescates_excel_celda(ws5, f5, 1, "TOTAL", bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws5, f5, 2, rd["total_deportado"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws5, f5, 3, rd["total_retornado"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws5, f5, 4, rd["total_general"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="informe_diario_{fecha_str}.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# Reporte de Rescates -- "Regiones" (CECO): Rio Bravo / Centro / Suchiate.
# Reglas de negocio reutilizadas TAL CUAL de estadistica.generar_pdfT (ver
# _rescates_regiones) -- no se alteran a peticion explicita, incluyendo el
# filtro .exclude(...) del original y el caso especial de Chiapas.
# =============================================================================

def rescates_reporte_regiones(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_regiones_reporte(fecha_str, oficina or None)
    datos["oficinas"] = RESCATES_OFICINAS
    datos["oficina_seleccionada"] = oficina
    return render(request, "mapa/rescates_reporte_regiones.html", datos)


def rescates_reporte_regiones_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_regiones_reporte(fecha_str, oficina or None)

    template = get_template("mapa/_rescates_regiones_pdf.html")
    html_string = template.render(datos)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="regiones_ceco_{fecha_str}.pdf"'
    return response


def rescates_reporte_regiones_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_regiones_reporte(fecha_str, oficina or None)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Por zona"
    ws1.merge_cells("A1:E1")
    _rescates_excel_celda(ws1, 1, 1, "INSTITUTO NACIONAL DE MIGRACIÓN — DIRECCIÓN GENERAL DE COORDINACIÓN DE OFICINAS DE REPRESENTACION", negrita=True, tam=10)
    ws1.merge_cells("A2:E2")
    _rescates_excel_celda(ws1, 2, 1, f"Regiones (CECO) — {datos['fecha_actual']} — {datos['oficina']}", bg=RESCATES_COLOR_FONDO[0], color_texto=RESCATES_COLOR_FONDO[1], negrita=True, tam=12)

    fila = 4
    _rescates_excel_fila(ws1, fila, ["CECO", "Oficina de representación", "Rescates primera vez", "Reincidentes", "Total CECO"], RESCATES_LETRA_T1, negrita=True)
    fila += 1

    def _bloque_zona(nombre_zona, zona, subtotal):
        nonlocal fila
        for of, d in zona.items():
            _rescates_excel_fila(ws1, fila, [nombre_zona, of, d["nuevos"], d["reincidentes"], d["total"]], RESCATES_LETRA_T2)
            fila += 1
        _rescates_excel_fila(ws1, fila, ["", f"SUBTOTAL {nombre_zona.upper()}", subtotal["nuevos"], subtotal["reincidentes"], subtotal["total"]], RESCATES_LETRA_T0, negrita=True)
        fila += 1

    _bloque_zona("Río Bravo", datos["zona_rio_bravo"], datos["subtotal_rio_bravo"])
    _bloque_zona("Suchiate", datos["zona_suchiate"], datos["subtotal_suchiate"])
    _bloque_zona("Centro", datos["zona_centro"], datos["subtotal_centro"])

    _rescates_excel_fila(ws1, fila, ["", "TOTAL", datos["total_regiones"]["nuevos"], datos["total_regiones"]["reincidentes"], datos["total_regiones"]["total"]], RESCATES_LETRA_T1, negrita=True)
    fila += 1

    # @FADAR -- cuadro resumen final, mismos colores de zona del mapa,
    # total en verde.
    fila += 2
    for etiqueta, valor, color in [
        ("SUCHIATE", datos["subtotal_suchiate"]["total"], "B45309"),
        ("RÍO BRAVO", datos["subtotal_rio_bravo"]["total"], "1D4ED8"),
        ("CENTRO", datos["subtotal_centro"]["total"], "7C3AED"),
        ("TOTAL", datos["total_regiones"]["total"], "16A34A"),
    ]:
        _rescates_excel_celda(ws1, fila, 1, etiqueta, bg=color, color_texto="FFFFFF", negrita=True, centrado=False)
        _rescates_excel_celda(ws1, fila, 2, valor, bg=color, color_texto="FFFFFF", negrita=True)
        fila += 1

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 24
    for col in ("C", "D", "E"):
        ws1.column_dimensions[col].width = 18

    ws2 = wb.create_sheet("Por nacionalidad")
    _rescates_excel_fila(ws2, 1, ["Nacionalidad", "Rescates primera vez", "Reincidentes", "Total rescates"], RESCATES_LETRA_T1, negrita=True)
    f = 2
    for nac, d in datos["nac_1_reinc"].items():
        _rescates_excel_celda(ws2, f, 1, nac, bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], centrado=False)
        _rescates_excel_celda(ws2, f, 2, d["nuevos"], bg=RESCATES_LETRA_T2[0], color_texto=RESCATES_LETRA_T2[1])
        _rescates_excel_celda(ws2, f, 3, d["reincidentes"], bg=RESCATES_LETRA_T2[0], color_texto=RESCATES_LETRA_T2[1])
        _rescates_excel_celda(ws2, f, 4, d["total"], bg=RESCATES_LETRA_T0[0], color_texto=RESCATES_LETRA_T0[1], negrita=True)
        f += 1
    _rescates_excel_fila(ws2, f, ["TOTAL", datos["total_nac_1_reinc"]["nuevos"], datos["total_nac_1_reinc"]["reincidentes"], datos["total_nac_1_reinc"]["total"]], RESCATES_LETRA_T1, negrita=True)
    ws2.column_dimensions["A"].width = 28

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="regiones_ceco_{fecha_str}.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# Reporte "CECO 2" -- listado real por punto (PRH capturado en
# usuario_rescatepunto.puntoEstra) con su tipo, para un solo dia. NO replica
# el catalogo TIPO que se llena a mano (no coincide con mapa_prhs/
# mapa_tipoprh, confirmado en conversacion -- lo que capturan a mano no
# tiene una fuente verificable en la base). En su lugar, "tipo" se deriva de
# las banderas de medio de rescate YA verificadas y usadas en el resto del
# modulo -- por eso el total aqui SIEMPRE cuadra exacto contra el total real
# de rescates del dia, a diferencia del Excel manual.
# =============================================================================

RESCATES_TIPO_CASE_SQL = """CASE
        WHEN r.aeropuerto THEN 'Aeropuerto'
        WHEN r.carretero THEN 'Carretero'
        WHEN r."centralAutobus" THEN 'Central de Autobuses'
        WHEN r.ferrocarril THEN 'Ferroviario'
        WHEN r."casaSeguridad" THEN 'Casa de Seguridad'
        WHEN r.hotel THEN 'Hotel'
        WHEN r."puestosADispo" THEN 'Puestos a Disposición'
        WHEN r.voluntarios THEN 'Voluntario'
        ELSE 'Otro'
    END"""


def _rescates_ceco2_detalle(fecha_str, oficina=None):
    """fecha_str: 'YYYY-MM-DD'. Un solo dia (igual que Cuadro de
    Datos/Informe Diario/CECO). Filas = (oficina, punto, tipo, rescates,
    primera_vez, reincidencias) -- primera_vez/reincidencias vienen de la
    misma vista materializada de reincidencia (>=2 apariciones historicas)
    ya usada en el resto del modulo, unida por identidad (nombre+apellidos+
    nacionalidad)."""
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    fecha_fmt = fecha_obj.strftime("%d-%m-%y")

    filtro = ""
    params = [fecha_fmt]
    if oficina:
        filtro = ' AND r."oficinaRepre" = %s'
        params.append(oficina)

    with connection.cursor() as cur:
        cur.execute(
            f'SELECT r."oficinaRepre", r."puntoEstra", {RESCATES_TIPO_CASE_SQL} AS tipo, COUNT(*), '
            f"  COUNT(*) FILTER (WHERE v.clasificacion = 'Rescate primera vez'), "
            f"  COUNT(*) FILTER (WHERE v.clasificacion = 'Reincidente') "
            f'FROM usuario_rescatepunto r '
            f'JOIN {RESCATES_MV_REINCIDENCIA} v '
            f'  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad '
            f'WHERE r.fecha = %s{filtro} '
            f'GROUP BY r."oficinaRepre", r."puntoEstra", {RESCATES_TIPO_CASE_SQL} '
            f'ORDER BY r."oficinaRepre", r."puntoEstra"',
            params,
        )
        filas = cur.fetchall()

    filas_tabla = [
        {"oficina": of, "prh": prh, "tipo": tipo, "rescates": r, "primera_vez": pv, "reincidencias": rc}
        for of, prh, tipo, r, pv, rc in filas
    ]
    return {
        "fecha_actual": f"{fecha_obj.day:02d}/{fecha_obj.month:02d}/{fecha_obj.year}",
        "fecha_iso": fecha_str,
        "oficina": oficina or "Nacional",
        "filas": filas_tabla,
        "total": sum(f["rescates"] for f in filas_tabla),
        "total_primera_vez": sum(f["primera_vez"] for f in filas_tabla),
        "total_reincidencias": sum(f["reincidencias"] for f in filas_tabla),
    }


def rescates_reporte_ceco2(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_ceco2_detalle(fecha_str, oficina or None)
    datos["oficinas"] = RESCATES_OFICINAS
    datos["oficina_seleccionada"] = oficina
    return render(request, "mapa/rescates_reporte_ceco2.html", datos)


def rescates_reporte_ceco2_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_ceco2_detalle(fecha_str, oficina or None)

    template = get_template("mapa/_rescates_ceco2_pdf.html")
    html_string = template.render(datos)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ceco2_{fecha_str}.pdf"'
    return response


def rescates_reporte_ceco2_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_ceco2_detalle(fecha_str, oficina or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CECO 2"
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20

    ws.merge_cells("A1:B2")
    _rescates_excel_celda(ws, 1, 1, "Fecha:", bg=RESCATES_COLOR_FONDO[0], color_texto="FFFFFF", negrita=True)
    ws.merge_cells("A3:B3")
    _rescates_excel_celda(ws, 3, 1, datos["fecha_actual"], bg="FFFFFF", negrita=True)

    ws.merge_cells("D1:E1")
    _rescates_excel_celda(ws, 1, 4, "TOTAL", bg=RESCATES_COLOR_FONDO[0], color_texto="FFFFFF", negrita=True)
    ws.merge_cells("D2:E2")
    _rescates_excel_celda(ws, 2, 4, datos["total"], bg="FFFFFF", negrita=True, tam=14)

    fila = 5
    _rescates_excel_fila(ws, fila, ["FECHA", "OR", "PRH", "RESCATES TOTAL", "PRIMERA VEZ/RESCATES REALES", "REINCIDENCIAS", "TIPO"], RESCATES_LETRA_T1, negrita=True)
    fila += 1
    for f in datos["filas"]:
        _rescates_excel_celda(ws, fila, 1, datos["fecha_actual"], centrado=True)
        _rescates_excel_celda(ws, fila, 2, f["oficina"], centrado=False)
        _rescates_excel_celda(ws, fila, 3, f["prh"] or "OTRA AUTORIDAD", centrado=False)
        _rescates_excel_celda(ws, fila, 4, f["rescates"])
        _rescates_excel_celda(ws, fila, 5, f["primera_vez"])
        _rescates_excel_celda(ws, fila, 6, f["reincidencias"])
        _rescates_excel_celda(ws, fila, 7, f["tipo"], centrado=False)
        fila += 1

    _rescates_excel_celda(ws, fila, 1, "TOTAL", bg="1F3864", color_texto="FFFFFF", negrita=True, centrado=False)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
    _rescates_excel_celda(ws, fila, 4, datos["total"], bg="1F3864", color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 5, datos["total_primera_vez"], bg="1F3864", color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 6, datos["total_reincidencias"], bg="1F3864", color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 7, "", bg="1F3864")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="ceco2_{fecha_str}.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# Reportes "CECO V1" / "CECO V2" -- nacionalidad x 8 categorias (mismas
# RESCATES_ETIQUETAS_CATEGORIA de Informe Diario), un bloque por entidad (32)
# mas un bloque "Nacional", replicando el formato de un archivo de
# referencia entregado por el area. Confirmado en conversacion:
#  - V1 = TODOS los rescates del dia (primera vez + reincidentes juntos en
#    un solo total, tal como indica el nombre del archivo original
#    "rescates+reincidentes"), incluye inadmitidos.
#  - V2 = "rescates reales" = SOLO primera vez (excluye reincidentes),
#    tambien incluye inadmitidos (no se excluyeron explicitamente).
#  - Lista de nacionalidades: UNA sola, calculada a nivel nacional (las 32
#    entidades juntas) para ese dia -- se repite igual en los 32 bloques,
#    aunque una entidad en particular tenga 0 en alguna.
#  - El PDF no puede replicar el ancho de la hoja de Excel (331 columnas) en
#    una sola pagina -- se generan paginas separadas por entidad, aceptado
#    explicitamente ("no importa si sale cortado o segmentado").
# =============================================================================

def _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=False):
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    fecha_fmt = fecha_obj.strftime("%d-%m-%y")

    if solo_primera_vez:
        with connection.cursor() as cur:
            cur.execute(
                f'SELECT r."oficinaRepre", r.nacionalidad, r.sexo, r.edad, r."numFamilia" '
                f'FROM usuario_rescatepunto r '
                f'JOIN {RESCATES_MV_REINCIDENCIA} v '
                f'  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad '
                f"WHERE r.fecha = %s AND v.clasificacion = 'Rescate primera vez'",
                [fecha_fmt],
            )
            datos_dia = [
                {"oficinaRepre": of, "nacionalidad": nac, "sexo": sexo, "edad": edad, "numFamilia": nf}
                for of, nac, sexo, edad, nf in cur.fetchall()
            ]
    else:
        datos_dia = list(
            RescatePunto.objects.filter(fecha=fecha_fmt)
            .values("oficinaRepre", "nacionalidad", "sexo", "edad", "numFamilia")
        )

    # Nacionalidades a nivel nacional (las 32 entidades juntas), ordenadas
    # por total descendente -- misma lista para los 32 bloques.
    conteo_nac = {}
    for d in datos_dia:
        conteo_nac[d["nacionalidad"]] = conteo_nac.get(d["nacionalidad"], 0) + 1
    nacionalidades = [n for n, _ in sorted(conteo_nac.items(), key=lambda x: x[1], reverse=True)]

    def _fila_vacia():
        fila = {clave: 0 for clave, _ in RESCATES_ETIQUETAS_CATEGORIA}
        fila["total"] = 0
        fila["nucleos"] = 0
        return fila

    bloques = {of: {nac: _fila_vacia() for nac in nacionalidades} for of in RESCATES_OFICINAS}
    totales_entidad = {of: _fila_vacia() for of in RESCATES_OFICINAS}
    nacional = {nac: _fila_vacia() for nac in nacionalidades}
    total_nacional = _fila_vacia()

    for d in datos_dia:
        of = d["oficinaRepre"]
        if of not in bloques:
            continue
        cat = _rescates_clasificar_categoria(d["sexo"], d["edad"], d["numFamilia"])
        nac = d["nacionalidad"]
        bloques[of][nac][cat] += 1
        bloques[of][nac]["total"] += 1
        totales_entidad[of][cat] += 1
        totales_entidad[of]["total"] += 1
        nacional[nac][cat] += 1
        nacional[nac]["total"] += 1
        total_nacional[cat] += 1
        total_nacional["total"] += 1

    # @FADAR -- "NUCLEOS FAMILIARES": columna adicional que se me habia
    # pasado en la primera revision del archivo de referencia. Misma
    # definicion ya usada en "Núcleos familiares detectados" del dashboard
    # (numFamilia>0, agrupado por oficina+hora+puntoEstra+numFamilia, con
    # >=2 integrantes) -- no se inventa una regla nueva.
    filtro_reinc_join = ""
    filtro_reinc_where = ""
    if solo_primera_vez:
        filtro_reinc_join = (
            f'JOIN {RESCATES_MV_REINCIDENCIA} v '
            f'  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad '
        )
        filtro_reinc_where = " AND v.clasificacion = 'Rescate primera vez'"
    with connection.cursor() as cur:
        cur.execute(
            f'SELECT oficina, nacionalidad, COUNT(*) FROM ( '
            f'  SELECT r."oficinaRepre" AS oficina, r.nacionalidad, r.hora, r."puntoEstra" AS punto, r."numFamilia" AS nf '
            f'  FROM ( '
            f'    SELECT r.*, COUNT(*) OVER (PARTITION BY r."oficinaRepre", r.hora, r."puntoEstra", r."numFamilia") AS integrantes '
            f'    FROM usuario_rescatepunto r '
            f'    {filtro_reinc_join}'
            f'    WHERE r.fecha = %s AND r."numFamilia" > 0{filtro_reinc_where} '
            f'  ) r '
            f'  WHERE integrantes >= 2 '
            f'  GROUP BY oficina, r.nacionalidad, r.hora, punto, nf '
            f') grupos '
            f'GROUP BY oficina, nacionalidad',
            [fecha_fmt],
        )
        for of, nac, n in cur.fetchall():
            if of not in bloques or nac not in bloques[of]:
                continue
            bloques[of][nac]["nucleos"] += n
            totales_entidad[of]["nucleos"] += n
            nacional[nac]["nucleos"] += n
            total_nacional["nucleos"] += n

    # Version "lista" de los mismos datos, para el template PDF (Django no
    # permite bloques.of con "of" como variable de loop -- solo listas).
    def _lista_bloque(nombre_bloque, filas_por_nac, total_bloque):
        return {
            "nombre": nombre_bloque,
            "filas": [{"nacionalidad": nac, **filas_por_nac[nac]} for nac in nacionalidades],
            "total": total_bloque,
        }

    bloques_lista = [_lista_bloque(of, bloques[of], totales_entidad[of]) for of in RESCATES_OFICINAS]
    bloques_lista.append(_lista_bloque("NACIONAL POR DÍA", nacional, total_nacional))

    # @FADAR -- misma info "transpuesta" para la vista en pantalla en
    # horizontal (una sola tabla ancha con scroll lateral, igual que el
    # Excel, en vez de un bloque apilado por entidad): cada fila es una
    # nacionalidad, con los datos de los 33 bloques ya en orden.
    filas_horizontal = [
        {"nacionalidad": nac, "por_bloque": [b["filas"][j] for b in bloques_lista]}
        for j, nac in enumerate(nacionalidades)
    ]
    fila_total_horizontal = [b["total"] for b in bloques_lista]

    # @FADAR -- cuadro "SUCHIATE / RÍO BRAVO / CENTRO / TOTAL": se arma
    # sumando los mismos totales_entidad ya calculados arriba (que ya
    # respetan el filtro solo_primera_vez de este reporte), en vez de
    # llamar a _rescates_regiones() -- esa funcion aplica una excepcion de
    # Chiapas (100% reincidente) ajena a la definicion de este reporte, y
    # dejaba fuera sus rescates de primera vez (causaba que el cuadro no
    # cuadrara contra "NACIONAL POR DÍA").
    resumen_zonas = {
        "rio_bravo": sum(totales_entidad[of]["total"] for of in RESCATES_ZONA_RIO_BRAVO if of in totales_entidad),
        "centro": sum(totales_entidad[of]["total"] for of in RESCATES_ZONA_CENTRO if of in totales_entidad),
        "suchiate": sum(totales_entidad[of]["total"] for of in RESCATES_ZONA_SUCHIATE if of in totales_entidad),
    }
    resumen_zonas["total"] = resumen_zonas["rio_bravo"] + resumen_zonas["centro"] + resumen_zonas["suchiate"]

    return {
        "resumen_zonas": resumen_zonas,
        "fecha_actual": f"{fecha_obj.day:02d}/{fecha_obj.month:02d}/{fecha_obj.year}",
        "fecha_iso": fecha_str,
        "filas_horizontal": filas_horizontal,
        "fila_total_horizontal": fila_total_horizontal,
        "nacionalidades": nacionalidades,
        "categorias": RESCATES_ETIQUETAS_CATEGORIA,
        "oficinas": RESCATES_OFICINAS,
        "bloques": bloques,
        "bloques_lista": bloques_lista,
        "totales_entidad": totales_entidad,
        "nacional": nacional,
        "total_nacional": total_nacional,
    }


# @FADAR -- colores/fuente extraidos EXACTOS del archivo de referencia
# (Rescatados_25AGO2026-CECO(rescates+reincidentes).xlsx, hoja "CECCO 2"),
# resueltos desde theme+tint a RGB real. La rotacion de 3 colores del
# encabezado de entidad en el original no sigue una regla identificable
# (parece asignacion manual, no ciclica) -- aqui se replica como una
# rotacion fija de esos mismos 3 colores, en vez de intentar adivinar el
# orden exacto sin base verificable.
RESCATES_CECOV_FUENTE = "Montserrat"
RESCATES_CECOV_VERDE = "70AD47"            # titulo "RESCATADOS" / bloque "NACIONAL POR DIA"
RESCATES_CECOV_AZUL = "4472C4"             # columna TOTAL
RESCATES_CECOV_NARANJA = "ED7D31"          # grupo "Mayores solos" / "Menores no acompañados"
RESCATES_CECOV_VERDE_OSCURO = "13322B"     # grupo "Mayores acompañan NNA" / "Menores acompañados"
RESCATES_CECOV_GRIS_CLARO = "E7E6E6"       # celda TOTAL por nacionalidad
RESCATES_CECOV_NEGRO = "000000"            # fila TOTAL de cada bloque
RESCATES_CECOV_ROTACION_ENTIDAD = ["808080", "ED7D31", "44546A"]


def _rescates_ceco_v_excel(datos, titulo, hoja_nombre="Hoja1"):
    # @FADAR -- estructura de encabezado EXACTA verificada celda por celda
    # contra el archivo de referencia (4 filas de encabezado, no 3; datos
    # empiezan en la fila 5; bloque de 10 columnas -- 9 categorias/total +
    # "NUCLEOS FAMILIARES" -- sin columna de separacion entre bloques).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja_nombre

    ws.column_dimensions["A"].width = 30
    col = 2  # B
    indice_bloque = 0
    fila_maxima = 0

    def _escribir_bloque(nombre_bloque, filas_por_nac, total_bloque, es_nacional=False):
        nonlocal col, indice_bloque, fila_maxima
        col_inicio = col
        color_header = RESCATES_CECOV_VERDE if es_nacional else RESCATES_CECOV_ROTACION_ENTIDAD[indice_bloque % 3]

        # Fila 1: nombre de la entidad (9 columnas) + celda decorativa (10a).
        _rescates_excel_celda(ws, 1, col_inicio, nombre_bloque, bg=color_header, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE, tam=14 if es_nacional else 12)
        ws.merge_cells(start_row=1, start_column=col_inicio, end_row=1, end_column=col_inicio + 8)
        _rescates_excel_celda(ws, 1, col_inicio + 9, None, bg=color_header)

        # Filas 2-4: TOTAL (3 filas).
        _rescates_excel_celda(ws, 2, col_inicio, "TOTAL", bg=RESCATES_CECOV_AZUL, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=2, start_column=col_inicio, end_row=4, end_column=col_inicio)

        # "Mayores de edad solos" / "...que acompañan NNA": 2 filas x 2 columnas.
        for nombre_grupo, c, color_grupo in [
            ("MAYORES DE EDAD SOLOS", col_inicio + 1, RESCATES_CECOV_NARANJA),
            ("MAYORES DE EDAD QUE ACOMPAÑAN A NNA'S", col_inicio + 3, RESCATES_CECOV_VERDE_OSCURO),
        ]:
            _rescates_excel_celda(ws, 2, c, nombre_grupo, bg=color_grupo, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
            ws.merge_cells(start_row=2, start_column=c, end_row=3, end_column=c + 1)

        # "Menores de edad" (fila 2, 4 columnas) subdividido en fila 3:
        # "Acompañados" (verde oscuro) / "No acompañados" (naranja).
        _rescates_excel_celda(ws, 2, col_inicio + 5, "MENORES DE EDAD", bg=RESCATES_CECOV_VERDE_OSCURO, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=2, start_column=col_inicio + 5, end_row=2, end_column=col_inicio + 8)
        _rescates_excel_celda(ws, 3, col_inicio + 5, "ACOMPAÑADOS", bg=RESCATES_CECOV_VERDE_OSCURO, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=3, start_column=col_inicio + 5, end_row=3, end_column=col_inicio + 6)
        _rescates_excel_celda(ws, 3, col_inicio + 7, "NO ACOMPAÑADOS", bg=RESCATES_CECOV_NARANJA, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=3, start_column=col_inicio + 7, end_row=3, end_column=col_inicio + 8)

        # Fila 4: H/M de las 8 categorias, con el color de su grupo.
        colores_categoria = [RESCATES_CECOV_NARANJA] * 2 + [RESCATES_CECOV_VERDE_OSCURO] * 4 + [RESCATES_CECOV_NARANJA] * 2
        for i, (clave, _) in enumerate(RESCATES_ETIQUETAS_CATEGORIA):
            letra = "H" if i % 2 == 0 else "M"
            _rescates_excel_celda(ws, 4, col_inicio + 1 + i, letra, bg=colores_categoria[i], color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)

        # "NUCLEOS FAMILIARES": 3 filas (2-4), 1 columna, verde.
        _rescates_excel_celda(ws, 2, col_inicio + 9, "NUCLEOS FAMILIARES", bg=RESCATES_CECOV_VERDE, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=2, start_column=col_inicio + 9, end_row=4, end_column=col_inicio + 9)

        # Datos: empiezan en la fila 5.
        f = 5
        for nac in datos["nacionalidades"]:
            fila_nac = filas_por_nac[nac]
            if col_inicio == 2:
                _rescates_excel_celda(ws, f, col_inicio - 1, nac, negrita=True, tam=12, centrado=False, fuente=RESCATES_CECOV_FUENTE)
            _rescates_excel_celda(ws, f, col_inicio, fila_nac["total"], bg=RESCATES_CECOV_GRIS_CLARO, color_texto="000000", negrita=True, fuente=RESCATES_CECOV_FUENTE)
            for i, (clave, _) in enumerate(RESCATES_ETIQUETAS_CATEGORIA):
                _rescates_excel_celda(ws, f, col_inicio + 1 + i, fila_nac[clave], fuente=RESCATES_CECOV_FUENTE)
            _rescates_excel_celda(ws, f, col_inicio + 9, fila_nac["nucleos"], fuente=RESCATES_CECOV_FUENTE)
            f += 1
        if col_inicio == 2:
            _rescates_excel_celda(ws, f, col_inicio - 1, "TOTAL", negrita=True, tam=12, fuente=RESCATES_CECOV_FUENTE)
        _rescates_excel_celda(ws, f, col_inicio, total_bloque["total"], bg=RESCATES_CECOV_NEGRO, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE)
        for i, (clave, _) in enumerate(RESCATES_ETIQUETAS_CATEGORIA):
            _rescates_excel_celda(ws, f, col_inicio + 1 + i, total_bloque[clave], bg=RESCATES_CECOV_NEGRO, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE)
        _rescates_excel_celda(ws, f, col_inicio + 9, total_bloque["nucleos"], bg=RESCATES_CECOV_NEGRO, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE)
        fila_maxima = max(fila_maxima, f)

        for c in range(col_inicio, col_inicio + 10):
            ws.column_dimensions[get_column_letter(c)].width = 9
        col = col_inicio + 10  # 10 columnas de datos, sin separacion (verificado contra el original)
        indice_bloque += 1

    ws.merge_cells("A1:A2")
    _rescates_excel_celda(ws, 1, 1, "RESCATADOS", bg=RESCATES_CECOV_VERDE, color_texto="FFFFFF", negrita=True, tam=14, fuente=RESCATES_CECOV_FUENTE)
    ws.merge_cells("A3:A4")
    _rescates_excel_celda(ws, 3, 1, datos["fecha_actual"], bg=RESCATES_CECOV_VERDE, color_texto="FFFFFF", negrita=True, tam=14, fuente=RESCATES_CECOV_FUENTE)

    for of in datos["oficinas"]:
        _escribir_bloque(of, datos["bloques"][of], datos["totales_entidad"][of])
    _escribir_bloque("NACIONAL POR DÍA", datos["nacional"], datos["total_nacional"], es_nacional=True)

    # @FADAR -- cuadro "SUCHIATE / RÍO BRAVO / CENTRO / TOTAL", mismos
    # colores exactos del archivo de referencia (no los de RESCATES_ZONA_COLOR
    # del mapa -- este cuadro es tal como aparece en el original).
    fz = fila_maxima + 2
    rz = datos["resumen_zonas"]
    for etiqueta, valor, color, color_texto in [
        ("SUCHIATE", rz["suchiate"], "44546A", "FFFFFF"),
        ("RÍO BRAVO", rz["rio_bravo"], "ED7D31", "FFFFFF"),
        ("CENTRO", rz["centro"], "808080", "FFFFFF"),
        ("TOTAL", rz["total"], "70AD47", "44546A"),
    ]:
        _rescates_excel_celda(ws, fz, 1, etiqueta, bg=color, color_texto=color_texto, negrita=True, tam=18, centrado=False, fuente=RESCATES_CECOV_FUENTE)
        _rescates_excel_celda(ws, fz, 2, valor, bg=color, color_texto=color_texto, negrita=True, tam=18, fuente=RESCATES_CECOV_FUENTE)
        fz += 1

    return wb


def rescates_reporte_cecov1(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=False)
    return render(request, "mapa/rescates_reporte_cecov.html", {**datos, "titulo": "CECO V1", "url_pdf": "rescates_reporte_cecov1_pdf", "url_excel": "rescates_reporte_cecov1_excel"})


def rescates_reporte_cecov1_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=False)
    template = get_template("mapa/_rescates_cecov_pdf.html")
    html_string = template.render({**datos, "titulo": "CECO V1 — Rescates + Reincidentes"})
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ceco_v1_{fecha_str}.pdf"'
    return response


def rescates_reporte_cecov1_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=False)
    wb = _rescates_ceco_v_excel(datos, "CECO V1 — Rescates + Reincidentes", "CECO V1")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="ceco_v1_{fecha_str}.xlsx"'
    wb.save(response)
    return response


def rescates_reporte_cecov2(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=True)
    return render(request, "mapa/rescates_reporte_cecov.html", {**datos, "titulo": "CECO V2", "url_pdf": "rescates_reporte_cecov2_pdf", "url_excel": "rescates_reporte_cecov2_excel"})


def rescates_reporte_cecov2_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=True)
    template = get_template("mapa/_rescates_cecov_pdf.html")
    html_string = template.render({**datos, "titulo": "CECO V2 — Rescates Reales (primera vez)"})
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ceco_v2_{fecha_str}.pdf"'
    return response


def rescates_reporte_cecov2_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=True)
    wb = _rescates_ceco_v_excel(datos, "CECO V2 — Rescates Reales (primera vez)", "CECO V2")
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="ceco_v2_{fecha_str}.xlsx"'
    wb.save(response)
    return response

