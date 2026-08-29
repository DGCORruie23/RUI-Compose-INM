# @FADAR -- Fase 1 del traslado del modulo de rescates: dashboard.
# Codigo copiado TAL CUAL desde mapa/views.py (misma logica, mismas reglas,
# misma estructura) -- mapa/views.py se deja intacto por ahora, el
# descarte de esa copia y del boton que la enlaza queda para una fase
# posterior, a peticion explicita.
from bokeh.plotting import figure
from bokeh.models import ColumnDataSource, HoverTool, DatetimeTickFormatter, LabelSet, NumeralTickFormatter, RangeTool
from bokeh.layouts import column
from bokeh.embed import components
from bokeh.transform import cumsum
from math import pi, cos, sin

import json
import openpyxl
import os
import unicodedata
from collections import Counter
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, date, timedelta

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.core.cache import cache  # @FADAR
from django.db import connection
from django.db.models import Count, Max, Q, Sum
from django.db.models.functions import TruncDay, TruncWeek
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.template.loader import get_template

from weasyprint import HTML

from usuario.models import RescatePunto
from mapa.models import CATALOGO_PUNTOS_MEX_EXT, Recibidos, RegistroMexExtPunto, Repatriados  # @FADAR


def normalizar_nombre(texto):
    if not texto: return ""
    texto = unicodedata.normalize('NFD', texto)
    texto = texto.encode('ascii', 'ignore').decode("utf-8")
    return str(texto).strip().upper()


# @FADAR -- clasificacion de nacionalidades "atipicas" (Medio Oriente/
# Europa/otras poco conocidas), armada a partir de los codigos ISO3 que
# realmente aparecen en la tabla -- no una lista generica adivinada. Para
# volver a verificar contra los datos reales:
#   SELECT iso3, MAX(nacionalidad), COUNT(*) FROM usuario_rescatepunto
#   GROUP BY iso3 ORDER BY 3 DESC;
RESCATES_ISO3_MEDIO_ORIENTE = {
    'TUR', 'AFG', 'EGY', 'JOR', 'IRN', 'IRQ', 'SYR', 'YEM', 'ISR', 'LBN',
    'SAU', 'ARE', 'PSE', 'KWT',
}
RESCATES_ISO3_EUROPA = {
    'POL', 'DEU', 'FRA', 'NLD', 'ITA', 'ESP', 'ROU', 'GBR', 'CHE', 'ALB',
    'BEL', 'SVK', 'CZE', 'HUN', 'SVN', 'DNK', 'IRL', 'AUT', 'MKD', 'PRT',
    'LTU', 'BGR', 'FIN', 'BLR', 'SWE', 'GRC', 'NOR', 'EST', 'HRV', 'LVA',
    'KOS', 'SRB', 'SMR', 'LUX', 'CYP', 'AND', 'MCO', 'RUS', 'UKR', 'MDA',
    'ARM', 'GEO', 'AZE',
}
# El grueso "normal" de la operacion (America Latina/Caribe) -- nunca cuenta
# como atipica, sin importar el volumen.
RESCATES_ISO3_LATAM_CARIBE = {
    'VEN', 'ECU', 'HND', 'GTM', 'COL', 'SLV', 'CUB', 'NIC', 'HTI', 'PER',
    'DOM', 'BOL', 'BRA', 'CHL', 'PAN', 'BLZ', 'CRI', 'URY', 'JAM', 'ARG',
    'DMA', 'GUY', 'TTO', 'MEX', 'ANT', 'PRI', 'GLP', 'SUR', 'VGB', 'VCT',
    'TCA', 'BHS', 'LCA', 'GRD', 'GUF', 'PRY',
}
# Paises grandes/muy conocidos fuera de Medio Oriente, Europa o LATAM (ej.
# India, EUA) -- se excluyen a proposito de "poco conocidas" para no
# etiquetarlos como obscuros solo por tener pocos registros en esta muestra.
# China (CHN) se pidio explicitamente que SI cuente como atipica, por eso
# no esta en este set -- cae en "Otras / poco conocidas" por defecto.
RESCATES_ISO3_CONOCIDOS_NO_ATIPICOS = {
    'IND', 'USA', 'CAN', 'JPN', 'KOR', 'VNM', 'PAK', 'PHL', 'IDN',
    'BGD', 'AUS', 'NZL', 'THA', 'MYS', 'SGP',
}

RESCATES_OFICINAS = [
    "AGUASCALIENTES", "BAJA CALIFORNIA", "BAJA CALIFORNIA SUR", "CAMPECHE",
    "CDMX", "CHIAPAS", "CHIHUAHUA", "COAHUILA", "COLIMA", "DURANGO",
    "EDOMEX", "GUANAJUATO", "GUERRERO", "HIDALGO", "JALISCO", "MICHOACÁN",
    "MORELOS", "NAYARIT", "NUEVO LEÓN", "OAXACA", "PUEBLA", "QUERÉTARO",
    "QUINTANA ROO", "SAN LUIS POTOSÍ", "SINALOA", "SONORA", "TABASCO",
    "TAMAULIPAS", "TLAXCALA", "VERACRUZ", "YUCATÁN", "ZACATECAS",
]

# Las 3 "zonas migratorias" (Rio Bravo/Centro/Suchiate) tal como estan
# definidas en estadistica.views.generar_pdfT (ORs_CECO_N/C/S) -- se copian
# aqui identicas, sin reordenar ni completar/corregir nada, porque la
# seccion "Regiones" del dashboard reutiliza esa clasificacion tal cual.
RESCATES_ZONA_RIO_BRAVO = [
    "BAJA CALIFORNIA", "CHIHUAHUA", "COAHUILA", "DURANGO", "NUEVO LEÓN",
    "SAN LUIS POTOSÍ", "SINALOA", "SONORA", "TAMAULIPAS",
]
RESCATES_ZONA_CENTRO = [
    "AGUASCALIENTES", "BAJA CALIFORNIA SUR", "COLIMA", "CDMX", "GUANAJUATO",
    "GUERRERO", "JALISCO", "MICHOACÁN", "MORELOS", "NAYARIT", "QUERÉTARO",
    "ZACATECAS",
]
RESCATES_ZONA_SUCHIATE = [
    "CAMPECHE", "CHIAPAS", "HIDALGO", "EDOMEX", "OAXACA", "PUEBLA",
    "QUINTANA ROO", "TABASCO", "TLAXCALA", "VERACRUZ", "YUCATÁN",
]

# @FADAR -- continente americano completo (LATAM/Caribe + Norteamerica),
# para marcar nacionalidades "extracontinentales" (fuera de America).
# Distinto del set de "atipicas" de arriba -- ese es solo para el widget
# del dashboard, este es "es de America si/no".
RESCATES_ISO3_AMERICA = RESCATES_ISO3_LATAM_CARIBE | {'USA', 'CAN'}


def _rescates_es_extracontinental(iso3):
    """True si el pais NO pertenece al continente americano."""
    return iso3 not in RESCATES_ISO3_AMERICA


def _rescates_region_atipica(iso3):
    """None si la nacionalidad NO cuenta como atipica (LATAM/Caribe o un
    pais grande/conocido); si no, el nombre de la region atipica."""
    if iso3 in RESCATES_ISO3_MEDIO_ORIENTE:
        return "Medio Oriente"
    if iso3 in RESCATES_ISO3_EUROPA:
        return "Europa"
    if iso3 in RESCATES_ISO3_LATAM_CARIBE or iso3 in RESCATES_ISO3_CONOCIDOS_NO_ATIPICOS:
        return None
    return "Otras / poco conocidas"


RESCATES_MV_REINCIDENCIA = "mapa_mv_reincidencia_rescates"


def _rescates_set_duplicados_historicos():
    """Set de (nombre, apellidos, nacionalidad) clasificados 'Reincidente'
    en mapa_mv_reincidencia_rescates -- usada en los 3 lugares que la
    necesitan (dashboard -> seccion "Regiones", reporte Cuadro de Datos,
    reporte Informe Diario). Misma interfaz que antes, ahora respaldada por
    la vista materializada en vez de recalcularse en Python."""
    with connection.cursor() as cur:
        cur.execute(
            f"SELECT nombre, apellidos, nacionalidad FROM {RESCATES_MV_REINCIDENCIA} "
            f"WHERE clasificacion = 'Reincidente'"
        )
        return set(cur.fetchall())


def _rescates_regiones(fecha_inicio, fecha_fin, oficina=None):
    """Reutiliza TAL CUAL la logica de generar_pdfT (estadistica/views.py,
    seccion "Cuadro de Registros" / CECO 2) -- 3 zonas migratorias (Rio
    Bravo/Centro/Suchiate) que cubren las 32 oficinas. Ninguna regla de
    negocio se altera respecto al original:
     - Mismo filtro .exclude(...) del reporte oficial (NO es la regla de
       "inadmitido" que usa el resto de este modulo -- se deja identico a
       proposito, tal como esta en el codigo fuente).
     - CHIAPAS aparte: el 100% de sus registros del rango cuenta como
       reincidente, SIN comparar contra el historico (asi esta en el
       reporte oficial -- comportamiento real y aceptado, no un descuido).
     - "Reincidente" para el resto = coincide (nombre, apellidos,
       nacionalidad) con otro registro en TODO el historico de
       RescatePunto.
    Unico cambio respecto al original: acepta un rango de fechas (el
    reporte original solo trabajaba un dia a la vez) y respeta el filtro
    de entidad si ya hay una oficina seleccionada -- son parte del
    "wrapper" de filtros compartido por el dashboard y los reportes, no de
    la regla de negocio en si. Usada tanto por el dashboard (rango) como
    por el reporte "Regiones" (fecha_inicio == fecha_fin)."""
    fecha_ini_obj = datetime.strptime(fecha_inicio, "%Y-%m-%d")
    fecha_fin_obj = datetime.strptime(fecha_fin, "%Y-%m-%d")
    dias_rango = (fecha_fin_obj - fecha_ini_obj).days
    array_fechas = [
        (fecha_ini_obj + timedelta(days=d)).strftime("%d-%m-%y")
        for d in range(max(dias_rango, 0) + 1)
    ]

    campos_valores = (
        'nombre', 'apellidos', 'nacionalidad', 'oficinaRepre',
        'puntoEstra', 'fecha', 'sexo', 'edad', 'numFamilia', 'iso3',
    )
    exclude_original = dict(
        aeropuerto=False, carretero=True, casaSeguridad=False, centralAutobus=False,
        ferrocarril=False, hotel=False, puestosADispo=False, voluntarios=True, otro=True,
    )

    oficinas_sin_chiapas = [o for o in RESCATES_OFICINAS if o != "CHIAPAS"]
    if oficina:
        oficinas_sin_chiapas = [oficina] if oficina in oficinas_sin_chiapas else []
    datos_no_chiapas = list(
        RescatePunto.objects.filter(fecha__in=array_fechas, oficinaRepre__in=oficinas_sin_chiapas)
        .exclude(**exclude_original)
        .values(*campos_valores)
        .annotate(total=Count('idRescate'))
    ) if oficinas_sin_chiapas else []

    incluir_chiapas = (not oficina) or (oficina == "CHIAPAS")
    datos_chiapas = list(
        RescatePunto.objects.filter(fecha__in=array_fechas, oficinaRepre="CHIAPAS")
        .exclude(**exclude_original)
        .values(*campos_valores)
        .annotate(total=Count('idRescate'))
    ) if incluir_chiapas else []

    set_duplicados_local = _rescates_set_duplicados_historicos()

    nuevos_local, reincidentes_local = [], []
    for dato in datos_no_chiapas:
        clave = (dato['nombre'], dato['apellidos'], dato['nacionalidad'])
        if clave in set_duplicados_local:
            reincidentes_local.append(dato)
        else:
            nuevos_local.append(dato)
    # Chiapas: 100% reincidente, sin excepcion (regla oficial).
    reincidentes_local.extend(datos_chiapas)

    zona_por_oficina_local = {}
    zona_rio_bravo_local = {o: {"nuevos": 0, "reincidentes": 0, "total": 0} for o in RESCATES_ZONA_RIO_BRAVO}
    zona_centro_local = {o: {"nuevos": 0, "reincidentes": 0, "total": 0} for o in RESCATES_ZONA_CENTRO}
    zona_suchiate_local = {o: {"nuevos": 0, "reincidentes": 0, "total": 0} for o in RESCATES_ZONA_SUCHIATE}
    for o in RESCATES_ZONA_RIO_BRAVO:
        zona_por_oficina_local[o] = zona_rio_bravo_local
    for o in RESCATES_ZONA_CENTRO:
        zona_por_oficina_local[o] = zona_centro_local
    for o in RESCATES_ZONA_SUCHIATE:
        zona_por_oficina_local[o] = zona_suchiate_local

    subtotal_rio_bravo_local = {"nuevos": 0, "reincidentes": 0, "total": 0}
    subtotal_centro_local = {"nuevos": 0, "reincidentes": 0, "total": 0}
    subtotal_suchiate_local = {"nuevos": 0, "reincidentes": 0, "total": 0}
    subtotal_por_oficina_local = {}
    for o in RESCATES_ZONA_RIO_BRAVO:
        subtotal_por_oficina_local[o] = subtotal_rio_bravo_local
    for o in RESCATES_ZONA_CENTRO:
        subtotal_por_oficina_local[o] = subtotal_centro_local
    for o in RESCATES_ZONA_SUCHIATE:
        subtotal_por_oficina_local[o] = subtotal_suchiate_local

    nac_1_reinc_local = {}
    total_nac_1_reinc_local = {"nuevos": 0, "reincidentes": 0, "total": 0}
    nac_iso3_local = {}

    def _clasificar(lista, campo):
        # @FADAR -- "dato['total']" (no +=1): los datos vienen de un
        # .values(...).annotate(total=Count(...)), asi que una fila
        # agrupada puede representar >1 registro real (ej. un duplicado
        # tecnico con los mismos 9 campos) -- sumar solo +=1 los contaba
        # como uno solo. Por ahora se cuentan como registros reales
        # independientes (sin deduplicar), a peticion explicita.
        for dato in lista:
            cantidad = dato['total']
            of = dato['oficinaRepre']
            tabla_zona = zona_por_oficina_local.get(of)
            if tabla_zona is not None:
                tabla_zona[of][campo] += cantidad
                tabla_zona[of]['total'] += cantidad
                subtotal_por_oficina_local[of][campo] += cantidad
                subtotal_por_oficina_local[of]['total'] += cantidad
            nac = str(dato['nacionalidad']).upper()
            if nac not in nac_1_reinc_local:
                nac_1_reinc_local[nac] = {"nuevos": 0, "reincidentes": 0, "total": 0}
            nac_1_reinc_local[nac][campo] += cantidad
            nac_1_reinc_local[nac]['total'] += cantidad
            total_nac_1_reinc_local[campo] += cantidad
            total_nac_1_reinc_local['total'] += cantidad
            nac_iso3_local.setdefault(nac, dato['iso3'])

    _clasificar(nuevos_local, 'nuevos')
    _clasificar(reincidentes_local, 'reincidentes')

    nac_1_reinc_ordenado_local = dict(
        sorted(nac_1_reinc_local.items(), key=lambda x: x[1]['total'], reverse=True)
    )
    total_regiones_local = {
        "nuevos": subtotal_rio_bravo_local['nuevos'] + subtotal_centro_local['nuevos'] + subtotal_suchiate_local['nuevos'],
        "reincidentes": subtotal_rio_bravo_local['reincidentes'] + subtotal_centro_local['reincidentes'] + subtotal_suchiate_local['reincidentes'],
        "total": subtotal_rio_bravo_local['total'] + subtotal_centro_local['total'] + subtotal_suchiate_local['total'],
    }
    # @FADAR -- nacionalidades extracontinentales (fuera de America), misma
    # regla que en el Informe Diario.
    nacionalidades_extracontinentales_local = {
        n for n, iso3 in nac_iso3_local.items() if _rescates_es_extracontinental(str(iso3).upper())
    }
    return (
        zona_rio_bravo_local, zona_centro_local, zona_suchiate_local,
        subtotal_rio_bravo_local, subtotal_centro_local, subtotal_suchiate_local,
        total_regiones_local, nac_1_reinc_ordenado_local, total_nac_1_reinc_local,
        nacionalidades_extracontinentales_local,
    )


RESCATES_ALIAS_GEOJSON = {
    "CDMX": "Ciudad de México",
    "EDOMEX": "Estado de México",
}

RESCATES_ZONA_COLOR = {
    "Río Bravo": "#1D4ED8",
    "Centro": "#7C3AED",
    "Suchiate": "#B45309",
}


def _rescates_geojson_regiones(zona_rio_bravo, zona_centro, zona_suchiate):
    """Reutiliza el mismo geojson de estados que ya usa mapa_interactivo/
    mapa_informacion (mapa/static/mapa/data/inegi_latlon_mexico.geojson) --
    se sigue leyendo desde la app "mapa" a proposito: es un dataset
    geografico compartido, no logica propia de rescates, no se duplica.
    Le anota, por estado, la zona (Rio Bravo/Centro/Suchiate) y sus
    numeros -- para que el mapa se pinte y etiquete directo desde las
    propiedades del geojson, sin logica aparte del lado del cliente."""
    geojson_path = os.path.join(settings.BASE_DIR, 'mapa', 'static', 'mapa', 'data', 'inegi_latlon_mexico.geojson')
    with open(geojson_path, 'r', encoding='utf-8') as f:
        geo_data = json.load(f)

    datos_por_oficina = {}
    for of, d in zona_rio_bravo.items():
        datos_por_oficina[of] = {"zona": "Río Bravo", **d}
    for of, d in zona_centro.items():
        datos_por_oficina[of] = {"zona": "Centro", **d}
    for of, d in zona_suchiate.items():
        datos_por_oficina[of] = {"zona": "Suchiate", **d}

    datos_por_clave = {}
    for of, d in datos_por_oficina.items():
        alias = RESCATES_ALIAS_GEOJSON.get(of, of)
        datos_por_clave[normalizar_nombre(alias)] = {"oficina": of, **d}

    for feature in geo_data.get("features", []):
        props = feature.setdefault("properties", {})
        clave = normalizar_nombre(props.get("name", ""))
        info = datos_por_clave.get(clave)
        if info:
            props["oficina_rescates"] = info["oficina"]
            props["zona_rescates"] = info["zona"]
            props["color_zona_rescates"] = RESCATES_ZONA_COLOR[info["zona"]]
            props["nuevos_rescates"] = info["nuevos"]
            props["reincidentes_rescates"] = info["reincidentes"]
            props["total_rescates"] = info["total"]
        else:
            props["oficina_rescates"] = props.get("name", "")
            props["zona_rescates"] = None
            props["color_zona_rescates"] = "#D1D5DB"
            props["nuevos_rescates"] = 0
            props["reincidentes_rescates"] = 0
            props["total_rescates"] = 0

    return geo_data


def rescates_dashboard(request):
    """Tablero de Rescates/Operación: diagrama de dispersión de rescates por
    día, barras por hora del día, desglose hombres/mujeres/niños/niñas
    (mayoría de edad = 18 años), y pastel de nacionalidades atípicas.
    Filtrable por rango de fechas (default: día actual) y por entidad
    federativa (oficina de representación)."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    hoy = date.today().isoformat()
    fecha_inicio = request.GET.get('fecha_inicio', hoy)
    fecha_fin = request.GET.get('fecha_fin', hoy)
    oficina = request.GET.get('oficina', '').strip()

    filtro_oficina_sql = ''
    params_base = [fecha_inicio, fecha_fin]
    if oficina:
        filtro_oficina_sql = ' AND "oficinaRepre" = %s'
        params_base = params_base + [oficina]

    # Optimización (sin tocar el esquema/índices, a petición explícita):
    # de ~7 consultas secuenciales, cada una escaneando toda la tabla por
    # separado, se pasa a 3 -- usando GROUPING SETS para calcular varias
    # agregaciones en una sola pasada -- corriendo en paralelo (cada una en
    # su propia conexión), en vez de una tras otra. El tiempo total pasa de
    # ser la suma de las 7 a ser el de la más lenta de las 3.

    def _consulta_sexo_edad():
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT "
                f"  COUNT(*) FILTER (WHERE sexo=true  AND edad>=18) AS hombres, "
                f"  COUNT(*) FILTER (WHERE sexo=false AND edad>=18) AS mujeres, "
                f"  COUNT(*) FILTER (WHERE sexo=true  AND edad<18)  AS ninos, "
                f"  COUNT(*) FILTER (WHERE sexo=false AND edad<18)  AS ninas "
                f"FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql}",
                params_base,
            )
            resultado = cur.fetchone()
        connection.close()
        return resultado

    def _consulta_dia_hora():
        # "hora" es texto libre y no siempre viene como HH:MM (hay valores
        # sueltos como "06") -- se descarta al formato inválido dentro del
        # CASE, para no tronar el cast a time por un registro mal capturado.
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT GROUPING(dia) AS g_dia, dia, hr, COUNT(*) FROM ( "
                f"  SELECT TO_DATE(fecha,'DD-MM-YY') AS dia, "
                f"         CASE WHEN hora ~ '^[0-2][0-9]:[0-5][0-9]$' THEN EXTRACT(HOUR FROM hora::time)::int END AS hr "
                f"  FROM usuario_rescatepunto "
                f"  WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f") sub "
                f"GROUP BY GROUPING SETS ((dia), (hr))",
                params_base,
            )
            filas = cur.fetchall()
        connection.close()
        filas_dia_local = sorted(
            [(dia, total) for g_dia, dia, hr, total in filas if g_dia == 0],
            key=lambda x: x[0],
        )
        conteo_por_hora_local = {
            int(hr): total for g_dia, dia, hr, total in filas if g_dia == 1 and hr is not None
        }
        return filas_dia_local, conteo_por_hora_local

    def _consulta_nacionalidad():
        # Nacionalidad por iso3 (nombre en mayúsculas, reduce duplicados de
        # capitalización). (Parentesco y Punto Estratégico "crudos" se
        # quitaron del dashboard a petición explícita -- era vista de
        # análisis interno, no algo para usuario final. numFamilia ahora
        # se resume como conteo de familias, no como distribución cruda.)
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT iso3, UPPER(MAX(nacionalidad)), COUNT(*) FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f"GROUP BY iso3",
                params_base,
            )
            iso_rows_local = cur.fetchall()
        connection.close()
        return iso_rows_local

    def _consulta_reincidentes():
        # Reincidente = coincide (nombre, apellidos, nacionalidad) con otro
        # registro en TODO el historico de usuario_rescatepunto (>=2
        # apariciones), vista mapa_mv_reincidencia_rescates -- reemplaza el
        # conteo que antes venia de mapa_extrescatados.reincidente.
        #
        # Dos apartados pedidos explícitamente:
        #  1. Por día, respetando el filtro de fecha/entidad actual.
        #  2. General histórico: TODOS los registros, sin importar el
        #     filtro de fecha/entidad seleccionado.
        filtro_estado_sql = ""
        params_reinc = [fecha_inicio, fecha_fin]
        if oficina:
            filtro_estado_sql = ' AND r."oficinaRepre" = %s'
            params_reinc = params_reinc + [oficina]
        with connection.cursor() as cur:
            # Reincidentes y primera vez del rango filtrado en UNA sola
            # pasada sobre el JOIN (antes eran 2 consultas identicas salvo
            # por el filtro de clasificacion -- redundante).
            cur.execute(
                f"SELECT "
                f"  COUNT(*) FILTER (WHERE v.clasificacion = 'Reincidente'), "
                f"  COUNT(*) FILTER (WHERE v.clasificacion = 'Rescate primera vez') "
                f"FROM usuario_rescatepunto r "
                f"JOIN {RESCATES_MV_REINCIDENCIA} v "
                f"  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad "
                f"WHERE TO_DATE(r.fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_estado_sql}",
                params_reinc,
            )
            total_reincidentes_local, total_primera_vez_local = cur.fetchone()

            cur.execute(
                f"SELECT TO_DATE(r.fecha,'DD-MM-YY') AS dia, COUNT(*) FROM usuario_rescatepunto r "
                f"JOIN {RESCATES_MV_REINCIDENCIA} v "
                f"  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad "
                f"WHERE v.clasificacion = 'Reincidente' "
                f"AND TO_DATE(r.fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_estado_sql} "
                f"GROUP BY dia ORDER BY dia",
                params_reinc,
            )
            reincidentes_por_dia_local = cur.fetchall()

            # Misma logica (>=2 vs =1) para el historico, tambien en una
            # sola pasada -- la vista ya trae "veces" pre-agregado por
            # persona, asi que ni siquiera toca usuario_rescatepunto.
            cur.execute(
                f"SELECT "
                f"  COALESCE(SUM(veces) FILTER (WHERE clasificacion = 'Reincidente'), 0), "
                f"  COALESCE(SUM(veces) FILTER (WHERE clasificacion = 'Rescate primera vez'), 0) "
                f"FROM {RESCATES_MV_REINCIDENCIA}"
            )
            total_reincidentes_historico_local, total_primera_vez_historico_local = cur.fetchone()
        connection.close()
        return (
            total_reincidentes_local, reincidentes_por_dia_local,
            total_reincidentes_historico_local, total_primera_vez_local,
            total_primera_vez_historico_local,
        )

    def _consulta_por_entidad():
        # Solo tiene sentido comparar entidades cuando NO hay una entidad ya
        # seleccionada -- si se filtró por "TABASCO", no hay nada que
        # comparar. Se calcula aparte (no usa filtro_oficina_sql/params_base
        # porque esos ya vienen con el filtro de oficina fijo).
        with connection.cursor() as cur:
            cur.execute(
                'SELECT "oficinaRepre", COUNT(*) FROM usuario_rescatepunto '
                "WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s "
                'GROUP BY "oficinaRepre" ORDER BY 2 DESC',
                [fecha_inicio, fecha_fin],
            )
            filas = cur.fetchall()
        connection.close()
        return filas

    def _consulta_familias():
        # Agrupa por "numFamilia" (el contador real de familia que ya
        # verificamos) + hora + punto estratégico + agente -- numFamilia

        with connection.cursor() as cur:
            cur.execute(
                f'SELECT * FROM ( '
                f'  SELECT "oficinaRepre" AS oficina, TO_DATE(fecha,\'DD-MM-YY\') AS fecha, hora, '
                f'         "puntoEstra" AS punto, "numFamilia" AS nf, '
                f'         nombre, apellidos, parentesco, "fechaNacimiento" AS fecha_nac, edad, sexo, '
                f"         UPPER(nacionalidad) AS nacionalidad, "
                f'         COUNT(*) OVER (PARTITION BY "oficinaRepre", fecha, hora, "puntoEstra", "numFamilia") AS integrantes '
                f"  FROM usuario_rescatepunto "
                f"  WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f'    AND "numFamilia" > 0 '
                f") sub "
                f"WHERE integrantes >= 2 "
                f"ORDER BY integrantes DESC, oficina, fecha, hora, punto, nf, edad DESC",
                params_base,
            )
            filas = cur.fetchall()
        connection.close()

        # Agrupa las filas planas en núcleos familiares (mismo
        # oficina+fecha+hora+punto+numFamilia). El conteo total de familias
        # se calcula sobre TODAS (no solo las que se muestran); la lista
        # que se despliega se limita a los primeros 15 núcleos -- no a un
        # número fijo de filas, para no cortar una familia a la mitad.
        familias = []
        familia_actual = None
        clave_actual = None
        total_familias = 0
        for oficina, fecha, hora, punto, nf, nombre, apellidos, parentesco, fecha_nac, edad, sexo, nacionalidad, integrantes in filas:
            clave = (oficina, fecha, hora, punto, nf)
            if clave != clave_actual:
                clave_actual = clave
                total_familias += 1
                familia_actual = {
                    "oficina": oficina, "fecha": fecha, "hora": hora, "punto": punto,
                    "num_familia": nf, "integrantes": integrantes, "miembros": [],
                }
                if total_familias <= 15:
                    familias.append(familia_actual)
            if total_familias <= 15:
                familia_actual["miembros"].append({
                    "nombre": nombre, "apellidos": apellidos, "parentesco": parentesco,
                    "fecha_nac": fecha_nac, "edad": edad, "sexo": sexo, "nacionalidad": nacionalidad,
                })
        return familias, total_familias

    def _consulta_tipo_rescate():
        # Medio de rescate (carretero/aéreo/ferroviario/etc.) -- campos
        # booleanos ya existentes en la tabla, en una sola pasada.
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT "
                f"  COUNT(*) FILTER (WHERE carretero) AS carretero, "
                f"  COUNT(*) FILTER (WHERE aeropuerto) AS aereo, "
                f"  COUNT(*) FILTER (WHERE ferrocarril) AS ferroviario, "
                f'  COUNT(*) FILTER (WHERE "centralAutobus") AS central_autobus, '
                f'  COUNT(*) FILTER (WHERE "casaSeguridad") AS casa_seguridad, '
                f"  COUNT(*) FILTER (WHERE hotel) AS hotel, "
                f"  COUNT(*) FILTER (WHERE reclusorio) AS reclusorio "
                f"FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql}",
                params_base,
            )
            resultado = cur.fetchone()
        connection.close()
        return resultado

    def _consulta_retornados():
        # mapa_retornados: tabla distinta a RescatePunto, con fecha real
        # (no texto). Mismo criterio que reincidentes: rango filtrado +
        # general histórico (toda la tabla, sin importar el filtro).
        filtro_estado_sql = ""
        params_ret = [fecha_inicio, fecha_fin]
        if oficina:
            filtro_estado_sql = " AND estado_id = (SELECT id FROM mapa_estado WHERE nombre = %s)"
            params_ret = params_ret + [oficina]
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT COALESCE(SUM(retornados_total),0), COALESCE(SUM(deportado),0), COALESCE(SUM(retornado),0) "
                f"FROM mapa_retornados WHERE fecha BETWEEN %s AND %s{filtro_estado_sql}",
                params_ret,
            )
            total_local, deportado_local, retornado_local = cur.fetchone()

            cur.execute("SELECT COALESCE(SUM(retornados_total),0) FROM mapa_retornados")
            total_historico_local = cur.fetchone()[0]
        connection.close()
        return total_local, deportado_local, retornado_local, total_historico_local

    with ThreadPoolExecutor(max_workers=8) as executor:
        futuro_sexo = executor.submit(_consulta_sexo_edad)
        futuro_dia_hora = executor.submit(_consulta_dia_hora)
        futuro_nac = executor.submit(_consulta_nacionalidad)
        futuro_reincidentes = executor.submit(_consulta_reincidentes)
        futuro_por_entidad = executor.submit(_consulta_por_entidad) if not oficina else None
        futuro_familias = executor.submit(_consulta_familias)
        futuro_tipo_rescate = executor.submit(_consulta_tipo_rescate)
        futuro_retornados = executor.submit(_consulta_retornados)
        # Regiones (Rio Bravo/Centro/Suchiate) -- se calcula tambien aqui
        # para el mini-mapa embebido en el dashboard (la parte mas cara,
        # la deteccion de reincidentes historica, ya vive en la vista
        # materializada mapa_mv_reincidencia_rescates, asi que el costo
        # extra de traerla tambien en el dashboard es bajo). La pagina dedicada
        # /rescates/regiones sigue existiendo aparte para el detalle
        # completo (tablas).
        futuro_regiones = executor.submit(_rescates_regiones, fecha_inicio, fecha_fin, oficina)

        hombres, mujeres, ninos, ninas = futuro_sexo.result()
        filas_dia, conteo_por_hora = futuro_dia_hora.result()
        iso_rows = futuro_nac.result()
        filas_por_entidad = futuro_por_entidad.result() if futuro_por_entidad else []
        (
            total_reincidentes, reincidentes_por_dia,
            total_reincidentes_historico, total_primera_vez,
            total_primera_vez_historico,
        ) = futuro_reincidentes.result()
        familias_detectadas, total_familias = futuro_familias.result()
        carretero, aereo, ferroviario, central_autobus, casa_seguridad, hotel, reclusorio = futuro_tipo_rescate.result()
        total_retornados, retornados_deportado, retornados_retornado, total_retornados_historico = futuro_retornados.result()
        (
            zona_rio_bravo, zona_centro, zona_suchiate,
            subtotal_rio_bravo, subtotal_centro, subtotal_suchiate,
            total_regiones, _nac_1_reinc_sin_usar, _total_nac_1_reinc_sin_usar,
            _nac_extracontinentales_sin_usar,
        ) = futuro_regiones.result()

    # El total ya no necesita su propia consulta -- se deriva de la suma de
    # las 4 categorías de sexo/edad (sexo y edad son campos obligatorios en
    # el modelo, así que la suma siempre coincide con el COUNT(*) real).
    total_rango = hombres + mujeres + ninos + ninas

    # --- Clasificación de atípicas (regla geográfica, no de frecuencia) ---
    atipicas_por_region = {"Medio Oriente": 0, "Europa": 0, "Otras / poco conocidas": 0}
    detalle_atipicas = []
    for iso3, nombre, total in iso_rows:
        region = _rescates_region_atipica(iso3)
        if region:
            atipicas_por_region[region] += total
            detalle_atipicas.append({"iso3": iso3, "nombre": nombre, "total": total, "region": region})
    detalle_atipicas.sort(key=lambda d: d["total"], reverse=True)
    total_atipicas = sum(atipicas_por_region.values())

    # --- Desglose (solo números) de hombres/mujeres/niños/niñas y núcleos
    # familiares, pero limitado a las nacionalidades atípicas -- para la
    # sección de detalle, no para el dashboard general. ---
    iso3_atipicas = [d["iso3"] for d in detalle_atipicas]
    if iso3_atipicas:
        placeholders_iso = ",".join(["%s"] * len(iso3_atipicas))
        with connection.cursor() as cur:
            cur.execute(
                f"SELECT "
                f"  COUNT(*) FILTER (WHERE sexo=true  AND edad>=18) AS hombres, "
                f"  COUNT(*) FILTER (WHERE sexo=false AND edad>=18) AS mujeres, "
                f"  COUNT(*) FILTER (WHERE sexo=true  AND edad<18)  AS ninos, "
                f"  COUNT(*) FILTER (WHERE sexo=false AND edad<18)  AS ninas "
                f"FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f"  AND iso3 IN ({placeholders_iso})",
                params_base + iso3_atipicas,
            )
            atip_hombres, atip_mujeres, atip_ninos, atip_ninas = cur.fetchone()

            cur.execute(
                f"SELECT COUNT(DISTINCT (\"oficinaRepre\", fecha, hora, \"puntoEstra\", \"numFamilia\")) "
                f"FROM usuario_rescatepunto "
                f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_oficina_sql} "
                f"  AND iso3 IN ({placeholders_iso}) AND \"numFamilia\" > 0",
                params_base + iso3_atipicas,
            )
            atip_nucleos = cur.fetchone()[0]
    else:
        atip_hombres = atip_mujeres = atip_ninos = atip_ninas = atip_nucleos = 0

    # --- Gráfica 1: diagrama de dispersión (rescates por día) ---
    x_disp = [datetime.combine(d, datetime.min.time()) for d, _ in filas_dia]
    y_disp = [c for _, c in filas_dia]
    etiquetas_disp = [f"{c:,}" for _, c in filas_dia]
    source_disp = ColumnDataSource(data=dict(x=x_disp, y=y_disp, label=etiquetas_disp))
    p_disp = figure(
        height=320, sizing_mode="stretch_width", x_axis_type="datetime",
        toolbar_location="right", tools="pan,box_zoom,reset",
        background_fill_color="#f7f7f7", border_fill_color=None,
        outline_line_color="#666666",
        title="Rescates por día (cada punto = un día)",
    )
    p_disp.scatter(x='x', y='y', size=9, color="#285C4D", alpha=0.75, source=source_disp)
    p_disp.y_range.start = 0
    p_disp.add_layout(LabelSet(
        x='x', y='y', text='label', source=source_disp,
        x_offset=0, y_offset=8, text_font_size="9px", text_color="#285C4D",
        text_align="center",
    ))
    p_disp.xaxis.formatter = DatetimeTickFormatter(days="%d %b", months="%b %Y", years="%Y")
    p_disp.xgrid.grid_line_color = "#ffffff"
    p_disp.ygrid.grid_line_color = "#ffffff"
    p_disp.add_tools(HoverTool(tooltips=[("Día", "@x{%d/%b/%y}"), ("Rescates", "@y{0,0}")], formatters={'@x': 'datetime'}))

    # --- Gráfica 2: barras por hora del día ---
    horas_lbl = [f"{h:02d}:00" for h in range(24)]
    valores_hora = [conteo_por_hora.get(h, 0) for h in range(24)]
    etiquetas_hora = [f"{v:,}" for v in valores_hora]
    source_horas = ColumnDataSource(data=dict(hora=horas_lbl, total=valores_hora, label=etiquetas_hora))
    p_horas = figure(
        x_range=horas_lbl, height=300, sizing_mode="stretch_width",
        toolbar_location=None, tools="",
        background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
        title="Rescates por hora del día (00:00 a 23:59:59)",
    )
    p_horas.vbar(x='hora', top='total', width=0.8, color="#9A0A38", source=source_horas)
    p_horas.y_range.start = 0
    p_horas.add_layout(LabelSet(
        x='hora', y='total', text='label', source=source_horas,
        x_offset=0, y_offset=4, text_font_size="8px", text_color="#9A0A38",
        text_align="center",
    ))
    p_horas.xaxis.major_label_orientation = 0.9
    p_horas.xgrid.grid_line_color = None
    p_horas.ygrid.grid_line_color = "#ffffff"
    p_horas.add_tools(HoverTool(tooltips=[("Hora", "@hora"), ("Rescates", "@total{0,0}")]))

    # --- Gráfica 3: barras hombres / mujeres / niños / niñas ---
    categorias_sexo = ["Hombres", "Mujeres", "Niños", "Niñas"]
    valores_sexo = [hombres, mujeres, ninos, ninas]
    etiquetas_sexo = [f"{v:,}" for v in valores_sexo]
    colores_sexo = ["#0EA5E9", "#EC4899", "#38BDF8", "#F9A8D4"]
    source_sexo = ColumnDataSource(data=dict(
        categoria=categorias_sexo, total=valores_sexo, color=colores_sexo, label=etiquetas_sexo,
    ))
    p_sexo = figure(
        x_range=categorias_sexo, height=300, sizing_mode="stretch_width",
        toolbar_location=None, tools="",
        background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
        title="Hombres, mujeres, niños y niñas (mayoría de edad = 18 años)",
    )
    p_sexo.vbar(x='categoria', top='total', width=0.6, color='color', source=source_sexo)
    p_sexo.y_range.start = 0
    p_sexo.add_layout(LabelSet(
        x='categoria', y='total', text='label', source=source_sexo,
        x_offset=0, y_offset=4, text_font_size="11px", text_color="#374151",
        text_align="center",
    ))
    p_sexo.xgrid.grid_line_color = None
    p_sexo.ygrid.grid_line_color = "#ffffff"
    p_sexo.add_tools(HoverTool(tooltips=[("Categoría", "@categoria"), ("Total", "@total{0,0}")]))

    # --- Gráfica 4: pastel de nacionalidades atípicas ---
    regiones_pie = ["Medio Oriente", "Europa", "Otras / poco conocidas"]
    valores_pie = [atipicas_por_region[r] for r in regiones_pie]
    colores_pie = ["#F59E0B", "#6366F1", "#10B981"]
    angulos = [(v / total_atipicas * 2 * pi) if total_atipicas else 0 for v in valores_pie]
    # Posición de cada etiqueta: a la mitad del ángulo de su rebanada, un
    # poco afuera del radio del pastel, para verse sin pasar el cursor.
    acumulado = 0
    x_etq_pie, y_etq_pie, etiquetas_pie = [], [], []
    for v, ang in zip(valores_pie, angulos):
        medio = acumulado + ang / 2
        x_etq_pie.append(0 + 0.55 * cos(medio))
        y_etq_pie.append(1 + 0.55 * sin(medio))
        etiquetas_pie.append(f"{v:,}")
        acumulado += ang
    source_pie = ColumnDataSource(data=dict(
        region=regiones_pie, total=valores_pie, color=colores_pie, angle=angulos,
    ))
    source_etq_pie = ColumnDataSource(data=dict(x=x_etq_pie, y=y_etq_pie, label=etiquetas_pie))
    p_pie = figure(
        height=320, sizing_mode="stretch_width", toolbar_location=None, tools="",
        background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
        title="Nacionalidades atípicas (Medio Oriente, Europa, otras poco conocidas)",
        x_range=(-0.9, 1.3),
    )
    p_pie.wedge(
        x=0, y=1, radius=0.4,
        start_angle=cumsum('angle', include_zero=True), end_angle=cumsum('angle'),
        line_color="#ffffff", fill_color='color', legend_field='region', source=source_pie,
    )
    p_pie.add_layout(LabelSet(
        x='x', y='y', text='label', source=source_etq_pie,
        text_font_size="11px", text_font_style="bold", text_color="#374151",
        text_align="center", text_baseline="middle",
    ))
    p_pie.axis.visible = False
    p_pie.grid.visible = False
    p_pie.add_tools(HoverTool(tooltips=[("Región", "@region"), ("Total", "@total{0,0}")]))

    # --- Gráfica 5: barras horizontales por entidad, solo cuando "Todas
    # las entidades" está seleccionado (si ya se filtró una, no hay nada
    # que comparar) -- ordenada de mayor a menor. ---
    figuras = {"dispersion": p_disp, "horas": p_horas, "sexo": p_sexo, "pie": p_pie}

    if not oficina and filas_por_entidad:
        entidades_ordenadas = [f[0] for f in filas_por_entidad]  # ya viene ORDER BY 2 DESC
        totales_entidad = [f[1] for f in filas_por_entidad]
        etiquetas_entidad = [f"{v:,}" for v in totales_entidad]
        source_entidad = ColumnDataSource(data=dict(
            entidad=entidades_ordenadas, total=totales_entidad, label=etiquetas_entidad,
        ))
        p_entidad = figure(
            y_range=list(reversed(entidades_ordenadas)),  # mayor arriba
            height=max(320, 24 * len(entidades_ordenadas)), sizing_mode="stretch_width",
            toolbar_location=None, tools="",
            background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
            title="Rescates por entidad federativa (mayor a menor)",
        )
        p_entidad.hbar(y='entidad', right='total', height=0.7, color="#285C4D", source=source_entidad)
        p_entidad.x_range.start = 0
        p_entidad.add_layout(LabelSet(
            x='total', y='entidad', text='label', source=source_entidad,
            x_offset=4, y_offset=-6, text_font_size="9px", text_color="#285C4D",
        ))
        p_entidad.ygrid.grid_line_color = None
        p_entidad.xgrid.grid_line_color = "#ffffff"
        p_entidad.add_tools(HoverTool(tooltips=[("Entidad", "@entidad"), ("Rescates", "@total{0,0}")]))
        figuras["entidad"] = p_entidad

    # --- Gráfica 6: reincidentes por día (dentro del rango filtrado) ---
    x_reinc = [datetime.combine(d, datetime.min.time()) for d, _ in reincidentes_por_dia]
    y_reinc = [int(v or 0) for _, v in reincidentes_por_dia]
    etiquetas_reinc = [f"{v:,}" for v in y_reinc]
    source_reinc = ColumnDataSource(data=dict(x=x_reinc, y=y_reinc, label=etiquetas_reinc))
    p_reinc = figure(
        height=280, sizing_mode="stretch_width", x_axis_type="datetime",
        toolbar_location="right", tools="pan,box_zoom,reset",
        background_fill_color="#f7f7f7", border_fill_color=None, outline_line_color="#666666",
        title="Reincidentes por día (dentro del rango filtrado)",
    )
    p_reinc.vbar(x='x', top='y', width=1000 * 60 * 60 * 20, color="#DC2626", source=source_reinc)
    p_reinc.y_range.start = 0
    p_reinc.add_layout(LabelSet(
        x='x', y='y', text='label', source=source_reinc,
        x_offset=0, y_offset=4, text_font_size="9px", text_color="#DC2626", text_align="center",
    ))
    p_reinc.xaxis.formatter = DatetimeTickFormatter(days="%d %b", months="%b %Y", years="%Y")
    p_reinc.xgrid.grid_line_color = "#ffffff"
    p_reinc.ygrid.grid_line_color = "#ffffff"
    p_reinc.add_tools(HoverTool(tooltips=[("Día", "@x{%d/%b/%y}"), ("Reincidentes", "@y{0,0}")], formatters={'@x': 'datetime'}))
    figuras["reincidentes_dia"] = p_reinc

    nombres = list(figuras.keys())
    plot_script, divs_lista = components(tuple(figuras[n] for n in nombres))
    divs = dict(zip(nombres, divs_lista))
    div_disp = divs["dispersion"]
    div_horas = divs["horas"]
    div_sexo = divs["sexo"]
    div_pie = divs["pie"]
    div_entidad = divs.get("entidad", "")
    div_reincidentes_dia = divs["reincidentes_dia"]

    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "oficina_seleccionada": oficina,
        "oficinas": RESCATES_OFICINAS,
        "total_rango": total_rango,
        "hombres": hombres,
        "mujeres": mujeres,
        "ninos": ninos,
        "ninas": ninas,
        "total_atipicas": total_atipicas,
        "atipicas_por_region": atipicas_por_region,
        "detalle_atipicas": detalle_atipicas[:20],
        "atip_hombres": atip_hombres,
        "atip_mujeres": atip_mujeres,
        "atip_ninos": atip_ninos,
        "atip_ninas": atip_ninas,
        "atip_nucleos": atip_nucleos,
        # Reincidentes (mapa_mv_reincidencia_rescates, >=2 apariciones de
        # nombre+apellidos+nacionalidad en usuario_rescatepunto). Dos
        # apartados: por día (respeta el filtro actual) y general histórico
        # (TODA la tabla, sin importar el filtro de fecha/entidad seleccionado).
        "total_reincidentes": total_reincidentes,
        "total_reincidentes_historico": total_reincidentes_historico,
        "total_primera_vez": total_primera_vez,
        "total_primera_vez_historico": total_primera_vez_historico,
        "familias_detectadas": familias_detectadas,
        "total_familias": total_familias,
        # Medio de rescate (campos booleanos de la tabla).
        "carretero": carretero,
        "aereo": aereo,
        "ferroviario": ferroviario,
        "central_autobus": central_autobus,
        "casa_seguridad": casa_seguridad,
        "hotel": hotel,
        "reclusorio": reclusorio,
        # Retornados (mapa_retornados) -- tabla distinta a RescatePunto.
        "total_retornados": total_retornados,
        "retornados_deportado": retornados_deportado,
        "retornados_retornado": retornados_retornado,
        "total_retornados_historico": total_retornados_historico,
        # Regiones -- mini-mapa embebido; el detalle completo (tablas) vive
        # en su propia pagina (rescates_regiones).
        "subtotal_rio_bravo": subtotal_rio_bravo,
        "subtotal_centro": subtotal_centro,
        "subtotal_suchiate": subtotal_suchiate,
        "total_regiones": total_regiones,
        "geo_data_regiones_json": json.dumps(_rescates_geojson_regiones(zona_rio_bravo, zona_centro, zona_suchiate)),
        "plot_script": plot_script,
        "div_dispersion": div_disp,
        "div_horas": div_horas,
        "div_sexo": div_sexo,
        "div_pie": div_pie,
        "div_entidad": div_entidad,
        "mostrar_seccion_entidad": not oficina,
        "div_reincidentes_dia": div_reincidentes_dia,
    }
    return render(request, "Reportes_Analisis/rescates.html", context)


def rescates_regiones(request):
    """Pagina dedicada a "Regiones" (Rio Bravo/Centro/Suchiate), separada
    del dashboard principal -- mismo filtro de fecha/entidad, con mapa
    interactivo (reutiliza el geojson/MapLibre de mapa_interactivo) que
    pinta cada estado segun su zona y le muestra el total encima."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    hoy = date.today().isoformat()
    fecha_inicio = request.GET.get('fecha_inicio', hoy)
    fecha_fin = request.GET.get('fecha_fin', hoy)
    oficina = request.GET.get('oficina', '').strip()

    (
        zona_rio_bravo, zona_centro, zona_suchiate,
        subtotal_rio_bravo, subtotal_centro, subtotal_suchiate,
        total_regiones, nac_1_reinc, total_nac_1_reinc,
        nacionalidades_extracontinentales,
    ) = _rescates_regiones(fecha_inicio, fecha_fin, oficina or None)

    geo_data = _rescates_geojson_regiones(zona_rio_bravo, zona_centro, zona_suchiate)

    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "oficina_seleccionada": oficina,
        "oficinas": RESCATES_OFICINAS,
        "zona_rio_bravo": zona_rio_bravo,
        "zona_centro": zona_centro,
        "zona_suchiate": zona_suchiate,
        "subtotal_rio_bravo": subtotal_rio_bravo,
        "subtotal_centro": subtotal_centro,
        "subtotal_suchiate": subtotal_suchiate,
        "total_regiones": total_regiones,
        "nac_1_reinc": nac_1_reinc,
        "total_nac_1_reinc": total_nac_1_reinc,
        "nacionalidades_extracontinentales": nacionalidades_extracontinentales,
        "geo_data_json": json.dumps(geo_data),
    }
    return render(request, "Reportes_Analisis/rescates_regiones.html", context)


RESCATES_MESES_ES = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
    7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic",
}


def _rescates_regiones_reporte(fecha_str, oficina=None):
    """Version 'reporte' (una sola fecha) de _rescates_regiones -- mismo
    patron que _rescates_cuadro_datos / _rescates_informe_diario: regresa
    un dict ya listo para el template/PDF/Excel."""
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    (
        zona_rio_bravo, zona_centro, zona_suchiate,
        subtotal_rio_bravo, subtotal_centro, subtotal_suchiate,
        total_regiones, nac_1_reinc, total_nac_1_reinc,
        nacionalidades_extracontinentales,
    ) = _rescates_regiones(fecha_str, fecha_str, oficina)
    return {
        "fecha_actual": f"{fecha_obj.day:02d} {RESCATES_MESES_ES[fecha_obj.month]} {fecha_obj.year}",
        "fecha_iso": fecha_str,
        "oficina": oficina or "Nacional",
        "zona_rio_bravo": zona_rio_bravo,
        "zona_centro": zona_centro,
        "zona_suchiate": zona_suchiate,
        "subtotal_rio_bravo": subtotal_rio_bravo,
        "subtotal_centro": subtotal_centro,
        "subtotal_suchiate": subtotal_suchiate,
        "total_regiones": total_regiones,
        "nac_1_reinc": nac_1_reinc,
        "total_nac_1_reinc": total_nac_1_reinc,
        "nacionalidades_extracontinentales": nacionalidades_extracontinentales,
    }


RESCATES_LETRA_T0 = ("B38E5D", "E9E8E8")  # fondo, texto
RESCATES_LETRA_T1 = ("761B36", "FFFFFF")
RESCATES_LETRA_T2 = ("D9D9D9", "000000")
RESCATES_COLOR_FONDO = ("4E1224", "F2F2F2")

RESCATES_BORDE_DELGADO = Border(*(Side(style="thin", color="000000"),) * 4)


def _rescates_excel_celda(ws, fila, col, valor, bg=None, color_texto="000000", negrita=False, tam=11, centrado=True, fuente=None):
    celda = ws.cell(row=fila, column=col, value=valor)
    kwargs_fuente = {"name": fuente} if fuente else {}
    celda.font = Font(bold=negrita, size=tam, color=color_texto, **kwargs_fuente)
    if bg:
        celda.fill = PatternFill("solid", fgColor=bg)
    celda.border = RESCATES_BORDE_DELGADO
    if centrado:
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # Formato de excel real (no texto) para cualquier valor numerico -- asi
    # la columna se puede sumar/importar en Excel en vez de quedar como
    # texto (p.ej. una celda con el valor -5 en vez del texto "-5").
    if isinstance(valor, (int, float)) and not isinstance(valor, bool):
        celda.number_format = "#,##0"
    return celda


def _rescates_excel_fila(ws, fila, valores, estilo, negrita=False, col_inicial=1):
    bg, color_txt = estilo
    for i, valor in enumerate(valores):
        _rescates_excel_celda(ws, fila, col_inicial + i, valor, bg=bg, color_texto=color_txt, negrita=negrita)


def rescates_reporte_regiones(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_regiones_reporte(fecha_str, oficina or None)
    datos["oficinas"] = RESCATES_OFICINAS
    datos["oficina_seleccionada"] = oficina
    return render(request, "Reportes_Analisis/rescates_reporte_regiones.html", datos)


def rescates_reporte_regiones_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_regiones_reporte(fecha_str, oficina or None)

    template = get_template("Reportes_Analisis/_rescates_regiones_pdf.html")
    html_string = template.render(datos)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="regiones_ceco_{fecha_str}.pdf"'
    return response


def rescates_reporte_regiones_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_regiones_reporte(fecha_str, oficina or None)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Por zona"
    ws1.merge_cells("A1:E1")
    _rescates_excel_celda(ws1, 1, 1, "INSTITUTO NACIONAL DE MIGRACIÓN — DIRECCIÓN GENERAL DE COORDINACIÓN DE OFICINAS DE REPRESENTACION", negrita=True, tam=10)
    ws1.merge_cells("A2:E2")
    _rescates_excel_celda(ws1, 2, 1, f"Regiones (CECO) — {datos['fecha_actual']} — {datos['oficina']}", bg=RESCATES_COLOR_FONDO[0], color_texto=RESCATES_COLOR_FONDO[1], negrita=True, tam=12)

    fila = 4
    _rescates_excel_fila(ws1, fila, ["CECO", "Oficina de representación", "Rescates primera vez", "Reincidentes", "Total CECO"], RESCATES_LETRA_T1, negrita=True)
    fila += 1

    def _bloque_zona(nombre_zona, zona, subtotal):
        nonlocal fila
        for of, d in zona.items():
            _rescates_excel_fila(ws1, fila, [nombre_zona, of, d["nuevos"], d["reincidentes"], d["total"]], RESCATES_LETRA_T2)
            fila += 1
        _rescates_excel_fila(ws1, fila, ["", f"SUBTOTAL {nombre_zona.upper()}", subtotal["nuevos"], subtotal["reincidentes"], subtotal["total"]], RESCATES_LETRA_T0, negrita=True)
        fila += 1

    _bloque_zona("Río Bravo", datos["zona_rio_bravo"], datos["subtotal_rio_bravo"])
    _bloque_zona("Suchiate", datos["zona_suchiate"], datos["subtotal_suchiate"])
    _bloque_zona("Centro", datos["zona_centro"], datos["subtotal_centro"])

    _rescates_excel_fila(ws1, fila, ["", "TOTAL", datos["total_regiones"]["nuevos"], datos["total_regiones"]["reincidentes"], datos["total_regiones"]["total"]], RESCATES_LETRA_T1, negrita=True)
    fila += 1

    # @FADAR -- cuadro resumen final, mismos colores de zona del mapa,
    # total en verde.
    fila += 2
    for etiqueta, valor, color in [
        ("SUCHIATE", datos["subtotal_suchiate"]["total"], "B45309"),
        ("RÍO BRAVO", datos["subtotal_rio_bravo"]["total"], "1D4ED8"),
        ("CENTRO", datos["subtotal_centro"]["total"], "7C3AED"),
        ("TOTAL", datos["total_regiones"]["total"], "16A34A"),
    ]:
        _rescates_excel_celda(ws1, fila, 1, etiqueta, bg=color, color_texto="FFFFFF", negrita=True, centrado=False)
        _rescates_excel_celda(ws1, fila, 2, valor, bg=color, color_texto="FFFFFF", negrita=True)
        fila += 1

    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 24
    for col in ("C", "D", "E"):
        ws1.column_dimensions[col].width = 18

    ws2 = wb.create_sheet("Por nacionalidad")
    _rescates_excel_fila(ws2, 1, ["Nacionalidad", "Rescates primera vez", "Reincidentes", "Total rescates"], RESCATES_LETRA_T1, negrita=True)
    f = 2
    for nac, d in datos["nac_1_reinc"].items():
        if nac in datos["nacionalidades_extracontinentales"]:
            bg_nac, color_nac = "B45309", "FFFFFF"
        else:
            bg_nac, color_nac = RESCATES_LETRA_T1
        _rescates_excel_celda(ws2, f, 1, nac, bg=bg_nac, color_texto=color_nac, centrado=False)
        _rescates_excel_celda(ws2, f, 2, d["nuevos"], bg=RESCATES_LETRA_T2[0], color_texto=RESCATES_LETRA_T2[1])
        _rescates_excel_celda(ws2, f, 3, d["reincidentes"], bg=RESCATES_LETRA_T2[0], color_texto=RESCATES_LETRA_T2[1])
        _rescates_excel_celda(ws2, f, 4, d["total"], bg=RESCATES_LETRA_T0[0], color_texto=RESCATES_LETRA_T0[1], negrita=True)
        f += 1
    _rescates_excel_fila(ws2, f, ["TOTAL", datos["total_nac_1_reinc"]["nuevos"], datos["total_nac_1_reinc"]["reincidentes"], datos["total_nac_1_reinc"]["total"]], RESCATES_LETRA_T1, negrita=True)
    ws2.column_dimensions["A"].width = 28

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="regiones_ceco_{fecha_str}.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# Reportes de Rescates -- "Cuadro de Datos" (mismo formato/reglas de negocio
# que estadistica.generar_cuadro_diario, pero 100% automatico: nada se
# captura a mano. La unica diferencia real es "Retornados", que en el
# reporte original es un campo de texto libre que alguien llena -- aqui se
# calcula de mapa_retornados, igual que en el resto del dashboard.
#
# Reglas de negocio reutilizadas tal cual del reporte oficial:
#  - "Inadmitido" = un registro de RescatePunto donde NINGUNA bandera de
#    medio de rescate (aereo/carretero/casa de seguridad/central de
#    autobuses/ferroviario/hotel/puestos a disposicion/voluntarios/otro)
#    esta en True.
#  - "Reincidente" = coincide (nombre, apellidos, nacionalidad) con otro
#    registro en TODO el historico de RescatePunto (no solo el dia
#    seleccionado) -- vive en la vista materializada
#    mapa_mv_reincidencia_rescates (ver _rescates_set_duplicados_historicos),
#    refrescada cada 20 min, asi que el costo pesado del GROUP BY ya no se
#    paga por request.
#  - "Alojados en EM" (numFamilia=0/NULL, va solo) vs "Alojados en
#    DIF/Albergue" (numFamilia>0, viene en familia), solo sobre los
#    rescates NUEVOS (no reincidentes) del dia.
#
# Nota: el reporte oficial (estadistica.generar_cuadro_diario) trata a
# CHIAPAS aparte y ahi TODOS sus registros del dia caen en "reincidentes"
# sin comparar contra el historico -- no se replico esa parte porque no
# hay ninguna nota que explique si es una regla real o un descuido del
# reporte original; aqui CHIAPAS se evalua igual que cualquier otra
# entidad. Si el criterio oficial es a proposito, avisar para ajustarlo.
# =============================================================================

RESCATES_BANDERAS_MEDIO = dict(
    aeropuerto=False, carretero=False, casaSeguridad=False, centralAutobus=False,
    ferrocarril=False, hotel=False, puestosADispo=False, voluntarios=False, otro=False,
)


# @FADAR -- detalle de Retornados (mapa_retornados), independiente del
# numero que ya resta dentro de la tabla principal del Cuadro de Datos /
# Informe Diario. Por estado si no hay entidad filtrada, por nacionalidad
# si ya se filtro una entidad especifica (misma entidad ya fija el estado).
def _rescates_retornados_detalle(fecha_str, oficina=None):
    with connection.cursor() as cur:
        if oficina:
            cur.execute(
                "SELECT n.nombre, SUM(r.deportado), SUM(r.retornado), SUM(r.retornados_total) "
                "FROM mapa_retornados r "
                "JOIN mapa_estado e ON e.id = r.estado_id "
                "JOIN mapa_nacionalidad n ON n.id = r.nacionalidad_id "
                "WHERE r.fecha = %s AND e.nombre = %s "
                "GROUP BY n.nombre ORDER BY SUM(r.retornados_total) DESC",
                [fecha_str, oficina],
            )
            columna = "Nacionalidad"
        else:
            cur.execute(
                "SELECT e.nombre, SUM(r.deportado), SUM(r.retornado), SUM(r.retornados_total) "
                "FROM mapa_retornados r "
                "JOIN mapa_estado e ON e.id = r.estado_id "
                "WHERE r.fecha = %s "
                "GROUP BY e.nombre ORDER BY SUM(r.retornados_total) DESC",
                [fecha_str],
            )
            columna = "Entidad"
        filas = cur.fetchall()

    filas_tabla = [{"nombre": n, "deportado": d, "retornado": r, "total": t} for n, d, r, t in filas]
    return {
        "columna": columna,
        "filas": filas_tabla,
        "total_deportado": sum(f["deportado"] for f in filas_tabla),
        "total_retornado": sum(f["retornado"] for f in filas_tabla),
        "total_general": sum(f["total"] for f in filas_tabla),
    }


def _rescates_cuadro_datos(fecha_str, oficina=None):
    """fecha_str: 'YYYY-MM-DD'. Devuelve el dict con todos los datos del
    Cuadro de Datos para esa fecha (y opcionalmente una sola entidad)."""
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    fecha_rescate_fmt = fecha_obj.strftime("%d-%m-%y")

    qs_dia = RescatePunto.objects.filter(fecha=fecha_rescate_fmt)
    if oficina:
        qs_dia = qs_dia.filter(oficinaRepre=oficina)

    total_inadmitidos = qs_dia.filter(**RESCATES_BANDERAS_MEDIO).count()

    datos_dia = list(
        qs_dia.exclude(**RESCATES_BANDERAS_MEDIO)
        .values("nombre", "apellidos", "nacionalidad", "sexo", "edad", "numFamilia")
    )
    total_rescatados = len(datos_dia) + total_inadmitidos

    # Reincidencia: coincidencia de (nombre, apellidos, nacionalidad) en
    # TODO el historico de RescatePunto (consulta pesada, cacheada -- ver
    # _rescates_set_duplicados_historicos).
    set_duplicados = _rescates_set_duplicados_historicos()

    reincidentes = []
    nuevos = []
    for d in datos_dia:
        clave = (d["nombre"], d["apellidos"], d["nacionalidad"])
        if clave in set_duplicados:
            reincidentes.append(d)
        else:
            nuevos.append(d)

    conteo_reincidentes = len(reincidentes)
    conteo_nuevos = len(nuevos)
    subtotal1 = total_rescatados - conteo_reincidentes
    subtotal2 = subtotal1 - total_inadmitidos

    # Retornados: automatico desde mapa_retornados (no manual).
    filtro_estado_sql = ""
    params_ret = [fecha_str, fecha_str]
    if oficina:
        filtro_estado_sql = " AND estado_id = (SELECT id FROM mapa_estado WHERE nombre = %s)"
        params_ret = params_ret + [oficina]
    with connection.cursor() as cur:
        cur.execute(
            f"SELECT COALESCE(SUM(retornados_total),0) FROM mapa_retornados "
            f"WHERE fecha BETWEEN %s AND %s{filtro_estado_sql}",
            params_ret,
        )
        total_retornados = cur.fetchone()[0]

    rescates_nuevos_neto = conteo_nuevos - total_retornados

    # EM vs DIF/Albergue, por nacionalidad, solo sobre los "nuevos".
    nacionalidades = {d["nacionalidad"] for d in nuevos}
    datos_em = {n: 0 for n in nacionalidades}
    datos_dif = {n: 0 for n in nacionalidades}
    for d in nuevos:
        va_solo = d["numFamilia"] is None or d["numFamilia"] == 0
        if va_solo:
            datos_em[d["nacionalidad"]] += 1
        else:
            datos_dif[d["nacionalidad"]] += 1

    def _top_y_otros(datos, corte=5):
        ordenado = dict(sorted(datos.items(), key=lambda x: x[1], reverse=True))
        if len(ordenado) <= 7:
            corte = 7
        top = dict(list(ordenado.items())[:corte])
        resto = sum(list(ordenado.values())[corte:])
        if len(ordenado) > corte:
            top["Otras nacionalidades"] = resto
        total = sum(top.values())
        return top, total

    top_em, total_em = _top_y_otros(datos_em)
    top_dif, total_dif = _top_y_otros(datos_dif)

    return {
        "fecha": f"{fecha_obj.day:02d} {RESCATES_MESES_ES[fecha_obj.month]} {fecha_obj.year}",
        "fecha_iso": fecha_str,
        "oficina": oficina or "Nacional",
        "rescatados": total_rescatados,
        "reincidentes": conteo_reincidentes,
        "subtotal1": subtotal1,
        "inadmitidos": total_inadmitidos,
        "subtotal2": subtotal2,
        "retornados": total_retornados,
        "rescates_nuevos": rescates_nuevos_neto,
        "rescates_EM": top_em,
        "rescates_DIF": top_dif,
        "EM_total": total_em,
        "DIF_total": total_dif,
        "retornados_detalle": _rescates_retornados_detalle(fecha_str, oficina),
    }


def rescates_reporte_cuadro(request):
    """Vista previa en pantalla del Cuadro de Datos, con selector de fecha
    y entidad (menus desplegables) -- sin nada manual."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()

    datos = _rescates_cuadro_datos(fecha_str, oficina or None)
    datos["oficinas"] = RESCATES_OFICINAS
    datos["oficina_seleccionada"] = oficina
    return render(request, "Reportes_Analisis/rescates_reporte_cuadro.html", datos)


def rescates_reporte_cuadro_pdf(request):
    """Mismo cálculo que la vista previa, pero devuelto como PDF (mismo
    patrón WeasyPrint que ya usa el resto del proyecto)."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_cuadro_datos(fecha_str, oficina or None)

    template = get_template("Reportes_Analisis/_rescates_cuadro_pdf.html")
    html_string = template.render(datos)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="cuadro_datos_{fecha_str}.pdf"'
    return response


# Mismos colores exactos del PDF/pantalla oficiales, para los 3 formatos
# (pantalla, PDF, Excel) se vean consistentes.
RESCATES_COLOR_GUINDA = "621132"      # rgb(98,17,50) -- resumen del Cuadro de Datos
RESCATES_COLOR_NEGATIVO = "9D2449"    # rgb(157,36,73) -- reincidentes/inadmitidos/retornados
RESCATES_COLOR_GRIS = "D9D9D9"        # rgb(217,217,217) -- filas de subtotal
RESCATES_COLOR_EM = "285C4D"          # rgb(40,92,77) -- Alojados en EM
RESCATES_COLOR_EM_DATO = "8B6B41"     # rgb(139,107,65) -- texto/fondo de datos EM
RESCATES_COLOR_DIF = "7030A0"         # rgb(112,48,160) -- Alojados en DIF/Albergue
RESCATES_COLOR_CELESTE = "0EA5E9"     # rgb(14,165,233) -- Retornados (apartado independiente)


def rescates_reporte_cuadro_excel(request):
    """Mismo cálculo y mismos colores que la vista previa/PDF (guinda,
    verde EM, morado DIF, gris de subtotales), para que los tres formatos
    se vean consistentes."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_cuadro_datos(fecha_str, oficina or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuadro de Datos"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    ws.merge_cells("A1:B1")
    _rescates_excel_celda(ws, 1, 1, "INSTITUTO NACIONAL DE MIGRACIÓN", negrita=True, tam=13)
    ws.merge_cells("A2:B2")
    _rescates_excel_celda(ws, 2, 1, "DIRECCIÓN GENERAL DE COORDINACIÓN DE OFICINAS DE REPRESENTACION", negrita=True, tam=10)
    ws.merge_cells("A3:B3")
    _rescates_excel_celda(ws, 3, 1, f"Cuadro de Datos — {datos['fecha']} — {datos['oficina']}", negrita=True, tam=12)

    # Los valores numericos van como numero real de Excel (incluido el
    # signo negativo), no como texto formateado ("-5") -- para que la
    # columna se pueda sumar/importar directamente en Excel. "FECHA" y "De
    # estos" son las unicas dos filas inherentemente textuales (encabezado
    # y descripcion), el resto son datos.
    filas_resumen = [
        ("FECHA", datos["fecha"], None, "FFFFFF", True),
        ("Rescatados:", datos["rescatados"], None, "FFFFFF", False),
        ("Reincidentes:", -datos["reincidentes"], RESCATES_COLOR_NEGATIVO, "FFFFFF", False),
        ("Subtotal:", datos["subtotal1"], RESCATES_COLOR_GRIS, "000000", False),
        ("Inadmitidos:", -datos["inadmitidos"], RESCATES_COLOR_NEGATIVO, "FFFFFF", False),
        ("Subtotal:", datos["subtotal2"], RESCATES_COLOR_GRIS, "000000", False),
        ("Retornados:", -datos["retornados"], RESCATES_COLOR_NEGATIVO, "FFFFFF", False),
        ("Subtotal:", datos["rescates_nuevos"], RESCATES_COLOR_GRIS, "000000", True),
    ]
    fila = 5
    for etiqueta, valor, bg, color_txt, negrita in filas_resumen:
        _rescates_excel_celda(ws, fila, 1, etiqueta, bg=bg or None, color_texto=color_txt if bg else "000000", negrita=negrita, centrado=False)
        _rescates_excel_celda(ws, fila, 2, valor, bg=bg or None, color_texto=color_txt if bg else "000000", negrita=negrita)
        fila += 1

    # "De estos": en el PDF es una sola celda de texto descriptivo, pero
    # aqui se separa en 2 filas con valor numerico real (EM_total/DIF_total
    # como numero, no incrustados en una oracion) -- mismo dato, formato
    # importable.
    _rescates_excel_celda(ws, fila, 1, "De estos, alojados en EM:", bg=RESCATES_COLOR_GUINDA, color_texto="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws, fila, 2, datos["EM_total"], bg=RESCATES_COLOR_GUINDA, color_texto="FFFFFF", negrita=True)
    fila += 1
    _rescates_excel_celda(ws, fila, 1, "De estos, canalizados al DIF:", bg=RESCATES_COLOR_GUINDA, color_texto="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws, fila, 2, datos["DIF_total"], bg=RESCATES_COLOR_GUINDA, color_texto="FFFFFF", negrita=True)
    fila += 1

    fila += 1
    fila_em_inicio = fila
    _rescates_excel_celda(ws, fila, 1, "Alojados en EM", bg=RESCATES_COLOR_EM, color_texto="FFFFFF", negrita=True)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    fila += 1
    _rescates_excel_celda(ws, fila, 1, "NACIONALIDAD", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 2, "TOTAL", bg="FFFFFF", negrita=True)
    fila += 1
    for nombre, total in datos["rescates_EM"].items():
        _rescates_excel_celda(ws, fila, 1, nombre, color_texto=RESCATES_COLOR_EM_DATO, centrado=False)
        _rescates_excel_celda(ws, fila, 2, total, color_texto=RESCATES_COLOR_EM_DATO)
        fila += 1

    fila += 2
    _rescates_excel_celda(ws, fila, 1, "Alojados en DIF/Albergue", bg=RESCATES_COLOR_DIF, color_texto="FFFFFF", negrita=True)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=2)
    fila += 1
    _rescates_excel_celda(ws, fila, 1, "NACIONALIDAD", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 2, "TOTAL", bg="FFFFFF", negrita=True)
    fila += 1
    for nombre, total in datos["rescates_DIF"].items():
        _rescates_excel_celda(ws, fila, 1, nombre, color_texto=RESCATES_COLOR_DIF, centrado=False)
        _rescates_excel_celda(ws, fila, 2, total, color_texto=RESCATES_COLOR_DIF)
        fila += 1

    # @FADAR -- Retornados: hoja aparte, independiente del cuadro principal.
    ws2 = wb.create_sheet("Retornados")
    ws2.column_dimensions["A"].width = 28
    for col in "BCD":
        ws2.column_dimensions[col].width = 14
    _rescates_excel_celda(ws2, 1, 1, f"Retornados — detalle por {datos['retornados_detalle']['columna']}", bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True, centrado=False)
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    _rescates_excel_celda(ws2, 2, 1, "Retornados:", bg="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws2, 2, 2, datos["retornados_detalle"]["total_retornado"], bg="FFFFFF", color_texto=RESCATES_COLOR_CELESTE, negrita=True)
    fila_hdr = 4
    _rescates_excel_celda(ws2, fila_hdr, 1, datos["retornados_detalle"]["columna"].upper(), bg="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws2, fila_hdr, 2, "DEPORTADOS", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws2, fila_hdr, 3, "RETORNADOS", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws2, fila_hdr, 4, "TOTAL", bg="FFFFFF", negrita=True)
    fila2 = fila_hdr + 1
    for f in datos["retornados_detalle"]["filas"]:
        _rescates_excel_celda(ws2, fila2, 1, f["nombre"], centrado=False)
        _rescates_excel_celda(ws2, fila2, 2, f["deportado"])
        _rescates_excel_celda(ws2, fila2, 3, f["retornado"])
        _rescates_excel_celda(ws2, fila2, 4, f["total"])
        fila2 += 1
    _rescates_excel_celda(ws2, fila2, 1, "TOTAL", bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws2, fila2, 2, datos["retornados_detalle"]["total_deportado"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws2, fila2, 3, datos["retornados_detalle"]["total_retornado"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws2, fila2, 4, datos["retornados_detalle"]["total_general"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="cuadro_datos_{fecha_str}.xlsx"'
    wb.save(response)
    return response


# =============================================================================
# Reportes de Rescates -- "Informe Diario de Rescatados" (mismo formato y
# reglas de negocio que estadistica.generar_pdf/reporteA.html, tambien
# automatico: Retornados sale de mapa_retornados, no se captura a mano).
#
# 8 categorias (igual criterio que el reporte oficial, definidas solo con
# sexo + edad + numFamilia):
#   H_AS/M_AS = Hombre/Mujer Adulto(a) Solo(a)        (numFamilia 0/NULL)
#   H_AA/M_AA = Hombre/Mujer Adulto(a) Acompaña NNA    (numFamilia > 0)
#   H_mA/M_mA = NNA Acompañado(a), Hombre/Mujer        (numFamilia > 0)
#   H_mS/M_mS = NNA No Acompañado(a), Hombre/Mujer     (numFamilia 0/NULL)
# =============================================================================

RESCATES_ETIQUETAS_CATEGORIA = [
    ("H_AS", "Adultos Hombres Solos"),
    ("M_AS", "Adultas Mujeres Solas"),
    ("H_AA", "Adultos Hombres Acompañan NNA"),
    ("M_AA", "Adultas Mujeres Acompañan NNA"),
    ("H_mA", "NNA Acompañados Hombres"),
    ("M_mA", "NNA Acompañados Mujeres"),
    ("H_mS", "NNA No Acompañados Hombres"),
    ("M_mS", "NNA No Acompañados Mujeres"),
]


def _rescates_clasificar_categoria(sexo, edad, num_familia):
    solo = num_familia is None or num_familia == 0
    if sexo and edad >= 18:
        return "H_AS" if solo else "H_AA"
    if not sexo and edad >= 18:
        return "M_AS" if solo else "M_AA"
    if sexo and edad < 18:
        return "H_mS" if solo else "H_mA"
    return "M_mS" if solo else "M_mA"


def _rescates_informe_diario(fecha_str):
    """fecha_str: 'YYYY-MM-DD'. Reporte nacional (las 32 entidades, sin
    tratamiento especial para ninguna -- ver nota de Chiapas en el Cuadro
    de Datos)."""
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    fecha_rescate_fmt = fecha_obj.strftime("%d-%m-%y")

    qs_dia = RescatePunto.objects.filter(fecha=fecha_rescate_fmt)

    # --- Tabla 1: rescates validos por oficina y medio de rescate ---
    por_oficina_qs = (
        qs_dia.exclude(**RESCATES_BANDERAS_MEDIO)
        .values("oficinaRepre")
        .annotate(
            total=Count("idRescate"),
            total_aereos=Count("idRescate", filter=Q(aeropuerto=True)),
            total_carreteros=Count("idRescate", filter=Q(carretero=True)),
            total_central=Count("idRescate", filter=Q(centralAutobus=True)),
            total_ferro=Count("idRescate", filter=Q(ferrocarril=True)),
            total_puestos=Count("idRescate", filter=Q(puestosADispo=True)),
            total_otros=(
                Count("idRescate", filter=Q(voluntarios=True))
                + Count("idRescate", filter=Q(otro=True))
                + Count("idRescate", filter=Q(casaSeguridad=True))
                + Count("idRescate", filter=Q(hotel=True))
            ),
        )
    )
    rescates_por_oficina = {o: {"total": 0, "total_aereos": 0, "total_carreteros": 0, "total_central": 0,
                                 "total_ferro": 0, "total_puestos": 0, "total_otros": 0} for o in RESCATES_OFICINAS}
    rescates_por_oficina["Total"] = dict(rescates_por_oficina[RESCATES_OFICINAS[0]])
    for fila in por_oficina_qs:
        of = fila["oficinaRepre"]
        if of not in rescates_por_oficina:
            continue
        for campo in ("total", "total_aereos", "total_carreteros", "total_central", "total_ferro", "total_puestos", "total_otros"):
            rescates_por_oficina[of][campo] = fila[campo]
            rescates_por_oficina["Total"][campo] += fila[campo]

    # --- Datos persona por persona (para reincidencia y las 8 categorias) ---
    datos_dia = list(
        qs_dia.exclude(**RESCATES_BANDERAS_MEDIO)
        .values("nombre", "apellidos", "nacionalidad", "oficinaRepre", "sexo", "edad", "numFamilia", "iso3")
    )

    set_duplicados = _rescates_set_duplicados_historicos()

    reincidentes, nuevos = [], []
    for d in datos_dia:
        clave = (d["nombre"], d["apellidos"], d["nacionalidad"])
        (reincidentes if clave in set_duplicados else nuevos).append(d)

    # --- Tabla 2: reincidentes / nuevos por oficina ---
    reincidentes_por_oficina_ct = Counter(d["oficinaRepre"] for d in reincidentes)
    nuevos_por_oficina_ct = Counter(d["oficinaRepre"] for d in nuevos)
    reincidentes_por_oficina = {o: {"total_reincidentes": 0, "total_nuevos": 0, "total": 0} for o in RESCATES_OFICINAS}
    reincidentes_por_oficina["Total"] = {"total_reincidentes": 0, "total_nuevos": 0, "total": 0}
    for of in RESCATES_OFICINAS:
        r = reincidentes_por_oficina_ct.get(of, 0)
        n = nuevos_por_oficina_ct.get(of, 0)
        reincidentes_por_oficina[of] = {"total_reincidentes": r, "total_nuevos": n, "total": r + n}
        reincidentes_por_oficina["Total"]["total_reincidentes"] += r
        reincidentes_por_oficina["Total"]["total_nuevos"] += n
        reincidentes_por_oficina["Total"]["total"] += r + n

    # --- Tablas 3 y 4: nacionalidad x 8 categorias, nuevos y reincidentes ---
    def _tabla_nacionalidad_categorias(lista):
        nacionalidades = {d["nacionalidad"] for d in lista}
        tabla = {n: {clave: 0 for clave, _ in RESCATES_ETIQUETAS_CATEGORIA} for n in nacionalidades}
        for n in tabla:
            tabla[n]["total"] = 0
        for d in lista:
            cat = _rescates_clasificar_categoria(d["sexo"], d["edad"], d["numFamilia"])
            tabla[d["nacionalidad"]][cat] += 1
            tabla[d["nacionalidad"]]["total"] += 1
        return dict(sorted(tabla.items(), key=lambda x: x[1]["total"], reverse=True))

    nacionalidades_nuevos = _tabla_nacionalidad_categorias(nuevos)
    nacionalidades_reincidentes = _tabla_nacionalidad_categorias(reincidentes)

    # --- Inadmitidos: nacionalidad x 4 categorias (sexo x adulto/menor) ---
    datos_inadmitidos = list(
        qs_dia.filter(**RESCATES_BANDERAS_MEDIO).values("nacionalidad", "sexo", "edad", "iso3")
    )
    nacionalidades_inadm = {d["nacionalidad"] for d in datos_inadmitidos}
    tabla_inadm = {n: {"H_A": 0, "M_A": 0, "H_m": 0, "M_m": 0, "total": 0} for n in nacionalidades_inadm}
    for d in datos_inadmitidos:
        tabla_inadm[d["nacionalidad"]]["total"] += 1
        if d["sexo"] and d["edad"] >= 18:
            tabla_inadm[d["nacionalidad"]]["H_A"] += 1
        elif not d["sexo"] and d["edad"] >= 18:
            tabla_inadm[d["nacionalidad"]]["M_A"] += 1
        elif d["sexo"] and d["edad"] < 18:
            tabla_inadm[d["nacionalidad"]]["H_m"] += 1
        else:
            tabla_inadm[d["nacionalidad"]]["M_m"] += 1
    nacionalidades_inadmitidos = dict(sorted(tabla_inadm.items(), key=lambda x: x[1]["total"], reverse=True))

    # @FADAR -- nacionalidades extracontinentales (fuera de America), para
    # remarcarlas en las 3 tablas del informe. Un nombre de nacionalidad
    # puede repetirse en varias filas con el mismo iso3 -- basta el primero.
    nacio_a_iso3 = {}
    for d in datos_dia:
        nacio_a_iso3.setdefault(d["nacionalidad"], d["iso3"])
    for d in datos_inadmitidos:
        nacio_a_iso3.setdefault(d["nacionalidad"], d["iso3"])
    nacionalidades_extracontinentales = {
        n for n, iso3 in nacio_a_iso3.items() if _rescates_es_extracontinental(str(iso3).upper())
    }

    # --- Retornados: automatico desde mapa_retornados (no manual) ---
    with connection.cursor() as cur:
        cur.execute(
            "SELECT COALESCE(SUM(retornados_total),0) FROM mapa_retornados WHERE fecha = %s",
            [fecha_str],
        )
        total_retornados = cur.fetchone()[0]

    return {
        "fecha_actual": fecha_obj.strftime("%d/%m/%Y"),
        "fecha_iso": fecha_str,
        "rescates": rescates_por_oficina,
        "reincidentes": reincidentes_por_oficina,
        "nacionalidades": nacionalidades_nuevos,
        "nacionalidades_re": nacionalidades_reincidentes,
        "nacionalidades_inadm": nacionalidades_inadmitidos,
        "nacionalidades_extracontinentales": nacionalidades_extracontinentales,
        "dato": total_retornados,
        "categorias": RESCATES_ETIQUETAS_CATEGORIA,
        "retornados_detalle": _rescates_retornados_detalle(fecha_str),
    }


def rescates_reporte_informe(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_informe_diario(fecha_str)
    return render(request, "Reportes_Analisis/rescates_reporte_informe.html", datos)


def rescates_reporte_informe_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_informe_diario(fecha_str)

    template = get_template("Reportes_Analisis/_rescates_informe_pdf.html")
    html_string = template.render(datos)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="informe_diario_{fecha_str}.pdf"'
    return response


# Mismos colores exactos del PDF ("letraT0".."letraT6") para el Informe
# Diario, igual criterio que el Cuadro de Datos: pantalla/PDF/Excel
# consistentes entre si.
RESCATES_LETRA_T3 = ("4BACC6", "000000")
RESCATES_LETRA_T4 = ("13322B", "EDEDED")
RESCATES_LETRA_T6 = ("7030A0", "EAEAEA")


def rescates_reporte_informe_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_informe_diario(fecha_str)

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Por oficina"
    oficinas_cols = list(datos["rescates"].keys())
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(oficinas_cols) + 1)
    _rescates_excel_celda(ws1, 1, 1, "INSTITUTO NACIONAL DE MIGRACIÓN — DIRECCIÓN GENERAL DE COORDINACIÓN DE OFICINAS DE REPRESENTACION", negrita=True, tam=10)
    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(oficinas_cols) + 1)
    _rescates_excel_celda(ws1, 2, 1, f"Informe diario de operaciones — {datos['fecha_actual']}", bg=RESCATES_COLOR_FONDO[0], color_texto=RESCATES_COLOR_FONDO[1], negrita=True, tam=12)

    fila = 4
    _rescates_excel_fila(ws1, fila, ["Rubro"] + oficinas_cols, RESCATES_LETRA_T1, negrita=True)
    fila += 1
    for campo, etiqueta in [("total_aereos", "Aeropuertos"), ("total_carreteros", "Carreteros"),
                             ("total_central", "Central de Autobús"), ("total_ferro", "Ferroviarios"),
                             ("total_puestos", "Puestos a Disposición"), ("total_otros", "Otros")]:
        _rescates_excel_fila(ws1, fila, [etiqueta] + [datos["rescates"][o][campo] for o in oficinas_cols], RESCATES_LETRA_T2)
        fila += 1
    _rescates_excel_fila(ws1, fila, ["TOTAL MIGRANTES RESCATADOS"] + [datos["rescates"][o]["total"] for o in oficinas_cols], RESCATES_LETRA_T1, negrita=True)
    fila += 2

    _rescates_excel_fila(ws1, fila, ["Rubro"] + oficinas_cols, RESCATES_LETRA_T4, negrita=True)
    fila += 1
    for campo, etiqueta in [("total_reincidentes", "Reincidentes"), ("total_nuevos", "Registros Nuevos")]:
        _rescates_excel_fila(ws1, fila, [etiqueta] + [datos["reincidentes"][o][campo] for o in oficinas_cols], RESCATES_LETRA_T2)
        fila += 1
    _rescates_excel_fila(ws1, fila, ["TOTAL MIGRANTES RESCATADOS"] + [datos["reincidentes"][o]["total"] for o in oficinas_cols], RESCATES_LETRA_T0, negrita=True)

    ws1.column_dimensions["A"].width = 26
    for i in range(len(oficinas_cols)):
        ws1.column_dimensions[get_column_letter(2 + i)].width = 12

    # @FADAR -- nacionalidades extracontinentales (fuera de America):
    # mismo ambar que pantalla/PDF, sobre la celda de nacionalidad.
    extracontinentales = datos["nacionalidades_extracontinentales"]

    def _hoja_nacionalidad(nombre, tabla):
        ws = wb.create_sheet(nombre)
        encabezados = ["Nacionalidad"] + [e for _, e in RESCATES_ETIQUETAS_CATEGORIA] + ["Total"]
        _rescates_excel_fila(ws, 1, encabezados, RESCATES_LETRA_T1, negrita=True)
        f = 2
        for nac, datos_nac in tabla.items():
            if nac in extracontinentales:
                bg_nac, color_nac = "B45309", "FFFFFF"
            else:
                bg_nac, color_nac = RESCATES_LETRA_T1
            _rescates_excel_celda(ws, f, 1, nac, bg=bg_nac, color_texto=color_nac, centrado=False)
            for i, (c, _) in enumerate(RESCATES_ETIQUETAS_CATEGORIA):
                _rescates_excel_celda(ws, f, 2 + i, datos_nac[c], bg=RESCATES_LETRA_T2[0], color_texto=RESCATES_LETRA_T2[1])
            _rescates_excel_celda(ws, f, 2 + len(RESCATES_ETIQUETAS_CATEGORIA), datos_nac["total"], bg=RESCATES_LETRA_T0[0], color_texto=RESCATES_LETRA_T0[1], negrita=True)
            f += 1
        ws.column_dimensions["A"].width = 28
        return ws

    _hoja_nacionalidad("Nacionalidades Nuevos", datos["nacionalidades"])
    _hoja_nacionalidad("Nacionalidades Reincidentes", datos["nacionalidades_re"])

    ws4 = wb.create_sheet("Inadmitidos y Retornados")
    _rescates_excel_fila(ws4, 1, ["Nacionalidad", "Adultos Hombres", "Adultas Mujeres", "Menores Hombres", "Menores Mujeres", "Total"], RESCATES_LETRA_T1, negrita=True)
    f = 2
    for nac, d in datos["nacionalidades_inadm"].items():
        if nac in extracontinentales:
            bg_nac, color_nac = "B45309", "FFFFFF"
        else:
            bg_nac, color_nac = RESCATES_LETRA_T1
        _rescates_excel_celda(ws4, f, 1, nac, bg=bg_nac, color_texto=color_nac, centrado=False)
        for i, campo in enumerate(("H_A", "M_A", "H_m", "M_m")):
            _rescates_excel_celda(ws4, f, 2 + i, d[campo], bg=RESCATES_LETRA_T3[0], color_texto=RESCATES_LETRA_T3[1])
        _rescates_excel_celda(ws4, f, 6, d["total"], bg=RESCATES_LETRA_T0[0], color_texto=RESCATES_LETRA_T0[1], negrita=True)
        f += 1
    f += 1
    _rescates_excel_fila(ws4, f, ["TOTAL DE RETORNADOS A SU PAÍS DE ORIGEN", datos["dato"]], RESCATES_LETRA_T6, negrita=True)
    ws4.column_dimensions["A"].width = 32

    # @FADAR -- Retornados: hoja aparte, independiente de "Inadmitidos y Retornados".
    ws5 = wb.create_sheet("Retornados (detalle)")
    ws5.column_dimensions["A"].width = 28
    for col in "BCD":
        ws5.column_dimensions[col].width = 14
    rd = datos["retornados_detalle"]
    _rescates_excel_celda(ws5, 1, 1, f"Retornados — detalle por {rd['columna']}", bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True, centrado=False)
    ws5.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    _rescates_excel_celda(ws5, 2, 1, "Retornados:", bg="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws5, 2, 2, rd["total_retornado"], bg="FFFFFF", color_texto=RESCATES_COLOR_CELESTE, negrita=True)
    fila_hdr5 = 4
    _rescates_excel_celda(ws5, fila_hdr5, 1, rd["columna"].upper(), bg="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws5, fila_hdr5, 2, "DEPORTADOS", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws5, fila_hdr5, 3, "RETORNADOS", bg="FFFFFF", negrita=True)
    _rescates_excel_celda(ws5, fila_hdr5, 4, "TOTAL", bg="FFFFFF", negrita=True)
    f5 = fila_hdr5 + 1
    for fila_r in rd["filas"]:
        _rescates_excel_celda(ws5, f5, 1, fila_r["nombre"], centrado=False)
        _rescates_excel_celda(ws5, f5, 2, fila_r["deportado"])
        _rescates_excel_celda(ws5, f5, 3, fila_r["retornado"])
        _rescates_excel_celda(ws5, f5, 4, fila_r["total"])
        f5 += 1
    _rescates_excel_celda(ws5, f5, 1, "TOTAL", bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True, centrado=False)
    _rescates_excel_celda(ws5, f5, 2, rd["total_deportado"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws5, f5, 3, rd["total_retornado"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws5, f5, 4, rd["total_general"], bg=RESCATES_COLOR_CELESTE, color_texto="FFFFFF", negrita=True)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="informe_diario_{fecha_str}.xlsx"'
    wb.save(response)
    return response

# @FADAR -- Fase 2 del traslado: CECO 2, CECO V1/V2 y Reporte Personalizado.
# Codigo copiado TAL CUAL desde mapa/views.py (misma logica, mismas reglas,
# misma estructura) -- mapa/views.py se deja intacto por ahora, el
# descarte de esa copia queda para una fase posterior. Las constantes y
# helpers que ya trajo la Fase 1 (RESCATES_OFICINAS, RESCATES_ZONA_*,
# RESCATES_ISO3_*, _rescates_es_extracontinental, RESCATES_MV_REINCIDENCIA,
# RESCATES_LETRA_T0-T2, RESCATES_COLOR_FONDO, RESCATES_BORDE_DELGADO,
# _rescates_excel_celda/_fila, RESCATES_ETIQUETAS_CATEGORIA,
# _rescates_clasificar_categoria, RESCATES_MESES_ES) no se duplican aqui.

# =============================================================================
# Reporte Personalizado -- filtros libres (calendario + menus desplegables)
# para que el usuario arme su propio reporte, en vez de los formatos fijos
# oficiales. Los resultados se agrupan por oficina + nacionalidad (nunca se
# muestran registros individuales/datos personales -- mismo criterio que el
# resto del dashboard).
# =============================================================================

RESCATES_TIPO_RESCATE_CAMPOS = [
    ("carretero", "Carretero"),
    ("aeropuerto", "Aéreo"),
    ("ferrocarril", "Ferroviario"),
    ("centralAutobus", "Central de autobuses"),
    ("casaSeguridad", "Casa de seguridad"),
    ("hotel", "Hotel"),
    ("reclusorio", "Reclusorio"),
]

RESCATES_CATEGORIA_SQL = {
    "hombres": "sexo=true AND edad>=18",
    "mujeres": "sexo=false AND edad>=18",
    "ninos": "sexo=true AND edad<18",
    "ninas": "sexo=false AND edad<18",
}

RESCATES_ZONAS_POR_NOMBRE = {
    "Río Bravo": RESCATES_ZONA_RIO_BRAVO,
    "Centro": RESCATES_ZONA_CENTRO,
    "Suchiate": RESCATES_ZONA_SUCHIATE,
}

RESCATES_CACHE_KEY_NACIONALIDADES = "rescates_nacionalidades_disponibles"
RESCATES_CACHE_TTL_NACIONALIDADES = 1800  # 30 min


def _rescates_nacionalidades_disponibles():
    """Lista de nacionalidades distintas para el menu desplegable -- se
    cachea porque es un DISTINCT sobre 1.5M filas sin indice."""
    resultado = cache.get(RESCATES_CACHE_KEY_NACIONALIDADES)
    if resultado is None:
        with connection.cursor() as cur:
            cur.execute("SELECT DISTINCT UPPER(nacionalidad) FROM usuario_rescatepunto WHERE nacionalidad != '' ORDER BY 1")
            resultado = [fila[0] for fila in cur.fetchall()]
        cache.set(RESCATES_CACHE_KEY_NACIONALIDADES, resultado, RESCATES_CACHE_TTL_NACIONALIDADES)
    return resultado


def _rescates_personalizado(fecha_inicio, fecha_fin, oficina=None, zona=None, tipo_rescate=None, categoria=None, nacionalidad=None):
    """Reporte con filtros libres: rango de fechas (obligatorio) + entidad,
    zona, tipo de rescate, categoria (sexo/edad) y nacionalidad (todos
    opcionales). Agrupa por oficina + nacionalidad con el mismo desglose
    hombres/mujeres/ninos/ninas que ya se usa en el resto del dashboard."""
    filtros_sql = []
    params = [fecha_inicio, fecha_fin]

    if oficina:
        filtros_sql.append('"oficinaRepre" = %s')
        params.append(oficina)
    elif zona and zona in RESCATES_ZONAS_POR_NOMBRE:
        oficinas_zona = RESCATES_ZONAS_POR_NOMBRE[zona]
        placeholders = ",".join(["%s"] * len(oficinas_zona))
        filtros_sql.append(f'"oficinaRepre" IN ({placeholders})')
        params.extend(oficinas_zona)

    if tipo_rescate and tipo_rescate in dict(RESCATES_TIPO_RESCATE_CAMPOS):
        filtros_sql.append(f'"{tipo_rescate}" = true')

    if categoria and categoria in RESCATES_CATEGORIA_SQL:
        filtros_sql.append(RESCATES_CATEGORIA_SQL[categoria])

    if nacionalidad:
        filtros_sql.append("UPPER(nacionalidad) = %s")
        params.append(nacionalidad)

    filtro_extra = "".join(f" AND {f}" for f in filtros_sql)

    with connection.cursor() as cur:
        cur.execute(
            f'SELECT "oficinaRepre", UPPER(nacionalidad), '
            f"  COUNT(*) FILTER (WHERE sexo=true  AND edad>=18) AS hombres, "
            f"  COUNT(*) FILTER (WHERE sexo=false AND edad>=18) AS mujeres, "
            f"  COUNT(*) FILTER (WHERE sexo=true  AND edad<18)  AS ninos, "
            f"  COUNT(*) FILTER (WHERE sexo=false AND edad<18)  AS ninas, "
            f"  COUNT(*) AS total, MAX(iso3) AS iso3 "
            f"FROM usuario_rescatepunto "
            f"WHERE TO_DATE(fecha,'DD-MM-YY') BETWEEN %s AND %s{filtro_extra} "
            f'GROUP BY "oficinaRepre", UPPER(nacionalidad) '
            f"ORDER BY total DESC",
            params,
        )
        filas = cur.fetchall()

    filas_tabla = [
        {"oficina": of, "nacionalidad": nac, "hombres": h, "mujeres": m, "ninos": n, "ninas": ni, "total": t, "iso3": iso3}
        for of, nac, h, m, n, ni, t, iso3 in filas
    ]
    total_general = sum(f["total"] for f in filas_tabla)
    # @FADAR -- nacionalidades extracontinentales (fuera de America), misma
    # regla que en el Informe Diario.
    nac_iso3 = {}
    for f in filas_tabla:
        nac_iso3.setdefault(f["nacionalidad"], f["iso3"])
    nacionalidades_extracontinentales = {
        n for n, iso3 in nac_iso3.items() if _rescates_es_extracontinental(str(iso3).upper())
    }
    return filas_tabla, total_general, nacionalidades_extracontinentales


def rescates_reporte_personalizado(request):
    """Vista previa en pantalla del reporte personalizado (filtros libres)."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    hoy = date.today().isoformat()
    fecha_inicio = request.GET.get('fecha_inicio', hoy)
    fecha_fin = request.GET.get('fecha_fin', hoy)
    oficina = request.GET.get('oficina', '').strip()
    zona = request.GET.get('zona', '').strip()
    tipo_rescate = request.GET.get('tipo_rescate', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    nacionalidad = request.GET.get('nacionalidad', '').strip()

    filas_tabla, total_general, nacionalidades_extracontinentales = _rescates_personalizado(
        fecha_inicio, fecha_fin, oficina or None, zona or None,
        tipo_rescate or None, categoria or None, nacionalidad or None,
    )

    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "oficina_seleccionada": oficina,
        "zona_seleccionada": zona,
        "tipo_rescate_seleccionado": tipo_rescate,
        "categoria_seleccionada": categoria,
        "nacionalidad_seleccionada": nacionalidad,
        "oficinas": RESCATES_OFICINAS,
        "zonas": list(RESCATES_ZONAS_POR_NOMBRE.keys()),
        "tipos_rescate": RESCATES_TIPO_RESCATE_CAMPOS,
        "nacionalidades": _rescates_nacionalidades_disponibles(),
        "filas_tabla": filas_tabla[:200],
        "total_filas": len(filas_tabla),
        "total_general": total_general,
        "nacionalidades_extracontinentales": nacionalidades_extracontinentales,
    }
    return render(request, "Reportes_Analisis/rescates_reporte_personalizado.html", context)


def rescates_reporte_personalizado_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_inicio = request.GET.get('fecha_inicio', date.today().isoformat())
    fecha_fin = request.GET.get('fecha_fin', date.today().isoformat())
    oficina = request.GET.get('oficina', '').strip()
    zona = request.GET.get('zona', '').strip()
    tipo_rescate = request.GET.get('tipo_rescate', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    nacionalidad = request.GET.get('nacionalidad', '').strip()

    filas_tabla, total_general, nacionalidades_extracontinentales = _rescates_personalizado(
        fecha_inicio, fecha_fin, oficina or None, zona or None,
        tipo_rescate or None, categoria or None, nacionalidad or None,
    )
    etiquetas_tipo = dict(RESCATES_TIPO_RESCATE_CAMPOS)
    context = {
        "fecha_inicio": fecha_inicio,
        "fecha_fin": fecha_fin,
        "oficina": oficina or "Todas",
        "zona": zona or "Todas",
        "tipo_rescate": etiquetas_tipo.get(tipo_rescate, "Todos"),
        "categoria": categoria or "Todas",
        "nacionalidad": nacionalidad or "Todas",
        "filas_tabla": filas_tabla,
        "total_general": total_general,
        "nacionalidades_extracontinentales": nacionalidades_extracontinentales,
    }
    template = get_template("Reportes_Analisis/_rescates_personalizado_pdf.html")
    html_string = template.render(context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="reporte_personalizado_{fecha_inicio}_a_{fecha_fin}.pdf"'
    return response


def rescates_reporte_personalizado_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)

    fecha_inicio = request.GET.get('fecha_inicio', date.today().isoformat())
    fecha_fin = request.GET.get('fecha_fin', date.today().isoformat())
    oficina = request.GET.get('oficina', '').strip()
    zona = request.GET.get('zona', '').strip()
    tipo_rescate = request.GET.get('tipo_rescate', '').strip()
    categoria = request.GET.get('categoria', '').strip()
    nacionalidad = request.GET.get('nacionalidad', '').strip()

    filas_tabla, total_general, nacionalidades_extracontinentales = _rescates_personalizado(
        fecha_inicio, fecha_fin, oficina or None, zona or None,
        tipo_rescate or None, categoria or None, nacionalidad or None,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Personalizado"
    etiquetas_tipo = dict(RESCATES_TIPO_RESCATE_CAMPOS)

    ws.merge_cells("A1:G1")
    _rescates_excel_celda(
        ws, 1, 1,
        f"Reporte Personalizado — {fecha_inicio} a {fecha_fin} — "
        f"Entidad: {oficina or 'Todas'} · Zona: {zona or 'Todas'} · "
        f"Tipo: {etiquetas_tipo.get(tipo_rescate, 'Todos')} · "
        f"Categoría: {categoria or 'Todas'} · Nacionalidad: {nacionalidad or 'Todas'}",
        negrita=True, tam=11, centrado=False,
    )

    fila = 3
    _rescates_excel_fila(ws, fila, ["Oficina", "Nacionalidad", "Hombres", "Mujeres", "Niños", "Niñas", "Total"], RESCATES_LETRA_T1, negrita=True)
    fila += 1
    for f in filas_tabla:
        # @FADAR -- nacionalidades extracontinentales (fuera de America)
        colores_fila = RESCATES_LETRA_T2
        if f["nacionalidad"] in nacionalidades_extracontinentales:
            colores_fila = ("B45309", "FFFFFF")
        _rescates_excel_fila(ws, fila, [f["oficina"], f["nacionalidad"], f["hombres"], f["mujeres"], f["ninos"], f["ninas"], f["total"]], colores_fila)
        fila += 1
    _rescates_excel_fila(ws, fila, ["", "TOTAL", "", "", "", "", total_general], RESCATES_LETRA_T0, negrita=True)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 24
    for col in ("C", "D", "E", "F", "G"):
        ws.column_dimensions[col].width = 12

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="reporte_personalizado_{fecha_inicio}_a_{fecha_fin}.xlsx"'
    wb.save(response)
    return response


# @FADAR -- nombre completo del mes, para los nombres de archivo descargables
# (REPORTE CECO DIA-MES-AÑO(PRHs).xlsx). RESCATES_MESES_ES (corto) ya viene
# de la Fase 1.
RESCATES_MESES_ES_LARGO = {
    1: "ENERO", 2: "FEBRERO", 3: "MARZO", 4: "ABRIL", 5: "MAYO", 6: "JUNIO",
    7: "JULIO", 8: "AGOSTO", 9: "SEPTIEMBRE", 10: "OCTUBRE", 11: "NOVIEMBRE", 12: "DICIEMBRE",
}




# =============================================================================
# Reporte "CECO 2" -- listado real por punto (PRH capturado en
# usuario_rescatepunto.puntoEstra) con su tipo, para un solo dia. NO replica
# el catalogo TIPO que se llena a mano (no coincide con mapa_prhs/
# mapa_tipoprh, confirmado en conversacion -- lo que capturan a mano no
# tiene una fuente verificable en la base). En su lugar, "tipo" se deriva de
# las banderas de medio de rescate YA verificadas y usadas en el resto del
# modulo -- por eso el total aqui SIEMPRE cuadra exacto contra el total real
# de rescates del dia, a diferencia del Excel manual.
# =============================================================================

RESCATES_TIPO_CASE_SQL = """CASE
        WHEN r.aeropuerto THEN 'Aeropuerto'
        WHEN r.carretero THEN 'Carretero'
        WHEN r."centralAutobus" THEN 'Central de Autobuses'
        WHEN r.ferrocarril THEN 'Ferroviario'
        WHEN r."casaSeguridad" THEN 'Casa de Seguridad'
        WHEN r.hotel THEN 'Hotel'
        WHEN r."puestosADispo" THEN 'Puestos a Disposición'
        WHEN r.voluntarios THEN 'Voluntario'
        ELSE 'Otro'
    END"""


def _rescates_ceco2_detalle(fecha_str, oficina=None):
    """fecha_str: 'YYYY-MM-DD'. Un solo dia (igual que Cuadro de
    Datos/Informe Diario/CECO). Filas = (oficina, punto, tipo, rescates,
    primera_vez, reincidencias) -- primera_vez/reincidencias vienen de la
    misma vista materializada de reincidencia (>=2 apariciones historicas)
    ya usada en el resto del modulo, unida por identidad (nombre+apellidos+
    nacionalidad)."""
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    fecha_fmt = fecha_obj.strftime("%d-%m-%y")

    filtro = ""
    params = [fecha_fmt]
    if oficina:
        filtro = ' AND r."oficinaRepre" = %s'
        params.append(oficina)

    with connection.cursor() as cur:
        cur.execute(
            f'SELECT r."oficinaRepre", r."puntoEstra", {RESCATES_TIPO_CASE_SQL} AS tipo, COUNT(*), '
            f"  COUNT(*) FILTER (WHERE v.clasificacion = 'Rescate primera vez'), "
            f"  COUNT(*) FILTER (WHERE v.clasificacion = 'Reincidente') "
            f'FROM usuario_rescatepunto r '
            f'JOIN {RESCATES_MV_REINCIDENCIA} v '
            f'  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad '
            f'WHERE r.fecha = %s{filtro} '
            f'GROUP BY r."oficinaRepre", r."puntoEstra", {RESCATES_TIPO_CASE_SQL} '
            f'ORDER BY r."oficinaRepre", r."puntoEstra"',
            params,
        )
        filas = cur.fetchall()

    filas_tabla = [
        {"oficina": of, "prh": prh, "tipo": tipo, "rescates": r, "primera_vez": pv, "reincidencias": rc}
        for of, prh, tipo, r, pv, rc in filas
    ]
    return {
        "fecha_actual": f"{fecha_obj.day:02d}/{fecha_obj.month:02d}/{fecha_obj.year}",
        "fecha_iso": fecha_str,
        "oficina": oficina or "Nacional",
        "filas": filas_tabla,
        "total": sum(f["rescates"] for f in filas_tabla),
        "total_primera_vez": sum(f["primera_vez"] for f in filas_tabla),
        "total_reincidencias": sum(f["reincidencias"] for f in filas_tabla),
    }


def rescates_reporte_ceco2(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_ceco2_detalle(fecha_str, oficina or None)
    datos["oficinas"] = RESCATES_OFICINAS
    datos["oficina_seleccionada"] = oficina
    return render(request, "Reportes_Analisis/rescates_reporte_ceco2.html", datos)


def rescates_reporte_ceco2_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_ceco2_detalle(fecha_str, oficina or None)

    template = get_template("Reportes_Analisis/_rescates_ceco2_pdf.html")
    html_string = template.render(datos)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ceco2_{fecha_str}.pdf"'
    return response


def rescates_reporte_ceco2_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    oficina = request.GET.get("oficina", "").strip()
    datos = _rescates_ceco2_detalle(fecha_str, oficina or None)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CECO 2"
    ws.column_dimensions["A"].width = 13
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 46
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20

    ws.merge_cells("A1:B2")
    _rescates_excel_celda(ws, 1, 1, "Fecha:", bg=RESCATES_COLOR_FONDO[0], color_texto="FFFFFF", negrita=True)
    ws.merge_cells("A3:B3")
    _rescates_excel_celda(ws, 3, 1, datos["fecha_actual"], bg="FFFFFF", negrita=True)

    ws.merge_cells("D1:E1")
    _rescates_excel_celda(ws, 1, 4, "TOTAL", bg=RESCATES_COLOR_FONDO[0], color_texto="FFFFFF", negrita=True)
    ws.merge_cells("D2:E2")
    _rescates_excel_celda(ws, 2, 4, datos["total"], bg="FFFFFF", negrita=True, tam=14)

    fila = 5
    _rescates_excel_fila(ws, fila, ["FECHA", "OR", "PRH", "RESCATES TOTAL", "PRIMERA VEZ/RESCATES REALES", "REINCIDENCIAS", "TIPO"], RESCATES_LETRA_T1, negrita=True)
    fila += 1
    for f in datos["filas"]:
        _rescates_excel_celda(ws, fila, 1, datos["fecha_actual"], centrado=True)
        _rescates_excel_celda(ws, fila, 2, f["oficina"], centrado=False)
        _rescates_excel_celda(ws, fila, 3, f["prh"] or "OTRA AUTORIDAD", centrado=False)
        _rescates_excel_celda(ws, fila, 4, f["rescates"])
        _rescates_excel_celda(ws, fila, 5, f["primera_vez"])
        _rescates_excel_celda(ws, fila, 6, f["reincidencias"])
        _rescates_excel_celda(ws, fila, 7, f["tipo"], centrado=False)
        fila += 1

    _rescates_excel_celda(ws, fila, 1, "TOTAL", bg="1F3864", color_texto="FFFFFF", negrita=True, centrado=False)
    ws.merge_cells(start_row=fila, start_column=1, end_row=fila, end_column=3)
    _rescates_excel_celda(ws, fila, 4, datos["total"], bg="1F3864", color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 5, datos["total_primera_vez"], bg="1F3864", color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 6, datos["total_reincidencias"], bg="1F3864", color_texto="FFFFFF", negrita=True)
    _rescates_excel_celda(ws, fila, 7, "", bg="1F3864")

    # @FADAR -- nombre de archivo pedido: REPORTE CECO DIA-MES-AÑO(PRHs).xlsx
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    nombre_archivo = f"REPORTE CECO {fecha_obj.day:02d}-{RESCATES_MESES_ES_LARGO[fecha_obj.month]}-{fecha_obj.year}(PRHs).xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


# =============================================================================
# Reportes "CECO V1" / "CECO V2" -- nacionalidad x 8 categorias (mismas
# RESCATES_ETIQUETAS_CATEGORIA de Informe Diario), un bloque por entidad (32)
# mas un bloque "Nacional", replicando el formato de un archivo de
# referencia entregado por el area. Confirmado en conversacion:
#  - V1 = TODOS los rescates del dia (primera vez + reincidentes juntos en
#    un solo total, tal como indica el nombre del archivo original
#    "rescates+reincidentes"), incluye inadmitidos.
#  - V2 = "rescates reales" = SOLO primera vez (excluye reincidentes),
#    tambien incluye inadmitidos (no se excluyeron explicitamente).
#  - Lista de nacionalidades: UNA sola, calculada a nivel nacional (las 32
#    entidades juntas) para ese dia -- se repite igual en los 32 bloques,
#    aunque una entidad en particular tenga 0 en alguna.
#  - El PDF no puede replicar el ancho de la hoja de Excel (331 columnas) en
#    una sola pagina -- se generan paginas separadas por entidad, aceptado
#    explicitamente ("no importa si sale cortado o segmentado").
# =============================================================================

def _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=False):
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    fecha_fmt = fecha_obj.strftime("%d-%m-%y")

    if solo_primera_vez:
        with connection.cursor() as cur:
            cur.execute(
                f'SELECT r."oficinaRepre", r.nacionalidad, r.sexo, r.edad, r."numFamilia", r.iso3 '
                f'FROM usuario_rescatepunto r '
                f'JOIN {RESCATES_MV_REINCIDENCIA} v '
                f'  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad '
                f"WHERE r.fecha = %s AND v.clasificacion = 'Rescate primera vez'",
                [fecha_fmt],
            )
            datos_dia = [
                {"oficinaRepre": of, "nacionalidad": nac, "sexo": sexo, "edad": edad, "numFamilia": nf, "iso3": iso3}
                for of, nac, sexo, edad, nf, iso3 in cur.fetchall()
            ]
    else:
        datos_dia = list(
            RescatePunto.objects.filter(fecha=fecha_fmt)
            .values("oficinaRepre", "nacionalidad", "sexo", "edad", "numFamilia", "iso3")
        )

    # Nacionalidades a nivel nacional (las 32 entidades juntas), ordenadas
    # por total descendente -- misma lista para los 32 bloques.
    conteo_nac = {}
    nac_iso3 = {}
    for d in datos_dia:
        conteo_nac[d["nacionalidad"]] = conteo_nac.get(d["nacionalidad"], 0) + 1
        nac_iso3.setdefault(d["nacionalidad"], d["iso3"])
    nacionalidades = [n for n, _ in sorted(conteo_nac.items(), key=lambda x: x[1], reverse=True)]
    # @FADAR -- nacionalidades extracontinentales (fuera de America), misma
    # regla que en el Informe Diario.
    nacionalidades_extracontinentales = {
        n for n, iso3 in nac_iso3.items() if _rescates_es_extracontinental(str(iso3).upper())
    }

    def _fila_vacia():
        fila = {clave: 0 for clave, _ in RESCATES_ETIQUETAS_CATEGORIA}
        fila["total"] = 0
        fila["nucleos"] = 0
        return fila

    bloques = {of: {nac: _fila_vacia() for nac in nacionalidades} for of in RESCATES_OFICINAS}
    totales_entidad = {of: _fila_vacia() for of in RESCATES_OFICINAS}
    nacional = {nac: _fila_vacia() for nac in nacionalidades}
    total_nacional = _fila_vacia()

    for d in datos_dia:
        of = d["oficinaRepre"]
        if of not in bloques:
            continue
        cat = _rescates_clasificar_categoria(d["sexo"], d["edad"], d["numFamilia"])
        nac = d["nacionalidad"]
        bloques[of][nac][cat] += 1
        bloques[of][nac]["total"] += 1
        totales_entidad[of][cat] += 1
        totales_entidad[of]["total"] += 1
        nacional[nac][cat] += 1
        nacional[nac]["total"] += 1
        total_nacional[cat] += 1
        total_nacional["total"] += 1

    # @FADAR -- "NUCLEOS FAMILIARES": columna adicional que se me habia
    # pasado en la primera revision del archivo de referencia. Misma
    # definicion ya usada en "Núcleos familiares detectados" del dashboard
    # (numFamilia>0, agrupado por oficina+hora+puntoEstra+numFamilia, con
    # >=2 integrantes) -- no se inventa una regla nueva.
    filtro_reinc_join = ""
    filtro_reinc_where = ""
    if solo_primera_vez:
        filtro_reinc_join = (
            f'JOIN {RESCATES_MV_REINCIDENCIA} v '
            f'  ON r.nombre = v.nombre AND r.apellidos = v.apellidos AND r.nacionalidad = v.nacionalidad '
        )
        filtro_reinc_where = " AND v.clasificacion = 'Rescate primera vez'"
    with connection.cursor() as cur:
        cur.execute(
            f'SELECT oficina, nacionalidad, COUNT(*) FROM ( '
            f'  SELECT r."oficinaRepre" AS oficina, r.nacionalidad, r.hora, r."puntoEstra" AS punto, r."numFamilia" AS nf '
            f'  FROM ( '
            f'    SELECT r.*, COUNT(*) OVER (PARTITION BY r."oficinaRepre", r.hora, r."puntoEstra", r."numFamilia") AS integrantes '
            f'    FROM usuario_rescatepunto r '
            f'    {filtro_reinc_join}'
            f'    WHERE r.fecha = %s AND r."numFamilia" > 0{filtro_reinc_where} '
            f'  ) r '
            f'  WHERE integrantes >= 2 '
            f'  GROUP BY oficina, r.nacionalidad, r.hora, punto, nf '
            f') grupos '
            f'GROUP BY oficina, nacionalidad',
            [fecha_fmt],
        )
        for of, nac, n in cur.fetchall():
            if of not in bloques or nac not in bloques[of]:
                continue
            bloques[of][nac]["nucleos"] += n
            totales_entidad[of]["nucleos"] += n
            nacional[nac]["nucleos"] += n
            total_nacional["nucleos"] += n

    # Version "lista" de los mismos datos, para el template PDF (Django no
    # permite bloques.of con "of" como variable de loop -- solo listas).
    def _lista_bloque(nombre_bloque, filas_por_nac, total_bloque):
        return {
            "nombre": nombre_bloque,
            "filas": [{"nacionalidad": nac, **filas_por_nac[nac]} for nac in nacionalidades],
            "total": total_bloque,
        }

    bloques_lista = [_lista_bloque(of, bloques[of], totales_entidad[of]) for of in RESCATES_OFICINAS]
    bloques_lista.append(_lista_bloque("NACIONAL POR DÍA", nacional, total_nacional))

    # @FADAR -- misma info "transpuesta" para la vista en pantalla en
    # horizontal (una sola tabla ancha con scroll lateral, igual que el
    # Excel, en vez de un bloque apilado por entidad): cada fila es una
    # nacionalidad, con los datos de los 33 bloques ya en orden.
    filas_horizontal = [
        {"nacionalidad": nac, "por_bloque": [b["filas"][j] for b in bloques_lista]}
        for j, nac in enumerate(nacionalidades)
    ]
    fila_total_horizontal = [b["total"] for b in bloques_lista]

    # @FADAR -- cuadro "SUCHIATE / RÍO BRAVO / CENTRO / TOTAL": se arma
    # sumando los mismos totales_entidad ya calculados arriba (que ya
    # respetan el filtro solo_primera_vez de este reporte), en vez de
    # llamar a _rescates_regiones() -- esa funcion aplica una excepcion de
    # Chiapas (100% reincidente) ajena a la definicion de este reporte, y
    # dejaba fuera sus rescates de primera vez (causaba que el cuadro no
    # cuadrara contra "NACIONAL POR DÍA").
    resumen_zonas = {
        "rio_bravo": sum(totales_entidad[of]["total"] for of in RESCATES_ZONA_RIO_BRAVO if of in totales_entidad),
        "centro": sum(totales_entidad[of]["total"] for of in RESCATES_ZONA_CENTRO if of in totales_entidad),
        "suchiate": sum(totales_entidad[of]["total"] for of in RESCATES_ZONA_SUCHIATE if of in totales_entidad),
    }
    resumen_zonas["total"] = resumen_zonas["rio_bravo"] + resumen_zonas["centro"] + resumen_zonas["suchiate"]

    return {
        "resumen_zonas": resumen_zonas,
        "fecha_actual": f"{fecha_obj.day:02d}/{fecha_obj.month:02d}/{fecha_obj.year}",
        "fecha_iso": fecha_str,
        "filas_horizontal": filas_horizontal,
        "fila_total_horizontal": fila_total_horizontal,
        "nacionalidades": nacionalidades,
        "nacionalidades_extracontinentales": nacionalidades_extracontinentales,
        "categorias": RESCATES_ETIQUETAS_CATEGORIA,
        "oficinas": RESCATES_OFICINAS,
        "bloques": bloques,
        "bloques_lista": bloques_lista,
        "totales_entidad": totales_entidad,
        "nacional": nacional,
        "total_nacional": total_nacional,
    }


# @FADAR -- colores/fuente extraidos EXACTOS del archivo de referencia
# (Rescatados_25AGO2026-CECO(rescates+reincidentes).xlsx, hoja "CECCO 2"),
# resueltos desde theme+tint a RGB real. La rotacion de 3 colores del
# encabezado de entidad en el original no sigue una regla identificable
# (parece asignacion manual, no ciclica) -- aqui se replica como una
# rotacion fija de esos mismos 3 colores, en vez de intentar adivinar el
# orden exacto sin base verificable.
RESCATES_CECOV_FUENTE = "Montserrat"
RESCATES_CECOV_VERDE = "70AD47"            # titulo "RESCATADOS" / bloque "NACIONAL POR DIA"
RESCATES_CECOV_AZUL = "4472C4"             # columna TOTAL
RESCATES_CECOV_NARANJA = "ED7D31"          # grupo "Mayores solos" / "Menores no acompañados"
RESCATES_CECOV_VERDE_OSCURO = "13322B"     # grupo "Mayores acompañan NNA" / "Menores acompañados"
RESCATES_CECOV_GRIS_CLARO = "E7E6E6"       # celda TOTAL por nacionalidad
RESCATES_CECOV_NEGRO = "000000"            # fila TOTAL de cada bloque
RESCATES_CECOV_ROTACION_ENTIDAD = ["808080", "ED7D31", "44546A"]


def _rescates_ceco_v_excel(datos, titulo, hoja_nombre="Hoja1"):
    # @FADAR -- estructura de encabezado EXACTA verificada celda por celda
    # contra el archivo de referencia (4 filas de encabezado, no 3; datos
    # empiezan en la fila 5; bloque de 10 columnas -- 9 categorias/total +
    # "NUCLEOS FAMILIARES" -- sin columna de separacion entre bloques).
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja_nombre

    ws.column_dimensions["A"].width = 30
    col = 2  # B
    indice_bloque = 0
    fila_maxima = 0

    def _escribir_bloque(nombre_bloque, filas_por_nac, total_bloque, es_nacional=False):
        nonlocal col, indice_bloque, fila_maxima
        col_inicio = col
        color_header = RESCATES_CECOV_VERDE if es_nacional else RESCATES_CECOV_ROTACION_ENTIDAD[indice_bloque % 3]

        # Fila 1: nombre de la entidad (9 columnas) + celda decorativa (10a).
        _rescates_excel_celda(ws, 1, col_inicio, nombre_bloque, bg=color_header, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE, tam=14 if es_nacional else 12)
        ws.merge_cells(start_row=1, start_column=col_inicio, end_row=1, end_column=col_inicio + 8)
        _rescates_excel_celda(ws, 1, col_inicio + 9, None, bg=color_header)

        # Filas 2-4: TOTAL (3 filas).
        _rescates_excel_celda(ws, 2, col_inicio, "TOTAL", bg=RESCATES_CECOV_AZUL, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=2, start_column=col_inicio, end_row=4, end_column=col_inicio)

        # "Mayores de edad solos" / "...que acompañan NNA": 2 filas x 2 columnas.
        for nombre_grupo, c, color_grupo in [
            ("MAYORES DE EDAD SOLOS", col_inicio + 1, RESCATES_CECOV_NARANJA),
            ("MAYORES DE EDAD QUE ACOMPAÑAN A NNA'S", col_inicio + 3, RESCATES_CECOV_VERDE_OSCURO),
        ]:
            _rescates_excel_celda(ws, 2, c, nombre_grupo, bg=color_grupo, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
            ws.merge_cells(start_row=2, start_column=c, end_row=3, end_column=c + 1)

        # "Menores de edad" (fila 2, 4 columnas) subdividido en fila 3:
        # "Acompañados" (verde oscuro) / "No acompañados" (naranja).
        _rescates_excel_celda(ws, 2, col_inicio + 5, "MENORES DE EDAD", bg=RESCATES_CECOV_VERDE_OSCURO, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=2, start_column=col_inicio + 5, end_row=2, end_column=col_inicio + 8)
        _rescates_excel_celda(ws, 3, col_inicio + 5, "ACOMPAÑADOS", bg=RESCATES_CECOV_VERDE_OSCURO, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=3, start_column=col_inicio + 5, end_row=3, end_column=col_inicio + 6)
        _rescates_excel_celda(ws, 3, col_inicio + 7, "NO ACOMPAÑADOS", bg=RESCATES_CECOV_NARANJA, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=3, start_column=col_inicio + 7, end_row=3, end_column=col_inicio + 8)

        # Fila 4: H/M de las 8 categorias, con el color de su grupo.
        colores_categoria = [RESCATES_CECOV_NARANJA] * 2 + [RESCATES_CECOV_VERDE_OSCURO] * 4 + [RESCATES_CECOV_NARANJA] * 2
        for i, (clave, _) in enumerate(RESCATES_ETIQUETAS_CATEGORIA):
            letra = "H" if i % 2 == 0 else "M"
            _rescates_excel_celda(ws, 4, col_inicio + 1 + i, letra, bg=colores_categoria[i], color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)

        # "NUCLEOS FAMILIARES": 3 filas (2-4), 1 columna, verde.
        _rescates_excel_celda(ws, 2, col_inicio + 9, "NUCLEOS FAMILIARES", bg=RESCATES_CECOV_VERDE, color_texto="FFFFFF", negrita=True, tam=8, fuente=RESCATES_CECOV_FUENTE)
        ws.merge_cells(start_row=2, start_column=col_inicio + 9, end_row=4, end_column=col_inicio + 9)

        # Datos: empiezan en la fila 5.
        f = 5
        for nac in datos["nacionalidades"]:
            fila_nac = filas_por_nac[nac]
            if col_inicio == 2:
                # @FADAR -- nacionalidades extracontinentales (fuera de America)
                if nac in datos["nacionalidades_extracontinentales"]:
                    _rescates_excel_celda(ws, f, col_inicio - 1, nac, bg="B45309", color_texto="FFFFFF", negrita=True, tam=12, centrado=False, fuente=RESCATES_CECOV_FUENTE)
                else:
                    _rescates_excel_celda(ws, f, col_inicio - 1, nac, negrita=True, tam=12, centrado=False, fuente=RESCATES_CECOV_FUENTE)
            _rescates_excel_celda(ws, f, col_inicio, fila_nac["total"], bg=RESCATES_CECOV_GRIS_CLARO, color_texto="000000", negrita=True, fuente=RESCATES_CECOV_FUENTE)
            for i, (clave, _) in enumerate(RESCATES_ETIQUETAS_CATEGORIA):
                _rescates_excel_celda(ws, f, col_inicio + 1 + i, fila_nac[clave], fuente=RESCATES_CECOV_FUENTE)
            _rescates_excel_celda(ws, f, col_inicio + 9, fila_nac["nucleos"], fuente=RESCATES_CECOV_FUENTE)
            f += 1
        if col_inicio == 2:
            _rescates_excel_celda(ws, f, col_inicio - 1, "TOTAL", negrita=True, tam=12, fuente=RESCATES_CECOV_FUENTE)
        _rescates_excel_celda(ws, f, col_inicio, total_bloque["total"], bg=RESCATES_CECOV_NEGRO, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE)
        for i, (clave, _) in enumerate(RESCATES_ETIQUETAS_CATEGORIA):
            _rescates_excel_celda(ws, f, col_inicio + 1 + i, total_bloque[clave], bg=RESCATES_CECOV_NEGRO, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE)
        _rescates_excel_celda(ws, f, col_inicio + 9, total_bloque["nucleos"], bg=RESCATES_CECOV_NEGRO, color_texto="FFFFFF", negrita=True, fuente=RESCATES_CECOV_FUENTE)
        fila_maxima = max(fila_maxima, f)

        for c in range(col_inicio, col_inicio + 10):
            ws.column_dimensions[get_column_letter(c)].width = 9
        col = col_inicio + 10  # 10 columnas de datos, sin separacion (verificado contra el original)
        indice_bloque += 1

    ws.merge_cells("A1:A2")
    _rescates_excel_celda(ws, 1, 1, "RESCATADOS", bg=RESCATES_CECOV_VERDE, color_texto="FFFFFF", negrita=True, tam=14, fuente=RESCATES_CECOV_FUENTE)
    ws.merge_cells("A3:A4")
    _rescates_excel_celda(ws, 3, 1, datos["fecha_actual"], bg=RESCATES_CECOV_VERDE, color_texto="FFFFFF", negrita=True, tam=14, fuente=RESCATES_CECOV_FUENTE)

    for of in datos["oficinas"]:
        _escribir_bloque(of, datos["bloques"][of], datos["totales_entidad"][of])
    _escribir_bloque("NACIONAL POR DÍA", datos["nacional"], datos["total_nacional"], es_nacional=True)

    # @FADAR -- cuadro "SUCHIATE / RÍO BRAVO / CENTRO / TOTAL", mismos
    # colores exactos del archivo de referencia (no los de RESCATES_ZONA_COLOR
    # del mapa -- este cuadro es tal como aparece en el original).
    fz = fila_maxima + 2
    rz = datos["resumen_zonas"]
    for etiqueta, valor, color, color_texto in [
        ("SUCHIATE", rz["suchiate"], "44546A", "FFFFFF"),
        ("RÍO BRAVO", rz["rio_bravo"], "ED7D31", "FFFFFF"),
        ("CENTRO", rz["centro"], "808080", "FFFFFF"),
        ("TOTAL", rz["total"], "70AD47", "44546A"),
    ]:
        _rescates_excel_celda(ws, fz, 1, etiqueta, bg=color, color_texto=color_texto, negrita=True, tam=18, centrado=False, fuente=RESCATES_CECOV_FUENTE)
        _rescates_excel_celda(ws, fz, 2, valor, bg=color, color_texto=color_texto, negrita=True, tam=18, fuente=RESCATES_CECOV_FUENTE)
        fz += 1

    # @FADAR -- "congelar paneles": columnas A+B (nacionalidad/etiqueta +
    # su valor) fijas al desplazarse por los 33 bloques de entidad.
    # Confirmado: congelar fila+columna a la vez ("B5") rompe la vista en
    # Google Sheets en esta hoja tan ancha (331 columnas) -- se congela
    # solo columnas, sin filas de encabezado (la hoja completa mide 28
    # filas, no hay mucho scroll vertical que perder). Se corta en C, no
    # en B, porque el cuadro SUCHIATE/RÍO BRAVO/CENTRO/TOTAL usa A para la
    # etiqueta y B para el numero -- cortar en B partia ese cuadro a la
    # mitad (una celda fija, la otra no).
    ws.freeze_panes = "C1"

    return wb


def rescates_reporte_cecov1(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=False)
    return render(request, "Reportes_Analisis/rescates_reporte_cecov.html", {**datos, "titulo": "CECO V1", "url_pdf": "Reportes_Analisis:rescates_reporte_cecov1_pdf", "url_excel": "Reportes_Analisis:rescates_reporte_cecov1_excel"})


def rescates_reporte_cecov1_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=False)
    template = get_template("Reportes_Analisis/_rescates_cecov_pdf.html")
    html_string = template.render({**datos, "titulo": "CECO V1 — Rescates + Reincidentes"})
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ceco_v1_{fecha_str}.pdf"'
    return response


def rescates_reporte_cecov1_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=False)
    wb = _rescates_ceco_v_excel(datos, "CECO V1 — Rescates + Reincidentes", "CECO V1")
    # @FADAR -- nombre pedido: Rescatados_DIAMESAÑO-CECO(rescates+reincidentes).xlsx
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    nombre_archivo = f"Rescatados_{fecha_obj.day:02d}{RESCATES_MESES_ES[fecha_obj.month].upper()}{fecha_obj.year}-CECO(rescates+reincidentes).xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


def rescates_reporte_cecov2(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=True)
    return render(request, "Reportes_Analisis/rescates_reporte_cecov.html", {**datos, "titulo": "CECO V2", "url_pdf": "Reportes_Analisis:rescates_reporte_cecov2_pdf", "url_excel": "Reportes_Analisis:rescates_reporte_cecov2_excel"})


def rescates_reporte_cecov2_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=True)
    template = get_template("Reportes_Analisis/_rescates_cecov_pdf.html")
    html_string = template.render({**datos, "titulo": "CECO V2 — Rescates Reales (primera vez)"})
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()
    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="ceco_v2_{fecha_str}.pdf"'
    return response


def rescates_reporte_cecov2_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    fecha_str = request.GET.get("fecha", date.today().isoformat())
    datos = _rescates_ceco_v_detalle(fecha_str, solo_primera_vez=True)
    wb = _rescates_ceco_v_excel(datos, "CECO V2 — Rescates Reales (primera vez)", "CECO V2")
    # @FADAR -- nombre pedido: Rescatados_DIAMESAÑO-CECO2(rescates_reales).xlsx
    fecha_obj = datetime.strptime(fecha_str, "%Y-%m-%d")
    nombre_archivo = f"Rescatados_{fecha_obj.day:02d}{RESCATES_MESES_ES[fecha_obj.month].upper()}{fecha_obj.year}-CECO2(rescates_reales).xlsx"
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


# @FADAR -- Fase 3 del traslado: Mexicanos vs Extranjeros. Nunca vivio en
# esta ubicacion (ni en mapa) -- traido completo desde el clon Desarrollo.
# Codigo copiado TAL CUAL (misma logica, mismas reglas, misma estructura).
def get_global_update_date():
    """Calcula el mínimo de los máximos para determinar la fecha de integridad total."""
    from mapa.models import Repatriados, Recibidos, ExtRescatados, Ingresos, Tramites, Retornados, Inadmitidos

    models_to_check = [
        Repatriados, Recibidos, ExtRescatados, Ingresos,
        Tramites, Retornados, Inadmitidos
    ]

    max_dates = []
    for model in models_to_check:
        res = model.objects.aggregate(max_f=Max('fecha'))['max_f']
        if res:
            max_dates.append(res)

    if not max_dates:
        return None

    # Retornamos el Mínimo de los Máximos (Integridad Total)
    return min(max_dates)

def _build_card_mex_ext(start, end, label):
    mex = Repatriados.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('mex_rep'))['total'] or 0
    ext = Recibidos.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('ext_rec'))['total'] or 0
    total = mex + ext
    p_mex = round((mex / total * 100)) if total > 0 else 0
    p_ext = 100 - p_mex if total > 0 else 0
    days = (end - start).days + 1
    avg = round(total / days) if days > 0 else 0
    if label == "Semana":
        fmt_periodo = f"del {start.strftime('%d')} al {end.strftime('%d de %b.')} de {end.year}"
    else:
        fmt_periodo = f"{start.strftime('%d/%b/%y')} a {end.strftime('%d/%b/%y')}"
    return {
        'total': total, 'total_fmt': f"{total:,}",
        'promedio': f"{avg:,}", 'periodo': fmt_periodo, 'label': label,
        'mex': mex, 'ext': ext, 'mex_fmt': f"{mex:,}", 'ext_fmt': f"{ext:,}",
        'p_mex': p_mex, 'p_ext': p_ext,
    }


def _top10_mex_ext(start, end):
    res_rec = Recibidos.objects.filter(fecha__range=[start, end]) \
        .values('nacionalidad__nombre').annotate(total=Sum('ext_rec')).order_by('-total')[:10]
    mex_total = Repatriados.objects.filter(fecha__range=[start, end]).aggregate(total=Sum('mex_rep'))['total'] or 0
    lista = [{'n': 'MÉXICO', 'v': int(mex_total)}]
    for item in res_rec:
        lista.append({'n': item['nacionalidad__nombre'] or 'DESCONOCIDA', 'v': int(item['total'] or 0)})
    return sorted(lista, key=lambda x: x['v'], reverse=True)[:10]


# @FADAR -- entidades tal como aparecen en FORMATO_MEX-.xlsm ("REPORTE
# GENERAL"), no las 32. Los puntos (12 ciudades fijas) ya no se duplican
# aqui -- vienen de CATALOGO_PUNTOS_MEX_EXT (mapa/models.py), unica fuente
# compartida con RegistroMexExtPunto.
RESCATES_ESTADOS_MEX_EXT = list(CATALOGO_PUNTOS_MEX_EXT.keys())


def _puntos_por_estado():
    return CATALOGO_PUNTOS_MEX_EXT


# @FADAR -- captura real por punto (RegistroMexExtPunto). Hombres+Mujeres
# se suman como "Adultos" y Ninos+Ninas como "Menores" para que la fila del
# punto use las mismas columnas que la fila del estado (Adultos/Menores/
# Total) -- el detalle fino (H/M/Ninos/Ninas) sigue intacto en la tabla,
# solo se resume asi para esta tabulacion.
def _capturas_por_punto(start, end):
    filas = RegistroMexExtPunto.objects.filter(fecha__range=[start, end], estado__in=RESCATES_ESTADOS_MEX_EXT) \
        .values('estado', 'punto', 'categoria') \
        .annotate(hombres=Sum('hombres'), mujeres=Sum('mujeres'), ninos=Sum('ninos'), ninas=Sum('ninas'))
    resultado = {}
    for f in filas:
        clave = (f['estado'], f['punto'])
        adultos = (f['hombres'] or 0) + (f['mujeres'] or 0)
        menores = (f['ninos'] or 0) + (f['ninas'] or 0)
        resultado.setdefault(clave, {})[f['categoria']] = {'ad': adultos, 'me': menores, 'total': adultos + menores}
    return resultado


def _tabla_mex_ext_por_estado(start, end):
    rep = Repatriados.objects.filter(fecha__range=[start, end], estado__nombre__in=RESCATES_ESTADOS_MEX_EXT) \
        .values('estado__nombre').annotate(mex=Sum('mex_rep'), mex_ad=Sum('adultos'), mex_me=Sum('menores'))
    rec = Recibidos.objects.filter(fecha__range=[start, end], estado__nombre__in=RESCATES_ESTADOS_MEX_EXT) \
        .values('estado__nombre').annotate(ext=Sum('ext_rec'), ext_ad=Sum('adultos'), ext_me=Sum('menores'))
    filas = {e: {'estado': e, 'mex': 0, 'mex_ad': 0, 'mex_me': 0, 'ext': 0, 'ext_ad': 0, 'ext_me': 0} for e in RESCATES_ESTADOS_MEX_EXT}
    for r in rep:
        e = r['estado__nombre']
        filas[e]['mex'] = r['mex'] or 0
        filas[e]['mex_ad'] = r['mex_ad'] or 0
        filas[e]['mex_me'] = r['mex_me'] or 0
    for r in rec:
        e = r['estado__nombre']
        filas[e]['ext'] = r['ext'] or 0
        filas[e]['ext_ad'] = r['ext_ad'] or 0
        filas[e]['ext_me'] = r['ext_me'] or 0

    catalogo = _puntos_por_estado()
    capturas = _capturas_por_punto(start, end)
    for f in filas.values():
        f['total'] = f['mex'] + f['ext']
        puntos_estado = []
        for nombre in catalogo.get(f['estado'], []):
            dato = capturas.get((f['estado'], nombre))
            mex_p = dato.get('MEX') if dato else None
            ext_p = dato.get('EXT') if dato else None
            puntos_estado.append({
                'nombre': nombre,
                'con_datos': dato is not None,
                'mex_ad': mex_p['ad'] if mex_p else None, 'mex_me': mex_p['me'] if mex_p else None, 'mex': mex_p['total'] if mex_p else None,
                'ext_ad': ext_p['ad'] if ext_p else None, 'ext_me': ext_p['me'] if ext_p else None, 'ext': ext_p['total'] if ext_p else None,
                'total': (mex_p['total'] if mex_p else 0) + (ext_p['total'] if ext_p else 0) if dato else None,
            })
        f['puntos'] = puntos_estado
    return sorted(filas.values(), key=lambda f: f['total'], reverse=True)


def _tabla_mex_ext_nacionalidades(start, end, limite=15):
    rec = Recibidos.objects.filter(fecha__range=[start, end], estado__nombre__in=RESCATES_ESTADOS_MEX_EXT) \
        .values('nacionalidad__nombre').annotate(ext=Sum('ext_rec'), ext_ad=Sum('adultos'), ext_me=Sum('menores')).order_by('-ext')[:limite]
    return [
        {'nacionalidad': r['nacionalidad__nombre'] or 'DESCONOCIDA', 'ext': r['ext'] or 0,
         'ext_ad': r['ext_ad'] or 0, 'ext_me': r['ext_me'] or 0}
        for r in rec
    ]


# @FADAR
def _sumar_filas_mex_ext(filas, campos):
    total = {c: 0 for c in campos}
    for f in filas:
        for c in campos:
            total[c] += f[c]
    return total


def _leer_rango_mex_ext(request, fecha_max):
    fecha_inicio = request.GET.get('fecha_inicio', (fecha_max - timedelta(days=6)).isoformat())
    fecha_fin = request.GET.get('fecha_fin', fecha_max.isoformat())
    fi = datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
    ff = datetime.strptime(fecha_fin, '%Y-%m-%d').date()
    return fecha_inicio, fecha_fin, fi, ff


# @FADAR
def _datos_mex_extranjeros():
    fecha_max = get_global_update_date() or date.today()
    CSP_START = date(2024, 10, 1)
    TRUMP_START = date(2025, 1, 20)
    SEMANA_START = fecha_max - timedelta(days=6)
    INICIO_2026 = date(2026, 1, 1)

    card_semana = _build_card_mex_ext(SEMANA_START, fecha_max, "Semana")
    card_csp = _build_card_mex_ext(CSP_START, fecha_max, "CSP")
    card_trump = _build_card_mex_ext(TRUMP_START, fecha_max, "Trump")

    # Evolucion diaria (desde CSP_START) -- mex_rep + ext_rec combinados.
    rep_qs = Repatriados.objects.filter(fecha__range=[CSP_START, fecha_max]) \
        .annotate(day=TruncDay('fecha')).values('day').annotate(total=Sum('mex_rep'))
    rec_qs = Recibidos.objects.filter(fecha__range=[CSP_START, fecha_max]) \
        .annotate(day=TruncDay('fecha')).values('day').annotate(total=Sum('ext_rec'))
    combined = {}
    for item in rep_qs:
        combined[item['day']] = item['total']
    for item in rec_qs:
        combined[item['day']] = combined.get(item['day'], 0) + item['total']
    evolucion_diaria = [(d, combined[d]) for d in sorted(combined.keys())]

    # Evolucion semanal 2026.
    rep_w = Repatriados.objects.filter(fecha__range=[INICIO_2026, fecha_max]) \
        .annotate(week=TruncWeek('fecha')).values('week').annotate(total=Sum('mex_rep'))
    rec_w = Recibidos.objects.filter(fecha__range=[INICIO_2026, fecha_max]) \
        .annotate(week=TruncWeek('fecha')).values('week').annotate(total=Sum('ext_rec'))
    combined_w = {}
    for item in rep_w:
        combined_w[item['week']] = item['total']
    for item in rec_w:
        combined_w[item['week']] = combined_w.get(item['week'], 0) + item['total']
    evolucion_semanal = [(w, combined_w[w]) for w in sorted(combined_w.keys())]
    total_2026 = sum(t for _, t in evolucion_semanal)

    start_top = max(CSP_START, fecha_max - timedelta(days=180))
    top_nacionalidades = _top10_mex_ext(start_top, fecha_max)

    return {
        'fecha_max': fecha_max,
        'card_semana': card_semana, 'card_csp': card_csp, 'card_trump': card_trump,
        'evolucion_diaria': evolucion_diaria,
        'evolucion_semanal': evolucion_semanal,
        'total_2026': total_2026,
        'top_nacionalidades': top_nacionalidades,
        'start_top': start_top,
    }


def reporte_mex_extranjeros(request):
    """Vista previa en pantalla -- version descargable del rubro 'Recibidos'
    del tablero /mapa/reportes (que no tiene boton de descarga)."""
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    datos = _datos_mex_extranjeros()
    fecha_inicio, fecha_fin, fi, ff = _leer_rango_mex_ext(request, datos['fecha_max'])
    card_personalizado = _build_card_mex_ext(fi, ff, "Personalizado")
    top_personalizado = _top10_mex_ext(fi, ff)
    tabla_estados = _tabla_mex_ext_por_estado(fi, ff)
    tabla_nacionalidades = _tabla_mex_ext_nacionalidades(fi, ff)
    total_estados = _sumar_filas_mex_ext(tabla_estados, ['mex_ad', 'mex_me', 'mex', 'ext_ad', 'ext_me', 'ext', 'total'])
    total_nacionalidades = _sumar_filas_mex_ext(tabla_nacionalidades, ['ext_ad', 'ext_me', 'ext'])

    # --- Grafica de evolucion diaria (linea + navegador) ---
    x_data = [datetime.combine(d, datetime.min.time()) for d, _ in datos['evolucion_diaria']]
    y_data = [t for _, t in datos['evolucion_diaria']]
    source_bar = ColumnDataSource(data=dict(x=x_data, y=y_data))
    fecha_max_dt = datetime.combine(datos['fecha_max'], datetime.min.time())
    rango_inicial = datetime.combine(datos['start_top'], datetime.min.time())
    p1 = figure(
        height=300, sizing_mode="stretch_width", x_axis_type="datetime", x_axis_location="above",
        x_range=(rango_inicial, fecha_max_dt), toolbar_location="right", tools="pan,box_zoom,xwheel_zoom,reset",
        background_fill_color="#efefef", border_fill_color=None, outline_line_color="#666666",
        title="Evolución diaria — Mexicanos repatriados + Extranjeros recibidos",
    )
    p1.y_range.start = 0
    p1.line(x='x', y='y', line_width=2, color="#285C4D", source=source_bar)
    p1.xgrid.grid_line_color = "#ffffff"
    p1.ygrid.grid_line_color = "#ffffff"
    p1.yaxis.formatter = NumeralTickFormatter(format="0a")
    p1.xaxis.formatter = DatetimeTickFormatter(days="%d %b", months="%b %Y", years="%Y")
    p1.add_tools(HoverTool(tooltips=[("Fecha", "@x{%d/%b/%y}"), ("Total", "@y{0,0}")], formatters={'@x': 'datetime'}))

    select = figure(
        title="Arrastra el recuadro para navegar por el tiempo", height=100, sizing_mode="stretch_width",
        x_axis_type="datetime", y_axis_type=None, tools="", toolbar_location=None,
        background_fill_color="#f9f9f9", outline_line_color="#e5e7eb",
    )
    select.line(x='x', y='y', color="#285C4D", alpha=0.5, source=source_bar)
    select.ygrid.grid_line_color = None
    select.xgrid.grid_line_color = None
    range_tool = RangeTool(x_range=p1.x_range)
    range_tool.overlay.fill_color = "#285C4D"
    range_tool.overlay.fill_alpha = 0.2
    select.add_tools(range_tool)
    layout_p1 = column(p1, select, sizing_mode="stretch_width")

    # --- Grafica semanal 2026 ---
    x_line = [w.strftime('%d-%m') for w, _ in datos['evolucion_semanal']]
    y_line = [t for _, t in datos['evolucion_semanal']]
    source_line = ColumnDataSource(data=dict(x=x_line, y=y_line))
    p2_args = dict(height=300, sizing_mode="stretch_width", toolbar_location=None, tools="",
                    background_fill_color=None, border_fill_color=None, outline_line_color=None,
                    title="Evolución semanal 2026")
    if x_line:
        p2_args['x_range'] = x_line
    p2 = figure(**p2_args)
    p2.line(x='x', y='y', line_width=2, color="#86895D", source=source_line)
    p2.scatter(x='x', y='y', size=8, color="#86895D", fill_color="white", source=source_line)
    p2.xgrid.grid_line_color = None
    p2.ygrid.grid_line_color = None
    p2.yaxis.visible = False
    p2.xaxis.major_label_orientation = 1.5708
    p2.xaxis.major_label_text_font_size = "7pt"
    p2.add_tools(HoverTool(tooltips=[("Semana", "@x"), ("Total", "@y{0,0}")]))

    # --- Grafica Top 10 nacionalidades (Mexico incluido, rango filtrado) ---
    top_names = [d['n'] for d in top_personalizado][::-1]
    top_values = [d['v'] for d in top_personalizado][::-1]
    source_top = ColumnDataSource(data=dict(names=top_names, values=top_values))
    p_top = figure(y_range=top_names, height=380, toolbar_location=None, tools="", sizing_mode="stretch_width",
                   title=f"Top 10 nacionalidades (México incluido) — {fecha_inicio} a {fecha_fin}")
    p_top.hbar(y='names', right='values', height=0.7, color="#285C4D", source=source_top)
    p_top.x_range.start = 0
    p_top.xaxis.formatter = NumeralTickFormatter(format="0a")
    p_top.grid.grid_line_color = None
    p_top.add_tools(HoverTool(tooltips=[("País", "@names"), ("Total", "@values{0,0}")]))

    plot_script, plot_divs = components((layout_p1, p2, p_top))
    plot_bar_div, plot_line_div, plot_top_div = plot_divs

    context = {
        'card_semana': datos['card_semana'], 'card_csp': datos['card_csp'], 'card_trump': datos['card_trump'],
        'card_personalizado': card_personalizado,
        'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin,
        'total_2026': f"{datos['total_2026']:,}",
        'plot_script': plot_script, 'plot_bar_div': plot_bar_div,
        'plot_line_div': plot_line_div, 'plot_top_div': plot_top_div,
        'tabla_estados': tabla_estados, 'tabla_nacionalidades': tabla_nacionalidades,
        'total_estados': total_estados, 'total_nacionalidades': total_nacionalidades,
    }
    return render(request, 'Reportes_Analisis/reporte_mex_extranjeros.html', context)


def reporte_mex_extranjeros_pdf(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    fecha_max = get_global_update_date() or date.today()
    fecha_inicio, fecha_fin, fi, ff = _leer_rango_mex_ext(request, fecha_max)
    tabla_estados = _tabla_mex_ext_por_estado(fi, ff)
    tabla_nacionalidades = _tabla_mex_ext_nacionalidades(fi, ff)
    context = {
        'fecha_inicio': fecha_inicio, 'fecha_fin': fecha_fin,
        'tabla_estados': tabla_estados, 'tabla_nacionalidades': tabla_nacionalidades,
        'total_estados': _sumar_filas_mex_ext(tabla_estados, ['mex_ad', 'mex_me', 'mex', 'ext_ad', 'ext_me', 'ext', 'total']),
        'total_nacionalidades': _sumar_filas_mex_ext(tabla_nacionalidades, ['ext_ad', 'ext_me', 'ext']),
    }
    template = get_template("Reportes_Analisis/_reporte_mex_extranjeros_pdf.html")
    html_string = template.render(context)
    pdf_file = HTML(string=html_string, base_url=request.build_absolute_uri()).write_pdf()

    response = HttpResponse(pdf_file, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="mexicanos_extranjeros_{fecha_max.isoformat()}.pdf"'
    return response


def reporte_mex_extranjeros_excel(request):
    if not request.user.is_authenticated:
        return redirect('/log-in/?next=%s' % request.path)
    if not request.user.is_superuser:
        return render(request, 'base/error404.html')

    datos = _datos_mex_extranjeros()
    fecha_inicio, fecha_fin, fi, ff = _leer_rango_mex_ext(request, datos['fecha_max'])
    card_personalizado = _build_card_mex_ext(fi, ff, "Personalizado")
    top_personalizado = _top10_mex_ext(fi, ff)
    tabla_estados = _tabla_mex_ext_por_estado(fi, ff)
    tabla_nacionalidades = _tabla_mex_ext_nacionalidades(fi, ff)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Resumen"
    ws1.merge_cells("A1:E1")
    _rescates_excel_celda(ws1, 1, 1, "Mexicanos Repatriados vs Extranjeros Recibidos", negrita=True, tam=13)
    fila = 3
    _rescates_excel_fila(ws1, fila, ["Periodo", "Rango", "Mexicanos", "Extranjeros", "Total", "% Mex", "% Ext"][:5], RESCATES_LETRA_T1, negrita=True)
    _rescates_excel_celda(ws1, fila, 6, "% Mex", bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], negrita=True)
    _rescates_excel_celda(ws1, fila, 7, "% Ext", bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], negrita=True)
    fila += 1
    for card in (datos['card_semana'], datos['card_csp'], datos['card_trump'], card_personalizado):
        _rescates_excel_fila(ws1, fila, [card['label'], card['periodo'], card['mex'], card['ext'], card['total'], f"{card['p_mex']}%", f"{card['p_ext']}%"], RESCATES_LETRA_T2)
        fila += 1
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 30
    for col in ("C", "D", "E", "F", "G"):
        ws1.column_dimensions[col].width = 14

    ws3 = wb.create_sheet("Evolución diaria")
    _rescates_excel_fila(ws3, 1, ["Fecha", "Total (Mex + Ext)"], RESCATES_LETRA_T1, negrita=True)
    f = 2
    for d, t in datos['evolucion_diaria']:
        _rescates_excel_fila(ws3, f, [d.strftime('%d/%m/%Y'), t], RESCATES_LETRA_T2)
        f += 1
    ws3.column_dimensions["A"].width = 14

    ws4 = wb.create_sheet(f"Top 10 ({fecha_inicio} a {fecha_fin})"[:31])
    _rescates_excel_fila(ws4, 1, ["Nacionalidad", "Total"], RESCATES_LETRA_T1, negrita=True)
    f = 2
    for item in top_personalizado:
        _rescates_excel_celda(ws4, f, 1, item['n'], bg=RESCATES_LETRA_T1[0], color_texto=RESCATES_LETRA_T1[1], centrado=False)
        _rescates_excel_celda(ws4, f, 2, item['v'], bg=RESCATES_LETRA_T0[0], color_texto=RESCATES_LETRA_T0[1], negrita=True)
        f += 1
    ws4.column_dimensions["A"].width = 28

    # Hojas "Reporte General"/"Nacionalidades" -- mismo formato visual exacto
    # que FORMATO_MEX-.xlsm (negro/blanco en titulos y encabezados, gris en
    # filas de dato, verde #235B4E en el total), tabulado al nivel real
    # disponible (estado, no punto; adultos/menores, no H/M/ninos/ninas).
    NEGRO, BLANCO, VERDE, GRIS = "000000", "FFFFFF", "235B4E", "C9B182"

    ws5 = wb.create_sheet("Reporte General")
    ws5.merge_cells("A1:H1")
    _rescates_excel_celda(ws5, 1, 1, "MEXICANOS - EXTRANJEROS RECIBIDOS", bg=NEGRO, color_texto=BLANCO, tam=20, centrado=True)
    ws5.merge_cells("A2:H2")
    _rescates_excel_celda(ws5, 2, 1, f"{fecha_inicio} a {fecha_fin}", tam=12, centrado=True)

    ws5.merge_cells("A4:A5")
    _rescates_excel_celda(ws5, 4, 1, "ESTADO", bg=NEGRO, color_texto=BLANCO, tam=14, negrita=True)
    ws5.merge_cells("B4:D4")
    _rescates_excel_celda(ws5, 4, 2, "MEXICANOS", bg=NEGRO, color_texto=BLANCO)
    ws5.merge_cells("E4:G4")
    _rescates_excel_celda(ws5, 4, 5, "EXTRANJEROS", bg=NEGRO, color_texto=BLANCO)
    ws5.merge_cells("H4:H5")
    _rescates_excel_celda(ws5, 4, 8, "TOTAL GENERAL", bg=NEGRO, color_texto=BLANCO, negrita=True)
    _rescates_excel_fila(ws5, 5, ["ADULTOS", "MENORES", "TOTAL", "ADULTOS", "MENORES", "TOTAL"], (None, "000000"), col_inicial=2)

    f = 6
    total_gral = {'mex_ad': 0, 'mex_me': 0, 'mex': 0, 'ext_ad': 0, 'ext_me': 0, 'ext': 0, 'total': 0}
    for e in tabla_estados:
        _rescates_excel_celda(ws5, f, 1, e['estado'], bg=GRIS)
        for i, campo in enumerate(('mex_ad', 'mex_me', 'mex', 'ext_ad', 'ext_me', 'ext', 'total')):
            _rescates_excel_celda(ws5, f, 2 + i, e[campo], bg=GRIS)
            total_gral[campo] = total_gral.get(campo, 0) + e[campo]
        f += 1
        for punto in e.get('puntos', []):
            _rescates_excel_celda(ws5, f, 1, punto['nombre'], color_texto="6B7280", tam=9, centrado=False)
            if punto['con_datos']:
                valores_punto = (punto['mex_ad'], punto['mex_me'], punto['mex'], punto['ext_ad'], punto['ext_me'], punto['ext'], punto['total'])
                for i, v in enumerate(valores_punto):
                    _rescates_excel_celda(ws5, f, 2 + i, v, color_texto="6B7280", tam=9)
            else:
                for col in range(2, 9):
                    _rescates_excel_celda(ws5, f, col, "—", color_texto="6B7280", tam=9)
            f += 1
    _rescates_excel_celda(ws5, f, 1, "TOTAL", bg=VERDE, color_texto=BLANCO, negrita=True)
    for i, campo in enumerate(('mex_ad', 'mex_me', 'mex', 'ext_ad', 'ext_me', 'ext', 'total')):
        _rescates_excel_celda(ws5, f, 2 + i, total_gral[campo], bg=VERDE, color_texto=BLANCO, negrita=True)
    ws5.column_dimensions["A"].width = 22

    ws6 = wb.create_sheet("Nacionalidades")
    ws6.merge_cells("A1:D1")
    _rescates_excel_celda(ws6, 1, 1, "NACIONALIDADES", bg=NEGRO, color_texto=BLANCO, tam=20, centrado=True)
    ws6.merge_cells("A2:D2")
    _rescates_excel_celda(ws6, 2, 1, f"{fecha_inicio} a {fecha_fin}", tam=12, centrado=True)
    ws6.merge_cells("A4:A5")
    _rescates_excel_celda(ws6, 4, 1, "PAÍS", bg=NEGRO, color_texto=BLANCO, tam=14, negrita=True)
    ws6.merge_cells("B4:D4")
    _rescates_excel_celda(ws6, 4, 2, "EXTRANJEROS", bg=NEGRO, color_texto=BLANCO)
    _rescates_excel_fila(ws6, 5, ["ADULTOS", "MENORES", "TOTAL"], (None, "000000"), col_inicial=2)
    f = 6
    total_nac = {'ext_ad': 0, 'ext_me': 0, 'ext': 0}
    for n in tabla_nacionalidades:
        _rescates_excel_celda(ws6, f, 1, n['nacionalidad'], bg=GRIS)
        _rescates_excel_celda(ws6, f, 2, n['ext_ad'], bg=GRIS)
        _rescates_excel_celda(ws6, f, 3, n['ext_me'], bg=GRIS)
        _rescates_excel_celda(ws6, f, 4, n['ext'], bg=GRIS)
        total_nac['ext_ad'] += n['ext_ad']
        total_nac['ext_me'] += n['ext_me']
        total_nac['ext'] += n['ext']
        f += 1
    _rescates_excel_celda(ws6, f, 1, "TOTAL", bg=VERDE, color_texto=BLANCO, negrita=True)
    _rescates_excel_celda(ws6, f, 2, total_nac['ext_ad'], bg=VERDE, color_texto=BLANCO, negrita=True)
    _rescates_excel_celda(ws6, f, 3, total_nac['ext_me'], bg=VERDE, color_texto=BLANCO, negrita=True)
    _rescates_excel_celda(ws6, f, 4, total_nac['ext'], bg=VERDE, color_texto=BLANCO, negrita=True)
    ws6.column_dimensions["A"].width = 22

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="mexicanos_extranjeros_{datos["fecha_max"].isoformat()}.xlsx"'
    wb.save(response)
    return response
