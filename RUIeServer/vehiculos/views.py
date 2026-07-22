"""
Vistas de solo lectura sobre los datos reales de VehiculosOR (definido en
la app 'mapa'). No se crea, edita ni elimina nada aquí — ese trabajo ya lo
hace mapa/views.py (guardar_vehiculo, guardar_kilometraje, etc.). Este
módulo es una capa de consulta/visualización separada, pensada para un
público distinto (lectura, sin permisos de gestión).

Nada de esto usa un CSV ni una base de datos propia — todo viene en vivo
de las mismas tablas que ya administra 'mapa'.
"""

from collections import Counter
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from django.http import Http404, HttpResponse, JsonResponse
from django.shortcuts import render
from django.views.decorators.http import require_GET

from mapa.models import (
    Capufe,
    CombustibleExt,
    Estado,
    Kilometraje,
    PrestadoDe,
    Siniestros,
    SituacionVeh,
    TipoAsignacionVeh,
    TipoVeh,
    VehiculosOR,
)


# --- Helpers de conversión (VehiculosOR -> dict amigable para templates) --

def _color_situacion(situacion_nombre):
    """Color consistente en todo el módulo: verde para Activo, amarillo
    para Mantenimiento, rojo para cualquier otra cosa (ej. Posible baja,
    o valores no reconocidos). Centralizado aquí para no repetir la
    lógica de comparación de texto en cada template por separado."""
    nombre = (situacion_nombre or "").strip().upper()
    if nombre == "ACTIVO":
        return "activo"
    if nombre == "MANTENIMIENTO":
        return "mantenimiento"
    return "baja"


def _vehiculo_a_dict(v):
    """Convierte una instancia de VehiculosOR a un dict con nombres
    estables, para no acoplar los templates a los nombres de campo reales
    (que pueden tener sus propias particularidades, ej. 'tipoVeh')."""
    situacion_nombre = v.situacion.situacion if v.situacion_id else "Sin especificar"
    return {
        "id": v.id,
        "placa": v.placa,
        "marca": v.marca,
        "modelo": v.modelo,
        "anio": v.anio.year if v.anio else "",
        "no_motor": v.no_motor,
        "tipo_vehiculo": v.tipoVeh.tipo_veh if v.tipoVeh else "Sin especificar",
        "tipo_asignacion": v.asignacion.tipo if v.asignacion_id else "Sin especificar",
        "situacion": situacion_nombre,
        "situacion_color": _color_situacion(situacion_nombre),
        "estado": v.estado.nombre if v.estado_id else "",
        "inmueble_destino": v.inmueble.nombre_inmueble if v.inmueble_id else "Sin asignar",
        "tarjeta": v.tarjeta_asig or "",
    }


def _queryset_base():
    return VehiculosOR.objects.select_related("tipoVeh", "asignacion", "situacion", "estado", "inmueble")


def _es_activo(v):
    return bool(v.situacion_id) and v.situacion.situacion.strip().upper() == "ACTIVO"


# --- Filtros compartidos (listado, popover, excel) ----------------------

def _aplicar_filtros(qs, request):
    tipo = request.GET.get("tipo", "").strip()
    situacion = request.GET.get("situacion", "").strip()
    asignacion = request.GET.get("asignacion", "").strip()
    estado = request.GET.get("estado", "").strip()
    placa = request.GET.get("placa", "").strip()

    if tipo:
        qs = qs.filter(tipoVeh__tipo_veh__iexact=tipo)
    if situacion:
        qs = qs.filter(situacion__situacion__iexact=situacion)
    if asignacion:
        qs = qs.filter(asignacion__tipo__iexact=asignacion)
    if estado:
        qs = qs.filter(estado__nombre__iexact=estado)
    if placa:
        qs = qs.filter(placa__icontains=placa)
    return qs


# --- Vista general (dashboard) -------------------------------------------

def dashboard(request):
    filas = list(_queryset_base())
    total = len(filas)
    activos = sum(1 for v in filas if _es_activo(v))
    mantenimiento = sum(1 for v in filas if _color_situacion(v.situacion.situacion if v.situacion_id else "") == "mantenimiento")
    posible_baja = total - activos - mantenimiento

    # SituacionVeh es texto libre en la base real (sin catálogo fijo de 3
    # valores) — se cuenta todo lo que NO sea "Activo" como un solo grupo,
    # y se muestra el desglose real de esas otras situaciones aparte.
    otras_situaciones = {}
    for v in filas:
        if not _es_activo(v):
            nombre = v.situacion.situacion if v.situacion_id else "Sin especificar"
            otras_situaciones[nombre] = otras_situaciones.get(nombre, 0) + 1

    tarjetas_asignadas = sum(1 for v in filas if v.tarjeta_asig)

    monto_combustible = sum((c.monto for c in CombustibleExt.objects.only("monto")), start=0)

    # "Total kilómetros" = suma de la lectura MÁS RECIENTE de cada
    # vehículo (no la suma de todo el historial, que sobrecontaría).
    ultima_lectura_por_vehiculo = {}
    for k in Kilometraje.objects.order_by("vehiculo_id", "-fecha").only("vehiculo_id", "odometro"):
        if k.vehiculo_id not in ultima_lectura_por_vehiculo:
            ultima_lectura_por_vehiculo[k.vehiculo_id] = k.odometro
    total_km = sum(ultima_lectura_por_vehiculo.values(), start=0)

    conteo_tipos = Counter(v.tipoVeh.tipo_veh for v in filas if v.tipoVeh_id)
    tipos_catalogo = TipoVeh.objects.all().order_by("tipo_veh")
    tipos_vehiculo = [
        {"nombre": t.tipo_veh, "cantidad": conteo_tipos.get(t.tipo_veh, 0)} for t in tipos_catalogo
    ]

    # Igual que con situación: TipoAsignacionVeh es texto libre en la base
    # real, así que no asumimos que solo existen "Propio"/"Arrendado".
    asignacion_conteo = {}
    for v in filas:
        if v.asignacion_id:
            nombre = v.asignacion.tipo
            asignacion_conteo[nombre] = asignacion_conteo.get(nombre, 0) + 1

    context = {
        "resumen_estado": {
            "activos": activos,
            "mantenimiento": mantenimiento,
            "posible_baja": posible_baja,
            "total": total,
            "otras_situaciones": otras_situaciones,
        },
        "modelos_distintos": len({v.modelo for v in filas if v.modelo}),
        "asignacion_conteo": asignacion_conteo,
        "tarjetas_asignadas": tarjetas_asignadas,
        "monto_combustible": monto_combustible,
        "total_kilometros": total_km,
        "tipos_vehiculo": tipos_vehiculo,
        "situacion_opciones": sorted({v.situacion.situacion for v in filas if v.situacion_id}),
        "asignacion_opciones": sorted(asignacion_conteo.keys()),
        "estado_opciones": sorted({v.estado.nombre for v in filas if v.estado_id}),
    }
    return render(request, "vehiculos/dashboard.html", context)


@require_GET
def filtrar_dashboard(request):
    qs = _aplicar_filtros(_queryset_base(), request)
    conteo_tipos = Counter(v.tipoVeh.tipo_veh for v in qs if v.tipoVeh_id)
    tipos_catalogo = TipoVeh.objects.all().order_by("tipo_veh")
    tipos_vehiculo = [
        {"nombre": t.tipo_veh, "cantidad": conteo_tipos.get(t.tipo_veh, 0)} for t in tipos_catalogo
    ]
    if request.GET.get("solo_con_unidades") == "1":
        tipos_vehiculo = [t for t in tipos_vehiculo if t["cantidad"] > 0]
    return render(request, "vehiculos/_tipos_grid.html", {"tipos_vehiculo": tipos_vehiculo})


# --- Listado filtrable -----------------------------------------------------

def listado(request):
    context = {
        "vehiculos": [],  # se llena vía AJAX al cargar, igual que el resto
        "tipos_opciones": TipoVeh.objects.all().order_by("tipo_veh").values_list("tipo_veh", flat=True),
        "situacion_opciones": SituacionVeh.objects.all().order_by("situacion").values_list("situacion", flat=True),
        "asignacion_opciones": TipoAsignacionVeh.objects.all().order_by("tipo").values_list("tipo", flat=True),
        "estado_opciones": Estado.objects.all().order_by("nombre").values_list("nombre", flat=True),
    }
    return render(request, "vehiculos/listado.html", context)


@require_GET
def filtrar_listado(request):
    qs = _aplicar_filtros(_queryset_base(), request).order_by("estado__nombre", "marca", "modelo")
    vehiculos = [_vehiculo_a_dict(v) for v in qs]
    return render(request, "vehiculos/_listado_filas.html", {"vehiculos": vehiculos})


@require_GET
def listado_fragmento(request):
    """Listado completo, pero como fragmento para abrir DENTRO del mismo
    popup del mapa (VehiculosModal), no como página aparte. Se puede
    llegar con '?estado=<nombre>' para llegar ya filtrado a ese estado
    (por ejemplo, viniendo del botón 'Ver listado completo' del resumen)."""
    estado_inicial = request.GET.get("estado", "").strip()
    qs = _aplicar_filtros(_queryset_base(), request).order_by("estado__nombre", "marca", "modelo")
    vehiculos = [_vehiculo_a_dict(v) for v in qs]

    context = {
        "vehiculos": vehiculos,
        "estado_inicial": estado_inicial,
        "tipos_opciones": TipoVeh.objects.all().order_by("tipo_veh").values_list("tipo_veh", flat=True),
        "situacion_opciones": SituacionVeh.objects.all().order_by("situacion").values_list("situacion", flat=True),
        "asignacion_opciones": TipoAsignacionVeh.objects.all().order_by("tipo").values_list("tipo", flat=True),
        "estado_opciones": Estado.objects.all().order_by("nombre").values_list("nombre", flat=True),
    }
    return render(request, "vehiculos/_listado_modal.html", context)


# --- Ficha de detalle (solo lectura) -------------------------------------

def _historial_vehiculo(vehiculo_obj):
    return {
        "kilometraje": Kilometraje.objects.filter(vehiculo=vehiculo_obj).order_by("-fecha"),
        "combustible": CombustibleExt.objects.filter(vehiculo=vehiculo_obj).order_by("-fecha"),
        "siniestros": Siniestros.objects.filter(vehiculo=vehiculo_obj).order_by("-fecha"),
        "capufe": Capufe.objects.filter(vehiculo=vehiculo_obj).order_by("-fecha_inicio"),
        "prestamos": PrestadoDe.objects.filter(vehiculo=vehiculo_obj).select_related("estado", "inmueble").order_by("-fecha_prestamo"),
    }


def _obtener_vehiculo_o_404(placa):
    vehiculo_obj = _queryset_base().filter(placa__iexact=placa).first()
    if vehiculo_obj is None:
        raise Http404("No se encontró un vehículo con esa placa.")
    return vehiculo_obj


def detalle_vehiculo(request, placa):
    vehiculo_obj = _obtener_vehiculo_o_404(placa)
    context = {"vehiculo": _vehiculo_a_dict(vehiculo_obj), **_historial_vehiculo(vehiculo_obj)}
    return render(request, "vehiculos/detalle.html", context)


def detalle_fragmento(request, placa):
    vehiculo_obj = _obtener_vehiculo_o_404(placa)
    context = {"vehiculo": _vehiculo_a_dict(vehiculo_obj), **_historial_vehiculo(vehiculo_obj)}
    return render(request, "vehiculos/_detalle_contenido.html", context)


# --- Popovers (hold-menu / mapa) -----------------------------------------

@require_GET
def popover_vehiculos(request):
    qs = _aplicar_filtros(_queryset_base(), request)
    total = qs.count()
    vehiculos = [_vehiculo_a_dict(v) for v in qs[:8]]
    return render(request, "vehiculos/_popover_lista.html", {
        "vehiculos": vehiculos,
        "total": total,
        "query": request.META.get("QUERY_STRING", ""),
    })


@require_GET
def popover_kilometraje(request):
    lecturas_qs = Kilometraje.objects.select_related("vehiculo").order_by("-fecha")[:8]
    total = Kilometraje.objects.count()
    lecturas = [
        {"placa": k.vehiculo.placa, "fecha": k.fecha, "km": k.odometro}
        for k in lecturas_qs
    ]
    return render(request, "vehiculos/_popover_kilometraje.html", {"lecturas": lecturas, "total": total})


@require_GET
def resumen_estado_fragmento(request):
    """Fragmento para el botón 'Parque Vehicular' del mapa orgánico.
    Sin parámetros -> resumen nacional, todos los vehículos.
    Con 'estado=<nombre>' -> solo los de ese estado.
    Con 'inmueble=<nombre>' -> solo los de ese inmueble específico
    (tiene prioridad sobre 'estado' si ambos llegaran a mandarse).
    Se abre dentro del mismo popup (#infoModal / VehiculosModal) que ya
    usa el resto del mapa, así que no hace falta un modal aparte."""
    estado_nombre = request.GET.get("estado", "").strip()
    inmueble_nombre = request.GET.get("inmueble", "").strip()

    # "Total Nacional" (o variantes) no es el nombre real de ningún estado
    # en la base — es la etiqueta que usa el mapa para "sin filtro". Si
    # llega tal cual, se trata como si no hubiera parámetro de estado.
    if estado_nombre.upper() in ("TOTAL NACIONAL", "NACIONAL", "TOTAL_NACIONAL"):
        estado_nombre = ""

    qs = _queryset_base()
    if inmueble_nombre:
        qs = qs.filter(inmueble__nombre_inmueble__iexact=inmueble_nombre)
        titulo = inmueble_nombre
        filtro_extra = "inmueble=" + quote(inmueble_nombre)
    elif estado_nombre:
        qs = qs.filter(estado__nombre__iexact=estado_nombre)
        titulo = estado_nombre
        filtro_extra = "estado=" + quote(estado_nombre)
    else:
        titulo = "Total Nacional"
        filtro_extra = "estado="

    filas = list(qs)
    total = len(filas)
    activos = sum(1 for v in filas if _es_activo(v))
    mantenimiento = sum(1 for v in filas if _color_situacion(v.situacion.situacion if v.situacion_id else "") == "mantenimiento")
    posible_baja = total - activos - mantenimiento

    # Desglose por situación individual (igual que el dashboard general),
    # no solo un bulto de "otras situaciones" — así cada categoría real
    # (Mantenimiento, Posible baja, etc.) tiene su propio botón filtrable.
    otras_situaciones = {}
    for v in filas:
        if not _es_activo(v):
            nombre = v.situacion.situacion if v.situacion_id else "Sin especificar"
            otras_situaciones[nombre] = otras_situaciones.get(nombre, 0) + 1

    # --- Mismas tarjetas que el dashboard general, pero filtradas ---
    conteo_tipos = Counter(v.tipoVeh.tipo_veh for v in filas if v.tipoVeh_id)
    tipos_catalogo = TipoVeh.objects.all().order_by("tipo_veh")
    tipos_vehiculo = [
        {"nombre": t.tipo_veh, "cantidad": conteo_tipos.get(t.tipo_veh, 0)} for t in tipos_catalogo
    ]
    tipos_vehiculo = [t for t in tipos_vehiculo if t["cantidad"] > 0]

    modelos_distintos = len({v.modelo for v in filas if v.modelo})
    tarjetas_asignadas = sum(1 for v in filas if v.tarjeta_asig)

    combustible_qs = CombustibleExt.objects.filter(vehiculo__in=filas)
    monto_combustible = sum((c.monto for c in combustible_qs.only("monto")), start=0)

    ultima_lectura_por_vehiculo = {}
    km_qs = Kilometraje.objects.filter(vehiculo__in=filas).order_by("vehiculo_id", "-fecha")
    for k in km_qs.only("vehiculo_id", "odometro"):
        if k.vehiculo_id not in ultima_lectura_por_vehiculo:
            ultima_lectura_por_vehiculo[k.vehiculo_id] = k.odometro
    total_km = sum(ultima_lectura_por_vehiculo.values(), start=0)

    vehiculos = [_vehiculo_a_dict(v) for v in qs.order_by("marca", "modelo")[:15]]

    return render(request, "vehiculos/_resumen_estado.html", {
        "estado_nombre": titulo,
        "filtro_extra": filtro_extra,  # ej. "estado=Jalisco" o "inmueble=Oficina%20X", ya codificado
        "total": total,
        "activos": activos,
        "mantenimiento": mantenimiento,
        "posible_baja": posible_baja,
        "inactivos": total - activos,
        "otras_situaciones": otras_situaciones,
        "modelos_distintos": modelos_distintos,
        "tipos_vehiculo": tipos_vehiculo,
        "tarjetas_asignadas": tarjetas_asignadas,
        "monto_combustible": monto_combustible,
        "total_kilometros": total_km,
        "vehiculos": vehiculos,
        "hay_mas": total > 15,
    })


# --- Exportación a Excel ---------------------------------------------------

ENCABEZADOS_EXCEL = [
    "Placa", "Marca", "Modelo", "Año", "No. Motor", "Tipo de vehículo",
    "Tipo de asignación", "Situación", "Estado", "Inmueble", "Tarjeta",
]
CAMPOS_EXCEL = [
    "placa", "marca", "modelo", "anio", "no_motor", "tipo_vehiculo",
    "tipo_asignacion", "situacion", "estado", "inmueble_destino", "tarjeta",
]


@require_GET
def exportar_excel(request):
    qs = _aplicar_filtros(_queryset_base(), request).order_by("estado__nombre", "marca")
    filas = [_vehiculo_a_dict(v) for v in qs]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehículos"

    ws.append(ENCABEZADOS_EXCEL)
    relleno_encabezado = PatternFill("solid", fgColor="9A0A38")
    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = relleno_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"

    for fila in filas:
        ws.append([fila.get(campo, "") for campo in CAMPOS_EXCEL])

    for indice, encabezado in enumerate(ENCABEZADOS_EXCEL, start=1):
        letra = get_column_letter(indice)
        ws.column_dimensions[letra].width = max(len(encabezado), 12) + 2

    partes_nombre = ["vehiculos"]
    for etiqueta in ("tipo", "situacion", "asignacion", "estado"):
        valor = request.GET.get(etiqueta, "")
        if valor:
            partes_nombre.append(valor.replace(" ", "_"))
    nombre_archivo = "_".join(partes_nombre) + ".xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}"'
    wb.save(response)
    return response


# --- Ubicaciones para el mapa (íconos por inmueble) -----------------------

@require_GET
def mapa_ubicaciones_geojson(request):
    """GeoJSON con un punto por cada inmueble que tenga vehículos asignados
    y coordenadas capturadas, con el conteo de vehículos en cada uno — para
    pintar íconos en el mapa (MapLibre). Los vehículos sin inmueble
    asignado, o cuyo inmueble no tenga latitud/longitud, no aparecen aquí
    (a propósito: no hay dónde ubicarlos con precisión)."""
    qs = _queryset_base().filter(
        inmueble__isnull=False,
        inmueble__latitud__isnull=False,
        inmueble__longitud__isnull=False,
    )

    conteo_por_inmueble = {}
    for v in qs:
        inmueble = v.inmueble
        clave = inmueble.id
        if clave not in conteo_por_inmueble:
            conteo_por_inmueble[clave] = {
                "nombre": inmueble.nombre_inmueble,
                "lat": float(inmueble.latitud),
                "lng": float(inmueble.longitud),
                "total": 0,
                "placa_unica": v.placa,  # se usa solo si total termina en 1
            }
        else:
            conteo_por_inmueble[clave]["placa_unica"] = None  # ya hay más de uno
        conteo_por_inmueble[clave]["total"] += 1

    features = [
        {
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [datos["lng"], datos["lat"]]},
            "properties": {
                "nombre": datos["nombre"],
                "total": datos["total"],
                # Si el inmueble tiene un único vehículo, se manda su placa
                # para poder ir directo a la ficha sin pasar por la lista.
                "placa_unica": datos["placa_unica"] if datos["total"] == 1 else None,
            },
        }
        for datos in conteo_por_inmueble.values()
    ]

    return JsonResponse({"type": "FeatureCollection", "features": features})
